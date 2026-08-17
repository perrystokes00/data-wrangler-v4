"""
dataview/file_catalog/extract_dump.py
====================================
Run the extractors against documents and write, for each one, an Excel workbook
you can actually read: a sheet per data type. Nothing touches the database.

WHY A WORKBOOK PER DOCUMENT
---------------------------
Counts tell you a loader is broken. Only the rows tell you why — a tops row with
the formation in the depth column is a column-mapping fault, and no summary line
will ever say that. One workbook per file, one sheet per table, is the shape
that lets you look.

Each workbook holds:

    SUMMARY          classification, UWI, well name, and every section found
    X_<section>      what the EXTRACTOR produced  (X_strat, X_stations, …)
    RAW_<table>      what is IN THE DOCUMENT, before any extractor touched it

The RAW sheets are the control, and they're the point. When a document yields
no rows there are two very different causes — the document had nothing, or the
extractor missed what was there — and they need opposite fixes. Put the raw
grid beside the extracted grid and the answer is visible rather than inferred.

WHAT IT COVERS
--------------
PDF   : classify_pdf THEN extended_classify_pdf, using the same precedence as
        _load_rows_to_catalog, so the type reported here is the type capture
        will act on. Every section the extractor returns is kept — a casing
        record gives casing AND cement AND the CBL evaluation, not the single
        list _do_extract flattens them into.
LAS   : header + per-curve metadata (header only; curve samples are never read).
DOCX  : every table.       XLSX : every sheet.
RAW   : pdfplumber extract_tables / python-docx tables / openpyxl sheets.

USAGE (from the repo root)
--------------------------
    py -m dataview.file_catalog.extract_dump --in C:\\synth_tx_docs --out C:\\probe
    py -m dataview.file_catalog.extract_dump --in <one.pdf> --out C:\\probe --show
    py -m dataview.file_catalog.extract_dump --in C:\\synth_tx_docs --out C:\\probe --ext .pdf
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import traceback
from collections import defaultdict

# Scalars about the file, and extractor diagnostics — neither is data, and both
# were previously written out as though they were tables of it.
_SKIP_KEYS = {"ok", "error", "text", "text_all", "report_type", "confidence",
              "pages", "page_count", "had_annotations", "annotation_uwi",
              "col_map", "columns_found", "file_path", "file_name",
              "station_count"}


# --------------------------------------------------------------------------- #
# Normalising extractor output
# --------------------------------------------------------------------------- #
def _rows_from(obj):
    """Turn an extractor section into a list of dicts.

    Extractors are inconsistent by nature: a section may be a list of row dicts,
    a single dict of fields (a header), or a scalar. All three are legitimate;
    this makes each writable without pretending a scalar is a row.
    """
    if obj is None:
        return []
    if isinstance(obj, dict):
        return [obj] if obj else []
    if isinstance(obj, (list, tuple)):
        out = []
        for x in obj:
            if isinstance(x, dict):
                out.append(x)
            elif x is not None:
                out.append({"value": x})
        return out
    return [{"value": obj}]


def _sections(res):
    return {k: _rows_from(v) for k, v in (res or {}).items()
            if k not in _SKIP_KEYS}


# --------------------------------------------------------------------------- #
# Extractors
# --------------------------------------------------------------------------- #
def extract_pdf(path):
    from dataview.file_catalog.pdf_survey_catalog import (
        classify_pdf, extract_stations, extract_eowr, extract_rft_data,
        extract_well_test, extract_petrophysical, extract_casing_cement,
        extract_ddr, extract_scout_ticket,
        RT_DIRECTIONAL, RT_EOWR, RT_FORMATION, RT_RFT, RT_WELL_TEST,
        RT_PETRO, RT_CASING, RT_DDR, RT_SCOUT,
    )
    try:
        from dataview.file_catalog.pdf_survey_catalog import extract_core, RT_CORE
    except ImportError:
        extract_core, RT_CORE = None, "CORE_ANALYSIS"

    # TWO CLASSIFIERS. classify_pdf knows only 5 types and mis-routes the rest —
    # scout tickets land under its COMPLETION keyword, EOW reports under its
    # survey keywords, casing and well tests fall to UNKNOWN. This mirrors
    # _load_rows_to_catalog's precedence exactly so the type reported here is
    # the type capture will act on: a confident EOW (>=0.5) overrides even
    # DIRECTIONAL, other extended types override any non-DIRECTIONAL base
    # result, and a genuine directional survey still wins.
    cl = classify_pdf(path) or {}
    rt = cl.get("report_type", "UNKNOWN")
    base_rt = rt
    ex_rt = None
    try:
        from dataview.file_catalog.pdf_survey_catalog import extended_classify_pdf
        ex = extended_classify_pdf(path) or {}
        ex_rt = ex.get("report_type", "UNKNOWN")
        _EXT = {RT_SCOUT, RT_EOWR, RT_WELL_TEST, RT_DDR, RT_RFT, RT_CASING}
        if ex_rt == RT_EOWR and ex.get("confidence", 0) >= 0.5:
            rt = RT_EOWR
        elif ex_rt in _EXT and rt != RT_DIRECTIONAL:
            rt = ex_rt
        if rt == ex_rt:
            for k in ("well_name", "operator", "uwi"):
                if not cl.get(k) and ex.get(k):
                    cl[k] = ex.get(k)
    except Exception:
        pass

    meta = {"report_type": rt, "base_type": base_rt, "extended_type": ex_rt,
            "uwi": cl.get("uwi") or "", "well_name": cl.get("well_name") or ""}

    fn = None
    if rt == RT_DIRECTIONAL:
        fn = extract_stations
    elif rt in (RT_EOWR, RT_FORMATION):
        fn = extract_eowr
    elif rt == RT_RFT:
        fn = extract_rft_data
    elif rt == RT_WELL_TEST:
        fn = extract_well_test
    elif rt == RT_CASING:
        fn = extract_casing_cement
    elif rt == RT_DDR:
        fn = extract_ddr
    elif rt == RT_SCOUT:
        fn = extract_scout_ticket
    elif rt == RT_CORE and extract_core:
        fn = extract_core
    elif rt in (RT_PETRO, "PETROPHYSICAL"):
        try:
            from dataview.file_catalog.extract_petro import extract_petro
            r = extract_petro(path)
            if r.get("ok"):
                return meta, _sections(r)
        except Exception:
            pass
        fn = extract_petrophysical

    # No handler for this type: return no sections rather than raising, so
    # "classified, extracted nothing" stays visible in the summary.
    if fn is None:
        return meta, {}
    res = fn(path) or {}
    if not isinstance(res, dict):
        return meta, {"rows": _rows_from(res)}
    return meta, _sections(res)


def extract_las(path):
    from dataview.file_catalog.las_header_loader import parse_las
    res = parse_las(path) or {}
    meta = {"report_type": "LAS", "base_type": "LAS", "extended_type": None,
            "uwi": res.get("uwi") or "", "well_name": res.get("well_name") or ""}
    out = {}
    hdr = {k: v for k, v in res.items()
           if k not in _SKIP_KEYS and not isinstance(v, (list, tuple, dict))}
    if hdr:
        out["header"] = [hdr]
    for k, v in res.items():
        if isinstance(v, (list, tuple)) and k not in _SKIP_KEYS:
            out[k] = _rows_from(v)
    return meta, out


def _docx_tables(path):
    from docx import Document
    out = {}
    for i, tbl in enumerate(Document(path).tables, start=1):
        if not tbl.rows:
            continue
        hdr = [c.text.strip() or f"col{j}"
               for j, c in enumerate(tbl.rows[0].cells, start=1)]
        rows = [dict(zip(hdr, [c.text.strip() for c in r.cells]))
                for r in tbl.rows[1:]]
        if rows:
            out[f"table_{i:02d}"] = rows
    return out


def _xlsx_sheets(path):
    from openpyxl import load_workbook
    out = {}
    for ws in load_workbook(path, read_only=True, data_only=True).worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            hdr = next(it)
        except StopIteration:
            continue
        hdr = [str(h).strip() if h is not None else f"col{j}"
               for j, h in enumerate(hdr, start=1)]
        rows = [dict(zip(hdr, ["" if v is None else v for v in vals]))
                for vals in it]
        if rows:
            out[f"sheet_{ws.title}"] = rows
    return out


def extract_docx(path):
    # Generic on purpose: dv_office_loader parses and writes in one step, so
    # there's no read-only seam to borrow. For the problem this was built to
    # investigate — a multi-well tops study losing seven of its eight wells —
    # the plain table IS the evidence, because the per-row UWI column is right
    # there in it.
    return ({"report_type": "DOCX", "base_type": "DOCX", "extended_type": None,
             "uwi": "", "well_name": ""}, _docx_tables(path))


def extract_xlsx(path):
    return ({"report_type": "XLSX", "base_type": "XLSX", "extended_type": None,
             "uwi": "", "well_name": ""}, _xlsx_sheets(path))


_DISPATCH = {".pdf": extract_pdf, ".las": extract_las,
             ".docx": extract_docx, ".xlsx": extract_xlsx}


# --------------------------------------------------------------------------- #
# The document's own tables — the control
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# Borderless tables — vendor reports that align with whitespace, not rules
# --------------------------------------------------------------------------- #
# pdfplumber's default table detection follows RULED LINES. ReportLab exports
# (and most generated documents) draw those, so the line strategy works. Real
# vendor reports frequently don't: a Baker Hughes survey has lines=0 and its
# whole station table collapses into a single cell.
#
# So recover the grid from WORD POSITIONS instead. Data rows are found first —
# they're the ones that are mostly numeric — their x-centres define the
# columns, and header words above are then assigned to the nearest column.
# That handles multi-word headers ("Meas Depth", "Dog Leg Sev") and the very
# common second header line carrying units, which is joined onto the first.
_NUM_RE = __import__("re").compile(r"[+\-]?[\d,]*\.?\d+")


def _text_lines(page, ytol=2.5):
    """Words grouped into visual lines, left to right."""
    rows = {}
    for w in page.extract_words(use_text_flow=False, keep_blank_chars=False):
        rows.setdefault(round(w["top"] / ytol), []).append(w)
    return [sorted(v, key=lambda w: w["x0"]) for _k, v in sorted(rows.items())]


def _is_number(tok):
    return bool(_NUM_RE.fullmatch(str(tok).replace(",", "")))


def _centre(w):
    return (w["x0"] + w["x1"]) / 2.0


def text_tables(page, min_cols=3, min_rows=3, header_lookback=3):
    """[(header, rows)] for whitespace-aligned tables on one page."""
    lines = _text_lines(page)
    flags = [bool(l) and len(l) >= min_cols
             and sum(_is_number(w["text"]) for w in l) / len(l) > 0.6
             for l in lines]

    out, i = [], 0
    while i < len(lines):
        if not flags[i]:
            i += 1
            continue
        j = i
        while j + 1 < len(lines) and flags[j + 1] and \
                len(lines[j + 1]) == len(lines[i]):
            j += 1
        block = lines[i:j + 1]
        if len(block) < min_rows:
            i = j + 1
            continue

        # Column positions from the data itself — the median x-centre at each
        # position across the block, so one ragged row can't shift a column.
        ncol = len(block[0])
        import statistics
        centres = [statistics.median([_centre(r[c]) for r in block])
                   for c in range(ncol)]

        def assign(words):
            cells = [[] for _ in range(ncol)]
            for w in words:
                x = _centre(w)
                cells[min(range(ncol), key=lambda c: abs(centres[c] - x))]\
                    .append(w["text"])
            return [" ".join(c).strip() for c in cells]

        header = [""] * ncol
        for back in range(1, header_lookback + 1):
            k = i - back
            if k < 0 or flags[k] or not lines[k]:
                break
            part = assign(lines[k])
            # Stop climbing at the metadata block. A real header line spans
            # most columns; "CONTRACTOR: Baker Hughes  REFERENCE: KB" sits
            # above the table, fills only a few, and carries colons. Without
            # this the column names come out as "CONTRACTOR: Meas Depth (ft)".
            filled = sum(1 for c in part if c)
            if filled < max(2, int(ncol * 0.6)):
                break
            if any(":" in c for c in part):
                break
            # Nearer lines are more specific; a units line sits directly above
            # the data and the names line above that, so prepend as we climb.
            header = [f"{p} {h}".strip() if p else h
                      for p, h in zip(part, header)]
        if not any(header):
            header = [f"col{c + 1}" for c in range(ncol)]
        out.append((header, [[w["text"] for w in r] for r in block]))
        i = j + 1
    return out


def raw_tables(path):
    """Every table in the file, before any extractor. No classification."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".pdf":
            import pdfplumber
            out, n = {}, 0
            with pdfplumber.open(path) as pdf:
                for pno, page in enumerate(pdf.pages, start=1):
                    found = 0
                    for t in (page.extract_tables() or []):
                        if not t or len(t) < 2:
                            continue
                        # A "table" whose every row is one cell is the ruled
                        # strategy finding a text block, not a grid — that's
                        # the borderless case, so don't count it as found.
                        if max(len(r) for r in t) < 2:
                            continue
                        n += 1
                        found += 1
                        hdr = [(c or "").strip() or f"col{j}"
                               for j, c in enumerate(t[0], start=1)]
                        out[f"p{pno}_t{n:02d}"] = [
                            dict(zip(hdr, [(c or "").strip() for c in r]))
                            for r in t[1:]]
                    if found:
                        continue
                    # Nothing gridded on this page — fall back to word
                    # positions for whitespace-aligned tables.
                    for hdr, rows in text_tables(page):
                        n += 1
                        out[f"p{pno}_x{n:02d}"] = [
                            dict(zip(hdr, r)) for r in rows]
            return out
        if ext == ".docx":
            return _docx_tables(path)
        if ext in (".xlsx", ".xls"):
            return _xlsx_sheets(path)
    except Exception as e:
        return {"_error": [{"error": f"{type(e).__name__}: {e}"}]}
    return {}


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #
_BAD_SHEET = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(prefix, name, used):
    """Excel sheet names: <=31 chars, no []:*?/\\, unique within the book."""
    base = f"{prefix}{_BAD_SHEET.sub('_', str(name))}"[:31] or "sheet"
    nm, i = base, 2
    while nm.lower() in used:
        suffix = f"~{i}"
        nm = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(nm.lower())
    return nm


def _write_sheet(wb, title, rows):
    from openpyxl.styles import Font
    ws = wb.create_sheet(title)
    if not rows:
        ws["A1"] = "(no rows)"
        return ws
    keys = []
    for r in rows:
        for k in r:
            if k not in keys:
                keys.append(k)
    for j, k in enumerate(keys, start=1):
        c = ws.cell(row=1, column=j, value=str(k))
        c.font = Font(bold=True)
    for i, r in enumerate(rows, start=2):
        for j, k in enumerate(keys, start=1):
            v = r.get(k)
            ws.cell(row=i, column=j,
                    value="" if v is None else
                          (v if isinstance(v, (int, float)) else str(v)))
    from openpyxl.utils import get_column_letter
    for j, k in enumerate(keys, start=1):
        width = max([len(str(k))] +
                    [len(str(r.get(k, ""))) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(9, width + 2))
    ws.freeze_panes = "A2"
    return ws


def write_workbook(path, meta, sections, raw, out_path):
    """One workbook: SUMMARY, a sheet per extracted section, per raw table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "SUMMARY"
    rows = [("file", os.path.basename(path)),
            ("path", path),
            ("classified as", meta.get("report_type", "")),
            ("base classifier", meta.get("base_type", "")),
            ("extended classifier", meta.get("extended_type") or ""),
            ("uwi", meta.get("uwi", "")),
            ("well name", meta.get("well_name", "")),
            ("", ""),
            ("EXTRACTED SECTIONS", "rows")]
    rows += [(k, len(v)) for k, v in sorted(sections.items())] or [("(none)", 0)]
    rows += [("", ""), ("RAW TABLES IN FILE", "rows")]
    rows += [(k, len(v)) for k, v in sorted(raw.items())] or [("(none)", 0)]
    for i, (a, b) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=a).font = Font(
            bold=a.isupper() and bool(a))
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60

    used = {"summary"}
    for k, v in sorted(sections.items()):
        _write_sheet(wb, _sheet_name("X_", k, used), v)
    for k, v in sorted(raw.items()):
        _write_sheet(wb, _sheet_name("RAW_", k, used), v)
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #
def dump(paths, out_dir, log=print, workbooks=True, with_raw=True, csvs=True):
    os.makedirs(out_dir, exist_ok=True)
    wb_dir = os.path.join(out_dir, "per_document")
    if workbooks:
        os.makedirs(wb_dir, exist_ok=True)

    buckets = defaultdict(list)
    summary, errors = [], []

    for p in paths:
        ext = os.path.splitext(p)[1].lower()
        fn = _DISPATCH.get(ext)
        base = os.path.basename(p)
        if fn is None:
            summary.append({"file": base, "ext": ext, "report_type": "",
                            "uwi": "", "sections": "", "rows": 0,
                            "raw_tables": 0, "note": "no extractor"})
            continue
        try:
            meta, sections = fn(p)
        except Exception as e:
            errors.append({"file": base, "ext": ext,
                           "error": f"{type(e).__name__}: {e}",
                           "traceback": traceback.format_exc()[-1500:]})
            summary.append({"file": base, "ext": ext, "report_type": "",
                            "uwi": "", "sections": "", "rows": 0,
                            "raw_tables": 0,
                            "note": f"ERROR {type(e).__name__}"})
            log(f"  !! {base}: {type(e).__name__}: {e}")
            continue

        raw = raw_tables(p) if with_raw else {}
        total = sum(len(v) for v in sections.values())
        parts = [f"{k}:{len(v)}" for k, v in sorted(sections.items()) if v]

        if csvs:
            for sec, rows in sections.items():
                for r in rows:
                    buckets[sec].append({
                        "_file": base,
                        "_report_type": meta.get("report_type", ""),
                        "_uwi": meta.get("uwi", ""),
                        **{k: ("" if v is None else v) for k, v in r.items()}})

        if workbooks:
            try:
                write_workbook(p, meta, sections, raw,
                               os.path.join(wb_dir,
                                            os.path.splitext(base)[0] + ".xlsx"))
            except Exception as e:
                log(f"  !! workbook for {base}: {type(e).__name__}: {e}")

        note = ""
        if not total and raw:
            # The finding worth stating: the document HAS tables, the extractor
            # returned none of them.
            note = f"extracted nothing — but the file has {len(raw)} table(s)"
        elif not total:
            note = "extracted nothing (no tables in the file either)"
        summary.append({"file": base, "ext": ext,
                        "report_type": meta.get("report_type", ""),
                        "uwi": meta.get("uwi", ""),
                        "sections": " ".join(parts), "rows": total,
                        "raw_tables": len(raw), "note": note})
        log(f"  {base:46} {meta.get('report_type',''):20}"
            f"{total:>5} row(s)  raw:{len(raw):<3} {' '.join(parts)}")

    if csvs:
        for sec, rows in sorted(buckets.items()):
            keys = []
            for r in rows:
                for k in r:
                    if k not in keys:
                        keys.append(k)
            with open(os.path.join(out_dir, f"{sec}.csv"), "w", newline="",
                      encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
                w.writeheader()
                w.writerows(rows)

    with open(os.path.join(out_dir, "_summary.csv"), "w", newline="",
              encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["file", "ext", "report_type", "uwi",
                                          "sections", "rows", "raw_tables",
                                          "note"])
        w.writeheader()
        w.writerows(summary)
    if errors:
        with open(os.path.join(out_dir, "_errors.csv"), "w", newline="",
                  encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["file", "ext", "error",
                                              "traceback"])
            w.writeheader()
            w.writerows(errors)

    log("")
    log(f"-- {len(summary):,} file(s) -> {out_dir}")
    if workbooks:
        log(f"   workbooks: {wb_dir}")
    blind = [s for s in summary if s["note"].startswith("extracted nothing — but")]
    if blind:
        log(f"\n!! {len(blind)} file(s) where the DOCUMENT HAS TABLES and the "
            f"extractor returned none:")
        for s in blind:
            log(f"     {s['file']:44} {s['report_type']:20} "
                f"{s['raw_tables']} raw table(s)")
        log("   open the workbook's RAW_ sheets — the data is there and the "
            "extractor isn't seeing it")
    if errors:
        log(f"\n!! {len(errors)} file(s) raised — see _errors.csv")
    return {"files": len(summary), "errors": len(errors)}


def collect(target, exts=None):
    if os.path.isfile(target):
        return [target]
    out = []
    for root, _dirs, files in os.walk(target):
        for f in sorted(files):
            e = os.path.splitext(f)[1].lower()
            if e in _DISPATCH and (not exts or e in exts):
                out.append(os.path.join(root, f))
    return out


def _main() -> int:
    ap = argparse.ArgumentParser(
        description="Extract documents to a per-document Excel workbook — a "
                    "sheet per data type, plus the file's own raw tables. "
                    "Writes no database.")
    ap.add_argument("--in", dest="inp", required=True,
                    help="a document, or a directory to walk")
    ap.add_argument("--out", default="extract_dump")
    ap.add_argument("--ext", action="append",
                    help="limit to an extension (repeatable), e.g. --ext .pdf")
    ap.add_argument("--limit", type=int)
    ap.add_argument("--no-raw", action="store_true",
                    help="skip the RAW_ sheets (faster on big PDFs)")
    ap.add_argument("--no-csv", action="store_true",
                    help="skip the pooled per-data-type CSVs")
    a = ap.parse_args()

    exts = {("." + e.lstrip(".")).lower() for e in (a.ext or [])} or None
    paths = collect(a.inp, exts)
    if a.limit:
        paths = paths[:a.limit]
    if not paths:
        print(f"-- nothing to do under {a.inp}")
        print(f"   supported: {', '.join(sorted(_DISPATCH))}")
        return 1
    print(f"-- {len(paths)} file(s) -> {a.out}\n")
    res = dump(paths, a.out, workbooks=True, with_raw=not a.no_raw,
               csvs=not a.no_csv)
    return 0 if res["files"] else 1


if __name__ == "__main__":
    raise SystemExit(_main())
