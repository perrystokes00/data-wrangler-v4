"""
page_workbench.py
=================
File Catalog & Workbench -- three tabs:

  1. Scan & Extract -- fast scan, bulk insert, Phase 2 header extraction
  2. Browse & View  -- filter catalog, view/plot files, extract and load data
  3. Header Files   -- query FILE_WELL_HEADER / FILE_SEIS_HEADER, export CSV
"""
import os
import re
import uuid
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime

# ── Extension sets ─────────────────────────────────────────────────────────
# Canonical definitions live in extract_core (shared with the process-pool
# parser) and are imported back here, so there is one source of truth. Add a
# new format extension in extract_core, not here.
from dataview.file_catalog.extract_core import (
    PDF_EXTS, LAS_EXTS, DLIS_EXTS, LIS_EXTS, SEGY_EXTS, P190_EXTS, SHP_EXTS,
    OFFICE_EXTS, CSV_EXTS, IMAGE_EXTS, WITSML_EXTS, JSON_LOG_EXTS, LOG_EXTS,
    _extract_fields,
)

# TABULAR_EXTS — delimited and spreadsheet tables. NOT scanned by the File
# Catalog at all: they belong to the Bulk Tabular Loader, which loads them into
# dv_* with mapping and FK resolution. There is no extractor for them here, so
# inventorying one creates a row that can never drain — it sits in "pending"
# forever and makes a finished run look unfinished. Single source of truth is
# promotion_lineage so the pipeline's default_exts() agrees with this page.
from dataview.file_catalog.promotion_lineage import TABULAR_EXTS

# Canonicalising a pasted scan root is shared with the pipeline (which scopes a
# forced re-extract to that root) and must stay importable without streamlit.
from dataview.core.path_identity import canon_root as _pi_canon_root

# ALL_EXTS is the DEFAULT scan universe — what a BLANK Formats-to-scan box
# walks. Tabular types are subtracted rather than never-added, because
# OFFICE_EXTS legitimately carries .xlsx alongside .docx and only the
# spreadsheet half is the loader's job.
ALL_EXTS = (PDF_EXTS | LOG_EXTS | SEGY_EXTS | P190_EXTS |
            SHP_EXTS | OFFICE_EXTS | IMAGE_EXTS |
            WITSML_EXTS | JSON_LOG_EXTS) - TABULAR_EXTS

# KNOWN_EXTS is everything the run will RECOGNIZE if typed. Tabular types are
# NOT in it: they used to be opt-in via the Formats-to-scan box, which is
# exactly how 167 dead .csv rows got inventoried. Typing one now gets an
# explicit redirect to the Bulk Tabular Loader instead of silently arming a
# scan that cannot finish.
KNOWN_EXTS = ALL_EXTS

# The capture path lives in extract_core (streamlit-free) so pipeline_run
# can import it without dragging the UI into the CLI or the process pool.
# Imported back here under the same names every call site already uses.
from dataview.file_catalog.extract_core import (
    EXT_GROUP, ENRICH_CHUNK, SELF_PARSING_EXTS,
    _norm_uwi, _safe_num, _safe_coord, _safe_int, _trunc,
    _safe_sample_interval, _safe_trace_count, _safe_epsg,
    _valid_date, _score, _issues, _clamp_well,
    _SQL_GFC_UPDATE, _SQL_WELL_MERGE, _SQL_SEIS_MERGE,
    _gfc_params, _well_params, _seis_params, ensure_seis_columns,
    _write_enrichment_on, _write_enrichment_batch,
    _set_readiness_cataloged, _load_rows_to_catalog,
    _do_extract as _ec_do_extract,
)


def _do_extract(fpath: str, fext: str) -> tuple:
    """UI wrapper — the extractor with page-visible errors."""
    return _ec_do_extract(fpath, fext, log=st.error)




# =============================================================================
# Entry point
# =============================================================================

# MOVED to dataview.core.path_identity, aliased here so this page, selftest and
# every existing caller keep working. The pipeline needs the same
# canonicalisation to scope a forced re-extract to the scan root, and it cannot
# import this module — page_workbench pulls in streamlit, which the CLI and the
# detached pipeline child must never load. One definition, two importers.
_canon_root = _pi_canon_root


def run(engine=None, dialect: str = "mssql"):
    st.title("🗂️ File Catalog & Workbench")
    st.caption(
        "Scan & extract files · Browse & view · "
        "Extract data · Load to DB"
    )

    if engine is None:
        st.warning("No database connection.")
        return

    # Phase 2 enrichment — runs at top of every rerun while active
    if st.session_state.get("wb_enriching"):
        _enrich_chunk(engine, dialect)
    elif st.session_state.get("wb_enrich_done"):
        st.success(
            f"✅ Extraction complete — "
            f"{st.session_state.get('wb_enrich_total',0):,} files processed."
        )

    # Section selector — a keyed radio instead of st.tabs, because st.tabs
    # resets to the first tab on every rerun, which snaps you off the Pipeline
    # section the moment you click a button in it. The radio's selection is
    # persisted in session_state ("wb_active_tab"), so it survives reruns.
    _TAB_LABELS = [
        "🗂 File Catalog Pipeline",
        "▶ Run Pipeline",
        "📂 Browse and View",
        "🚦 Status & Backlog",
        "📦 Vault",
    ]
    # drop any stale section value persisted from a previous layout
    if st.session_state.get("wb_active_tab") not in _TAB_LABELS:
        st.session_state.pop("wb_active_tab", None)
    active = st.radio(
        "Workbench section",
        _TAB_LABELS,
        key="wb_active_tab",
        horizontal=True,
        label_visibility="collapsed",
    )
    st.divider()

    if active == _TAB_LABELS[0]:
        _tab_pipeline(engine, dialect)
    elif active == _TAB_LABELS[1]:
        # Headless run page — launches pipeline_run.py in a fresh process and
        # tails its log live (replaces the old Fast Track / monitor embed).
        try:
            from dataview.file_catalog import page_run
            page_run.render(engine)
        except Exception as e:
            st.error(f"Run Pipeline unavailable: {e}")
            import traceback
            with st.expander("details"):
                st.code(traceback.format_exc())
    elif active == _TAB_LABELS[2]:
        _tab_browse(engine, dialect)
    elif active == _TAB_LABELS[3]:
        _tab_status(engine, dialect)
    elif active == _TAB_LABELS[4]:
        try:
            from dataview.file_catalog import page_vault
            page_vault.render(engine)
        except Exception as e:
            st.error(f"Vault unavailable: {e}")
            import traceback
            with st.expander("details"):
                st.code(traceback.format_exc())
    else:
        # EVERY SECTION IS MATCHED BY INDEX, none by falling through. Vault used
        # to be _TAB_LABELS[3] with Browse & View as the `else`; inserting a
        # label ahead of Vault silently repointed [3] at the new section and the
        # Vault button would have opened it instead. An explicit branch per
        # label turns that into a visible edit rather than a silent one.
        _tab_browse(engine, dialect)


# =============================================================================
# Tab 1 -- Scan & Extract
# =============================================================================

def _tab_scan(engine, dialect):
    from sqlalchemy import text as _t

    st.markdown("#### 🔍 Scan & Extract")

    # ── Config ────────────────────────────────────────────────────────────────
    scan_path = st.text_input(
        "Root folder",
        value=st.session_state.get(
            "wb_last_scan_path",
            r"C:\Users\perry\OneDrive\Documents\PPDM\claude_use_ai"
            r"\data_wrangler\training\test_crawl"),
        placeholder=r"\\server\share\WellData  or  C:\WellData",
        key="wb_scan_path",
    )
    ext_groups = st.multiselect(
        "File types to scan",
        options=[
            "PDF",
            "Well Log",
            "Seismic",
            "Shapefile",
            "Office",
            "WITSML",
            "OSDU / JSON Well Log",
            "Image",
        ],
        default=[
            "PDF",
            "Well Log",
            "Seismic",
            "Shapefile",
            "Office",
            "WITSML",
            "OSDU / JSON Well Log",
        ],
        key="wb_scan_exts",
        help=(
            "**PDF** — .pdf\n\n"
            "**Well Log** — .las  ·  .dlis  ·  .dlf  ·  .dis  ·  .lis\n\n"
            "**Seismic** — .segy  ·  .sgy  ·  .seg  ·  .p190  ·  .p90  ·  .p1\n\n"
            "**Shapefile** — .shp  ·  .gpkg  ·  .kml  ·  .kmz\n\n"
            "**Office** — .xlsx  ·  .xls  ·  .xlsm  ·  .docx  ·  .doc  ·  .csv  ·  .tsv\n\n"
            "**WITSML** — .xml\n"
            "*(WITSML 1.3.1 / 1.4.1 — trajectory, log, mudLog, well, wellbore)*\n\n"
            "**OSDU / JSON Well Log** — .json\n"
            "*(16 OSDU schemas: Well, Wellbore, WellLog, WellboreTrajectory, "
            "WellboreMarkerSet, WellborePressureData, WellboreCompletion, "
            "WellCoreAnalysis, ProductionVolume, RockFluidOrganisation/SCAL, "
            "Field, Reservoir, SeismicAcquisitionSurvey, SeismicHorizon, "
            "SeismicFault, Document — plus JSON Well Log Format/JSONWLF)*\n\n"
            "**Image** — .tif  ·  .tiff  ·  .png  ·  .jpg  ·  .jpeg\n"
            "*(Phase 1 scan only — no extractor. Useful for inventorying "
            "core photos, well plat images, seismic sections etc.)*"
        ),
    )
    _exts = {e for e, g in EXT_GROUP.items() if g in ext_groups}

    c1, c2, c3 = st.columns(3)

    # Phase 1
    if c1.button("🔍 Scan (Phase 1)", type="primary",
                 key="wb_p1", use_container_width=True):
        if not scan_path:
            st.error("Enter a folder path.")
        elif not Path(scan_path).exists():
            st.error(f"Not found: `{scan_path}`")
        else:
            st.session_state["wb_last_scan_path"] = scan_path
            _run_scan(engine, dialect, scan_path, _exts)

    # Phase 2
    if c2.button("⚙️ Extract (Phase 2)", type="secondary",
                 key="wb_p2", use_container_width=True):
        st.session_state["wb_enriching"]     = True
        st.session_state["wb_enrich_offset"] = 0
        st.rerun()

    if c3.button("⏹ Stop", key="wb_stop", use_container_width=True):
        st.session_state["wb_enriching"] = False

    # Phase 2 thread count — tune parallelism. Default 8 works for mixed
    # extraction (PDF/DLIS/Office). Drop to 2-4 for DLIS-heavy batches
    # that load big chunks into memory; raise to 16 for many small files.
    _w = st.slider(
        "Phase 2 threads",
        min_value=1, max_value=16,
        value=int(st.session_state.get("wb_phase2_workers", 8)),
        key="wb_phase2_workers_slider",
        help="Files per chunk extracted in parallel. Lower for "
             "DLIS-heavy batches; higher for many small files. Default 8.",
    )
    st.session_state["wb_phase2_workers"] = _w

    # ── Catalog summary ───────────────────────────────────────────────────────
    st.divider()
    try:
        with engine.connect() as con:
            rows = con.execute(_t("""
                SELECT
                    FILE_TYPE_GROUP,
                    COUNT(*)                                               total,
                    SUM(CASE WHEN HEADER_EXTRACTED='Y' THEN 1 ELSE 0 END) extracted,
                    SUM(CASE WHEN CATALOG_READINESS='READY'     THEN 1 ELSE 0 END) ready,
                    SUM(CASE WHEN CATALOG_READINESS='NEEDS_UWI' THEN 1 ELSE 0 END) needs_uwi,
                    SUM(CASE WHEN CATALOG_READINESS='ATTENTION' THEN 1 ELSE 0 END) attention,
                    SUM(CASE WHEN CATALOG_READINESS='CATALOGED' THEN 1 ELSE 0 END) cataloged,
                    SUM(CASE WHEN ISNULL(FLAG_DELETE,'N')='Y'   THEN 1 ELSE 0 END) flagged,
                    SUM(CASE WHEN HEADER_EXTRACTED='S'  THEN 1 ELSE 0 END) skipped
                FROM file_catalog.GLOBAL_FILE_CATALOG
                GROUP BY FILE_TYPE_GROUP
                ORDER BY total DESC
            """)).fetchall()

        if rows:
            df = pd.DataFrame(rows, columns=[
                "Type","Total","Extracted","Ready","Needs UWI","Attention",
                "Cataloged","Flagged","Skipped"])
            tot  = df["Total"].sum()
            enr  = df["Extracted"].sum()
            skip = int(df["Skipped"].sum())
            m1,m2,m3,m4,m5,m6 = st.columns(6)
            m1.metric("Total cataloged",    f"{tot:,}")
            m2.metric("Extracted",          f"{enr:,}")
            m3.metric("Pending extraction", f"{tot-enr-skip:,}")
            m4.metric("Skipped",            skip,
                      help="Files skipped — too large for extraction. "
                           "HEADER_EXTRACTED='S' in catalog.")
            m5.metric("Cataloged",          int(df["Cataloged"].sum()),
                      help="Files whose rows are captured in the cat_* mirrors "
                           "(CATALOG_READINESS='CATALOGED').")
            m6.metric("Flagged",            int(df["Flagged"].sum()))
            st.dataframe(df, hide_index=True, use_container_width=True)
        else:
            st.info("Catalog is empty — run Phase 1 scan.")
    except Exception as e:
        st.caption(f"Catalog summary: {e}")

    # ── Delete tools ──────────────────────────────────────────────────────────
    st.divider()

    # Extractor reference — behind an expander so it doesn't dominate the page
    with st.expander("📋 Extractor reference — what Phase 2 captures", expanded=False):
        st.caption(
            "Phase 2 reads each file's internal header and writes identifying "
            "metadata to FILE_WELL_HEADER or FILE_SEIS_HEADER. Files on disk "
            "are never modified. File types with no extractor are cataloged in "
            "Phase 1 but skipped in Phase 2 — deleting them before running "
            "extraction keeps the batch fast.\n\n"
            "**Well logs —** "
            "LAS (.las): UWI/API, well name, operator, field, state, county, "
            "lat/lon, total depth, spud date, contractor, curve mnemonics, depth range. "
            "DLIS (.dlis .dlf .dis): well name, field, operator from origins block, "
            "channel count, frame count. "
            "LIS (.lis): well name, UWI, operator, field, state, county, contractor, "
            "depth range, curve mnemonics via dlisio; raw byte scan fallback for "
            "non-standard files.\n\n"
            "**Seismic —** "
            "SEG-Y (.segy .sgy .seg): trace count, sample interval, 2D/3D "
            "classification (updates FILE_TYPE_GROUP), survey name and contractor "
            "from text header, bounding box and survey outline polygon from 400 "
            "strided trace headers with CRS detection and WGS84 reprojection, "
            "inline/crossline range for 3D surveys. "
            "P190 (.p190 .p90 .p1): survey name, contractor, shot count, "
            "bounding box from S-records.\n\n"
            "**Documents —** "
            "PDF (.pdf): report type classification (directional survey, mud log, "
            "scout ticket, completion, DST, core, well proposal), UWI, well name, "
            "operator, field, state, county, lat/lon, total depth, spud date, "
            "rig release, survey type, contractor, confidence score.\n\n"
            "**Spatial —** "
            "Shapefile / GeoPackage / KML (.shp .gpkg .kml .kmz): "
            "feature type classification (well, seismic 2D/3D, field, lease, "
            "pipeline, facility, boundary), CRS, bounding box, DBF column mapping, "
            "sample UWIs/well names/operators/field names/status codes, date ranges. "
            "Well shapefiles promote UWI and operator into the catalog record.\n\n"
            "**Office —** "
            "Excel (.xlsx .xls .xlsm): sheet classification (BOEM borehole, KGS well, "
            "production, completion, formation tops, well header, core, pressure, "
            "survey, reserves), row count, column headers, UWI from data. "
            "Known schemas (BOEM_BOREHOLE, KGS_WELL, RRC_WELL) detected before "
            "generic classification. Read via openpyxl streaming — no hang on "
            "large files. "
            "Word (.docx .doc): document type classification (completion report, "
            "geological, DST, well proposal, regulatory, formation tops, HSE), "
            "headings, table classification, UWI and well name from text. "
            "CSV/TSV: column classification, row count, UWI from data.\n\n"
            "**WITSML (.xml) —** "
            "Trajectory: well name, UWI, survey tool type, station count, depth "
            "range, contractor from commonData. "
            "Log: curve mnemonics from logCurveInfo, depth range, service company, "
            "run number. "
            "MudLog: formation interval count, gas show summary from chromatograph "
            "elements, comments. "
            "File must contain the witsml.org/schemas namespace — other XML files "
            "(config, SVG, RSS) are skipped automatically.\n\n"
            "**JSON Well Log / OSDU (.json) —** "
            "16 OSDU schemas detected by the 'kind' field: "
            "Well (name, UWI, operator, lat/lon, field, spud, TD), "
            "WellLog (curves, depth range, contractor), "
            "WellboreTrajectory (KOP, landing, lateral length, max inc, max DLS), "
            "WellboreMarkerSet (full formation tops list with MD/TVD/subsea/quality), "
            "WellborePressureData (DST pressures, flow rates, permeability, skin), "
            "WellboreCompletion (stages, clusters, fluid/proppant volumes, formations), "
            "WellCoreAnalysis (plug count, porosity/perm stats, full plug list), "
            "ProductionVolume (monthly records, cumulative volumes, peak rate), "
            "RockFluidOrganisation / SCAL (system types, end-point saturations, "
            "capillary pressure method), "
            "Field (discovery year/well, basin, fluid type, bbox, cumulative production), "
            "Reservoir (porosity, perm, net pay, pressure, temperature, GOR, OOIP), "
            "SeismicAcquisitionSurvey (2D/3D, bbox, inline/crossline, fold, bin size), "
            "SeismicHorizon (geologic unit, depth stats, node count, well control), "
            "SeismicFault (fault type, strike/dip, max throw, length, horizons cut), "
            "Document (document type, author, file format, page count). "
            "Also handles JSON Well Log Format (JSONWLF) from NORCE/NPD. "
            "Non-petroleum JSON (config files, package.json) skipped automatically "
            "by a 512-byte header check."
        )

    # Delete section — heading first, then Select All, then grid
    st.markdown("**Delete from catalog**")
    st.caption(
        "Removes selected file types from the catalog index only — "
        "files on disk are never touched."
    )

    # Human-readable label for every extension the catalog might contain.
    # Covers all known sets plus a fallback for anything unexpected.
    _EXT_LABEL = {
        # Well logs
        ".las":   "LAS — Log ASCII Standard well log",
        ".dlis":  "DLIS — Digital Log Interchange Standard",
        ".dlf":   "DLF — DLIS variant",
        ".dis":   "DIS — DLIS variant",
        ".lis":   "LIS — Log Information Standard",
        # Seismic
        ".segy":  "SEG-Y — Seismic data (classified as 2D or 3D after extraction)",
        ".sgy":   "SGY — SEG-Y seismic data (classified as 2D or 3D after extraction)",
        ".seg":   "SEG — SEG-Y seismic data (legacy extension)",
        ".p190":  "P190 — Navigation / shot point data",
        ".p90":   "P90 — P190 variant",
        ".p1":    "P1 — P190 variant",
        # Documents
        ".pdf":   "PDF — Portable Document (scout tickets, reports, surveys)",
        # Office
        ".xlsx":  "XLSX — Excel workbook",
        ".xls":   "XLS — Excel workbook (legacy)",
        ".xlsm":  "XLSM — Excel macro-enabled workbook",
        ".docx":  "DOCX — Word document",
        ".doc":   "DOC — Word document (legacy)",
        ".csv":   "CSV — Comma-separated values",
        ".tsv":   "TSV — Tab-separated values",
        # Shapefiles / spatial
        ".shp":     "SHP — Shapefile geometry",
        ".geojson": "GeoJSON — Geographic JSON",
        ".gpkg":    "GPKG — GeoPackage",
        ".kml":     "KML — Keyhole Markup Language",
        ".kmz":     "KMZ — Compressed KML",
        # WITSML
        ".xml":     "XML / WITSML — trajectory, log, mud log, well header",
        # JSON Well Log / OSDU
        ".json":    "JSON — OSDU (16 schemas: Well, Wellbore, WellLog, WellboreTrajectory, "
                   "WellboreMarkerSet, WellborePressureData, WellboreCompletion, "
                   "WellCoreAnalysis, ProductionVolume, RockFluidOrganisation/SCAL, "
                   "Field, Reservoir, SeismicAcquisitionSurvey, SeismicHorizon, "
                   "SeismicFault, Document) + JSON Well Log Format (JSONWLF)",
        ".tiff":  "TIFF — TIFF image",
        ".png":   "PNG — PNG image",
        ".jpg":   "JPG — JPEG image",
        ".jpeg":  "JPEG — JPEG image",
    }

    # Load distinct extensions currently in the catalog
    try:
        with engine.connect() as con:
            _ext_rows = con.execute(_t("""
                SELECT FILE_EXT, COUNT(*) AS n
                FROM file_catalog.GLOBAL_FILE_CATALOG
                WHERE FILE_EXT IS NOT NULL AND FILE_EXT <> ''
                GROUP BY FILE_EXT
                ORDER BY FILE_EXT
            """)).fetchall()
        _ext_counts = {r[0]: r[1] for r in _ext_rows}
    except Exception as _e:
        _ext_counts = {}
        st.caption(f"Extension list unavailable: {_e}")

    if _ext_counts:
        _total_all_files = sum(_ext_counts.values())

        # Select ALL checkbox — sits above the grid, turns on every Delete
        # checkbox without being a separate toggle-style control. Matches
        # the spreadsheet metaphor: header checkbox selects all rows.
        _sel_all = st.checkbox(
            f"☑ Select ALL  ({len(_ext_counts)} extension types · "
            f"{_total_all_files:,} files)",
            key="wb_sel_all",
            help="Check to mark every extension for deletion. "
                 "Uncheck individual rows in the grid to exclude them.",
        )

        st.markdown("---")

        # Build the dataframe for the editor
        import pandas as _pd
        _rows = []
        for _ext in sorted(_ext_counts.keys()):
            _rows.append({
                "Delete": _sel_all,   # pre-check all rows when Select ALL is on
                "Extension": _ext,
                "Description": _EXT_LABEL.get(
                    _ext.lower(),
                    f"{_ext.lstrip('.').upper()} file"),
                "Files": _ext_counts[_ext],
            })
        _df_exts = _pd.DataFrame(_rows)

        _edited = st.data_editor(
            _df_exts,
            use_container_width=True,
            hide_index=True,
            disabled=["Extension", "Description", "Files"],
            column_config={
                "Delete": st.column_config.CheckboxColumn(
                    "Delete", width="small"),
                "Extension": st.column_config.TextColumn(
                    "Extension", width="small"),
                "Description": st.column_config.TextColumn(
                    "Description"),
                "Files": st.column_config.NumberColumn(
                    "Files", width="small", format="%d"),
            },
            key="wb_ext_editor",
        )

        _checked_exts = _edited.loc[
            _edited["Delete"] == True, "Extension"].tolist()

        st.markdown("---")

        if _checked_exts:
            _total_sel = sum(_ext_counts.get(e, 0) for e in _checked_exts)
            st.caption(
                f"{len(_checked_exts)} extension(s) selected · "
                f"{_total_sel:,} files will be removed from the catalog index "
                f"(files on disk are NOT deleted)"
            )
            if st.button(
                f"🗑️ Delete {_total_sel:,} files from catalog",
                key="wb_del_ext_go",
                type="primary",
            ):
                st.session_state["wb_del_ext_confirm"] = True

            if st.session_state.get("wb_del_ext_confirm"):
                st.warning(
                    f"Remove **{_total_sel:,} files** with "
                    f"{len(_checked_exts)} extension(s) from catalog? "
                    "This cannot be undone."
                )
                _cc1, _cc2 = st.columns(2)
                if _cc1.button("✅ Yes, delete", key="wb_del_ext_yes",
                               type="primary"):
                    try:
                        _deleted = 0
                        with engine.begin() as con:
                            for _ext in _checked_exts:
                                _deleted += con.execute(_t("""
                                    DELETE FROM file_catalog.GLOBAL_FILE_CATALOG
                                    WHERE FILE_EXT = :e
                                """), {"e": _ext}).rowcount
                        st.success(f"Deleted {_deleted:,} files from catalog.")
                        st.session_state.pop("wb_del_ext_confirm", None)
                        st.rerun()
                    except Exception as _de:
                        st.error(f"Delete failed: {_de}")
                if _cc2.button("✗ Cancel", key="wb_del_ext_cancel"):
                    st.session_state.pop("wb_del_ext_confirm", None)
                    st.rerun()
    else:
        st.caption("Catalog is empty — nothing to delete.")

    # ── Delete all flagged ────────────────────────────────────────────────────
    st.divider()
    try:
        with engine.connect() as con:
            nf = con.execute(_t("""
                SELECT COUNT(*) FROM file_catalog.GLOBAL_FILE_CATALOG
                WHERE FLAG_DELETE='Y'
            """)).scalar() or 0
        if nf > 0:
            if st.button(f"🗑️ Delete all {nf:,} flagged files",
                         type="primary", key="wb_del_flagged"):
                st.session_state["wb_del_flagged_confirm"] = True

            if st.session_state.get("wb_del_flagged_confirm"):
                st.warning(
                    f"Remove **{nf:,} flagged files** from catalog? "
                    "Files on disk are NOT deleted."
                )
                cc1, cc2 = st.columns(2)
                if cc1.button("✅ Yes, delete", key="wb_del_flag_yes",
                              type="primary"):
                    with engine.begin() as con:
                        n = con.execute(_t("""
                            DELETE FROM file_catalog.GLOBAL_FILE_CATALOG
                            WHERE FLAG_DELETE='Y'
                        """)).rowcount
                    st.success(f"Deleted {n:,} flagged files from catalog.")
                    st.session_state.pop("wb_del_flagged_confirm", None)
                    st.rerun()
                if cc2.button("❌ Cancel", key="wb_del_flag_no"):
                    st.session_state.pop("wb_del_flagged_confirm", None)
                    st.rerun()
        else:
            st.caption("No files flagged for deletion.")
    except Exception as e:
        st.caption(f"Flag check: {e}")


def _run_scan(engine, dialect, root: str, exts: set):
    """Phase 1: fast os.scandir walk + BULK INSERT to GLOBAL_FILE_CATALOG."""
    import csv, tempfile, hashlib
    from datetime import datetime, timezone
    from sqlalchemy import text as _t
    from dataview.core.fingerprint import file_fingerprint, DEDUPE_SQL   # single source of truth

    prog = st.progress(0.0, text="Walking...")
    found = []
    folders = 0
    stack = [root]

    while stack:
        dirpath = stack.pop()
        folders += 1
        try:
            with os.scandir(dirpath) as it:
                for entry in it:
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            stack.append(entry.path)
                        else:
                            ext = os.path.splitext(entry.name)[1].lower()
                            if ext in exts:
                                # JSON peek — only catalog .json files that
                                # look like OSDU or JSONWLF petroleum data.
                                # Reads the first 100 bytes only (single disk
                                # read) so Phase 1 speed is not affected.
                                # Skips package.json, settings.json, tsconfig,
                                # Streamlit config, and any other non-petroleum
                                # JSON files before they enter the catalog.
                                if ext == ".json":
                                    try:
                                        with open(entry.path, "rb") as _jf:
                                            _peek = _jf.read(100)
                                        # OSDU files always have "kind" near
                                        # the top. JSONWLF files have "header".
                                        # Any other JSON is not petroleum data.
                                        if (b'"kind"'   not in _peek and
                                                b'"header"' not in _peek):
                                            continue
                                    except OSError:
                                        continue
                                st_res = entry.stat()
                                found.append((
                                    entry.path, entry.name, ext,
                                    round(st_res.st_size/1024, 2),
                                    datetime.fromtimestamp(
                                        st_res.st_mtime,
                                        tz=timezone.utc
                                    ).strftime("%Y-%m-%d %H:%M:%S"),
                                    EXT_GROUP.get(ext, "Other"),
                                    root,
                                    file_fingerprint(entry.path,
                                                     st_res.st_size,
                                                     st_res.st_mtime),
                                ))
                    except OSError:
                        pass
        except (PermissionError, OSError):
            pass
        if folders % 2000 == 0:
            prog.progress(0.3, text=f"{folders:,} folders · {len(found):,} files")

    if not found:
        st.warning("No files found.")
        return

    # Blocklist: files fingerprinted as "bad" (junk / unparseable) are skipped
    # so they never re-enter the catalog. Keyed on INVENTORY_ID = SHA1(path) —
    # the same fingerprint computed below — so the check is a cheap set lookup.
    bad_set = set()
    try:
        with engine.connect() as con:
            bad_set = {r[0] for r in con.execute(_t(
                "SELECT INVENTORY_ID FROM file_catalog.BAD_FILE")).fetchall()}
    except Exception:
        bad_set = set()   # table not created yet — nothing to skip

    prog.progress(0.5, text=f"{len(found):,} files — writing CSV...")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False,
        newline="", encoding="utf-8"
    )
    csv_path = tmp.name
    # NO escapechar — it doubled every separator in a Windows path and BULK
    # INSERT stored the doubled form. The id below is hashed from the CLEAN
    # fpath, so the escaped write left INVENTORY_ID and FILE_PATH describing
    # different strings. See path_identity.bulk_csv_writer.
    from dataview.core.path_identity import bulk_csv_writer, bulk_field
    writer = bulk_csv_writer(tmp)
    n_bad = 0   # files skipped because they're on the bad-file blocklist
    n_sanitised = 0
    for (fpath, fname, fext, size_kb, mod_dt, grp, rpath, fhash) in found:
        inv_id = hashlib.sha1(
            fpath.upper().encode("utf-8")).hexdigest().upper()
        if inv_id in bad_set:
            n_bad += 1
            continue
        row = []
        for v in (inv_id, fpath[:900], fname[:260], fext[:20],
                  grp[:50], size_kb if size_kb else "",
                  fhash, "", "UNCATALOGED", "",
                  rpath[:900], now, now, now):
            val, changed = bulk_field(v)
            row.append(val)
            n_sanitised += bool(changed)
        writer.writerow(row)
    if n_sanitised:
        st.warning(f"{n_sanitised} field(s) contained a tab, quote or newline "
                   f"and were rewritten to load — the stored value differs "
                   f"from the value on disk.")
    tmp.close()

    prog.progress(0.7, text="Bulk inserting...")
    try:
        with engine.begin() as con:
            con.execute(_t("""
                IF OBJECT_ID('file_catalog.fc_stage','U') IS NOT NULL
                    DROP TABLE file_catalog.fc_stage;
                CREATE TABLE file_catalog.fc_stage (
                    INVENTORY_ID     NVARCHAR(40),
                    FILE_PATH        NVARCHAR(900),
                    FILE_NAME        NVARCHAR(260),
                    FILE_EXT         NVARCHAR(20),
                    FILE_TYPE_GROUP  NVARCHAR(50),
                    FILE_SIZE_KB     NVARCHAR(30),
                    FILE_HASH        NVARCHAR(40),
                    DUPLICATE_GROUP  NVARCHAR(64),
                    CATALOG_STATUS   NVARCHAR(20),
                    CATALOG_TABLE    NVARCHAR(100),
                    ROOT_PATH        NVARCHAR(900),
                    SCAN_DATE        NVARCHAR(30),
                    ROW_CREATED_DATE NVARCHAR(30),
                    ROW_CHANGED_DATE NVARCHAR(30)
                );
            """))
            con.execute(_t(f"""
                BULK INSERT file_catalog.fc_stage
                FROM '{csv_path}'
                WITH (FIELDTERMINATOR='\\t', ROWTERMINATOR='0x0D0A',
                      CODEPAGE='65001', FIRSTROW=1, TABLOCK);
            """))
            con.execute(_t("""
                MERGE file_catalog.GLOBAL_FILE_CATALOG AS tgt
                USING file_catalog.fc_stage AS src
                ON tgt.INVENTORY_ID = src.INVENTORY_ID
                WHEN MATCHED THEN UPDATE SET
                    FILE_SIZE_KB     = TRY_CAST(src.FILE_SIZE_KB AS DECIMAL(15,2)),
                    FILE_HASH        = src.FILE_HASH,
                    SCAN_DATE        = TRY_CAST(src.SCAN_DATE AS DATETIME2),
                    ROW_CHANGED_DATE = TRY_CAST(src.ROW_CHANGED_DATE AS DATETIME2)
                WHEN NOT MATCHED THEN INSERT (
                    INVENTORY_ID,FILE_PATH,FILE_NAME,FILE_EXT,
                    FILE_TYPE_GROUP,FILE_SIZE_KB,FILE_HASH,
                    DUPLICATE_GROUP,CATALOG_STATUS,CATALOG_TABLE,
                    ROOT_PATH,SCAN_DATE,ROW_CREATED_DATE,ROW_CHANGED_DATE
                ) VALUES (
                    src.INVENTORY_ID,src.FILE_PATH,src.FILE_NAME,src.FILE_EXT,
                    src.FILE_TYPE_GROUP,
                    TRY_CAST(src.FILE_SIZE_KB AS DECIMAL(15,2)),
                    src.FILE_HASH,src.DUPLICATE_GROUP,src.CATALOG_STATUS,
                    src.CATALOG_TABLE,src.ROOT_PATH,
                    TRY_CAST(src.SCAN_DATE AS DATETIME2),
                    TRY_CAST(src.ROW_CREATED_DATE AS DATETIME2),
                    TRY_CAST(src.ROW_CHANGED_DATE AS DATETIME2)
                );
            """))
            # Content-dedupe via the shared rule (identical to the CLI + File
            # Manager scans): one canonical per FILE_HASH stays processable;
            # redundant copies get DUPLICATE_GROUP set and are skipped by
            # extract + capture. Idempotent — full recompute each scan.
            con.execute(_t(DEDUPE_SQL))
            n_dup = con.execute(_t(
                "SELECT COUNT(*) FROM file_catalog.GLOBAL_FILE_CATALOG "
                "WHERE DUPLICATE_GROUP IS NOT NULL")).scalar() or 0
            con.execute(_t(
                "DROP TABLE IF EXISTS file_catalog.fc_stage;"))

        prog.progress(1.0, text="Done.")
        st.success(
            f"✅ Phase 1 complete — {len(found) - n_bad:,} files "
            f"across {folders:,} folders"
            + (f" · {n_bad:,} bad file(s) skipped" if n_bad else "")
            + (f" · {n_dup:,} dup(s) skipped" if n_dup else "")
            + ". Click **Extract** to extract headers."
        )
    except Exception as e:
        st.error(f"Bulk insert failed: {e}")
    finally:
        try:
            os.unlink(csv_path)
        except Exception:
            pass


def _enrich_chunk(engine, dialect):
    """
    Phase 2: process ENRICH_CHUNK files per rerun.
    Extracts headers → FILE_WELL_HEADER / FILE_SEIS_HEADER.
    Shows a persistent progress bar at the top of the page.

    Extraction within each chunk runs in parallel across PHASE2_WORKERS
    threads. DB writes stay sequential in the main thread — single-row
    UPDATEs are microseconds each, so the bottleneck is file parsing
    (PDFs/DLIS can take seconds), which is what we parallelize.

    The chunked-rerun pattern is preserved: each chunk finishes, Streamlit
    reruns, the next chunk starts. The user can still hit Stop between
    chunks to pause without losing progress.
    """
    from sqlalchemy import text as _t
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # ── Ensure new 3D columns exist in FILE_SEIS_HEADER ──────────────────────
    # Added when the 3D extractor was upgraded (inline/crossline range and
    # survey outline polygon). ALTER TABLE is a no-op if column already
    # exists — guarded by the sys.columns check so it's safe to run every
    # time and costs a single metadata query per session.
    try:
        with engine.begin() as _con:
            for _col, _def in [
                ("IL_MIN",         "INT NULL"),
                ("IL_MAX",         "INT NULL"),
                ("XL_MIN",         "INT NULL"),
                ("XL_MAX",         "INT NULL"),
                ("SURVEY_OUTLINE", "NVARCHAR(MAX) NULL"),
            ]:
                _con.execute(_t(f"""
                    IF NOT EXISTS (
                        SELECT 1 FROM sys.columns
                        WHERE object_id = OBJECT_ID(
                            'file_catalog.FILE_SEIS_HEADER')
                          AND name = '{_col}'
                    )
                    ALTER TABLE file_catalog.FILE_SEIS_HEADER
                        ADD [{_col}] {_def}
                """))
    except Exception:
        pass  # Non-fatal — extraction proceeds; new fields just won't write

    # Phase 2 worker count from session state. Default 8 — balanced for
    # mixed extraction (PDF/DLIS/Office). User can tune via the slider in
    # the Scan tab.
    PHASE2_WORKERS = int(st.session_state.get("wb_phase2_workers", 8))

    try:
        with engine.connect() as con:
            total_all = con.execute(_t("""
                SELECT COUNT(*) FROM file_catalog.GLOBAL_FILE_CATALOG
            """)).scalar() or 1

            pending = con.execute(_t("""
                SELECT COUNT(*) FROM file_catalog.GLOBAL_FILE_CATALOG
                WHERE HEADER_EXTRACTED IS NULL OR HEADER_EXTRACTED='N'
            """)).scalar() or 0

            rows = con.execute(_t(f"""
                SELECT TOP {ENRICH_CHUNK}
                    INVENTORY_ID, FILE_PATH, FILE_EXT
                FROM file_catalog.GLOBAL_FILE_CATALOG
                WHERE (HEADER_EXTRACTED IS NULL OR HEADER_EXTRACTED='N')
                  AND ISNULL(HEADER_EXTRACTED,'') <> 'S'
                ORDER BY SCAN_DATE DESC
            """)).fetchall()
    except Exception as e:
        st.error(f"Extraction query failed: {e}")
        st.session_state["wb_enriching"] = False
        return

    if not rows:
        st.success("✅ Phase 2 complete — all files extracted.")
        st.session_state["wb_enriching"] = False
        st.session_state["wb_enrich_done"] = True
        st.session_state["wb_enrich_total"] = total_all
        return

    # Live progress bar — kept as a handle and advanced as each file finishes,
    # so it moves continuously during the chunk instead of only on the rerun
    # between chunks. Reused for the final update after the writes commit.
    _done_before = max(0, total_all - pending)
    _bar = st.progress(
        min(1.0, _done_before / max(total_all, 1)),
        text=f"⚙️ Extracting — {_done_before:,} / {total_all:,} · next {len(rows)}…",
    )

    # ── Parallel extraction within this chunk ────────────────────────────
    # Each worker handles one file: read it, extract fields, return result.
    # Workers don't touch the DB — writes happen in the main thread below.
    # Each result carries its own elapsed time so we can spot slow formats
    # and compare sum-of-times (would-be sequential) vs wall-clock (parallel).
    import time as _time
    _t_chunk_start = _time.monotonic()

    def _worker(row):
        inv_id, fpath, fext = row
        _t0 = _time.monotonic()
        try:
            fields = _extract_fields(fpath, (fext or "").lower())
            # Size-gate skip — surface as a distinct status so it's written
            # as HEADER_EXTRACTED='S' and never re-attempted.
            if fields.get("skip_reason"):
                return ("skip", inv_id, fpath, fext, fields,
                        fields["skip_reason"],
                        _time.monotonic() - _t0)
            return ("ok", inv_id, fpath, fext, fields, None,
                    _time.monotonic() - _t0)
        except Exception as e:
            return ("err", inv_id, fpath, fext, None,
                    f"{type(e).__name__}: {e}",
                    _time.monotonic() - _t0)

    results = []
    _t_pool_start = _time.monotonic()
    with ThreadPoolExecutor(max_workers=PHASE2_WORKERS) as pool:
        futures = [pool.submit(_worker, row) for row in rows]
        _seen = 0
        for fut in as_completed(futures):
            try:
                # Per-file timeout: 60s ceiling. Hung extractor doesn't
                # block the chunk.
                results.append(fut.result(timeout=60))
            except Exception as e:
                results.append(("err", None, "", "", None,
                                f"worker died: {e}", 0.0))
            # Move the bar as each file lands.
            _seen += 1
            _bar.progress(
                min(1.0, (_done_before + _seen) / max(total_all, 1)),
                text=(f"⚙️ Extracting — {_done_before + _seen:,} / "
                      f"{total_all:,} · this batch {_seen}/{len(rows)}…"),
            )
    _t_pool_end = _time.monotonic()
    _pool_elapsed = _t_pool_end - _t_pool_start

    # ── Batched DB writes via executemany (one round-trip per statement) ─
    # Previous attempt at "batched" was just one transaction wrapping per-
    # row execute() calls. Diagnostic showed that didn't help — cost is
    # per-statement, not per-transaction. ODBC was still doing N round-
    # trips, just with one commit at the end.
    #
    # Real fix: build parameter LISTS, call executemany() once per
    # statement-shape. pyodbc has fast_executemany=True enabled in our
    # engine config (db.py line 159), which makes executemany pack many
    # parameter sets into a single network round-trip.
    #
    # Expected: 3 round-trips per chunk (one for GLOBAL_FILE_CATALOG,
    # one for FILE_WELL_HEADER MERGE, one for FILE_SEIS_HEADER MERGE)
    # instead of N. For chunks with mixed file types, the WELL/SEIS
    # MERGEs only run if there's data of that category.
    _t_writes_start = _time.monotonic()

    # Build parameter lists by statement shape
    update_params: list = []     # for GLOBAL_FILE_CATALOG (success path)
    error_params:  list = []     # for GLOBAL_FILE_CATALOG (error path - 'E')
    skip_params:   list = []     # for GLOBAL_FILE_CATALOG (skip path - 'S')
    well_params:   list = []     # for FILE_WELL_HEADER MERGE
    seis_params:   list = []     # for FILE_SEIS_HEADER MERGE
    done = 0
    last = ""
    per_ext_times: dict = {}

    for outcome, inv_id, fpath, fext, fields, err, elapsed in results:
        if fpath:
            last = Path(fpath).name
        ext_key = (fext or "?").lower()
        per_ext_times.setdefault(ext_key, []).append(elapsed)

        if outcome == "ok" and inv_id is not None:
            # Success: queue the GLOBAL_FILE_CATALOG UPDATE and the
            # matching MERGE for whichever header table this category lives in.
            score, readiness = _score(fields)
            category = fields.get("file_category", "UNKNOWN")
            # Extension gate for the seismic table: FILE_SEIS_HEADER is for
            # SEG-Y / P190 files ONLY. A non-seismic file (e.g. a lease/blocks
            # shapefile with a "survey" column, or a JSON) that gets classified
            # SEIS must NOT write a bogus survey there. If the category says SEIS
            # but the extension isn't a real seismic type, drop it out of the
            # seismic path (it stays inventoried; it just doesn't fabricate a
            # seismic survey). Belt-and-suspenders alongside the shapefile
            # classifier fix.
            _SEIS_EXTS = {".segy", ".sgy", ".seg", ".p190", ".p90", ".p1"}
            if category == "SEIS" and (fext or "").lower() not in _SEIS_EXTS:
                category = "UNKNOWN"
            # Canonicalize UWI to bare-14 before any catalog write, so a
            # display-formatted API (scout-ticket PDFs print '42-999-00001-00-00')
            # can't leak into MATCHED_UWI / FILE_WELL_HEADER.UWI and false-fail
            # the dv_well FK gate. One mutation covers both writes below.
            fields["uwi"] = _norm_uwi(fields.get("uwi"))
            # Refine FILE_TYPE_GROUP for seismic files now that we know
            # whether the file is 2D or 3D. Phase 1 sets "Seismic" for all
            # SEG-Y; Phase 2 upgrades it to "Seismic 2D" or "Seismic 3D"
            # based on the trace count heuristic in _extract_fields.
            _seis_type = fields.get("seis_set_type")  # "2D", "3D", or None
            if category == "SEIS" and _seis_type in ("2D", "3D"):
                _type_group = f"Seismic {_seis_type}"
            else:
                # Keep the existing group from Phase 1 scan for non-seismic
                # files and seismic files where type couldn't be determined.
                _type_group = EXT_GROUP.get((fext or "").lower(), "Other")
            update_params.append({
                "score":      score,
                "readiness":  readiness,
                "uwi":        _trunc(fields.get("uwi"), 40),
                "issues":     "; ".join(_issues(fields)),
                "type_group": _type_group,
                "id":         inv_id,
            })
            if category == "WELL":
                well_params.append({
                    "hid":     uuid.uuid5(uuid.NAMESPACE_URL, inv_id).hex.upper(),
                    "inv_id":  inv_id,
                    "uwi":     _trunc(fields.get("uwi"),40),
                    "wn":      _trunc(fields.get("well_name"),255),
                    "op":      _trunc(fields.get("operator"),255),
                    "fld":     _trunc(fields.get("well_field"),100),
                    "st":      _trunc(fields.get("state"),50),
                    "co":      _trunc(fields.get("county"),100),
                    "lat":     _trunc(fields.get("latitude"),30),
                    "lon":     _trunc(fields.get("longitude"),30),
                    "td":      _trunc(fields.get("total_depth"),20),
                    "spud":    _trunc(fields.get("spud_date"),20),
                    "rig":     _trunc(fields.get("rig_release"),20),
                    "rt":      _trunc(fields.get("report_type"),50),
                    "stype":   _trunc(fields.get("survey_type"),50),
                    "contr":   _trunc(fields.get("contractor"),255),
                    "conf":    _safe_num(fields.get("confidence")),
                })
            elif category == "SEIS":
                # _seis_params, not a sixth hand-written copy — it is
                # what carries SURVEY_NAME_SOURCE into the MERGE.
                seis_params.append(_seis_params(inv_id, fields))
            done += 1
        elif outcome == "skip" and inv_id is not None:
            # Size-gate or other deliberate skip — write 'S' so the file
            # is never re-attempted. The skip_reason is stored in err.
            skip_params.append({"id": inv_id, "reason": (err or "SKIPPED")[:500]})
        elif inv_id is not None:
            # Extraction errored — queue the error-marker UPDATE
            error_params.append({"id": inv_id})

    # Execute the batched writes. Each executemany() is a single round-trip
    # with fast_executemany=True. We isolate each statement-shape in its
    # own try/except so a failure in one (e.g. a triggered constraint
    # violation in FILE_WELL_HEADER) doesn't poison the others.
    try:
        with engine.begin() as con:
            if update_params:
                con.execute(_t("""
                    UPDATE file_catalog.GLOBAL_FILE_CATALOG SET
                        CATALOG_SCORE     = :score,
                        CATALOG_READINESS = :readiness,
                        MATCHED_UWI       = :uwi,
                        CATALOG_ISSUES    = :issues,
                        FILE_TYPE_GROUP   = :type_group,
                        HEADER_EXTRACTED  = 'Y',
                        ROW_CHANGED_DATE  = GETUTCDATE()
                    WHERE INVENTORY_ID = :id
                """), update_params)
            if well_params:
                con.execute(_t("""
                    MERGE file_catalog.FILE_WELL_HEADER AS tgt
                    USING (SELECT :hid AS WELL_HEADER_ID) src
                    ON tgt.WELL_HEADER_ID = src.WELL_HEADER_ID
                    WHEN MATCHED THEN UPDATE SET
                        UWI=:uwi, WELL_NAME=:wn, OPERATOR=:op,
                        WELL_FIELD=:fld, STATE=:st, COUNTY=:co,
                        LATITUDE=:lat, LONGITUDE=:lon,
                        TOTAL_DEPTH=:td, SPUD_DATE=:spud,
                        RIG_RELEASE=:rig, REPORT_TYPE=:rt,
                        SURVEY_TYPE=:stype, CONTRACTOR=:contr,
                        CONFIDENCE=:conf, EXTRACTED_DATE=GETUTCDATE()
                    WHEN NOT MATCHED THEN INSERT (
                        WELL_HEADER_ID,INVENTORY_ID,
                        UWI,WELL_NAME,OPERATOR,WELL_FIELD,
                        STATE,COUNTY,LATITUDE,LONGITUDE,
                        TOTAL_DEPTH,SPUD_DATE,RIG_RELEASE,
                        REPORT_TYPE,SURVEY_TYPE,CONTRACTOR,CONFIDENCE,
                        EXTRACTED_DATE,EXTRACTED_BY
                    ) VALUES (
                        :hid,:inv_id,
                        :uwi,:wn,:op,:fld,
                        :st,:co,:lat,:lon,
                        :td,:spud,:rig,
                        :rt,:stype,:contr,:conf,
                        GETUTCDATE(),'DataWrangler'
                    );
                """), well_params)
            if seis_params:
                ensure_seis_columns(con)
                con.execute(_t(_SQL_SEIS_MERGE), seis_params)
            if error_params:
                con.execute(_t("""
                    UPDATE file_catalog.GLOBAL_FILE_CATALOG
                    SET HEADER_EXTRACTED='E',
                        ROW_CHANGED_DATE=GETUTCDATE()
                    WHERE INVENTORY_ID=:id
                """), error_params)
            if skip_params:
                # 'S' = deliberately skipped (too large, format limit).
                # Never re-attempted by the extraction loop.
                # Skip reason stored in CATALOG_READINESS so it's visible
                # in the Browse tab without a separate column.
                con.execute(_t("""
                    UPDATE file_catalog.GLOBAL_FILE_CATALOG
                    SET HEADER_EXTRACTED='S',
                        CATALOG_READINESS='SKIPPED',
                        ROW_CHANGED_DATE=GETUTCDATE()
                    WHERE INVENTORY_ID=:id
                """), skip_params)
    except Exception as e:
        st.error(f"Chunk transaction failed (rolled back): {e}")
        st.session_state["wb_enriching"] = False
        return
    _t_writes_end = _time.monotonic()
    _writes_elapsed = _t_writes_end - _t_writes_start

    # Sum of per-file extraction times. If sum_per_file >> pool_elapsed,
    # parallelism IS helping (threads overlapping). If sum ≈ pool, the
    # GIL or some serializing call is preventing real parallelism.
    _sum_per_file = sum(e for grp in per_ext_times.values() for e in grp)
    _speedup_ratio = (_sum_per_file / _pool_elapsed) if _pool_elapsed > 0 else 0.0

    total_done = total_all - pending + done
    pct = min(1.0, total_done / max(total_all, 1))

    # Final chunk? Flip to done and rerun so the page redraws with ONLY the
    # completion message — no leftover progress bar, "remaining" caption, or
    # chunk diagnostics lingering after extraction finishes.
    if not (len(rows) == ENRICH_CHUNK and pending > done):
        st.session_state["wb_enriching"]    = False
        st.session_state["wb_enrich_done"]  = True
        st.session_state["wb_enrich_total"] = total_all
        st.rerun()

    # Mid-run: advance the bar, show what's left and the chunk diagnostics,
    # then loop to the next chunk.
    _bar.progress(pct, text=(
        f"⚙️ Extracting — {total_done:,} / {total_all:,} "
        f"({pct*100:.0f}%) · {last}"
    ))
    st.caption(
        f"{pending - done:,} remaining · "
        f"{PHASE2_WORKERS} threads · "
        "click **⏹ Stop** to pause"
    )

    # ── Diagnostic: where did time go this chunk? ────────────────────────
    # Sum of per-file extraction times vs wall-clock pool time tells us
    # whether parallelism is actually working.
    _per_ext_summary = " · ".join(
        f"{ext}:{sum(times):.2f}s({len(times)})"
        for ext, times in sorted(per_ext_times.items(),
                                 key=lambda kv: -sum(kv[1]))
    )
    st.caption(
        f"⏱ chunk={_pool_elapsed + _writes_elapsed:.2f}s "
        f"(extract:{_pool_elapsed:.2f}s, writes:{_writes_elapsed:.2f}s) · "
        f"sum-of-files={_sum_per_file:.2f}s · "
        f"speedup={_speedup_ratio:.1f}× (ideal: {PHASE2_WORKERS}×) · "
        f"by-ext: {_per_ext_summary}"
    )

    import time
    time.sleep(0.1)
    st.rerun()


# =============================================================================
# Field extraction and enrichment write
# =============================================================================

# _extract_fields now lives in extract_core (imported above) so the
# pipeline process-pool workers can parse without importing streamlit.
# Single source of truth — edit it there.


def _write_enrichment(engine, inv_id: str, fields: dict):
    """Write extracted header fields to catalog tables.

    Single-file write — opens its own transaction. Kept for the few
    callers outside the Phase 2 chunk loop. The chunk loop uses
    _write_enrichment_on() with a shared connection for batched-commit
    performance.
    """
    with engine.begin() as con:
        _write_enrichment_on(con, inv_id, fields)








def _date_only(v):
    """Keep just the date part (drops any trailing time), so a mm/dd/yyyy
    value stays ~10 chars and never overflows the 20-char date columns.
    Preserves the source date format; only strips the time."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s.replace("T", " ").split(" ", 1)[0][:20]












# The 17 bind names one header row needs, in the order the VALUES list
# below emits them.
_WELL_MERGE_COLS = ("hid", "inv_id", "uwi", "wn", "op", "fld", "st", "co",
                    "lat", "lon", "td", "spud", "rig", "rt", "stype",
                    "contr", "conf")

# SQL Server allows 2,100 parameters per statement; 17 per row means 123 is
# the ceiling. 100 leaves room and keeps the statement readable in a trace.
_WELL_MERGE_CHUNK = 100


def _sql_well_merge_many(n):
    """One MERGE that handles n header rows instead of one.

    WHY NOT executemany, which is what everything else here uses: the
    single-row form is `USING (SELECT :hid AS WELL_HEADER_ID)`, and that
    subquery defeats pyodbc's fast_executemany column sizing — it sizes
    string buffers from the FIRST row, under-sizes them, and silently
    truncates a later longer value. That is why this write has always been
    per row, and it was the right call: correctness over speed.

    Measured cost of that choice: header_write 136.4s of a 181.7s extract
    stage, 22 chunks of 50 files, ~124ms per file — about a thousand
    round trips.

    `USING (VALUES (...),(...),…)` keeps every value a SEPARATELY BOUND
    parameter, sized individually exactly as the per-row path does, so the
    truncation cannot come back — while collapsing a hundred round trips
    into one statement. Same MERGE body, same semantics, same clamping.
    """
    rows = ", ".join(
        "(" + ", ".join(f":{c}{i}" for c in _WELL_MERGE_COLS) + ")"
        for i in range(n))
    cols = ", ".join(_WELL_MERGE_COLS)
    return f"""
    MERGE file_catalog.FILE_WELL_HEADER AS tgt
    USING (VALUES {rows}) AS src ({cols})
    ON tgt.WELL_HEADER_ID = src.hid
    WHEN MATCHED THEN UPDATE SET
        UWI=src.uwi, WELL_NAME=src.wn, OPERATOR=src.op,
        WELL_FIELD=src.fld, STATE=src.st, COUNTY=src.co,
        LATITUDE=src.lat, LONGITUDE=src.lon,
        TOTAL_DEPTH=src.td, SPUD_DATE=src.spud,
        RIG_RELEASE=src.rig, REPORT_TYPE=src.rt,
        SURVEY_TYPE=src.stype, CONTRACTOR=src.contr,
        CONFIDENCE=src.conf, EXTRACTED_DATE=GETUTCDATE()
    WHEN NOT MATCHED THEN INSERT (
        WELL_HEADER_ID,INVENTORY_ID,
        UWI,WELL_NAME,OPERATOR,WELL_FIELD,
        STATE,COUNTY,LATITUDE,LONGITUDE,
        TOTAL_DEPTH,SPUD_DATE,RIG_RELEASE,
        REPORT_TYPE,SURVEY_TYPE,CONTRACTOR,CONFIDENCE,
        EXTRACTED_DATE,EXTRACTED_BY
    ) VALUES (
        src.hid,src.inv_id,
        src.uwi,src.wn,src.op,src.fld,
        src.st,src.co,src.lat,src.lon,
        src.td,src.spud,src.rig,
        src.rt,src.stype,src.contr,src.conf,
        GETUTCDATE(),'DataWrangler'
    );
    """


def _merge_wells_chunked(con, wells):
    """MERGE header rows in chunks, falling back to per-row on any failure.

    DEDUPED BY hid, LAST WINS: a MERGE raises if the source offers two rows
    matching the same target row, and two entries for one file in a single
    chunk would do exactly that. The per-row loop never had to care because
    each statement saw one row.
    """
    from sqlalchemy import text as _t
    seen = {}
    for w in wells:
        seen[w.get("hid")] = w
    rows = list(seen.values())

    for i in range(0, len(rows), _WELL_MERGE_CHUNK):
        chunk = rows[i:i + _WELL_MERGE_CHUNK]
        binds = {}
        for n, w in enumerate(chunk):
            for c in _WELL_MERGE_COLS:
                binds[f"{c}{n}"] = w.get(c)
        try:
            con.execute(_t(_sql_well_merge_many(len(chunk))), binds)
        except Exception:
            # One bad row must not cost the chunk. The per-row statement is
            # still here and still correct; this is the same fallback shape
            # the capture batch uses.
            for w in chunk:
                con.execute(_t(_SQL_WELL_MERGE), w)













def _xl_file_uri(p):
    """Turn a filesystem path into a file:// URI Excel will open as a hyperlink.
    Handles Windows drive paths (C:\\x → file:///C:/x), UNC shares
    (\\\\srv\\share → file://srv/share) and POSIX paths (/mnt/x → file:///mnt/x).
    Returns None for blanks or anything too long for an Excel hyperlink target."""
    s = str(p or "").strip()
    if not s:
        return None
    if s.startswith("\\\\"):                       # UNC \\server\share\...
        uri = "file:" + s.replace("\\", "/")       #   → file://server/share/...
    else:
        f = s.replace("\\", "/")
        uri = ("file://" + f) if f.startswith("/") else ("file:///" + f)
    return uri if len(uri) <= 2000 else None       # Excel hyperlink target cap


def _scorecard_xlsx(df):
    """Stage-scorecard DataFrame -> xlsx bytes, FILE NAME cell hyperlinked.

    Mirrors the Browse & View exports: the name is the link, the path column
    stays plain text so it can still be copied, and a UWI is written as TEXT so
    Excel can't render a 14-digit API as 1.50012E+13 and destroy it on save.

    Raises ImportError when openpyxl isn't installed — the caller turns that
    into a one-line hint rather than losing the CSV path as well.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    cols = [c for c in df.columns if c != "path"] + (
        ["path"] if "path" in df.columns else [])
    wb = Workbook()
    ws = wb.active
    ws.title = "scorecard"

    hdr = Font(bold=True)
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = hdr

    link = Font(color="0563C1", underline="single")
    name_i = cols.index("file") + 1 if "file" in cols else None
    uwi_i = cols.index("uwi") + 1 if "uwi" in cols else None

    for i, rec in enumerate(df.to_dict("records"), start=2):
        for j, c in enumerate(cols, start=1):
            v = rec.get(c)
            cell = ws.cell(row=i, column=j,
                           value="" if v is None else str(v))
            if j == uwi_i:
                cell.number_format = "@"      # keep 14-digit UWIs as text
        if name_i:
            target = _xl_file_uri(rec.get("path"))
            if target:
                c = ws.cell(row=i, column=name_i)
                c.hyperlink = target
                c.font = link

    for j, c in enumerate(cols, start=1):
        width = max([len(c)] + [len(str(r.get(c) or ""))
                                for r in df.to_dict("records")[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(60, max(9, width + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (f"A1:{get_column_letter(len(cols))}"
                          f"{max(2, len(df) + 1)}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Manual survey-name assignment (single owner) ───────────────────────────
# Assigning a SURVEY_NAME has to touch TWO tables, and writing only the first
# one is a silent dead end:
#   * GLOBAL_FILE_CATALOG.SURVEY_NAME — display/override only. Nothing in the
#     promote path reads it.
#   * FILE_SEIS_HEADER.SURVEY_NAME    — what promote actually joins on
#     (seis_done: dv_seis_set.seis_set_name = FILE_SEIS_HEADER.SURVEY_NAME),
#     so a survey assigned only in GFC would flip readiness to READY and then
#     never be credited.
# The seis header row is upserted on the SAME derived key extraction uses
# (uuid5(NAMESPACE_URL, inventory_id + "_s")), so this updates the extracted
# row when one exists and creates a matching one when it doesn't — extraction
# re-running later updates that same row rather than duplicating it.
_SQL_SURVEY_GFC = """
    UPDATE file_catalog.GLOBAL_FILE_CATALOG SET
        SURVEY_NAME       = :sn,
        MATCH_METHOD      = 'MANUAL',
        CATALOG_READINESS = 'READY',
        ROW_CHANGED_DATE  = GETUTCDATE()
    WHERE INVENTORY_ID = :iid
"""

_SQL_SURVEY_SEIS = """
    MERGE file_catalog.FILE_SEIS_HEADER AS tgt
    USING (SELECT :hid AS SEIS_HEADER_ID) src
    ON tgt.SEIS_HEADER_ID = src.SEIS_HEADER_ID
    WHEN MATCHED THEN UPDATE SET
        SURVEY_NAME=:sn, SURVEY_NAME_SOURCE='manual',
        EXTRACTED_DATE=GETUTCDATE()
    WHEN NOT MATCHED THEN INSERT (
        SEIS_HEADER_ID, INVENTORY_ID, SURVEY_NAME, SURVEY_NAME_SOURCE,
        EXTRACTED_DATE, EXTRACTED_BY
    ) VALUES (
        :hid, :iid, :sn, 'manual', GETUTCDATE(), 'DataWrangler'
    );
"""


def _normalize_uwi14(raw):
    """Canonicalize a hand-entered UWI to bare-14, or None if it doesn't.

    Returning None (rather than a truncated/padded string) lets callers report
    the value as bad instead of writing a malformed key that would fail the
    dv_well FK later — the '070009' DLIS case is what a missing guard looks
    like downstream.
    """
    try:
        from dataview.core import path_identity as _pi
    except Exception:
        _pi = None
    s = str(raw or "").strip()
    if not s:
        return None
    u = (_pi.norm_uwi14(s) if _pi
         else "".join(ch for ch in s if ch.isalnum()))
    u = (u or "")[:14]
    return u if len(u) == 14 else None


# Manual UWI assignment. CAPTURED_HASH is deliberately cleared here: capture
# skips a file when CAPTURED_HASH = FILE_HASH, and that stamp keys on file
# CONTENT. Assigning a UWI doesn't change the file, so without this the very
# next run considers the file already captured and silently skips it — the
# recurring "+0 captured" false alarm. Clearing the stamp makes the assignment
# re-arm capture by itself, which is what the Apply toggle is (reasonably)
# expected to do.
#
# NOTE the asymmetry with _assign_survey_name, which does NOT clear it: the
# SEG-Y capture path re-parses the file into FILE_SEIS_HEADER, so re-arming a
# seismic file would overwrite a manually assigned SURVEY_NAME with whatever
# the header holds (nothing, for the files that needed assigning).
#
# Re-capture is safe to repeat: capture() replaces a file's rows scoped to
# INVENTORY_ID, so it refreshes rather than duplicates.
_SQL_ASSIGN_UWI = """
    UPDATE file_catalog.GLOBAL_FILE_CATALOG SET
        MATCHED_UWI       = :uwi,
        UWI14             = :uwi,
        MATCH_METHOD      = 'MANUAL',
        CATALOG_READINESS = 'READY',
        CAPTURED_HASH     = NULL,
        ROW_CHANGED_DATE  = GETUTCDATE()
    WHERE INVENTORY_ID = :iid
"""


def _assign_uwi(con, iid, uwi14):
    """Write a normalized 14-char UWI on an open connection and re-arm capture.
    Returns the rowcount (0 = no such INVENTORY_ID)."""
    from sqlalchemy import text as _t
    return con.execute(_t(_SQL_ASSIGN_UWI),
                       {"uwi": uwi14, "iid": str(iid)}).rowcount


def _assign_survey_name(con, iid, sn):
    """Write a manually assigned survey name to BOTH tables on an open
    connection. Returns the GFC rowcount (0 = no such INVENTORY_ID, so the
    caller can report a miss). The seis-header upsert is best-effort: a DB
    without FILE_SEIS_HEADER shouldn't break the GFC assignment."""
    from sqlalchemy import text as _t
    import uuid as _uuid
    _sn = str(sn or "").strip()[:255]
    _iid = str(iid)
    res = con.execute(_t(_SQL_SURVEY_GFC), {"sn": _sn, "iid": _iid})
    try:
        _hid = _uuid.uuid5(_uuid.NAMESPACE_URL, _iid + "_s").hex.upper()
        con.execute(_t(_SQL_SURVEY_SEIS),
                    {"hid": _hid, "iid": _iid, "sn": _sn})
    except Exception:
        pass          # display write already landed; promote join is best-effort
    return res.rowcount


# ── Bounded variants ────────────────────────────────────────────────────
# pyodbc's fast_executemany pre-checks numeric ranges and rejects the
# whole batch if any value overflows the target column's precision/scale.
# These helpers clamp values to known-safe ranges, dropping outliers to
# NULL rather than letting them poison the batch.






# =============================================================================
# Tab 2 -- Browse & View
# =============================================================================



def _backfill_cataloged(engine):
    """Set CATALOG_READINESS='CATALOGED' for every file already captured in a
    cat_* mirror. One set-based UPDATE per mirror table (EXISTS join — no
    per-row loops), so it stays fast on a large catalog. Returns rows updated."""
    from sqlalchemy import text as _t
    with engine.connect() as con:
        tabs = [r[0] for r in con.execute(_t("""
            SELECT t.name
            FROM sys.tables t
            JOIN sys.schemas s ON s.schema_id = t.schema_id
            JOIN sys.columns c ON c.object_id = t.object_id
            WHERE s.name = 'file_catalog'
              AND t.name LIKE 'cat[_]%'
              AND c.name = 'INVENTORY_ID'""")).fetchall()]
    if not tabs:
        return 0
    updated = 0
    with engine.begin() as con:
        for t in tabs:
            r = con.execute(_t(f"""
                UPDATE g SET g.CATALOG_READINESS = 'CATALOGED'
                FROM file_catalog.GLOBAL_FILE_CATALOG g
                WHERE ISNULL(g.CATALOG_READINESS,'') <> 'CATALOGED'
                  AND EXISTS (SELECT 1 FROM file_catalog.{t} m
                              WHERE m.INVENTORY_ID = g.INVENTORY_ID)"""))
            updated += r.rowcount or 0
    return updated


def _wb_run_search(engine):
    """Run the Browse & View query and refresh st.session_state['wb_results'].

    Extracted so the SAME query backs the Search button, the Refresh button and
    the post-action refresh after a batch load or reject. Previously the query
    lived inline under the Search button, so anything that changed a file's
    readiness left the on-screen list stale with no way to update it short of
    searching again.

    Filters are read from the widgets' session keys rather than passed in, so
    a refresh always reproduces exactly what the user last searched for.
    """
    import pandas as pd
    from sqlalchemy import text as _t

    grp = st.session_state.get("wb_grp") or []
    rd = st.session_state.get("wb_rd", "All")
    srch = st.session_state.get("wb_srch", "")
    flagged = st.session_state.get("wb_flagged", False)

    conditions = ["1=1"]
    params = {}
    if grp:
        _ph = ",".join(f":grp{i}" for i in range(len(grp)))
        conditions.append(f"FILE_TYPE_GROUP IN ({_ph})")
        for i, _g in enumerate(grp):
            params[f"grp{i}"] = _g
    if rd != "All":
        conditions.append("CATALOG_READINESS=:rd")
        params["rd"] = rd
    if srch:
        conditions.append("FILE_NAME LIKE :srch")
        params["srch"] = f"%{srch}%"
    if flagged:
        conditions.append("ISNULL(FLAG_DELETE,'N')='Y'")
    # SKIPPED is terminal. A rejected file now KEEPS its catalog row so its
    # history stays attributable (see _mark_bad), which means this list has to
    # hide it — otherwise rejecting a file would look like it did nothing.
    # Choosing SKIPPED in the Readiness filter still finds them, which is how
    # you review what was rejected.
    if rd != "SKIPPED":
        conditions.append("ISNULL(CATALOG_READINESS,'') <> 'SKIPPED'")
    try:
        with engine.connect() as con:
            rows = con.execute(_t(f"""
                SELECT TOP 500
                    g.INVENTORY_ID, g.FILE_PATH, g.FILE_NAME, g.FILE_EXT,
                    g.FILE_TYPE_GROUP, g.FILE_SIZE_KB,
                    g.CATALOG_READINESS, g.CATALOG_SCORE,
                    g.MATCHED_UWI,
                    (SELECT TOP 1 h.WELL_NAME
                       FROM file_catalog.FILE_WELL_HEADER h
                      WHERE h.INVENTORY_ID = g.INVENTORY_ID) AS WELL_NAME,
                    -- Survey name for seismic. The pipeline writes it ONLY to
                    -- FILE_SEIS_HEADER (that's the column promote joins on),
                    -- so reading GFC alone showed blank for every resolved
                    -- SEG-Y. Prefer a manual GFC override if one was set,
                    -- else fall back to the extracted seis header.
                    ISNULL(NULLIF(LTRIM(RTRIM(g.SURVEY_NAME)), ''),
                           (SELECT TOP 1 s.SURVEY_NAME
                              FROM file_catalog.FILE_SEIS_HEADER s
                             WHERE s.INVENTORY_ID = g.INVENTORY_ID)) AS SURVEY_NAME,
                    g.CATALOG_ISSUES,
                    ISNULL(g.FLAG_DELETE,'N') FLAG_DELETE,
                    g.HEADER_EXTRACTED
                FROM file_catalog.GLOBAL_FILE_CATALOG g
                WHERE {" AND ".join(conditions)}
                ORDER BY g.CATALOG_SCORE DESC, g.FILE_NAME
            """), params).fetchall()
        df = pd.DataFrame(rows, columns=[
            "INVENTORY_ID","FILE_PATH","FILE_NAME","FILE_EXT",
            "FILE_TYPE_GROUP","FILE_SIZE_KB",
            "CATALOG_READINESS","CATALOG_SCORE",
            "MATCHED_UWI","WELL_NAME","SURVEY_NAME","CATALOG_ISSUES",
            "FLAG_DELETE","HEADER_EXTRACTED",
        ])
        st.session_state["wb_results"] = df
        st.session_state.pop("wb_nav_idx", None)
    except Exception as e:
        st.error(f"Search failed: {e}")

    return st.session_state.get("wb_results")


def _tab_browse(engine, dialect):
    from sqlalchemy import text as _t

    st.markdown("#### 📂 Browse & View")

    # ── Filters ───────────────────────────────────────────────────────────────
    f1, f2, f3, f4 = st.columns(4)
    grp = f1.multiselect(
        "File type",
        ["PDF","Well Log","Seismic","Shapefile","Office","Image"],
        key="wb_grp",
        placeholder="All types",
    )
    rd = f2.selectbox(
        "Readiness",
        # SKIPPED is listed so rejected files stay reachable — _wb_run_search
        # hides them from every other selection, so this is the only way back to
        # them. Appended, not inserted: a fixed-key selectbox keeps whatever the
        # session already holds, and reordering would silently change it.
        ["All","CATALOGED","READY","REVIEW","NEEDS_UWI","ATTENTION","SKIPPED"],
        key="wb_rd",
    )
    srch = f3.text_input("Filename contains", key="wb_srch",
                          placeholder="partial name...")
    flagged = f4.checkbox("Flagged only", key="wb_flagged")

    b1, b2 = st.columns([1, 1])
    if b1.button("🔍 Search", type="primary", key="wb_search_btn",
                 use_container_width=True):
        _wb_run_search(engine)
    # Re-runs the same query. Readiness changes as files are captured and
    # rejected, so a result set goes stale the moment you act on it.
    if b2.button("↻ Refresh list", key="wb_refresh_btn",
                 use_container_width=True,
                 disabled=st.session_state.get("wb_results") is None,
                 help="Re-run the current search — picks up readiness changes "
                      "from loading or rejecting files."):
        _wb_run_search(engine)
        st.rerun()

    if st.button("🗂 Refresh cataloged status", key="wb_backfill_cat",
                 help="Reconcile every file's readiness to row reality: CATALOGED "
                      "where cat_* rows exist, PROMOTED where they reached dv_*, and "
                      "demote any CATALOGED with no rows behind it. Safe to re-run."):
        try:
            try:
                from dataview.file_catalog.catalog_readiness import reconcile_readiness
            except Exception:
                from dataview.file_catalog.catalog_readiness import reconcile_readiness
            _res = reconcile_readiness(engine, "mssql", log=lambda *_: None)
            _n = _res.get("changed", 0)
            st.success(f"Reconciled readiness — {_n:,} file(s) updated.")
            st.session_state.pop("wb_results", None)
            st.rerun()
        except Exception as e:
            st.error(f"Reconcile failed: {e}")

    df = st.session_state.get("wb_results")
    if df is None:
        st.info("Set filters and click Search.")
        return

    st.caption(f"{len(df):,} files (max 500)")

    # ── Selectable results grid ──────────────────────────────────────────────
    # Checkbox column so a review pass is one screen: tick the doubtful files,
    # open any of them in the viewer below to confirm, then reject the batch.
    # Everything except Select is read-only — this grid is for choosing files,
    # not editing them; UWI/SURVEY_NAME assignment has its own editor.
    _grid = df[["FILE_NAME", "FILE_EXT", "FILE_TYPE_GROUP",
                "CATALOG_READINESS", "CATALOG_SCORE",
                "FLAG_DELETE"]].rename(columns={"CATALOG_READINESS": "Readiness"})
    _grid.insert(0, "Select", False)
    # The editor and the confirm box are keyed with a NONCE rather than a fixed
    # name. Streamlit refuses any write to a widget's key once that widget has
    # been instantiated, so "reset the checkbox after rejecting" cannot be done
    # by assigning to session_state — it raises. Bumping the nonce retires the
    # old widgets and mints fresh ones, which default to unticked. It also
    # discards stale row-index edits, which matters because rejecting changes
    # the result set and a leftover tick would point at a different file.
    _nonce = st.session_state.get("wb_sel_nonce", 0)
    _ed = st.data_editor(
        _grid, hide_index=True, use_container_width=True,
        key=f"wb_sel_editor_{_nonce}",
        disabled=[c for c in _grid.columns if c != "Select"],
        column_config={
            "Select": st.column_config.CheckboxColumn(
                "✓", help="Tick files to view or reject", width="small"),
        })
    try:
        _picked = [i for i, v in enumerate(_ed["Select"].tolist()) if bool(v)]
    except Exception:
        _picked = []

    v1, v2, v3 = st.columns([1, 1, 2])
    if v1.button(f"👁 View ({len(_picked)})", key="wb_sel_view",
                 disabled=not _picked, use_container_width=True,
                 help="Open the first ticked file in the viewer below. "
                      "Use Prev/Next there to walk the rest."):
        st.session_state["wb_nav_idx"] = _picked[0]
        st.rerun()

    _reason = v3.text_input(
        "Reason", key="wb_sel_reason", label_visibility="collapsed",
        placeholder="why are these bad? (stored on the blocklist)")
    _confirm = v2.checkbox("Confirm", key=f"wb_sel_confirm_{_nonce}",
                           help="Rejecting removes the file from the catalog. "
                                "Tick to enable the button.")

    # Rejecting DELETES the rows a file staged (and retires its catalog row as
    # SKIPPED), so it needs both a selection and an explicit confirm — a stray
    # click on a 500-row result set would otherwise discard most of what the
    # catalog has captured. Failures are collected per file rather than aborting,
    # so one unreadable file doesn't strand the rest half-done.
    if st.button(f"🚫 Reject checked ({len(_picked)}) — blocklist + remove",
                 key=f"wb_sel_reject_{_nonce}", type="secondary",
                 disabled=not (_picked and _confirm),
                 use_container_width=True):
        _ok, _fail, _rows = 0, [], 0
        for _i in _picked:
            _r = df.iloc[_i]
            try:
                _rows += sum(_mark_bad(
                    engine, _r["INVENTORY_ID"], _r["FILE_PATH"],
                    _r.get("FILE_NAME"), _r.get("FILE_SIZE_KB"),
                    _reason or "rejected on review").values())
                _ok += 1
            except Exception as _be:
                _fail.append(f"{_r.get('FILE_NAME','?')}: "
                             f"{type(_be).__name__}: {_be}")
        if _ok:
            st.success(
                f"Rejected {_ok:,} file(s) — blocklisted, catalog row retired "
                f"as SKIPPED"
                + (f", {_rows:,} staged row(s) dropped" if _rows else "")
                + ". The next crawl will skip them.")
        for _m in _fail[:5]:
            st.warning(_m)
        if len(_fail) > 5:
            st.warning(f"…and {len(_fail) - 5:,} more failed.")
        st.session_state.pop("wb_results", None)
        st.session_state.pop("wb_nav_idx", None)
        # Retire this round's widgets (see the nonce comment above) — never
        # assign to wb_sel_confirm_* itself.
        st.session_state["wb_sel_nonce"] = _nonce + 1
        st.rerun()

    # ── Export / Import assignments (Excel) ───────────────────────────────────
    # Round-trip the whole result grid through Excel so UWI / SURVEY_NAME can be
    # assigned in bulk offline (fill-down, VLOOKUP against a master, etc.) then
    # imported back in one pass. INVENTORY_ID is the join key, so the export
    # writes it AND the UWI column as TEXT ('@' number format) — otherwise Excel
    # silently turns a 14-digit API into 4.2999E+13 or drops a leading zero and
    # the round trip corrupts every UWI. Import matches on INVENTORY_ID (the PK,
    # so it works regardless of what's currently searched), ignores blank cells,
    # normalizes UWI with the same norm_uwi14 recipe as the inline editor, and
    # flips readiness to READY — identical to a manual Save.
    with st.expander("📤 Export / 📥 Import assignments (Excel)", expanded=False):
        try:
            from dataview.core import path_identity as _pi
        except Exception:
            _pi = None

        # ---- Export -----------------------------------------------------------
        _exp = df.copy()
        if "SURVEY_NAME" not in _exp.columns:
            _exp["SURVEY_NAME"] = ""          # search SELECT omits it; blank to fill
        _cols = ["INVENTORY_ID", "FILE_NAME", "FILE_PATH", "FILE_TYPE_GROUP",
                 "FILE_EXT", "CATALOG_READINESS", "CATALOG_SCORE", "WELL_NAME",
                 "CATALOG_ISSUES", "MATCHED_UWI", "SURVEY_NAME"]
        _exp = _exp[[c for c in _cols if c in _exp.columns]].copy()
        for _c in _exp.columns:               # clean strings, no literal 'None'/'nan'
            _exp[_c] = (_exp[_c].fillna("").astype(str)
                        .replace({"None": "", "nan": ""}))
        try:
            import io
            from openpyxl.utils import get_column_letter
            _buf = io.BytesIO()
            with pd.ExcelWriter(_buf, engine="openpyxl") as _xw:
                _exp.to_excel(_xw, index=False, sheet_name="assignments")
                _ws = _xw.sheets["assignments"]
                _text_cols = {"INVENTORY_ID", "MATCHED_UWI", "SURVEY_NAME"}
                for _i, _name in enumerate(_exp.columns, start=1):
                    _letter = get_column_letter(_i)
                    _wid = _exp[_name].str.len().max()
                    _wid = 12 if pd.isna(_wid) else int(_wid)
                    _ws.column_dimensions[_letter].width = min(max(12, _wid + 2), 60)
                    if _name in _text_cols:          # keep IDs/UWIs verbatim
                        for _cell in _ws[_letter]:
                            _cell.number_format = "@"
                # make the FILE_NAME cell a clickable link to the file (target
                # built from FILE_PATH; the path column stays plain text).
                _cl = list(_exp.columns)
                if "FILE_NAME" in _cl and "FILE_PATH" in _cl:
                    from openpyxl.styles import Font as _Font
                    _nlet = get_column_letter(_cl.index("FILE_NAME") + 1)
                    _plet = get_column_letter(_cl.index("FILE_PATH") + 1)
                    _lfont = _Font(color="0563C1", underline="single")
                    for _nc, _pc in zip(_ws[_nlet][1:], _ws[_plet][1:]):  # data rows
                        _u = _xl_file_uri(_pc.value)
                        if _u:
                            _nc.hyperlink = _u
                            _nc.font = _lfont
                _ws.freeze_panes = "A2"
            st.download_button(
                "📤 Export grid to Excel",
                data=_buf.getvalue(),
                file_name=f"assignments_{datetime.now():%Y%m%d_%H%M}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
                use_container_width=True,
                help="Every row in the current results grid. Fill MATCHED_UWI "
                     "(wells) or SURVEY_NAME (seismic); leave a row blank to skip "
                     "it. Do NOT edit INVENTORY_ID — it's how import finds the row.")
        except Exception as _xx:
            st.error(f"Excel export unavailable ({type(_xx).__name__}: {_xx}). "
                     "Needs openpyxl — `pip install openpyxl`.")

        st.divider()

        # ---- Import -----------------------------------------------------------
        _up = st.file_uploader("Import filled Excel", type=["xlsx"],
                               key="wb_assign_xlsx",
                               help="Matches rows by INVENTORY_ID; applies every "
                                    "non-blank MATCHED_UWI / SURVEY_NAME cell.")
        if _up is not None and st.button("📥 Apply assignments from file",
                                         type="primary", key="wb_assign_import"):
            try:
                _imp = pd.read_excel(_up, dtype=str)          # dtype=str: no coercion
            except Exception as _rx:
                st.error(f"Could not read workbook: {_rx}")
                _imp = None
            if _imp is not None:
                _imp.columns = [str(_c).strip() for _c in _imp.columns]
                if "INVENTORY_ID" not in _imp.columns:
                    st.error("Workbook has no INVENTORY_ID column — export a fresh "
                             "template above and fill that one.")
                else:
                    # UWI + survey both go through the shared module-level
                    # helpers (_assign_uwi / _assign_survey_name) so the inline
                    # panel and this importer can't drift apart.

                    def _clean_uwi(v):
                        """Strip the trailing '.0' a number-typed cell can pick up
                        (lossless — 14-digit UWIs are exact as doubles). Anything
                        genuinely malformed is left as-is so norm_uwi14 rejects it
                        rather than a lossy reconstruction applying a wrong UWI."""
                        s = str(v or "").strip()
                        if not s or s.lower() == "nan":
                            return ""
                        if s.endswith(".0") and s[:-2].isdigit():
                            s = s[:-2]
                        return s

                    n_uwi = n_srv = notfound = 0
                    bad = []
                    try:
                        with engine.begin() as con:
                            for _, r in _imp.iterrows():
                                iid = str(r.get("INVENTORY_ID", "") or "").strip()
                                if not iid or iid.lower() == "nan":
                                    continue
                                raw_uwi = _clean_uwi(r.get("MATCHED_UWI", ""))
                                raw_srv = str(r.get("SURVEY_NAME", "") or "").strip()
                                if raw_srv.lower() == "nan":
                                    raw_srv = ""
                                fname = str(r.get("FILE_NAME", iid) or iid)
                                if raw_uwi:
                                    u = _normalize_uwi14(raw_uwi)
                                    if not u:
                                        bad.append(f"{fname}: '{raw_uwi}' → not 14 chars")
                                        continue
                                    if _assign_uwi(con, iid, u):
                                        n_uwi += 1
                                    else:
                                        notfound += 1
                                elif raw_srv:
                                    if _assign_survey_name(con, iid, raw_srv):
                                        n_srv += 1
                                    else:
                                        notfound += 1
                        _msg = []
                        if n_uwi: _msg.append(f"{n_uwi} UWI")
                        if n_srv: _msg.append(f"{n_srv} survey name")
                        if _msg:
                            st.success("Imported " + " + ".join(_msg) +
                                       " → readiness READY. Re-run capture to load them.")
                        else:
                            st.info("No non-blank UWI / SURVEY_NAME cells to apply.")
                        if notfound:
                            st.caption(f"{notfound} row(s) had no matching "
                                       f"INVENTORY_ID in the catalog (skipped).")
                        if bad:
                            st.warning("Skipped (fix and re-import): " + "; ".join(bad))
                        st.session_state.pop("wb_results", None)
                        st.rerun()
                    except Exception as _ex:
                        st.error(f"Import failed: {_ex}")

    # ── Assign UWI (well data) / SURVEY_NAME (seismic) to unresolved files ─────────
    # Edit MATCHED_UWI or SURVEY_NAME inline for NEEDS_UWI/REVIEW files, then Save.
    # UWI is normalized to the canonical 14 (path_identity.norm_uwi14) — the same
    # recipe the pipeline uses — and readiness flips to READY so capture picks it up.
    # CATALOGED + blank MATCHED_UWI belongs here too. Such a file DID produce
    # rows — which is exactly why readiness reads CATALOGED rather than
    # NEEDS_UWI — but what holds those rows is the missing well, and assigning
    # a UWI is the fix. Leaving it out meant the files with real staged data
    # behind them were the only ones the panel would not offer to resolve, so
    # "no UWI" looked like a reason to REJECT rather than to assign (17 Aug:
    # SCOUT_REPORT1797 with 25 staged rows and WELL_TEST_REPORT6155 with 19
    # were both invisible here).
    #
    # The blank-UWI qualifier is what keeps this honest: a CATALOGED file that
    # already HAS a MATCHED_UWI is merely pending promote, not unresolved, and
    # including those would crowd the grid with rows that need nothing.
    _unresolved = ["NEEDS_UWI", "REVIEW", "ATTENTION"]
    _blank_uwi = df["MATCHED_UWI"].fillna("").astype(str).str.strip().eq("")
    _assignable = df[df["CATALOG_READINESS"].isin(_unresolved)
                     | (df["CATALOG_READINESS"].eq("CATALOGED") & _blank_uwi)]
    if not _assignable.empty:
        with st.expander(f"✏️ Assign UWI / Survey — {len(_assignable)} unresolved file(s)",
                         expanded=False):
            try:
                from dataview.core import path_identity as _pi
            except Exception:
                _pi = None

            _edit_src = _assignable[[
                "INVENTORY_ID", "FILE_NAME", "FILE_TYPE_GROUP",
                "CATALOG_READINESS", "MATCHED_UWI", "SURVEY_NAME"
            ]].copy() if "SURVEY_NAME" in _assignable.columns else \
                _assignable[["INVENTORY_ID", "FILE_NAME", "FILE_TYPE_GROUP",
                             "CATALOG_READINESS", "MATCHED_UWI"]].copy()
            if "SURVEY_NAME" not in _edit_src.columns:
                _edit_src["SURVEY_NAME"] = ""
            _edit_src["MATCHED_UWI"] = _edit_src["MATCHED_UWI"].fillna("").astype(str).str.strip()
            _edit_src["SURVEY_NAME"] = _edit_src["SURVEY_NAME"].fillna("").astype(str).str.strip()

            st.caption("Well data → set **assign UWI** (14 digits, dashes ok — normalized on "
                       "save). Seismic → set **assign SURVEY_NAME**. Blank rows are left "
                       "untouched. Saving flips readiness to READY. A row already reading "
                       "**CATALOGED** has rows staged: saving re-arms capture, which refreshes "
                       "them, and its detail promotes once that well exists in dv_well.")
            # Keep the editable column named MATCHED_UWI so it matches the field the
            # user knows; only SURVEY_NAME gets a friendlier label. No decoy read-only
            # UWI column: it's dropped from the summary table above (see st.dataframe).
            _grid = _edit_src.rename(columns={"SURVEY_NAME": "assign SURVEY_NAME"})
            # Lock the identity/status columns; leave ONLY 'assign UWI' and
            # 'assign SURVEY_NAME' editable. `disabled` as a column-name LIST is the
            # reliable way to do this — per-column disabled=True in column_config can
            # render the whole grid read-only when combined with renamed columns.
            edited = st.data_editor(
                _grid,
                key="wb_assign_editor",
                hide_index=True,
                use_container_width=True,
                num_rows="fixed",
                disabled=["INVENTORY_ID", "FILE_NAME", "FILE_TYPE_GROUP",
                          "CATALOG_READINESS"],
                column_config={
                    "MATCHED_UWI": st.column_config.TextColumn(
                        "MATCHED_UWI ✏️", help="14-char UWI; dashes/spaces stripped, padded to 14"),
                    "assign SURVEY_NAME": st.column_config.TextColumn(
                        "assign SURVEY_NAME", help="seismic survey name (nvarchar 255)"),
                },
            )

            if st.button("💾 Save assignments", type="primary", key="wb_assign_save"):
                # UWI + survey both go through the shared module-level helpers
                # (_assign_uwi / _assign_survey_name) — one owner each.
                n_uwi = n_srv = 0
                bad = []
                try:
                    with engine.begin() as con:
                        for _, r in edited.iterrows():
                            iid = str(r["INVENTORY_ID"])
                            raw_uwi = str(r.get("MATCHED_UWI", "") or "").strip()
                            raw_srv = str(r.get("assign SURVEY_NAME", "") or "").strip()
                            if raw_uwi:
                                u = _normalize_uwi14(raw_uwi)
                                if not u:
                                    bad.append(f"{r['FILE_NAME']}: '{raw_uwi}' → not 14 chars")
                                    continue
                                _assign_uwi(con, iid, u)
                                n_uwi += 1
                            elif raw_srv:
                                _assign_survey_name(con, iid, raw_srv)
                                n_srv += 1
                    msg = []
                    if n_uwi: msg.append(f"{n_uwi} UWI")
                    if n_srv: msg.append(f"{n_srv} survey name")
                    if msg:
                        st.success("Assigned " + " + ".join(msg) +
                                   " → readiness READY. Re-run capture to load them.")
                    if bad:
                        st.warning("Skipped (fix and re-save): " + "; ".join(bad))
                    st.session_state.pop("wb_results", None)
                    st.rerun()
                except Exception as e:
                    st.error(f"Assignment save failed: {e}")

    if df.empty:
        return

    st.divider()
    mode = st.radio(
        "Mode",
        ["🔎 Interactive — view & load", "📦 Batch load"],
        horizontal=True, key="wb_mode", label_visibility="collapsed",
    )
    st.divider()
    if mode.startswith("🔎"):
        _wb_nav(engine, dialect, df)
    else:
        _wb_batch(engine, dialect, df)


def _wb_nav(engine, dialect, df):
    """Prev/Next nav + file detail + viewer + extract + load."""
    from sqlalchemy import text as _t

    n       = len(df)
    idx_key = "wb_nav_idx"
    if idx_key not in st.session_state:
        st.session_state[idx_key] = 0
    idx = max(0, min(st.session_state[idx_key], n-1))

    # Nav bar
    c_prev, c_info, c_next, c_jump = st.columns([1,4,1,2])
    with c_prev:
        if st.button("◀ Prev", key="wb_prev", disabled=(idx==0)):
            st.session_state[idx_key] = idx-1
            st.rerun()
    with c_next:
        if st.button("Next ▶", key="wb_next", disabled=(idx>=n-1)):
            st.session_state[idx_key] = idx+1
            st.rerun()
    with c_info:
        row = df.iloc[idx]
        badge = " 🚩" if row["FLAG_DELETE"]=="Y" else ""
        st.markdown(f"**{idx+1} / {n}**{badge}  `{row['FILE_NAME']}`")
    with c_jump:
        names  = df["FILE_NAME"].tolist()
        jumped = st.selectbox("Jump to", names, index=idx,
                              key="wb_jump",
                              label_visibility="collapsed")
        ji = names.index(jumped)
        if ji != idx:
            st.session_state[idx_key] = ji
            st.rerun()

    row    = df.iloc[idx]
    fpath  = row["FILE_PATH"]
    fext   = row["FILE_EXT"].lower()
    inv_id = row["INVENTORY_ID"]

    # ── Action bar ────────────────────────────────────────────────────────────
    a1, a2, a3, a4 = st.columns(4)

    is_flagged = row["FLAG_DELETE"] == "Y"
    if a1.button(
            "🚩 Unflag" if is_flagged else "🚩 Flag",
            key="wb_flag_btn"):
        _toggle_flag(engine, inv_id, not is_flagged)
        st.session_state.pop("wb_results", None)
        st.rerun()

    if a2.button("🔄 Re-extract", key="wb_reenrich"):
        with st.spinner("Extracting..."):
            try:
                fields = _extract_fields(fpath, fext)
                _write_enrichment(engine, inv_id, fields)
                st.success("Re-extracted.")
                st.session_state.pop("wb_results", None)
                st.rerun()
            except Exception as e:
                st.error(f"Re-extract failed: {e}")

    # Download
    if Path(fpath).exists():
        try:
            a4.download_button(
                "⬇ Download", data=Path(fpath).read_bytes(),
                file_name=row["FILE_NAME"], key="wb_dl_btn")
        except Exception:
            pass

    # ── Mark as bad ───────────────────────────────────────────────────────────
    # Fingerprint a junk / badly-formatted file so the next crawl skips it and
    # drop it from the catalog now.
    if "wb_bad_reason" not in st.session_state:
        st.session_state["wb_bad_reason"] = "junk / bad format"
    bc1, bc2 = st.columns([3, 1])
    bc1.text_input("Reason", key="wb_bad_reason",
                   label_visibility="collapsed",
                   placeholder="why is this file bad?")
    if bc2.button("🚫 Mark bad", key="wb_bad_btn",
                  help="Fingerprint this file as bad, drop the rows it staged, "
                       "and retire its catalog row — the next crawl will skip it."):
        try:
            _dropped = _mark_bad(engine, inv_id, fpath, row.get("FILE_NAME"),
                                 row.get("FILE_SIZE_KB"),
                                 st.session_state.get("wb_bad_reason"))
            _n = sum(_dropped.values())
            st.success(
                f"Marked bad — the next crawl will skip "
                f"`{row.get('FILE_NAME', '')}`."
                + (f" Dropped {_n:,} staged row(s) from {len(_dropped)} "
                   f"mirror(s): {', '.join(sorted(_dropped))}." if _n else ""))
            st.session_state.pop("wb_results", None)
            st.rerun()
        except Exception as e:
            st.error(f"Mark bad failed: {e}")

    st.divider()

    # ── Catalog attributes from header tables ─────────────────────────────────
    _show_header_attrs(engine, inv_id, row)

    # ── Universal viewer ──────────────────────────────────────────────────────
    if Path(fpath).exists():
        try:
            from dataview.file_catalog.file_viewer import view as _view
            _view(fpath, fext)
        except Exception as e:
            st.error(f"Viewer error: {e}")
    else:
        st.warning(f"File not found on disk: `{fpath}`")

    # ── Extract & Load (automatic) ───────────────────────────────────────────
    _extract_and_load(engine, dialect, fpath, fext, inv_id, row)


def _show_header_attrs(engine, inv_id: str, row):
    """Show enriched header from FILE_WELL_HEADER or FILE_SEIS_HEADER."""
    from sqlalchemy import text as _t

    grp = row.get("FILE_TYPE_GROUP","")
    is_seis = grp in ("Seismic","Shapefile")

    try:
        with engine.connect() as con:
            if is_seis:
                r = con.execute(_t("""
                    SELECT SURVEY_NAME, LINE_NAME, SEIS_SET_TYPE,
                           SURVEY_DATE, CONTRACTOR,
                           BBOX_MIN_LAT, BBOX_MAX_LAT,
                           BBOX_MIN_LON, BBOX_MAX_LON,
                           EPSG_CODE, SAMPLE_INTERVAL, TRACE_COUNT,
                           SHOT_FIRST, SHOT_LAST
                    FROM file_catalog.FILE_SEIS_HEADER
                    WHERE INVENTORY_ID=:id
                """), {"id": inv_id}).fetchone()
                if r:
                    attrs = dict(zip([
                        "Survey Name","Line Name","Set Type","Survey Date",
                        "Contractor","Min Lat","Max Lat","Min Lon","Max Lon",
                        "EPSG","Sample Interval","Trace Count",
                        "Shot First","Shot Last",
                    ], r))
                else:
                    attrs = {}
            else:
                r = con.execute(_t("""
                    SELECT UWI, WELL_NAME, OPERATOR, WELL_FIELD,
                           STATE, COUNTY, LATITUDE, LONGITUDE,
                           TOTAL_DEPTH, SPUD_DATE, RIG_RELEASE,
                           REPORT_TYPE, SURVEY_TYPE, CONTRACTOR, CONFIDENCE
                    FROM file_catalog.FILE_WELL_HEADER
                    WHERE INVENTORY_ID=:id
                """), {"id": inv_id}).fetchone()
                if r:
                    attrs = dict(zip([
                        "UWI","Well Name","Operator","Field",
                        "State","County","Latitude","Longitude",
                        "Total Depth","Spud Date","Rig Release",
                        "Report Type","Survey Type","Contractor","Confidence",
                    ], r))
                else:
                    attrs = {}

        if attrs:
            with st.expander("📋 Extracted header", expanded=True):
                adf = pd.DataFrame(
                    [{"Field": k, "Value": str(v)}
                     for k, v in attrs.items()
                     if v is not None and str(v).strip()
                     not in ("","None","nan")]
                )
                if not adf.empty:
                    st.dataframe(adf, hide_index=True,
                                 use_container_width=True)
        else:
            st.caption("No header extracted yet — click Re-extract.")
    except Exception as e:
        st.caption(f"Header lookup: {e}")


def _extract_and_load(engine, dialect, fpath, fext, inv_id, row):
    """Extract structured data rows from file and offer DB load."""
    from sqlalchemy import text as _t

    st.markdown("#### 📐 Extracted Data")

    rows, label = _do_extract(fpath, fext)

    if not rows:
        st.info(f"No structured data extracted for {fext} files.")
        return

    df = pd.DataFrame(rows).fillna("")
    st.metric(f"{label} extracted", len(df))
    st.dataframe(df, hide_index=True, use_container_width=True)

    c1, c2 = st.columns(2)
    c1.download_button(
        f"⬇ Download {label} CSV",
        data=df.to_csv(index=False),
        file_name=f"{Path(fpath).stem}_{label.lower().replace(' ','_')}.csv",
        mime="text/csv",
        key="wb_extract_dl",
    )

    # Load to DB — check well exists first
    uwi = row.get("MATCHED_UWI","") or ""
    if not uwi:
        c2.warning("No UWI — load wells from Header Files tab first.")
        return

    # Capture writes to the file_catalog.cat_* mirrors and does NOT require a
    # dv_well header to exist — the header is created later by promote_catalog
    # from cat_well. Key everything off the document UWI.
    _resolved = uwi

    if c2.button(
            f"🚀 Capture {len(rows)} records to catalog",
            type="primary", key="wb_load_btn"):
        _do_load(engine, dialect, fpath, fext, _resolved, rows)






def _do_load(engine, dialect, fpath, fext, uwi, rows):
    """Interactive single-file capture — renders Streamlit messages."""
    res = _load_rows_to_catalog(engine, dialect, fpath, fext, uwi, rows)

    for e in res["errors"]:
        if str(e).startswith("header capture:"):
            st.warning(f"Header capture skipped: "
                       f"{str(e).split(':', 1)[1].strip()}")
    real_errs = [e for e in res["errors"]
                 if not str(e).startswith("header capture:")]

    note = res.get("note", "")
    if note.startswith("petro_fail:"):
        st.error(f"Extraction failed: {note.split(':', 1)[1]}")
        return
    if note.startswith("not_impl:"):
        st.warning(f"Load not implemented for {note.split(':', 1)[1]}")
        return
    if note == "shapefile":
        st.success("✅ Shapefile loaded")
        return
    if note == "unsupported":
        st.info("Direct DB load not yet implemented for this file type.")
        return

    if real_errs:
        st.error(f"Load errors: {'; '.join(str(e) for e in real_errs[:3])}")
    else:
        st.success(f"✅ Loaded {res['loaded']} records")


def _wb_batch(engine, dialect, df):
    """Batch mode: checkbox table of files → capture all checked at once."""
    import pandas as _pd

    st.markdown("##### 📦 Batch load to catalog")
    st.caption("Check the files to capture, then load them all into the "
               "`cat_*` mirrors in one pass. A file needs a matched UWI to "
               "be captured.")

    n = len(df)
    _has_uwi = (df["MATCHED_UWI"].fillna("").astype(str).str.strip() != "")
    n_uwi = int(_has_uwi.sum())

    # Select-ALL checkbox above the grid (matches the spreadsheet metaphor
    # used elsewhere in this page). Only flips rows that have a UWI.
    _sel_all = st.checkbox(
        f"☑ Select ALL with a UWI  ({n_uwi} of {n} files loadable)",
        key="wb_batch_sel_all",
        help="Marks every file that has a matched UWI. Uncheck rows in the "
             "grid to exclude them.",
    )

    st.markdown("---")

    _tbl = _pd.DataFrame({
        "Load":      [bool(_sel_all and u) for u in _has_uwi],
        "Bad":       [False] * n,
        "File":      df["FILE_NAME"].values,
        "UWI":       df["MATCHED_UWI"].fillna("").astype(str).values,
        "Well name": df.get("WELL_NAME",
                            _pd.Series([""] * n)).fillna("").astype(str).values,
        "Ext":       df["FILE_EXT"].astype(str).values,
        "_inv":      df["INVENTORY_ID"].values,
        "_path":     df["FILE_PATH"].values,
        "_size":     df.get("FILE_SIZE_KB",
                            _pd.Series([None] * n)).values,
    })

    _edited = st.data_editor(
        _tbl,
        use_container_width=True,
        hide_index=True,
        disabled=["File", "UWI", "Well name", "Ext"],
        column_config={
            "Load": st.column_config.CheckboxColumn("Load", width="small"),
            "Bad": st.column_config.CheckboxColumn(
                "Bad", width="small",
                help="Fingerprint as junk / badly formatted — removed from the "
                     "catalog now and skipped on the next crawl."),
            "File": st.column_config.TextColumn("File"),
            "UWI": st.column_config.TextColumn("UWI", width="medium"),
            "Well name": st.column_config.TextColumn("Well name"),
            "Ext": st.column_config.TextColumn("Ext", width="small"),
            "_inv": None,
            "_path": None,
            "_size": None,
        },
        # Nonced: st.data_editor remembers ticks by row index, so after a load
        # or a reject the boxes stayed checked against rows that had moved or
        # gone. Bumping the nonce mints a fresh editor with everything cleared.
        # (It also can't be reset by writing to the key — Streamlit forbids
        # assigning to a widget key once the widget exists.)
        key=f"wb_batch_editor_{st.session_state.get('wb_batch_nonce', 0)}",
    )

    _checked = _edited[_edited["Load"] == True].copy()
    _checked["UWI"] = _checked["UWI"].fillna("").astype(str).str.strip()
    _loadable = _checked[_checked["UWI"] != ""]
    _skip = len(_checked) - len(_loadable)

    msg = f"**{len(_loadable)}** file(s) ready to capture"
    if _skip:
        msg += f" · {_skip} checked without UWI (skipped)"
    st.caption(msg)

    if st.button(
            f"🚀 Load {len(_loadable)} checked to catalog",
            type="primary", key="wb_batch_go",
            disabled=_loadable.empty):
        _run_batch_load(engine, dialect, _loadable)
        # Capture changes readiness, so the list on screen is now stale and the
        # ticks refer to work already done. Re-query and retire the editor —
        # the per-file report is held in session state and survives the rerun.
        st.session_state["wb_batch_nonce"] = (
            st.session_state.get("wb_batch_nonce", 0) + 1)
        _wb_run_search(engine)
        st.rerun()

    # ── Mark checked files as bad ─────────────────────────────────────────────
    # Sits directly under the Load button because both act on the SAME grid
    # ticks — it used to render below the results report and the Excel
    # expander, so ticking Bad appeared to do nothing until you scrolled
    # past two other sections to find its button.
    # Fingerprint the rows ticked "Bad" so the next crawl skips them, and drop
    # them from the catalog now (shares _mark_bad with the interactive viewer).
    _bad = _edited[_edited["Bad"] == True].copy()
    if not _bad.empty:
        if "wb_batch_bad_reason" not in st.session_state:
            st.session_state["wb_batch_bad_reason"] = "junk / bad format"
        _bnonce = st.session_state.get("wb_batch_nonce", 0)
        rcol, ccol, bcol = st.columns([3, 1, 1])
        rcol.text_input("Reason", key="wb_batch_bad_reason",
                        label_visibility="collapsed",
                        placeholder="why are these files bad?")
        # Confirm gate: this deletes catalog rows and blocklists the files, and
        # the batch grid can hold 500 of them. Nonced with the editor so it
        # resets after each action without an illegal write to a widget key.
        _bconfirm = ccol.checkbox("Confirm", key=f"wb_batch_bad_confirm_{_bnonce}",
                                  help="Rejecting removes these files from the "
                                       "catalog. Tick to enable.")
        if bcol.button(f"🚫 Mark {len(_bad)} bad", key="wb_batch_bad_btn",
                       disabled=not _bconfirm,
                       help="Fingerprint these files as bad and remove them "
                            "from the catalog — the next crawl will skip them."):
            reason = st.session_state.get("wb_batch_bad_reason")
            done, errs = 0, []
            for _, br in _bad.iterrows():
                try:
                    _mark_bad(engine, br["_inv"], br["_path"],
                              br.get("File"), br.get("_size"), reason)
                    done += 1
                except Exception as e:
                    errs.append(f"{br.get('File', '')}: {e}")
            if done:
                st.success(f"Marked {done} file(s) bad — the next crawl "
                           "will skip them.")
            if errs:
                st.error("Some files could not be marked:\n- "
                         + "\n- ".join(errs[:10]))
            # Re-query rather than dropping the result set: the rejected rows
            # are gone from the catalog so they vanish from the list, and
            # everything else stays on screen instead of sending you back to
            # the search form.
            st.session_state["wb_batch_nonce"] = _bnonce + 1
            _wb_run_search(engine)
            st.rerun()

    # Per-file results report (persists across reruns until cleared / next load).
    _render_batch_report()

    # ── Export / Import batch selection (Excel) ───────────────────────────────
    # Round-trip the batch candidate list through Excel: export every file in the
    # current grid with Load / Bad flag columns (plus its UWI), curate offline,
    # then import to drive the batch. Bad-flagged rows are marked bad + dropped;
    # Load-flagged rows that carry a UWI are captured into cat_* via the SAME
    # _run_batch_load path the button uses, so the per-file results report above
    # is identical to a manual run. INVENTORY_ID / UWI export as TEXT so Excel
    # can't reformat a 14-digit API into scientific notation.
    with st.expander("📤 Export / 📥 Import batch selection (Excel)", expanded=False):
        _bexp = _pd.DataFrame({
            "INVENTORY_ID": df["INVENTORY_ID"].astype(str).values,
            "FILE_NAME":    df["FILE_NAME"].astype(str).values,
            "FILE_EXT":     df["FILE_EXT"].astype(str).values,
            "WELL_NAME":    df.get("WELL_NAME", _pd.Series([""] * n))
                              .fillna("").astype(str).values,
            "MATCHED_UWI":  df["MATCHED_UWI"].fillna("").astype(str).values,
            "FILE_PATH":    df["FILE_PATH"].astype(str).values,
            "FILE_SIZE_KB": df.get("FILE_SIZE_KB", _pd.Series([""] * n))
                              .fillna("").astype(str).values,
            # Pre-fill Load for rows that already have a UWI (mirrors "Select ALL
            # with a UWI"); Bad starts blank — tick the ones to discard.
            "Load": ["Y" if str(u).strip() else ""
                     for u in df["MATCHED_UWI"].fillna("")],
            "Bad":  [""] * n,
        })
        try:
            import io
            from openpyxl.utils import get_column_letter
            _bbuf = io.BytesIO()
            with _pd.ExcelWriter(_bbuf, engine="openpyxl") as _bxw:
                _bexp.to_excel(_bxw, index=False, sheet_name="batch")
                _bws = _bxw.sheets["batch"]
                _btext = {"INVENTORY_ID", "MATCHED_UWI"}
                for _bi, _bname in enumerate(_bexp.columns, start=1):
                    _bl = get_column_letter(_bi)
                    _bw = _bexp[_bname].str.len().max()
                    _bw = 12 if _pd.isna(_bw) else int(_bw)
                    _bws.column_dimensions[_bl].width = min(max(10, _bw + 2), 60)
                    if _bname in _btext:
                        for _bc in _bws[_bl]:
                            _bc.number_format = "@"
                # make the FILE_NAME cell a clickable link to the file (target
                # built from FILE_PATH; the path column stays plain text).
                _bcl = list(_bexp.columns)
                if "FILE_NAME" in _bcl and "FILE_PATH" in _bcl:
                    from openpyxl.styles import Font as _Font
                    _bnlet = get_column_letter(_bcl.index("FILE_NAME") + 1)
                    _bplet = get_column_letter(_bcl.index("FILE_PATH") + 1)
                    _blfont = _Font(color="0563C1", underline="single")
                    for _bnc, _bpc in zip(_bws[_bnlet][1:], _bws[_bplet][1:]):
                        _bu = _xl_file_uri(_bpc.value)
                        if _bu:
                            _bnc.hyperlink = _bu
                            _bnc.font = _blfont
                _bws.freeze_panes = "A2"
            st.download_button(
                "📤 Export batch grid to Excel",
                data=_bbuf.getvalue(),
                file_name=f"batch_selection_{datetime.now():%Y%m%d_%H%M}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
                use_container_width=True,
                help="Set Load = Y to capture a file, Bad = Y to discard it "
                     "(Bad wins if both). Load needs a UWI in that row. Don't "
                     "edit INVENTORY_ID — it's how import finds the file.")
        except Exception as _bxx:
            st.error(f"Excel export unavailable ({type(_bxx).__name__}: {_bxx}). "
                     "Needs openpyxl — `pip install openpyxl`.")

        st.divider()

        _bup = st.file_uploader("Import curated batch Excel", type=["xlsx"],
                                key="wb_batch_xlsx",
                                help="Applies Bad = Y (mark bad + drop from "
                                     "catalog) and Load = Y (capture into cat_*) "
                                     "per row, matched by INVENTORY_ID.")
        st.text_input("Reason for Bad rows", value="junk / bad format",
                      key="wb_batch_xlsx_reason")
        if _bup is not None and st.button("📥 Run batch from file",
                                          type="primary", key="wb_batch_import"):
            try:
                _bimp = _pd.read_excel(_bup, dtype=str)
            except Exception as _brx:
                st.error(f"Could not read workbook: {_brx}")
                _bimp = None
            if _bimp is not None:
                _bimp.columns = [str(_c).strip() for _c in _bimp.columns]
                _need = {"INVENTORY_ID", "FILE_PATH", "FILE_NAME", "FILE_EXT"}
                _missing = _need - set(_bimp.columns)
                if _missing:
                    st.error("Workbook missing column(s): "
                             + ", ".join(sorted(_missing))
                             + " — export a fresh template above.")
                else:
                    for _c in ("MATCHED_UWI", "Load", "Bad", "FILE_SIZE_KB"):
                        if _c not in _bimp.columns:
                            _bimp[_c] = ""
                    _bimp = _bimp.fillna("")

                    def _yes(v):    # tolerant truthy for the flag columns
                        return str(v or "").strip().lower() in (
                            "y", "yes", "true", "1", "x", "✓")

                    _reason = st.session_state.get("wb_batch_xlsx_reason") \
                        or "junk / bad format"

                    # Bad first — and a Bad row is never also loaded.
                    _bad_rows = _bimp[_bimp["Bad"].map(_yes)]
                    _bad_ids  = set(_bad_rows["INVENTORY_ID"].astype(str))
                    n_bad, bad_errs = 0, []
                    for _, br in _bad_rows.iterrows():
                        try:
                            _mark_bad(engine, str(br["INVENTORY_ID"]),
                                      br["FILE_PATH"], br.get("FILE_NAME"),
                                      _safe_int(br.get("FILE_SIZE_KB")), _reason)
                            n_bad += 1
                        except Exception as _e:
                            bad_errs.append(f"{br.get('FILE_NAME', '')}: {_e}")

                    # Load rows with a UWI → feed the same _run_batch_load path.
                    _lr = _bimp[_bimp["Load"].map(_yes)].copy()
                    _lr = _lr[~_lr["INVENTORY_ID"].astype(str).isin(_bad_ids)]
                    _lr["UWI"] = _lr["MATCHED_UWI"].astype(str).str.strip()
                    _skip_no_uwi = int((_lr["UWI"] == "").sum())
                    _loadable = _lr[_lr["UWI"] != ""]

                    if n_bad:
                        st.success(f"Marked {n_bad} file(s) bad.")
                    if bad_errs:
                        st.error("Some Bad rows failed:\n- "
                                 + "\n- ".join(bad_errs[:10]))
                    if _skip_no_uwi:
                        st.caption(f"{_skip_no_uwi} Load row(s) skipped — no UWI.")

                    if not _loadable.empty:
                        _feed = _pd.DataFrame({
                            "_path": _loadable["FILE_PATH"].values,
                            "Ext":   _loadable["FILE_EXT"].values,
                            "UWI":   _loadable["UWI"].values,
                            "File":  _loadable["FILE_NAME"].values,
                            "_inv":  _loadable["INVENTORY_ID"].astype(str).values,
                        })
                        _run_batch_load(engine, dialect, _feed)   # sets the report
                        # same post-action refresh as the manual button
                        st.session_state["wb_batch_nonce"] = (
                            st.session_state.get("wb_batch_nonce", 0) + 1)
                        _wb_run_search(engine)
                        st.rerun()      # re-enter so _render_batch_report shows it
                    elif n_bad:
                        st.session_state["wb_batch_nonce"] = (
                            st.session_state.get("wb_batch_nonce", 0) + 1)
                        _wb_run_search(engine)
                        st.rerun()      # bad-only: refresh the grid like manual path
                    else:
                        st.info("Nothing to do — no Load=Y (with UWI) or "
                                "Bad=Y rows in the sheet.")

    # Second table: resolve a UWI for the files in these results that don't
    # have one yet — guess + match against the reference well master.
    _wb_resolve_unmatched(engine, df)


def _fmt_detail(d):
    """Format a {cat_table: count} map as a compact 'tops×12, dst×3' string."""
    if not d:
        return ""
    return ", ".join(f"{k.replace('cat_', '')}×{v}"
                     for k, v in d.items() if v)


def _run_batch_load(engine, dialect, loadable):
    """Capture each checked file into the cat_* mirrors, then build a per-file
    results report (status, record count, and which mirror tables got rows) and
    stash it in session so it survives reruns."""
    import pandas as _pd

    total = len(loadable)
    prog  = st.progress(0.0, text="Starting…")
    results = []
    ok_invs = set()   # inv ids that captured rows → mark CATALOGED in the grid

    for i, (_, r) in enumerate(loadable.iterrows(), start=1):
        fpath = r["_path"]
        fext  = str(r["Ext"]).lower()
        uwi   = str(r["UWI"]).strip()
        fname = r["File"]
        prog.progress(i / total, text=f"{i}/{total} · {fname}")

        try:
            rows, label = _do_extract(fpath, fext)
            if not rows and fext not in SELF_PARSING_EXTS:
                results.append({"File": fname, "UWI": uwi, "Type": "—",
                                "Status": "— no data", "Records": 0,
                                "Captured into": ""})
                continue
            res = _load_rows_to_catalog(engine, dialect, fpath, fext, uwi,
                                        rows or [])
            real_errs = [e for e in res["errors"]
                         if not str(e).startswith("header capture:")]
            hdr_errs = [e for e in res["errors"]
                        if str(e).startswith("header capture:")]
            note = res.get("note", "")
            if real_errs:
                status = f"❌ {str(real_errs[0])[:60]}"
            elif note.startswith("petro_fail:"):
                status = f"❌ extract: {note.split(':', 1)[1][:50]}"
            elif note.startswith("not_impl:"):
                status = f"⚠ no loader ({note.split(':', 1)[1]})"
            elif note == "unsupported":
                status = "⚠ unsupported type"
            elif note == "shapefile":
                status = "✅ shapefile"
            elif res["loaded"]:
                status = "✅ loaded"
                if hdr_errs:
                    status += " (header skipped)"
            else:
                status = "— nothing captured"
            if status.startswith("✅"):
                ok_invs.add(r["_inv"])
            results.append({"File": fname, "UWI": uwi,
                            "Type": res.get("rt") or label,
                            "Status": status,
                            "Records": res.get("loaded", 0),
                            "Captured into": _fmt_detail(res.get("detail"))})
        except Exception as e:
            results.append({"File": fname, "UWI": uwi, "Type": "—",
                            "Status": f"❌ {str(e)[:60]}", "Records": 0,
                            "Captured into": ""})

    prog.progress(1.0, text="Done")
    rdf = _pd.DataFrame(results, columns=[
        "File", "UWI", "Type", "Status", "Records", "Captured into"])
    n_ok = int(rdf["Status"].str.startswith("✅").sum()) if not rdf.empty else 0

    # Keep the cached search grid in sync: loaded files now read CATALOGED
    # (the DB rows were updated by _set_readiness_cataloged during capture).
    if ok_invs:
        _gdf = st.session_state.get("wb_results")
        if _gdf is not None and "CATALOG_READINESS" in _gdf.columns:
            _gdf.loc[_gdf["INVENTORY_ID"].isin(ok_invs),
                     "CATALOG_READINESS"] = "CATALOGED"

    # Persist so the report survives the reruns triggered by other widgets on
    # this page (resolver, mark-bad). Rendered by _render_batch_report().
    st.session_state["wb_batch_report"] = {
        "rdf": rdf, "n_ok": n_ok, "total": total,
        "tot_rec": int(rdf["Records"].sum()) if not rdf.empty else 0,
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def _render_batch_report():
    """Render the persisted batch results report with a CSV download."""
    rep = st.session_state.get("wb_batch_report")
    if not rep:
        return
    rdf = rep["rdf"]
    st.divider()
    rc1, rc2 = st.columns([5, 1])
    rc1.success(f"Batch complete · {rep['n_ok']}/{rep['total']} file(s) "
                f"captured · {rep['tot_rec']:,} record(s)  ·  {rep['ts']}")
    if rc2.button("✖ Clear", key="wb_report_clear"):
        st.session_state.pop("wb_batch_report", None)
        st.rerun()
    st.dataframe(rdf, hide_index=True, use_container_width=True)
    st.download_button(
        "⬇ Download report CSV",
        data=rdf.to_csv(index=False).encode("utf-8"),
        file_name=f"batch_load_report_{rep['ts'].replace(' ', '_').replace(':', '')}.csv",
        mime="text/csv", key="wb_report_dl")
    st.caption("Then run `py promote_catalog.py --apply` to create headers "
               "and promote the captured rows into the `dv_*` tables.")


# Reference well master used to resolve a UWI from a name/UWI guess.
# Default is the production master; the pipeline UI can override it (e.g. point
# at a mini master for a demo) via the "wb_ref_choice" session_state key.
_WB_REF_DEFAULT = "WELL_REF.well_ref.well_master_gold"
_WB_REF_OPTIONS = {
    "WELL_REF.well_ref.well_master_gold":      "Full master (production)",
    "WELL_REF.well_ref.WELL_MASTER_MINI": "Mini master (demo / test)",
}


# Columns the reference lookups want. The two masters don't carry the same set —
# the production gold table has no UWI or API_NUM — and a hardcoded SELECT fails
# the whole lookup with "Invalid column name" rather than degrading. So the list
# is reflected once per table and any column that isn't there is selected as
# NULL under its own name, which keeps every downstream m.<COL> access working
# without a single conditional at the call sites.
_WB_REF_WANT = ("WELL_NAME", "UWI14", "UWI", "API_NUM", "OPERATOR_NAME",
                "TOTAL_DEPTH", "SPUD_DATE", "COUNTY", "PROVINCE_STATE")

# WHERE clauses key on these, so their absence isn't recoverable by aliasing.
_WB_REF_REQUIRED = ("WELL_NAME", "UWI14")


def _wb_ref_columns(engine, ref: str) -> set:
    """Actual column names on a db.schema.table reference master, cached.

    Reads the REFERENCED database's own sys.* (three-part name), so it works
    across databases on the same instance without a linked server.
    """
    _cache = st.session_state.setdefault("_wb_ref_cols", {})
    if ref in _cache:
        return _cache[ref]
    from sqlalchemy import text as _t
    parts = [p.strip("[]") for p in str(ref).split(".")]
    if len(parts) == 3:
        db, sch, tbl = parts
    elif len(parts) == 2:
        db, sch, tbl = None, parts[0], parts[1]
    else:
        db, sch, tbl = None, "dbo", parts[-1]
    pfx = f"[{db}]." if db else ""
    cols = set()
    try:
        with engine.connect() as con:
            rows = con.execute(_t(
                f"SELECT c.name FROM {pfx}sys.columns c "
                f"JOIN {pfx}sys.objects o ON o.object_id = c.object_id "
                f"JOIN {pfx}sys.schemas s ON s.schema_id = o.schema_id "
                f"WHERE s.name = :s AND o.name = :t"),
                {"s": sch, "t": tbl}).fetchall()
        cols = {r[0].upper() for r in rows}
    except Exception:
        cols = set()                     # unreadable → caller falls back
    _cache[ref] = cols
    return cols


def _wb_ref_select(engine, ref: str) -> str:
    """SELECT list for `ref`: real columns as themselves, absent ones as NULL.

    Falls back to the full list when reflection fails, so a permissions problem
    on sys.* behaves exactly as before rather than silently nulling everything.
    """
    have = _wb_ref_columns(engine, ref)
    if not have:
        return ", ".join(_WB_REF_WANT)
    return ", ".join(c if c in have else f"CAST(NULL AS NVARCHAR(1)) AS {c}"
                     for c in _WB_REF_WANT)


def _wb_ref_missing_required(engine, ref: str):
    """Required columns the reference master doesn't have — [] when fine."""
    have = _wb_ref_columns(engine, ref)
    if not have:
        return []
    return [c for c in _WB_REF_REQUIRED if c not in have]


def _wb_ref() -> str:
    """The reference table enrich/triage resolve against. Overridable from the
    pipeline UI dropdown; falls back to the production master."""
    try:
        v = st.session_state.get("wb_ref_choice")
        if v and str(v).strip():
            return str(v).strip()
    except Exception:
        pass
    return _WB_REF_DEFAULT


def _wb_resolve_unmatched(engine, df):
    """Second table for the batch flow: the files in the current results that
    have NO matched UWI. Shows an editable well-name + UWI guess, matches each
    against the reference well master, and (on Save) writes the confirmed UWI
    back to MATCHED_UWI + the header so the file becomes loadable."""
    from sqlalchemy import text as _t
    try:
        from dataview.core import path_identity as _pi
    except Exception:
        _pi = None

    nomatch = df[df["MATCHED_UWI"].fillna("").astype(str).str.strip() == ""]
    if nomatch.empty:
        return   # nothing to resolve — stay quiet

    invs = [str(i) for i in nomatch["INVENTORY_ID"].tolist()][:_REVIEW_PAGE]

    # Pull header guesses (well name, internal UWI, corroborating attrs).
    g_by_inv = {}
    try:
        with engine.connect() as con:
            _ph = ",".join(f":i{n}" for n in range(len(invs)))
            q = _t(f"""
                SELECT g.INVENTORY_ID AS inv, g.FILE_NAME AS fname,
                       g.FILE_PATH AS path,
                       h.WELL_NAME AS wn, h.UWI AS uwi, h.OPERATOR AS oper,
                       h.TOTAL_DEPTH AS td, h.SPUD_DATE AS spud
                FROM file_catalog.GLOBAL_FILE_CATALOG g
                LEFT JOIN file_catalog.FILE_WELL_HEADER h
                       ON h.INVENTORY_ID = g.INVENTORY_ID
                WHERE g.INVENTORY_ID IN ({_ph})""")
            for r in con.execute(
                    q, {f"i{n}": v for n, v in enumerate(invs)}).fetchall():
                g_by_inv[str(r.inv)] = r
    except Exception as e:
        st.divider()
        st.error(f"Could not read header guesses: {e}")
        return

    st.divider()
    st.markdown("##### 🧭 Unmatched files — guess & match a UWI")
    st.caption(f"{len(nomatch)} file(s) here have no matched UWI"
               + (f" — showing the first {_REVIEW_PAGE}"
                  if len(nomatch) > _REVIEW_PAGE else "")
               + ". Well name and UWI are editable guesses; **Match against "
                 "reference** resolves them, then **Save** writes the UWI back "
                 "so the file can be captured.")

    # A confirmed UWI from a prior Match is promoted into the UWI field before
    # the widgets are built (Streamlit forbids setting widget state afterward).
    _pending = st.session_state.pop("_wb_um_fill", {})

    h1, h2, h3, h4 = st.columns([3, 2, 2, 3])
    h1.markdown("**File**")
    h2.markdown("**Well name (guess)**")
    h3.markdown("**UWI (guess)**")
    h4.markdown("**Match result**")

    inputs = []   # (inv, nkey, ukey, rkey, guess_row)
    for inv in invs:
        g = g_by_inv.get(inv)
        fname = (g.fname if g else None) or f"(inventory {inv})"
        nkey, ukey, rkey = f"wb_um_n_{inv}", f"wb_um_u_{inv}", f"wb_um_r_{inv}"
        if nkey not in st.session_state:
            st.session_state[nkey] = ((g.wn if g else "") or "").strip()
        if ukey not in st.session_state:
            guess = ""
            if g and _pi:
                guess = (_pi.norm_uwi14(g.uwi or "")
                         or (_pi.uwi14_from_path(g.path or "")[0]
                             if g.path else "") or "")
            elif g:
                guess = str(g.uwi or "").strip()
            st.session_state[ukey] = guess
        if inv in _pending:
            st.session_state[ukey] = _pending[inv]
        c1, c2, c3, c4 = st.columns([3, 2, 2, 3])
        c1.write(fname)
        c2.text_input("n", key=nkey, label_visibility="collapsed",
                      placeholder="well name")
        c3.text_input("u", key=ukey, label_visibility="collapsed",
                      placeholder="UWI / API")
        c4.write(st.session_state.get(rkey, "—"))
        inputs.append((inv, nkey, ukey, rkey, g))

    b1, b2 = st.columns(2)
    if b1.button("🔍 Match against reference", key="wb_um_match",
                 use_container_width=True):
        _wb_um_match(engine, inputs, _pi)
    if b2.button("💾 Save matched UWIs", type="primary", key="wb_um_save",
                 use_container_width=True):
        _wb_um_save(engine, inputs, _pi)


def _wb_um_match(engine, inputs, _pi):
    """Look each guess up in the reference well master (by UWI, else by name
    corroborated with TD / spud / operator) and write the result + a fill."""
    from sqlalchemy import text as _t, bindparam

    def _nu(u):
        return _pi.norm_uwi14(u) if _pi else str(u or "").strip()

    def _norm_op(s):
        return "".join(ch for ch in (s or "").upper() if ch.isalnum())

    def _td_close(a, b, tol=50.0):
        try:
            a, b = float(a), float(b)
        except (TypeError, ValueError):
            return False
        return a > 0 and b > 0 and abs(a - b) <= max(tol, 0.01 * max(a, b))

    def _same_spud(a, b):
        if not a or not b:
            return False
        try:
            da = a.date() if hasattr(a, "date") else a
            db = b.date() if hasattr(b, "date") else b
            return da == db
        except Exception:
            return False

    cur = [(inv, nkey, ukey, rkey, g,
            st.session_state.get(nkey, "").strip(),
            _nu(st.session_state.get(ukey, "")))
           for inv, nkey, ukey, rkey, g in inputs]
    names = sorted({nm for *_x, nm, _u in cur if nm})
    uwis  = sorted({u for *_x, _nm, u in cur if u})

    _missing = _wb_ref_missing_required(engine, _wb_ref())
    if _missing:
        st.error(f"Reference master `{_wb_ref()}` has no "
                 + ", ".join(_missing)
                 + " column — name/UWI matching can't run against it.")
        return
    cols = _wb_ref_select(engine, _wb_ref())
    name_map, uwi_map = {}, {}
    try:
        with engine.connect() as con:
            if names:
                qn = _t(f"SELECT {cols} FROM {_wb_ref()} WHERE WELL_NAME IN :v"
                        ).bindparams(bindparam("v", expanding=True))
                for m in con.execute(qn, {"v": names}).fetchall():
                    name_map.setdefault((m.WELL_NAME or "").strip().upper(),
                                        []).append(m)
            if uwis:
                qu = _t(f"SELECT {cols} FROM {_wb_ref()} WHERE UWI14 IN :v"
                        ).bindparams(bindparam("v", expanding=True))
                for m in con.execute(qu, {"v": uwis}).fetchall():
                    uwi_map[(m.UWI14 or "").strip()] = m
    except Exception as e:
        st.error(f"Reference lookup failed: {type(e).__name__}: {e}")
        return

    fills = {}
    for inv, nkey, ukey, rkey, g, nm, uw in cur:
        txt, fill = "— enter a name or UWI", ""
        if uw and uw in uwi_map:
            m = uwi_map[uw]
            loc = " · ".join(x for x in (m.COUNTY, m.PROVINCE_STATE) if x)
            txt = (f"✓ {uw} · in reference · {m.WELL_NAME or ''}"
                   + (f" ({loc})" if loc else ""))
            fill = uw
        elif uw:
            # THE FILE'S OWN UWI IS STILL AN ANSWER. The reference master
            # is one corpus, not the world -- a LAS whose header or name
            # carries a good API is not wrong because well_master_gold has
            # never heard of it. Reporting only "not found" left the guess
            # on screen, unusable, and Save wrote nothing, so files with a
            # perfectly good UWI could not be keyed at all.
            #
            # So the result IS the UWI, and it fills. The tick is reserved
            # for corroborated ones: unverified and verified must stay
            # tellable apart, which is the whole reason to show the mark.
            txt = f"● {uw} · not in reference — using the file's own"
            fill = uw
        elif nm:
            cands = name_map.get(nm.upper(), [])
            scored = []
            for m in cands:
                sigs = []
                if g and _td_close(g.td, m.TOTAL_DEPTH):
                    sigs.append("TD")
                if g and _same_spud(g.spud, m.SPUD_DATE):
                    sigs.append("spud")
                if g and g.oper and _norm_op(g.oper) == _norm_op(m.OPERATOR_NAME):
                    sigs.append("oper")
                scored.append((len(sigs), sigs, m))
            scored.sort(key=lambda x: x[0], reverse=True)
            if not scored:
                txt = "✗ name not found in reference"
            else:
                top_n, top_sigs, top_m = scored[0]
                tie = len(scored) > 1 and scored[1][0] == top_n
                u14 = top_m.UWI14 or top_m.UWI or top_m.API_NUM
                loc = " · ".join(x for x in (top_m.COUNTY, top_m.PROVINCE_STATE) if x)
                if top_n >= 1 and not tie:
                    txt = (f"✓ {u14} · name+{'+'.join(top_sigs)}"
                           + (f" ({loc})" if loc else ""))
                    fill = (_pi.norm_uwi14(u14 or "") if _pi else str(u14 or ""))
                elif top_n >= 1 and tie:
                    txt = f"⚠ {len(scored)} wells match name+attrs — verify"
                elif len(cands) == 1:
                    txt = (f"? {u14} · name only — verify"
                           + (f" ({loc})" if loc else ""))
                else:
                    txt = f"⚠ {len(cands)} share this name — verify"
        st.session_state[rkey] = txt
        if fill:
            fills[inv] = fill
    st.session_state["_wb_um_fill"] = fills
    st.rerun()


def _wb_um_save(engine, inputs, _pi):
    """Write the confirmed UWI to MATCHED_UWI (so the file is loadable) and to
    the header's UWI / UWI14 (so re-enrichment keeps it)."""
    from sqlalchemy import text as _t

    ups = []
    for inv, nkey, ukey, rkey, g in inputs:
        u = (_pi.norm_uwi14(st.session_state.get(ukey, "")) if _pi
             else str(st.session_state.get(ukey, "") or "").strip())
        if u:
            ups.append({"inv": inv, "u": u})
    if not ups:
        st.warning("No valid UWIs to write.")
        return
    try:
        with engine.begin() as con:
            for up in ups:
                con.execute(_t(
                    "UPDATE file_catalog.GLOBAL_FILE_CATALOG "
                    "SET MATCHED_UWI=:u, ROW_CHANGED_DATE=GETUTCDATE() "
                    "WHERE INVENTORY_ID=:inv"), up)
                con.execute(_t(
                    "UPDATE file_catalog.FILE_WELL_HEADER "
                    "SET UWI=:u, UWI14=:u WHERE INVENTORY_ID=:inv"), up)
    except Exception as e:
        st.error(f"Save failed: {e}")
        return

    # Clear the per-row widget state and the cached results so the batch table
    # rebuilds with the newly matched UWIs.
    for inv, nkey, ukey, rkey, g in inputs:
        for k in (nkey, ukey, rkey):
            st.session_state.pop(k, None)
    st.session_state.pop("wb_results", None)
    st.success(f"Wrote {len(ups)} UWI(s). Re-run Search to load them.")
    st.rerun()


def _toggle_flag(engine, inv_id: str, flag: bool):
    from sqlalchemy import text as _t
    try:
        with engine.begin() as con:
            con.execute(_t("""
                UPDATE file_catalog.GLOBAL_FILE_CATALOG
                SET FLAG_DELETE=:f, ROW_CHANGED_DATE=GETUTCDATE()
                WHERE INVENTORY_ID=:id
            """), {"f": "Y" if flag else "N", "id": inv_id})
    except Exception as e:
        st.error(f"Flag failed: {e}")


# Columns of FILE_WELL_HEADER that are NOT evidence the file said anything.
# Keys and provenance, obviously -- and two that look like content and are not:
#   REPORT_TYPE is assigned from the FILE TYPE, so every .doc carries 'OFFICE'
#               whether or not a word was read out of it.
#   CONFIDENCE  is a score ABOUT the extraction; 0.00 is not a fact about a well.
# Measured 24 Aug: teapot_3d_load.doc and Synthetic.doc each filled 1 of 25
# columns, and that one column was REPORT_TYPE='OFFICE'. Counting the ROW as
# extraction kept 16 of 20 blocklisted files, including every useless one --
# right question, wrong test, the shape this codebase keeps meeting.
_NOT_EVIDENCE = {"WELL_HEADER_ID", "INVENTORY_ID", "EXTRACTED_DATE",
                 "EXTRACTED_BY", "REPORT_TYPE", "CONFIDENCE"}


def _has_real_extraction(con, inv_id):
    """Did this file actually yield anything worth a catalog row?

    True when ANY of:
      * a dv_* row is attributed to it. This one is not about bloat, it is
        about provenance: delete the catalog row and those rows cite a source
        nothing can resolve -- the exact break that left 1,317 dv_well rows
        orphaned in August. Checked FIRST for that reason.
      * a FILE_SEIS_HEADER row exists. Seismic headers are substantive by
        construction; there is no label-only variant.
      * a FILE_WELL_HEADER row carries a value in any column that is not a key,
        provenance, the file-type label, or a confidence score.

    The dv_ sweep is derived from INFORMATION_SCHEMA rather than a hand-kept
    list, for the reason MIRROR_TABLES exists as a cautionary tale.
    """
    from sqlalchemy import text as _t

    dv = [r[0] for r in con.execute(_t("""
        SELECT t.name FROM sys.tables t
          JOIN sys.schemas s ON s.schema_id = t.schema_id
          JOIN sys.columns c ON c.object_id = t.object_id
         WHERE s.name = 'dataview' AND c.name = 'INVENTORY_ID'
         ORDER BY t.name""")).fetchall()]
    if dv:
        union = " UNION ALL ".join(
            f"SELECT 1 AS hit FROM dataview.[{t}] WHERE INVENTORY_ID = :id"
            for t in dv)
        if con.execute(_t(f"SELECT TOP 1 hit FROM ({union}) q"),
                       {"id": inv_id}).fetchone():
            return True

    if con.execute(_t("SELECT TOP 1 1 FROM file_catalog.FILE_SEIS_HEADER "
                      "WHERE INVENTORY_ID = :id"), {"id": inv_id}).fetchone():
        return True

    cols = [r[0] for r in con.execute(_t("""
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
         WHERE TABLE_SCHEMA = 'file_catalog'
           AND TABLE_NAME  = 'FILE_WELL_HEADER'""")).fetchall()]
    ev = [c for c in cols if c.upper() not in _NOT_EVIDENCE]
    if not ev:
        return False
    # CONVERT then NULLIF, so '' and '   ' count as absent the way NULL does.
    where = " OR ".join(
        f"NULLIF(LTRIM(RTRIM(CONVERT(nvarchar(4000), [{c}]))), '') IS NOT NULL"
        for c in ev)
    return bool(con.execute(
        _t(f"SELECT TOP 1 1 FROM file_catalog.FILE_WELL_HEADER "
           f"WHERE INVENTORY_ID = :id AND ({where})"),
        {"id": inv_id}).fetchone())


def _mark_bad(engine, inv_id, fpath, fname=None, size_kb=None, reason=None):
    """Fingerprint a file as bad (junk / badly formatted) so the next crawl
    skips it, DROP THE ROWS IT STAGED, and retire its catalog row as SKIPPED.

    The blocklist is keyed on INVENTORY_ID — the same SHA-1(uppercase path)
    fingerprint the crawler computes — so the skip is a cheap set lookup with
    no per-file hashing during the walk. A best-effort content signature
    (first 4 MB) is stored alongside for reference, but is never used to gate
    the crawl (so it can't stall on huge files).

    A PARENT THAT CANNOT BE PROMOTED MUST TAKE ITS CHILDREN WITH IT.
    Rejecting used to delete the GLOBAL_FILE_CATALOG row and nothing else, so
    every cat_* row the file had already captured was left with no catalog
    parent: invisible to every report that joins through GFC, and still
    PROMOTED = 0, so promote retried them forever. MEASURED 17 Aug: 10 rejected
    files had orphaned 50 rows across 6 mirrors (30 cat_prod_volume, 8
    cat_well_stimulation, 4 cat_well, 4 cat_well_casing, 2
    cat_well_formation_top, 2 cat_prod_entity). Those rows also read as a
    provenance break — a cat_ row whose INVENTORY_ID resolves nowhere looks
    exactly like a broken write path, and was diagnosed as one.

    dataview.dv_* is NOT touched. Rows that already promoted belong to whatever
    load inserted them (first one in wins); un-promoting them is a separate
    decision, not a side effect of blocklisting a file.

    Returns {mirror_table: rows_deleted} for the mirrors that had rows, so the
    caller can report what went. Empty dict = the file had staged nothing.
    """
    from sqlalchemy import text as _t
    import hashlib

    fhash = ""
    try:
        with open(fpath, "rb") as _f:
            fhash = hashlib.sha1(_f.read(4 * 1024 * 1024)).hexdigest().upper()
    except Exception:
        fhash = ""

    try:
        sz = float(size_kb) if size_kb not in (None, "") else None
    except (TypeError, ValueError):
        sz = None

    with engine.begin() as con:
        con.execute(_t("""
            IF OBJECT_ID('file_catalog.BAD_FILE','U') IS NULL
            CREATE TABLE file_catalog.BAD_FILE (
                INVENTORY_ID  NVARCHAR(40) NOT NULL PRIMARY KEY,
                FILE_PATH     NVARCHAR(900),
                FILE_NAME     NVARCHAR(260),
                FILE_SIZE_KB  DECIMAL(15,2),
                FILE_HASH     NVARCHAR(40),
                REASON        NVARCHAR(500),
                MARKED_BY     NVARCHAR(100),
                MARKED_DATE   DATETIME2
                    CONSTRAINT DF_BAD_FILE_DT DEFAULT SYSUTCDATETIME()
            );
        """))
        con.execute(_t("""
            MERGE file_catalog.BAD_FILE AS tgt
            USING (SELECT :id AS INVENTORY_ID) AS src
              ON tgt.INVENTORY_ID = src.INVENTORY_ID
            WHEN MATCHED THEN UPDATE SET
                FILE_PATH=:p, FILE_NAME=:n, FILE_SIZE_KB=:sz,
                FILE_HASH=:h, REASON=:r, MARKED_DATE=SYSUTCDATETIME()
            WHEN NOT MATCHED THEN INSERT
                (INVENTORY_ID, FILE_PATH, FILE_NAME, FILE_SIZE_KB,
                 FILE_HASH, REASON, MARKED_BY, MARKED_DATE)
                VALUES (:id, :p, :n, :sz, :h, :r, 'DataWrangler',
                        SYSUTCDATETIME());
        """), {"id": inv_id, "p": (fpath or "")[:900],
               "n": (fname or "")[:260], "sz": sz, "h": fhash,
               "r": (reason or "junk / bad format")[:500]})
        # CASCADE. The table list is a SWEEP, deliberately NOT
        # promotion_lineage.LINEAGE: LINEAGE answers "which cat_ -> dv_ pairs can
        # a report attribute?", while a cascade must answer the wider "what did
        # this file stage ANYWHERE" — and a mirror missing from LINEAGE is
        # precisely the case that would go on orphaning rows silently. Same
        # reasoning catalog_readiness._sources gives for widening its dv_ side.
        # 'cat[_]%' is bracketed because _ is a T-SQL LIKE wildcard: unbracketed
        # 'cat_%' also matches catalog_setting. The mirrors carry no FKs between
        # them (build_catalog_mirror.build_ddl makes only a PK on CAT_ROW_ID plus
        # two indexes), so delete order does not matter.
        mirrors = [r[0] for r in con.execute(_t("""
            SELECT t.name
              FROM sys.tables t
              JOIN sys.schemas s ON s.schema_id = t.schema_id
              JOIN sys.columns c ON c.object_id = t.object_id
             WHERE s.name = 'file_catalog'
               AND t.name LIKE 'cat[_]%'
               AND c.name = 'INVENTORY_ID'
             ORDER BY t.name
        """)).fetchall()]
        dropped = {}
        for _m in mirrors:
            _n = con.execute(
                _t(f"DELETE FROM file_catalog.[{_m}] WHERE INVENTORY_ID = :id"),
                {"id": inv_id}).rowcount or 0
            if _n:
                dropped[_m] = _n

        # RETIRE, don't delete. 'SKIPPED' is already the terminal state for
        # "leave this file alone", and both halves of the pipeline already honour
        # it: catalog_readiness._CASE tests SKIPPED FIRST (so a leftover row
        # cannot make the file read pending), and pipeline_run._already_done_filter
        # keeps it excluded even under --force — "forcing means 'ignore that this
        # was already processed', not 'ignore that I told you to leave it alone'".
        # Deleting the row instead threw away the file's entire catalog history
        # and the join path every report and lineage query needs. The crawl skip
        # reads BAD_FILE, not the row's absence, so blocklisting still works.
        # RETIRE WHAT HAS A HISTORY; REMOVE WHAT NEVER HAD ONE.
        #
        # Retiring rather than deleting is right for a file that produced
        # something: the row is the join path every lineage query needs, and
        # deleting it once left dv_ rows citing a source nothing could resolve.
        # None of that applies to a file that yielded nothing. Its row records
        # a path, a size and a list of everything it did not contain, which is
        # bloat wearing the costume of history -- and BAD_FILE already holds
        # the path AND the content hash, so the crawl still skips it (the skip
        # reads BAD_FILE, never the row's absence).
        #
        # Measured 24 Aug: of 20 blocklisted files, 4 had no header row at all
        # and 4 more had one filled solely with REPORT_TYPE='OFFICE'. Those 8
        # rows are what this removes; the other 12 carry real content and stay.
        #
        # Restore works either way: _unmark_bad clears BAD_FILE, and because
        # INVENTORY_ID is a deterministic hash of the path, a re-scan rebuilds
        # a removed row under the SAME id.
        if _has_real_extraction(con, inv_id):
            con.execute(_t("""
                UPDATE file_catalog.GLOBAL_FILE_CATALOG
                   SET CATALOG_READINESS = 'SKIPPED',
                       ROW_CHANGED_DATE  = GETUTCDATE()
                 WHERE INVENTORY_ID = :id
            """), {"id": inv_id})
        else:
            # The empty header row goes with it -- leaving it behind would
            # orphan the very thing this is meant to stop accumulating.
            con.execute(_t("DELETE FROM file_catalog.FILE_WELL_HEADER "
                           "WHERE INVENTORY_ID = :id"), {"id": inv_id})
            con.execute(_t("DELETE FROM file_catalog.GLOBAL_FILE_CATALOG "
                           "WHERE INVENTORY_ID = :id"), {"id": inv_id})
    return dropped


def _unmark_bad(engine, inv_id):
    """Undo _mark_bad: off the blocklist, out of SKIPPED, pending again.

    Returns True when the file was actually ON the blocklist, so the caller can
    report a real count instead of the size of the selection.

    THREE THINGS COME UNDONE, because _mark_bad did three:

      * the BAD_FILE row goes, or the next crawl skips the file again. This is
        the one everybody remembers.
      * CATALOG_READINESS stops being 'SKIPPED'. Set to NULL rather than to a
        guessed value: readiness is DERIVED (catalog_readiness._CASE) and the
        next pass recomputes it. Writing 'CATALOGED' here would state a fact
        nothing measured.
      * HEADER_EXTRACTED goes back to 'N'. THIS IS THE ONE THAT IS EASY TO
        MISS. _mark_bad DELETED the rows the file had staged in cat_*, but left
        its stamp reading 'Y' -- so a restore that stopped at the first two
        would leave a file claiming to be extracted, holding no staged rows,
        and never re-extracted to get them back. Data silently missing behind a
        completed stamp is the shape this codebase keeps getting bitten by.

    dv_* is not touched, for the same reason _mark_bad does not touch it: rows
    that already promoted belong to whichever load inserted them, and
    un-promoting them is a separate decision.
    """
    from sqlalchemy import text as _t

    with engine.begin() as con:
        was_bad = bool(con.execute(_t(
            "SELECT 1 FROM file_catalog.BAD_FILE WHERE INVENTORY_ID = :id"),
            {"id": inv_id}).fetchone())
        con.execute(_t("DELETE FROM file_catalog.BAD_FILE "
                       "WHERE INVENTORY_ID = :id"), {"id": inv_id})
        # was_bad is read BEFORE the DELETE and passed in as a parameter. An
        # EXISTS against BAD_FILE inside this UPDATE would run AFTER the delete
        # in the same transaction and so always be false, leaving a file that
        # was blocklisted but never stamped SKIPPED with a stale 'Y'.
        con.execute(_t("""
            UPDATE file_catalog.GLOBAL_FILE_CATALOG
               SET CATALOG_READINESS = NULL,
                   HEADER_EXTRACTED  = 'N',
                   ROW_CHANGED_DATE  = GETUTCDATE()
             WHERE INVENTORY_ID = :id
               AND (ISNULL(CATALOG_READINESS, '') = 'SKIPPED' OR :wasbad = 1)
        """), {"id": inv_id, "wasbad": 1 if was_bad else 0})
    return was_bad



# =============================================================================
# Status & Backlog -- every file, its state, and WHY it is stuck
# =============================================================================
# Browse & View owns the identity axis: files that never reached capture, fixed
# by assigning a UWI or a survey name. This section owns everything after that
# -- what is staged, what promoted, and what promote is REFUSING, with the
# reason named per file and the action that clears it next to it.
#
# The reason is derived, not stored. catalog_status explains why at length; the
# short version is that every one of these reasons is curable by an action that
# touches a different file's rows (seed one dv_r_uom code and thirty files stop
# being held), so a stamp would be stale the moment it was written.


def _status_run(engine, log=None):
    """Run catalog_status into session_state. Shared by the Run button, the
    post-action refresh and the first open, so nothing can act on a grid that
    no longer matches the database."""
    from dataview.file_catalog import catalog_status as _cs
    _notes = []
    try:
        res = _cs.file_status(
            engine,
            root=(None if st.session_state.get("sb_root") in (None, _SB_ALL)
                  else st.session_state.get("sb_root")),
            this_crawl=bool(st.session_state.get("sb_this_crawl")),
            log=_notes.append)
    except Exception as e:
        st.error(f"Status failed: {type(e).__name__}: {e}")
        return None
    st.session_state["sb_res"] = res
    st.session_state["sb_notes"] = _notes
    return res


def _run_coord_enrich(engine):
    """Fill cat_well surface coords from gold, then dv_well -- promote's own
    pre-gate, run on demand.

    This is the ONE hold reason with a genuine one-click cure, and the function
    that cures it already exists: promote_catalog._fill_cat_coords_from_gold is
    what promote itself calls before counting eligibility. Calling it here is
    not a shortcut past the gate -- the gate still has to pass afterwards. It
    just stops a well being held for a location the database already knows,
    which measured 26 of 29 held wells on 17 Aug.

    Returns (filled, note). Never raises: the note carries the failure, because
    a silent coord-enrich failure looks exactly like a well gold never heard of.
    """
    from dataview.file_catalog import promote_catalog as _pc
    raw = engine.raw_connection()
    try:
        cur = raw.cursor()
        filled, note = _pc._fill_cat_coords_from_gold(
            cur, "cat_well", "surface_latitude", "surface_longitude", "", [])
        raw.commit()
        return filled, note
    except Exception as e:
        try:
            raw.rollback()
        except Exception:
            pass
        return 0, f" · coord enrich FAILED ({type(e).__name__}: {e})"
    finally:
        try:
            raw.close()
        except Exception:
            pass


def _run_capture(engine, dialect, root=None, log=None):
    """Run the capture stage — stage pending files' rows into file_catalog.cat_*.

    Delegates to pipeline_run._stage_capture, the SAME function run_pipeline
    drives, rather than re-deriving what "pending" means. Six spellings of
    'pending' once gave six different answers on one catalog (31/43/190/1,319/
    2,190/3,876); this page is not becoming the seventh.

    NO EXTENSION FILTER, deliberately. _stage_capture's docstring says binaries
    are "left to the deep stage", but run_pipeline calls this very function with
    exts={.las,.segy,.sgy,.seg,.p190} for its binary lane — the restriction is
    in the CALLER, not the stage. Passing exts=None is the plain path
    run_pipeline itself takes when no recogniser pack is configured, and it is
    the only one that covers LAS and documents in a single press. Filtering here
    would silently capture the PDFs and skip every LAS.

    Idempotent: capture() replaces a file's rows scoped to INVENTORY_ID, so a
    re-run refreshes rather than duplicates.

    Returns the stage's counts dict, or raises — the caller reports.
    """
    from dataview.import_data import pipeline_run as _pr
    _lines = []
    _log = log or _lines.append
    res = _pr._stage_capture(engine, dialect, _log, exts=None, workers=8,
                             parallel=False, force=False, root=root)
    return (res or {}), _lines


_SB_ALL = "(whole catalog)"      # scope: the scan-root selector
_SB_ANY = "All"                  # filters: state / reason / type

# Column order for the grid. `path` is carried for the export's hyperlink and
# deliberately not shown -- a full Windows path per row pushes every useful
# column off the right edge, the same reason the stage scorecard drops it.
_SB_GRID_COLS = ["file", "state", "reason", "ext", "uwi",
                 "rows_cat", "rows_dv", "readiness"]


def _tab_status(engine, dialect):
    import pandas as pd
    from dataview.file_catalog import catalog_status as _cs

    st.markdown("#### 🚦 Status & Backlog")
    st.caption(
        "Every inventoried file, what state it reached, and — when promote is "
        "refusing it — the reason, per file. Reasons are re-derived from the "
        "mirrors each run, so they cannot go stale."
    )

    # ── Scope ────────────────────────────────────────────────────────────────
    try:
        from sqlalchemy import text as _t
        with engine.connect() as _c:
            _roots = [r[0] for r in _c.execute(_t(
                "SELECT DISTINCT ROOT_PATH FROM file_catalog.GLOBAL_FILE_CATALOG "
                "WHERE NULLIF(LTRIM(RTRIM(ROOT_PATH)),'') IS NOT NULL "
                "ORDER BY ROOT_PATH")).fetchall()]
    except Exception:
        _roots = []

    s1, s2, s3 = st.columns([3, 1, 1])
    s1.selectbox("Scan root", [_SB_ALL] + _roots, key="sb_root",
                 help="Limit to one scan root, or the whole catalog.")
    s2.checkbox("This crawl only", value=False, key="sb_this_crawl",
                help="Limit to files scanned today.")
    if s3.button("🔄 Run / refresh", type="primary", key="sb_run",
                 use_container_width=True):
        _status_run(engine)
        st.session_state["sb_nonce"] = st.session_state.get("sb_nonce", 0) + 1

    res = st.session_state.get("sb_res")
    if res is None or getattr(res, "df", None) is None:
        st.info("Choose a scope and click **Run / refresh**.")
        return
    df = res.df
    if df.empty:
        st.info("No files in scope. Run a crawl, or widen the scope.")
        return

    # ── Anything that narrowed the answer, said out loud ─────────────────────
    # A report that quietly covers less than it claims is the failure this whole
    # section exists to fix, so its own gaps are not allowed to be silent.
    _notes = st.session_state.get("sb_notes") or []
    if _notes:
        with st.expander(f"⚠ {len(_notes)} note(s) — this report could not see "
                         f"everything", expanded=False):
            for _n in _notes:
                st.caption(f"• {_n}")

    # ── Where everything stands ──────────────────────────────────────────────
    _n = df["state"].value_counts().to_dict()
    _held = int(_n.get(_cs.ST_HELD, 0))
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Files", f"{len(df):,}")
    m2.metric("Promoted", f"{int(_n.get(_cs.ST_PROMOTED, 0)):,}",
              help="Rows reached dataview.dv_*.")
    m3.metric("Staged", f"{int(_n.get(_cs.ST_STAGED, 0)):,}",
              help="Rows in cat_*, every gate passes — promote has not run yet.")
    m4.metric("Held", f"{_held:,}",
              help="Rows in cat_*, and promote refuses at least one of them. "
                   "The backlog.")
    m5.metric("Not captured",
              f"{int(_n.get(_cs.ST_INVENTORIED, 0)) + int(_n.get(_cs.ST_EXTRACTED, 0)):,}",
              help="Inventoried or extracted, nothing staged. Fix identity in "
                   "Browse & View.")
    m6.metric("Rejected", f"{int(_n.get(_cs.ST_SKIPPED, 0)):,}")

    # ── Which single fix clears the most ─────────────────────────────────────
    _rs = _cs.reason_summary(res)
    if not _rs.empty:
        st.markdown("**🔴 Held by reason** — worst first; the top row is the "
                    "fix that clears the most backlog.")
        st.dataframe(_rs, hide_index=True, use_container_width=True,
                     column_config={
                         "reason": st.column_config.TextColumn("Reason", width="medium"),
                         "files": st.column_config.NumberColumn("Files", width="small"),
                         "rows": st.column_config.NumberColumn("Rows", width="small"),
                         "clears_by": st.column_config.TextColumn(
                             "What clears it", width="large"),
                     })
    elif _held == 0:
        st.success("✅ Nothing held — every staged row passes promote's gates.")

    # ── Clear the biggest seismic hold from HERE ─────────────────────────────
    # THE FIX BELONGS NEXT TO THE DIAGNOSIS. CLEAR_ROUTE exists so "the fix is
    # one click from the diagnosis instead of a hunt through four pages", and
    # for unnamed surveys it pointed at a different tab — Perry followed it,
    # found nothing, and reasonably concluded the feature did not exist. This
    # is that route made into the actual control.
    #
    # 2D lines arrive as a SET: lineA..lineE are five files of ONE survey, and
    # the per-file path guess offers five DIFFERENT names (LINEA, LINEB, …),
    # so the path of least resistance was five surveys where there is one.
    # Hence the group control rather than a row-by-row edit.
    #
    # Rendered only when something is actually held this way — an expander's
    # body executes whether or not it is open (see the note further down), so
    # gating on the count keeps its query off the page when there is no work.
    _unnamed_n = sum(1 for _d in res.holds.values()
                     if _cs._HOLD_SEIS_UNNAMED in _d)
    if _unnamed_n:
        # Sibling of the "Open a file" expander below, NOT nested inside it —
        # expanders cannot nest. seis_survey_assign uses a bordered container
        # for its group box for the same reason.
        with st.expander(
                f"🎛 Name a survey — {_unnamed_n} seismic file(s) held for "
                f"having none. 2D lines usually share ONE survey.",
                expanded=False):
            from dataview.file_catalog.seis_survey_assign import (
                seis_survey_grid as _seis_assign)
            _seis_assign(engine)

    # ── files with no UWI: NOT a hold, and that is why they were invisible ───
    # These have an extracted header but NO staged rows, so no cat_ mirror gate
    # can fire and nothing above can report them. In the four states they are
    # nearer "Nothing" than "Held" — and a file that produced a header and then
    # stopped is exactly what this page exists to surface.
    #
    # Found 23 Aug tracing 15007205750000_2.las: READY, header extracted
    # (WELL_NAME 'WEST 1-3'), UWI and UWI14 both NULL, no cat_ rows, and no
    # reason anywhere — because the reason lives in a mirror it never reached.
    # 17 files were in that state, the filename carrying a perfectly good UWI
    # for most of them.
    #
    # Counted directly rather than from res.holds, since by construction these
    # are NOT holds. The count is one indexed read of a small table and gates
    # the expander, whose body would otherwise run its own query on every
    # render of this tab.
    _nouwi_n = 0
    try:
        with engine.connect() as _c:
            if _c.execute(_t(
                    "SELECT 1 FROM sys.columns WHERE name='UWI14' AND "
                    "object_id=OBJECT_ID('file_catalog.FILE_WELL_HEADER')")).fetchone():
                _nouwi_n = _c.execute(_t(
                    "SELECT COUNT(*) FROM file_catalog.FILE_WELL_HEADER "
                    "WHERE UWI14 IS NULL OR UWI14 = '00000000000000'")).scalar() or 0
    except Exception as _ue:
        # Say so rather than showing nothing: a swallowed count here would look
        # exactly like "no files need a UWI".
        st.caption(f"(could not count files needing a UWI: {type(_ue).__name__})")
    if _nouwi_n:
        with st.expander(
                f"🔑 Key a well — {_nouwi_n} file(s) have a header but no UWI, "
                f"so nothing is staged and no gate can name a reason.",
                expanded=False):
            st.caption(
                "A UWI is per-well, so there is no group shortcut here — but "
                "the UWI is usually IN THE FILENAME and is already filled in "
                "below as a guess. Check them and Save. **Match against "
                "reference** looks each row up in the well master, "
                "corroborated by total depth, spud date and operator, so a "
                "shared well name alone never fills the wrong UWI.")
            _well_key_grid(engine)

    # ── Filters ──────────────────────────────────────────────────────────────
    st.divider()
    _states = [_SB_ANY] + sorted(df["state"].unique().tolist())
    _reasons = [_SB_ANY] + sorted(
        {r for d in res.holds.values() for r in d})
    _types = [_SB_ANY] + sorted(
        {t for t in df["type"].fillna("").tolist() if t})

    f1, f2, f3, f4 = st.columns(4)
    f_state = f1.selectbox("State", _states, key="sb_f_state")
    f_reason = f2.selectbox("Reason", _reasons, key="sb_f_reason")
    f_type = f3.selectbox("File type", _types, key="sb_f_type")
    f_name = f4.text_input("Filename contains", key="sb_f_name",
                           placeholder="partial name…")

    view = df
    if f_state != _SB_ANY:
        view = view[view["state"] == f_state]
    if f_reason != _SB_ANY:
        # Match the REASON DICT, not the rendered string. The rendered cell is
        # "<label> (<rows>); <label> (<rows>)", and one label is a substring of
        # another -- "no UWI" is contained in "no UWI on detail row" -- so a
        # substring test silently returns the union of two different holds with
        # two different repairs. Keying on the dict is exact.
        _ids = {inv for inv, d in res.holds.items() if f_reason in d}
        view = view[view["inventory_id"].isin(_ids)]
    if f_type != _SB_ANY:
        view = view[view["type"] == f_type]
    if f_name:
        view = view[view["file"].str.contains(f_name, case=False, na=False)]

    st.caption(f"{len(view):,} of {len(df):,} file(s)")
    if view.empty:
        st.info("No files match these filters.")
        return

    # ── The grid ─────────────────────────────────────────────────────────────
    # Nonce-keyed for the same reason _tab_browse is: Streamlit refuses a write
    # to a widget's key once instantiated, so "untick everything after acting"
    # cannot be an assignment. Bumping the nonce retires the editor and mints a
    # fresh one, which also discards row-index ticks that would otherwise point
    # at different files once the result set changes.
    _nonce = st.session_state.get("sb_nonce", 0)
    _grid = view[_SB_GRID_COLS].copy()
    _grid.insert(0, "Select", False)
    _ed = st.data_editor(
        _grid, hide_index=True, use_container_width=True,
        key=f"sb_grid_{_nonce}_editor",
        disabled=[c for c in _grid.columns if c != "Select"],
        column_config={
            "Select": st.column_config.CheckboxColumn(
                "✓", help="Tick files to act on", width="small"),
            "file": st.column_config.TextColumn("File", width="medium"),
            "state": st.column_config.TextColumn("State", width="small"),
            "reason": st.column_config.TextColumn(
                "Why it is held", width="large"),
            "rows_cat": st.column_config.NumberColumn(
                "cat rows", width="small",
                help="Unpromoted rows staged in file_catalog.cat_*"),
            "rows_dv": st.column_config.NumberColumn(
                "dv rows", width="small",
                help="Rows this file has in dataview.dv_*"),
        })

    try:
        _picked = view.loc[[i for i, v in zip(view.index, _ed["Select"]) if v]]
    except Exception:
        _picked = view.iloc[0:0]
    st.caption(f"**{len(_picked)}** file(s) selected")

    # ── Inspect ──────────────────────────────────────────────────────────────
    # The UWI and the location you are about to type are IN the document. Making
    # you leave for Browse & View to read them, then come back to type them, is
    # how a supply-the-missing-value workflow turns into two-page ping-pong.
    #
    # file_viewer.view is nest-safe by construction (it uses _vsection, a
    # bordered container, precisely so it can be embedded), so this can live
    # inside the expander below without breaking the no-nested-expander rule.
    st.divider()
    with st.expander("🔍 Open a file — read the UWI or location off the document",
                     expanded=False):
        _ids = list(view["inventory_id"])
        _lbl = {r["inventory_id"]: f"{r['file']}  ·  {r['state']}"
                                   + (f"  ·  {r['reason']}" if r["reason"] else "")
                for _, r in view.iterrows()}

        # A prev/next click cannot assign the selectbox's key AFTER the widget
        # exists, so it parks the request here and this consumes it BEFORE the
        # widget is drawn. Same discipline as the map page's request flags.
        _req = st.session_state.pop("sb_view_req", None)
        if _req in _ids:
            st.session_state["sb_view_pick"] = _req
        # A fixed-key selectbox keeps a value the new filter may have removed —
        # drop it before instantiation rather than letting it raise.
        if st.session_state.get("sb_view_pick") not in _ids:
            st.session_state.pop("sb_view_pick", None)

        if not _ids:
            st.caption("No files in the current view.")
        else:
            v1, v2, v3 = st.columns([6, 1, 1])
            _pick = v1.selectbox(
                "File", _ids, key="sb_view_pick",
                format_func=lambda i: _lbl.get(i, str(i)),
                label_visibility="collapsed")
            _pos = _ids.index(_pick) if _pick in _ids else 0
            if v2.button("◀ Prev", key="sb_view_prev", use_container_width=True,
                         disabled=_pos == 0):
                st.session_state["sb_view_req"] = _ids[_pos - 1]
                st.rerun()
            if v3.button("Next ▶", key="sb_view_next", use_container_width=True,
                         disabled=_pos >= len(_ids) - 1):
                st.session_state["sb_view_req"] = _ids[_pos + 1]
                st.rerun()

            _row = view[view["inventory_id"] == _pick].iloc[0]
            _fpath = _row["path"]
            st.caption(f"`{_fpath}`")
            if _row["reason"]:
                st.warning(f"Held for: {_row['reason']}")

            # What extraction already read off this file. The well name and
            # operator are usually what identifies a document whose UWI did not
            # parse, so they belong next to the viewer, not a page away.
            try:
                from sqlalchemy import text as _t2
                with engine.connect() as _c2:
                    _h = _c2.execute(_t2("""
                        SELECT TOP 1 UWI, WELL_NAME, OPERATOR, API_NUM
                          FROM file_catalog.FILE_WELL_HEADER
                         WHERE INVENTORY_ID = :i"""), {"i": _pick}).fetchone()
                if _h and any(_h):
                    st.caption(
                        f"extracted header — UWI `{_h[0] or '—'}` · "
                        f"well `{_h[1] or '—'}` · operator `{_h[2] or '—'}`"
                        + (f" · API `{_h[3]}`" if _h[3] else ""))
            except Exception as _he:
                st.caption(f"(header lookup unavailable: {str(_he)[:80]})")

            if not _fpath:
                st.warning("No path recorded for this file.")
            elif not Path(_fpath).exists():
                # HEADER_EXTRACTED='M' is its own state for exactly this: the
                # catalog is stale, the file is not corrupt. Say which.
                st.warning(
                    f"Not on disk at that path. If this file was moved, "
                    f"re-scan its new location — the catalog row keeps its id "
                    f"and its history."
                    + ("  (Catalog already marks it MOVED.)"
                       if _row["state"] == _cs.ST_MOVED else ""))
            else:
                o1, o2 = st.columns([1, 4])
                if o1.button("↗ Open natively", key="sb_view_native",
                             use_container_width=True,
                             help="Open in the OS default application. Costs "
                                  "nothing here — the app does not parse it."):
                    try:
                        from dataview.file_catalog.catalog_docs import _open_native
                        _err = _open_native(_fpath)
                        if _err:
                            st.error(f"Could not open: {_err}")
                    except Exception as _oe:
                        st.error(f"Could not open: {type(_oe).__name__}: {_oe}")
                # RENDERING IS OPT-IN, and the toggle is not decoration.
                # Streamlit executes an expander's body whether or not it is
                # open, so an ungated viewer here re-parses a document on EVERY
                # rerun of this page — every filter change, every tick of a
                # checkbox, every Preview. For a PDF that means rasterising
                # pages through PyMuPDF each time. Off by default; the native
                # open above needs no parse at all.
                _show = o2.toggle(
                    "Render preview in-app", key="sb_view_render", value=False,
                    help="Parses the file to display it here. Leave off for "
                         "large PDFs or SEG-Y and use Open natively instead.")
                if _show:
                    try:
                        from dataview.file_catalog.file_viewer import view as _fview
                        _fview(_fpath,
                               _row["ext"] if _row["ext"] != "(none)" else None)
                    except Exception as _ve:
                        st.error(f"Viewer error: {type(_ve).__name__}: {_ve}")

    # ── Drain: supply what the gate is asking for ────────────────────────────
    # The backlog empties three ways and only three: supply a UWI, supply a
    # location, or reject the file. Everything below is one of those, followed
    # by promote. Nothing waives a gate.
    # HELD, PLUS ANYTHING THAT SIMPLY HAS NO IDENTITY YET.
    #
    # This panel was keyed on HELD alone, and a hold is DERIVED FROM cat_* rows
    # — so a file that could not stage anything BECAUSE it has no UWI is
    # exactly the file this panel exists for and exactly the file it could not
    # see. Measured 23 Aug: 0 HELD, 30 EXTRACTED, 17 of those needing a UWI.
    # The grid was empty, so Preview stayed disabled, so "Apply these fixes"
    # could never activate — reported as the button not working, when the
    # button was right and its input was missing.
    #
    # Safe for un-staged files: apply_fix writes MATCHED_UWI/UWI14 on the
    # catalog row and CLEARS CAPTURED_HASH so capture re-runs and stages the
    # rows — which is precisely the remedy these files need. Its mirror
    # updates touch 0 rows when nothing is staged, and header-minting is
    # already guarded on detail rows actually existing.
    _no_id = ((view["uwi"].fillna("").astype(str).str.strip() == "")
              & view["state"].isin([_cs.ST_EXTRACTED, _cs.ST_INVENTORIED]))
    _heldv = view[(view["state"] == _cs.ST_HELD) | _no_id]
    if not _heldv.empty:
        st.divider()
        st.markdown("##### ① Supply what is missing")
        st.caption(
            "Type a UWI and/or a surface location. Listed here: files **held** "
            "at a gate, and files that staged **nothing** because they have no "
            "UWI to key rows to — the second kind never appears in the hold "
            "counts above, since a hold is counted from staged rows. A UWI is "
            "written to the catalog **and to any rows already staged** — the "
            "step plain assignment skips — and clears the capture stamp so the "
            "next capture picks the file up. A location fills a coordless "
            "header, or creates one when detail rows name a well that has none."
        )

        # SUGGEST, DON'T ASSERT. The filename usually carries the UWI in this
        # corpus, and path_identity.uwi14_from_path reads it — measured against
        # this catalog it agreed with the known UWI on 145 files and disagreed
        # on none. But triage deliberately keeps a path-derived identity OUT of
        # READY (`IDENTITY_SOURCE NOT LIKE 'path%'`), because a name is a claim
        # about a file, not a fact from inside it. So it lands in an EDITABLE
        # cell with its provenance shown beside it, and still has to survive
        # Preview and Apply. Proposed by code, confirmed by you.
        try:
            from dataview.core.path_identity import uwi14_from_path as _u4p
        except Exception:
            _u4p = lambda _p: (None, None)

        _sug, _src = [], []
        for _, _hr in _heldv.iterrows():
            if str(_hr.get("uwi") or "").strip():
                _sug.append("")          # already identified — never overwrite
                _src.append("")
                continue
            try:
                _u, _s = _u4p(_hr["path"])
            except Exception:
                _u, _s = None, None
            _sug.append(_u or "")
            _src.append(_s or "")

        _fix = _heldv[["file", "reason", "uwi"]].copy()
        # A file that never staged has NO hold reason, because reasons come
        # from cat_* rows it does not have. Say why it is in this list instead
        # of showing an empty cell, which reads as a missing value rather than
        # as the actual answer.
        _fix["reason"] = _fix["reason"].fillna("").replace(
            "", "nothing staged — no UWI to key rows to")
        _fix.insert(0, "inventory_id", _heldv["inventory_id"])
        _fix["set_uwi"] = _sug
        _fix["from"] = _src
        _fix["set_lat"] = None
        _fix["set_lon"] = None
        _n_sug = sum(1 for s in _sug if s)
        if _n_sug:
            st.info(
                f"📋 Pre-filled **{_n_sug}** UWI(s) read from the file name. "
                f"Check them against the document before applying — the `from` "
                f"column says where each came from, and the viewer above opens "
                f"the file. Clear any cell you don't want written.")
        _fed = st.data_editor(
            _fix, hide_index=True, use_container_width=True,
            key=f"sb_fix_{_nonce}_editor",
            disabled=["inventory_id", "file", "reason", "uwi", "from"],
            column_config={
                "inventory_id": None,       # carried for the write, not shown
                "file": st.column_config.TextColumn("File", width="medium"),
                "reason": st.column_config.TextColumn("Held for", width="large"),
                "uwi": st.column_config.TextColumn("Current UWI", width="small"),
                "set_uwi": st.column_config.TextColumn(
                    "Set UWI", width="small",
                    help="14-character UWI. Anything that does not normalize to "
                         "exactly 14 is refused, not padded. Pre-filled from the "
                         "file name where one could be read — check it."),
                "from": st.column_config.TextColumn(
                    "read from", width="small",
                    help="Where a pre-filled UWI came from: 'filename' or "
                         "'folder'. Blank means you typed it, or nothing was "
                         "suggested. A folder-derived UWI is the weaker of the "
                         "two — the file may not belong to that folder's well."),
                "set_lat": st.column_config.NumberColumn(
                    "Latitude", width="small", format="%.6f"),
                "set_lon": st.column_config.NumberColumn(
                    "Longitude", width="small", format="%.6f"),
            })

        # Identify rows BY POSITION against _heldv, not by reading the hidden
        # inventory_id back out of the editor's return value. A column hidden
        # through column_config may not survive the round trip, and that would
        # be a KeyError that fires only once somebody actually types an edit —
        # invisible in any render that changes nothing.
        _inv_by_pos = list(_heldv["inventory_id"])
        _edits = []
        for _pos, (_, _r) in enumerate(_fed.iterrows()):
            if _pos >= len(_inv_by_pos):
                break
            if (str(_r.get("set_uwi") or "").strip()
                    or _r.get("set_lat") is not None
                    or _r.get("set_lon") is not None):
                _edits.append({
                    "inventory_id": _inv_by_pos[_pos],
                    "file": _r.get("file", ""),
                    "uwi": str(_r.get("set_uwi") or "").strip(),
                    "lat": _r.get("set_lat"),
                    "lon": _r.get("set_lon"),
                })

        p1, p2 = st.columns(2)
        # SAMPLE BEFORE APPLY. A coordinate backfill once came within a
        # twenty-row sample of writing 1,436 confidently wrong positions, every
        # one at quality_score 100. Nothing here writes until the exact list of
        # writes has been shown and the button under it pressed.
        if p1.button(f"🔎 Preview {len(_edits)} edit(s)", key="sb_preview",
                     use_container_width=True, disabled=not _edits):
            from dataview.file_catalog import catalog_status as _cs2
            _plan, _bad = [], False
            with engine.connect() as _con:
                for _e in _edits:
                    _p = _cs2.plan_fix(_con, _e["inventory_id"], _e["uwi"],
                                       _e["lat"], _e["lon"])
                    for _err in _p["errors"]:
                        _plan.append({"file": _e["file"], "will write": "",
                                      "⚠": _err})
                        _bad = True
                    for _act in _p["actions"]:
                        _plan.append({"file": _e["file"], "will write": _act,
                                      "⚠": ""})
                    for _nte in _p["notes"]:
                        _plan.append({"file": _e["file"], "will write": "",
                                      "⚠": _nte})
            # Preview is now INFORMATIONAL ONLY — Apply validates the current
            # edits itself, so nothing downstream reads sb_plan_ok or
            # sb_plan_edits to decide whether a write may happen. Kept so the
            # table renders and the flagged-row warning still shows.
            st.session_state["sb_plan"] = _plan
            st.session_state["sb_plan_ok"] = (not _bad) and bool(_plan)

        _plan = st.session_state.get("sb_plan")
        if _plan:
            st.markdown("**Exactly what will be written**")
            st.dataframe(pd.DataFrame(_plan), hide_index=True,
                         use_container_width=True)
            if not st.session_state.get("sb_plan_ok"):
                st.warning("Fix the flagged rows, then preview again. "
                           "Nothing has been written.")
        elif _edits:
            st.caption(f"**{len(_edits)} edit(s) ready.** Apply writes them; "
                       f"Preview first if you want to see the exact statements.")

        # APPLY IS ONE CLICK, AND VALIDATES ITSELF.
        #
        # It used to require a Preview first, on the "sample before apply"
        # rule that a coordinate backfill once came within twenty rows of
        # writing 1,436 confidently wrong positions. That rule is right, but
        # it was being served twice: THE GRID ABOVE IS THE SAMPLE. Every value
        # is on screen, editable, with a "read from" column saying whether a
        # UWI came from the filename or a folder. Making the operator press
        # Preview to be shown the same numbers again is ceremony, not a
        # decision — and Perry rejects files and keys wells all day.
        #
        # What Preview genuinely added was VALIDATION: plan_fix refuses a UWI
        # that does not normalise to 14, and reports it instead of writing.
        # That is kept — Apply runs the same plan_fix over the CURRENT edits
        # and writes nothing at all if any row fails. So the guarantee is
        # unchanged (nothing invalid is ever written) while the click is not.
        #
        # It also removes the staleness bug this block used to carry: Apply
        # wrote sb_plan_edits, the values as they were WHEN PREVIEWED, so
        # editing a cell afterwards wrote the old value while the screen
        # showed the new one. Validating the current edits at the moment of
        # the write makes that impossible rather than guarded against.
        if p2.button("✅ Apply these fixes", key="sb_apply", type="primary",
                     use_container_width=True, disabled=not _edits,
                     help="Validates every row first. If any is bad, nothing "
                          "is written and the problems are listed."):
            from dataview.file_catalog import catalog_status as _cs2
            _tot = {"uwi_rows": 0, "coord_rows": 0, "headers": 0}
            _fail = []
            _paths = dict(zip(_heldv["inventory_id"], _heldv["path"]))

            # VALIDATE EVERYTHING BEFORE WRITING ANYTHING. Same plan_fix the
            # Preview button runs, over the edits as they are RIGHT NOW. A
            # single bad row stops the whole batch: a half-applied repair
            # leaves rows carrying a UWI whose header was never minted, which
            # is worse than not having started.
            _bad, _plan_now = [], []
            with engine.connect() as _vcon:
                for _e in _edits:
                    _p = _cs2.plan_fix(_vcon, _e["inventory_id"], _e["uwi"],
                                       _e["lat"], _e["lon"])
                    for _err in _p["errors"]:
                        _bad.append(f"{_e['file']}: {_err}")
                    for _act in _p["actions"]:
                        _plan_now.append({"file": _e["file"], "will write": _act})
            if _bad:
                st.error("Nothing was written — these rows are not valid:")
                for _b in _bad[:20]:
                    st.write(f"• {_b}")
                st.stop()
            if not _plan_now:
                st.warning("Nothing to write — the values given change nothing.")
                st.stop()

            try:
                # ONE transaction for the whole batch, for the same reason.
                with engine.begin() as _con:
                    for _e in _edits:
                        try:
                            _d = _cs2.apply_fix(
                                _con, _e["inventory_id"], _e["uwi"],
                                _e["lat"], _e["lon"],
                                source_path=_paths.get(_e["inventory_id"]))
                            _tot["uwi_rows"] += _d["uwi_rows"]
                            _tot["coord_rows"] += _d["coord_rows"]
                            _tot["headers"] += int(_d["header_created"])
                        except Exception as _ex:
                            _fail.append(f"{_e['file']}: {_ex}")
                            raise
            except Exception as _ex:
                st.error(f"Nothing was written — the batch rolled back. {_ex}")
            else:
                st.success(
                    f"Applied: {_tot['uwi_rows']:,} row(s) given a UWI · "
                    f"{_tot['coord_rows']:,} header(s) given coordinates · "
                    f"{_tot['headers']} header(s) created. "
                    f"**Now run ② Run promote** — the rows are eligible but "
                    f"still staged until it lifts them, so the Held count "
                    f"below will not drop until you do.")
            for _k in ("sb_plan", "sb_plan_ok", "sb_plan_edits"):
                st.session_state.pop(_k, None)
            _status_run(engine)
            st.session_state["sb_nonce"] = _nonce + 1
            st.rerun()

    # ── Actions ──────────────────────────────────────────────────────────────
    st.markdown("##### ② Process, reject, or restore")
    _n_uncap = (int(_n.get(_cs.ST_INVENTORIED, 0))
                + int(_n.get(_cs.ST_EXTRACTED, 0)))
    st.caption(
        f"Capture stages a file's rows into `cat_*`; promote lifts them into "
        f"`dv_*`. **{_n_uncap:,}** file(s) here have never been captured, so "
        f"promote has nothing of theirs to lift — run capture first, then "
        f"promote. Capture is idempotent (rows are replaced per file)."
    )
    a0, a1, a2, a3, a4 = st.columns(5)

    # 0. Capture. Sits BEFORE promote because that is the order the pipeline
    #    runs them in, and a file that was never captured cannot be promoted or
    #    held — it is not in the backlog yet, it is upstream of it.
    if a0.button("📥 Run capture", key="sb_capture", use_container_width=True,
                 help="Stage pending files' rows into the cat_* mirrors. "
                      "Scoped to the Scan root chosen above. Files with no "
                      "identity may land in the backlog afterwards — that is "
                      "the pipeline working, and ① drains them."):
        _root = (None if st.session_state.get("sb_root") in (None, _SB_ALL)
                 else st.session_state.get("sb_root"))
        try:
            with st.spinner("Capturing…"):
                _cres, _clines = _run_capture(engine, dialect, root=_root)
        except Exception as _ce:
            st.error(f"Capture failed: {type(_ce).__name__}: {_ce}")
        else:
            st.success(
                f"Capture: {_cres.get('capture_files', 0):,} file(s) · "
                f"{_cres.get('capture_rows', 0):,} row(s) staged."
                + (f" {_cres['capture_missing']:,} file(s) missing on disk."
                   if _cres.get("capture_missing") else ""))
            if _clines:
                with st.expander("capture log", expanded=False):
                    st.code("\n".join(str(_l) for _l in _clines[-200:]))
        _status_run(engine)
        st.session_state["sb_nonce"] = _nonce + 1

    # 1. Cure the coord holds. The only reason with a real one-click fix.
    if a1.button("📍 Fill missing coords", key="sb_coords",
                 use_container_width=True,
                 help="Fill cat_well surface coordinates from the gold master, "
                      "then from dv_well — promote's own pre-gate. Wells whose "
                      "location is already known stop being held. The gate is "
                      "not waived."):
        _filled, _note = _run_coord_enrich(engine)
        if _filled:
            st.success(f"Filled {_filled:,} coordinate row(s).{_note}")
        else:
            _why = _note or " No held well matched gold or dv_well."
            st.warning(f"No coordinates filled.{_why}")
        _status_run(engine)
        st.session_state["sb_nonce"] = _nonce + 1
        st.rerun()

    # 2. Re-run promote so cleared rows actually lift.
    if a2.button("⬆ Run promote", key="sb_promote", use_container_width=True,
                 help="Lift every now-eligible row into dataview.dv_*. Run this "
                      "after clearing a reason."):
        from dataview.file_catalog import promote_catalog as _pc
        raw = engine.raw_connection()
        _lines, _err = [], None
        try:
            cur = raw.cursor()
            with st.spinner("Promoting…"):
                _pc.run_promote(cur, None, True, log=_lines.append)
            raw.commit()
        except Exception as e:
            try:
                raw.rollback()
            except Exception:
                pass
            _err = f"{type(e).__name__}: {e}"
        finally:
            try:
                raw.close()
            except Exception:
                pass
        if _err:
            st.error(_err)
        else:
            st.success("Promote complete.")
        st.code("\n".join(_lines) or "(no output)")
        _status_run(engine)
        st.session_state["sb_nonce"] = _nonce + 1

    # 3. Reject -- ONE CLICK, because it is no longer a one-way door. It was
    #    two (arm, then confirm) for the reason destructive actions usually
    #    are: _mark_bad blocklists the file, DROPS the rows it staged and
    #    retires its catalog row as SKIPPED, and nothing in the app could undo
    #    any of it -- nothing ever deleted from BAD_FILE. A mistake was
    #    permanent, so the ceremony bought something real.
    #
    #    Restore below undoes all three, so the ceremony now buys nothing and
    #    costs a click on the common path (Perry rejects files routinely).
    #    "Automation may skip ceremony, never a decision": the decision is
    #    still made -- tick the files, press the button -- what is gone is the
    #    transcription of it.
    #
    #    dv_* is still never touched: rows that already promoted belong to the
    #    load that inserted them, and un-promoting is a separate decision.
    _help = ("Blocklists the selected files, drops the rows they staged, and "
             "retires their catalog rows as SKIPPED. Rows already promoted "
             "into dv_* are left alone. Undo with Restore.")
    if a3.button(f"🚫 Reject {len(_picked)} file(s)", key="sb_reject",
                 use_container_width=True, disabled=_picked.empty,
                 help=_help):
        # RE-ASSERT THE PRECONDITION INSIDE THE HANDLER. `disabled=` is a UI
        # affordance, not a guarantee: it stops a human clicking and stops
        # nothing else. A replayed widget state, or a refactor that moves the
        # button, can reach this block with the gate visibly closed -- that
        # cost 8 files and 160 staged rows on 19 Aug, the button rendering
        # "Reject 0 file(s)" and disabled while the rows went anyway.
        if _picked.empty:
            st.error("Nothing ticked -- nothing was changed.")
            st.stop()
        _reason = "rejected from Status & Backlog"
        _tot, _rows, _fail = 0, 0, []
        for _, _r in _picked.iterrows():
            try:
                _dropped = _mark_bad(engine, _r["inventory_id"], _r["path"],
                                     _r["file"], _r.get("size_kb"), _reason)
                _tot += 1
                _rows += sum(_dropped.values())
            except Exception as e:
                _fail.append(f"{_r['file']}: {type(e).__name__}: {e}")
        st.success(f"Rejected {_tot} file(s); dropped {_rows:,} staged row(s). "
                   f"To undo: filter State = SKIPPED, tick them, press "
                   f"Restore.")
        if _fail:
            st.error("Some rejects failed: " + " | ".join(_fail))
        _status_run(engine)
        st.session_state["sb_nonce"] = _nonce + 1
        st.rerun()

    # 4. Restore -- the inverse of Reject, and the reason Reject is one click.
    #    Rejected files stay in the catalog (retired as SKIPPED, not deleted),
    #    which is what makes this possible at all: filter State = SKIPPED to
    #    find them.
    #
    #    _unmark_bad returns the file to PENDING, not to "done". Reject deleted
    #    its staged rows, so clearing only the blocklist would leave a file
    #    stamped extracted with nothing staged -- run capture afterwards and
    #    the rows come back.
    if a4.button(f"♻ Restore {len(_picked)} file(s)", key="sb_restore",
                 use_container_width=True, disabled=_picked.empty,
                 help="Takes the selected files off the blocklist, clears "
                      "SKIPPED and marks them pending, so the next run "
                      "re-extracts them. Files that were not rejected are "
                      "left alone."):
        if _picked.empty:
            st.error("Nothing ticked -- nothing was changed.")
            st.stop()
        _back, _left, _rfail = 0, 0, []
        for _, _r in _picked.iterrows():
            try:
                if _unmark_bad(engine, _r["inventory_id"]):
                    _back += 1
                else:
                    _left += 1
            except Exception as e:
                _rfail.append(f"{_r['file']}: {type(e).__name__}: {e}")
        if _back:
            st.success(f"Restored {_back} file(s) -- off the blocklist and "
                       f"pending again. Run capture to re-stage their rows."
                       + (f" {_left} were not rejected and were left alone."
                          if _left else ""))
        else:
            st.info(f"Nothing to restore -- none of the {len(_picked)} ticked "
                    f"file(s) were on the blocklist.")
        if _rfail:
            st.error("Some restores failed: " + " | ".join(_rfail))
        _status_run(engine)
        st.session_state["sb_nonce"] = _nonce + 1
        st.rerun()

    # ── The blocklist ───────────────────────────────────────
    # BAD_FILE is the AUTHORITATIVE list of rejected files, and it has to be:
    # rejecting a file that extracted nothing now removes its catalog row, so
    # the grid above cannot show them all any more. Filtering State = SKIPPED
    # finds only the rejects that kept a row; this panel finds every one, and
    # for a pruned file it is the only place it exists at all.
    #
    # A CONTAINER, NOT AN EXPANDER. Status & Backlog already opens expanders of
    # its own and Streamlit refuses to nest them -- the failure surfaces on
    # whatever page draws next, which is how a layout error once got reported
    # as a corrupt SEG-Y file.
    with st.container(border=True):
        try:
            _bl = pd.read_sql(
                "SELECT b.INVENTORY_ID, b.FILE_NAME, b.FILE_SIZE_KB, b.REASON, "
                "       b.MARKED_DATE, b.FILE_PATH, "
                "       CASE WHEN g.INVENTORY_ID IS NULL THEN 'row removed' "
                "            ELSE 'row kept' END AS CATALOG_ROW "
                "  FROM file_catalog.BAD_FILE b "
                "  LEFT JOIN file_catalog.GLOBAL_FILE_CATALOG g "
                "         ON g.INVENTORY_ID = b.INVENTORY_ID "
                " ORDER BY b.MARKED_DATE DESC", engine)
        except Exception as _be:
            _bl = pd.DataFrame()
            st.caption(f"Blocklist unavailable: {type(_be).__name__}: {_be}")

        if _bl.empty:
            st.markdown("**🚫 Blocklist** — empty. Nothing has been rejected.")
        else:
            st.markdown(f"**🚫 Blocklist — {len(_bl):,} rejected file(s)**")
            st.caption(
                "Every rejected file, whether or not it kept a catalog row. "
                "`row removed` means the file had extracted nothing, so its "
                "catalog entry was dropped rather than retired — restoring it "
                "clears the blocklist, and the next scan of its folder rebuilds "
                "the entry under the same id.")
            _blv = _bl.drop(columns=["INVENTORY_ID"]).copy()
            _blv.insert(0, "Select", False)
            _bled = st.data_editor(
                _blv, hide_index=True, use_container_width=True,
                key=f"sb_blocklist_{_nonce}_editor",
                disabled=[c for c in _blv.columns if c != "Select"],
                column_config={
                    "Select": st.column_config.CheckboxColumn("✓", width="small"),
                    "FILE_NAME": st.column_config.TextColumn("File", width="medium"),
                    "FILE_SIZE_KB": st.column_config.NumberColumn("KB", width="small"),
                    "REASON": st.column_config.TextColumn("Why", width="large"),
                    "MARKED_DATE": st.column_config.TextColumn("When", width="small"),
                    "FILE_PATH": st.column_config.TextColumn("Path", width="large"),
                    "CATALOG_ROW": st.column_config.TextColumn("Catalog", width="small"),
                })
            try:
                _blsel = [i for i, v in enumerate(_bled["Select"]) if v]
            except Exception:
                _blsel = []
            if st.button(f"♻ Restore {len(_blsel)} from blocklist",
                         key="sb_blrestore", use_container_width=True,
                         disabled=not _blsel,
                         help="Clears the blocklist entry and, where the catalog "
                              "row survived, marks the file pending again. A file "
                              "whose row was removed comes back on the next scan "
                              "of its folder."):
                if not _blsel:
                    st.error("Nothing ticked — nothing was changed.")
                    st.stop()
                _n_ok, _n_rescan, _bfail = 0, 0, []
                for _i in _blsel:
                    _row = _bl.iloc[_i]
                    try:
                        _unmark_bad(engine, str(_row["INVENTORY_ID"]))
                        _n_ok += 1
                        if _row["CATALOG_ROW"] == "row removed":
                            _n_rescan += 1
                    except Exception as _re:
                        _bfail.append(f"{_row['FILE_NAME']}: "
                                      f"{type(_re).__name__}: {_re}")
                st.success(
                    f"Restored {_n_ok} file(s) off the blocklist."
                    + (f" {_n_rescan} had no catalog row left — re-scan their "
                       f"folder to bring them back." if _n_rescan else
                       " Run capture to re-stage their rows."))
                if _bfail:
                    st.error("Some restores failed: " + " | ".join(_bfail))
                _status_run(engine)
                st.session_state["sb_nonce"] = _nonce + 1
                st.rerun()

    # ── Export ───────────────────────────────────────────────────────────────
    import time as _tm
    _ts = _tm.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        "⬇ Download this view (CSV)",
        data=view.to_csv(index=False).encode("utf-8"),
        file_name=f"catalog_status_{_ts}.csv", mime="text/csv",
        key="sb_csv", use_container_width=True)

    # ── Reference FK review ──────────────────────────────────────────────────
    # NOT inside an expander: promote_fk_review.render opens one per held group,
    # and Streamlit expanders cannot nest — nesting silently breaks the panel.
    if any(r.startswith("unresolved ") for d in res.holds.values() for r in d):
        st.divider()
        st.markdown("##### Resolve held reference codes")
        st.caption("These are the `unresolved …` reasons above. Add a code as "
                   "new vocabulary or map it to an existing one, then Run "
                   "promote.")
        try:
            from dataview.file_catalog.promote_fk_review import render as _rfk
            _rfk(engine, st)
        except Exception as _e:
            st.caption(f"(FK review unavailable: {str(_e)[:150]})")


# =============================================================================
# Tab 3 -- Well Map
# =============================================================================

def _tab_map(engine, dialect):
    from sqlalchemy import text as _t

    st.markdown("#### 🗺️ Well Map")
    st.caption(
        "Wells from FILE_WELL_HEADER with lat/lon. "
        "Click a marker for details and link to source file. "
        "Seismic survey footprints from FILE_SEIS_HEADER."
    )

    # ── Controls ──────────────────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    show_seis  = c1.checkbox("Show seismic footprints", value=True,
                              key="wm_seis")
    show_all   = c2.checkbox("Include suspect coords", value=False,
                              key="wm_all",
                              help="Include wells that may have wrong coordinates")
    tile_style = c3.selectbox("Base map",
        ["CartoDB positron","OpenStreetMap","CartoDB dark_matter"],
        key="wm_tiles")

    # ── Query wells ───────────────────────────────────────────────────────────
    try:
        with engine.connect() as con:
            well_rows = con.execute(_t("""
                SELECT
                    wh.UWI, wh.WELL_NAME, wh.OPERATOR,
                    wh.WELL_FIELD, wh.STATE, wh.COUNTY,
                    CAST(wh.LATITUDE  AS FLOAT) AS LAT,
                    CAST(wh.LONGITUDE AS FLOAT) AS LON,
                    wh.TOTAL_DEPTH, wh.SPUD_DATE,
                    wh.REPORT_TYPE, wh.CONTRACTOR,
                    wh.CONFIDENCE,
                    gfc.FILE_PATH, gfc.FILE_NAME, gfc.FILE_EXT,
                    gfc.CATALOG_READINESS
                FROM file_catalog.FILE_WELL_HEADER wh
                JOIN file_catalog.GLOBAL_FILE_CATALOG gfc
                    ON gfc.INVENTORY_ID = wh.INVENTORY_ID
                WHERE wh.LATITUDE  IS NOT NULL
                  AND wh.LONGITUDE IS NOT NULL
                  AND TRY_CAST(wh.LATITUDE  AS FLOAT) BETWEEN -90  AND 90
                  AND TRY_CAST(wh.LONGITUDE AS FLOAT) BETWEEN -180 AND 180
            """)).fetchall()

            seis_rows = []
            if show_seis:
                try:
                    seis_rows = con.execute(_t("""
                        SELECT
                            sh.SURVEY_NAME, sh.SEIS_SET_TYPE,
                            sh.CONTRACTOR, sh.TRACE_COUNT,
                            CAST(sh.BBOX_MIN_LAT AS FLOAT) AS MIN_LAT,
                            CAST(sh.BBOX_MAX_LAT AS FLOAT) AS MAX_LAT,
                            CAST(sh.BBOX_MIN_LON AS FLOAT) AS MIN_LON,
                            CAST(sh.BBOX_MAX_LON AS FLOAT) AS MAX_LON,
                            gfc.FILE_NAME, gfc.FILE_EXT
                        FROM file_catalog.FILE_SEIS_HEADER sh
                        JOIN file_catalog.GLOBAL_FILE_CATALOG gfc
                            ON gfc.INVENTORY_ID = sh.INVENTORY_ID
                        WHERE sh.BBOX_MIN_LAT IS NOT NULL
                          AND sh.BBOX_MAX_LAT IS NOT NULL
                          AND sh.BBOX_MIN_LON IS NOT NULL
                          AND sh.BBOX_MAX_LON IS NOT NULL
                          AND TRY_CAST(sh.BBOX_MIN_LAT AS FLOAT) BETWEEN -90 AND 90
                          AND TRY_CAST(sh.BBOX_MIN_LON AS FLOAT) BETWEEN -180 AND 0
                    """)).fetchall()
                except Exception:
                    pass

    except Exception as e:
        st.error(f"Query failed: {e}")
        return

    if not well_rows:
        st.warning(
            "No wells with coordinates found. "
            "Run Phase 2 extraction first."
        )
        return

    import pandas as pd
    wells = pd.DataFrame(well_rows, columns=[
        "uwi","well_name","operator","field","state","county",
        "lat","lon","total_depth","spud_date","report_type",
        "contractor","confidence","file_path","file_name",
        "file_ext","readiness",
    ])

    # Optionally filter suspect coordinates
    if not show_all:
        # US bounding box roughly
        wells = wells[
            wells["lat"].between(24, 50) &
            wells["lon"].between(-125, -65)
        ]

    m1, m2, m3 = st.columns(3)
    m1.metric("Wells on map",    len(wells))
    m2.metric("States",          wells["state"].nunique())
    m3.metric("Seismic surveys", len(seis_rows))

    if wells.empty:
        st.warning("No wells in valid US coordinates. "
                   "Check 'Include suspect coords' to show all.")
        return

    # ── Build folium map ──────────────────────────────────────────────────────
    import folium
    center_lat = wells["lat"].mean()
    center_lon = wells["lon"].mean()

    tile_map = {
        "CartoDB positron":    "CartoDB positron",
        "OpenStreetMap":       "OpenStreetMap",
        "CartoDB dark_matter": "CartoDB dark_matter",
    }

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=5,
        tiles=tile_map.get(tile_style, "CartoDB positron"),
    )

    # Color by report type
    type_colors = {
        "WELL_LOG":           "#378ADD",
        "DIRECTIONAL_SURVEY": "#C8922A",
        "OFFICE":             "#888780",
        "UNKNOWN":            "#B4B2A9",
    }

    # ── Well markers ──────────────────────────────────────────────────────────
    for _, w in wells.iterrows():
        color = type_colors.get(w["report_type"], "#378ADD")

        popup_html = f"""
        <div style="font-family:sans-serif;font-size:13px;min-width:200px">
            <b style="font-size:14px">{w['well_name'] or w['uwi']}</b><br>
            <hr style="margin:4px 0">
            <b>UWI:</b> {w['uwi'] or '—'}<br>
            <b>Field:</b> {w['field'] or '—'}<br>
            <b>Operator:</b> {w['operator'] or '—'}<br>
            <b>State:</b> {w['state'] or '—'} · {w['county'] or '—'}<br>
            <b>TD:</b> {w['total_depth'] or '—'} ft<br>
            <b>Type:</b> {w['report_type'] or '—'}<br>
            <b>Readiness:</b> {w['readiness'] or '—'}<br>
            <hr style="margin:4px 0">
            <b>Source:</b> {w['file_name']}<br>
            <small style="color:#666;word-break:break-all">{w['file_path']}</small>
        </div>
        """

        tooltip = f"{w['well_name'] or w['uwi']} · {w['field'] or ''}"

        folium.CircleMarker(
            location=[w["lat"], w["lon"]],
            radius=8,
            color="white",
            weight=1.5,
            fill=True,
            fill_color=color,
            fill_opacity=0.85,
            tooltip=folium.Tooltip(tooltip),
            popup=folium.Popup(popup_html, max_width=280),
        ).add_to(m)

    # ── Seismic footprints ────────────────────────────────────────────────────
    if seis_rows:
        for sr in seis_rows:
            try:
                (sname, stype, contr, traces,
                 min_lat, max_lat, min_lon, max_lon,
                 fname, fext) = sr

                if None in (min_lat, max_lat, min_lon, max_lon):
                    continue

                popup_html = f"""
                <div style="font-family:sans-serif;font-size:13px">
                    <b>{sname or fname}</b><br>
                    <b>Type:</b> {stype or '—'}<br>
                    <b>Contractor:</b> {contr or '—'}<br>
                    <b>Traces:</b> {traces or '—'}<br>
                    <b>File:</b> {fname}
                </div>
                """

                folium.Rectangle(
                    bounds=[[min_lat, min_lon],
                            [max_lat, max_lon]],
                    color="#1D9E75",
                    weight=2,
                    fill=True,
                    fill_color="#1D9E75",
                    fill_opacity=0.1,
                    tooltip=folium.Tooltip(sname or fname),
                    popup=folium.Popup(popup_html, max_width=240),
                ).add_to(m)
            except Exception:
                pass

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_html = """
    <div style="position:fixed;bottom:30px;left:30px;z-index:1000;
                background:white;padding:10px 14px;border-radius:8px;
                border:1px solid #ccc;font-size:12px;font-family:sans-serif">
        <b>Well type</b><br>
        <span style="color:#378ADD">&#9679;</span> Well log (LAS/DLIS)<br>
        <span style="color:#C8922A">&#9679;</span> Directional survey<br>
        <span style="color:#888780">&#9679;</span> Office / other<br>
        <span style="color:#1D9E75">&#9632;</span> Seismic footprint
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))

    # Fit bounds to data
    m.fit_bounds([
        [wells["lat"].min() - 0.5, wells["lon"].min() - 0.5],
        [wells["lat"].max() + 0.5, wells["lon"].max() + 0.5],
    ])

    # Render map HTML directly — avoids st_folium height issues in tabs
    map_html = m._repr_html_()
    st.components.v1.html(
        f'''<div style="width:100%;height:600px;">{map_html}</div>''',
        height=620,
        scrolling=False,
    )

    # ── Export options ────────────────────────────────────────────────────────
    st.divider()
    st.markdown("**Export**")
    ec1, ec2 = st.columns(2)

    # CSV
    ec1.download_button(
        "⬇ Download well locations CSV",
        data=wells.to_csv(index=False),
        file_name="well_locations.csv",
        mime="text/csv",
        key="wm_csv",
    )

    # GeoJSON
    try:
        import json
        features = []
        for _, w in wells.iterrows():
            features.append({
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [w["lon"], w["lat"]],
                },
                "properties": {
                    "uwi":         w["uwi"],
                    "well_name":   w["well_name"],
                    "operator":    w["operator"],
                    "field":       w["field"],
                    "state":       w["state"],
                    "county":      w["county"],
                    "total_depth": w["total_depth"],
                    "report_type": w["report_type"],
                    "file_path":   w["file_path"],
                    "file_name":   w["file_name"],
                    "readiness":   w["readiness"],
                },
            })
        gj = json.dumps({
            "type": "FeatureCollection",
            "features": features,
        }, indent=2)
        ec2.download_button(
            "⬇ Download GeoJSON",
            data=gj,
            file_name="well_locations.geojson",
            mime="application/geo+json",
            key="wm_geojson",
        )
    except Exception:
        pass


# =============================================================================
# Tab 4 -- Header Files
# =============================================================================

def _tab_headers(engine, dialect):
    from sqlalchemy import text as _t

    st.markdown("#### 📋 Header Files")
    st.caption(
        "Query extracted headers from FILE_WELL_HEADER and "
        "FILE_SEIS_HEADER. Export flat CSV for well creation "
        "or load seismic to dv_seis_set."
    )

    sub = st.tabs(["📋 Well Headers", "📡 Seis Headers"])

    # ── Well Headers ──────────────────────────────────────────────────────────
    with sub[0]:
        st.markdown("**Well header flat file**")
        f1, f2 = st.columns(2)
        has_uwi   = f1.checkbox("Has UWI only",     key="wh2_has_uwi")
        has_coord = f2.checkbox("Has Lat/Lon only",  key="wh2_has_coord")

        if st.button("🔍 Query", type="primary", key="wh2_query"):
            try:
                conds = ["1=1"]
                params = {}
                if has_uwi:
                    conds.append("wh.UWI IS NOT NULL AND wh.UWI!=''")
                if has_coord:
                    conds.append(
                        "wh.LATITUDE IS NOT NULL "
                        "AND wh.LONGITUDE IS NOT NULL")
                with engine.connect() as con:
                    rows = con.execute(_t(f"""
                        SELECT
                            gfc.FILE_PATH, gfc.FILE_NAME, gfc.FILE_EXT,
                            gfc.FILE_TYPE_GROUP,
                            gfc.CATALOG_READINESS, gfc.CATALOG_SCORE,
                            wh.UWI, wh.WELL_NAME, wh.OPERATOR,
                            wh.WELL_FIELD, wh.STATE, wh.COUNTY,
                            wh.LATITUDE, wh.LONGITUDE,
                            wh.TOTAL_DEPTH, wh.SPUD_DATE, wh.RIG_RELEASE,
                            wh.REPORT_TYPE, wh.SURVEY_TYPE,
                            wh.CONTRACTOR, wh.CONFIDENCE
                        FROM file_catalog.FILE_WELL_HEADER wh
                        JOIN file_catalog.GLOBAL_FILE_CATALOG gfc
                            ON gfc.INVENTORY_ID = wh.INVENTORY_ID
                        WHERE {" AND ".join(conds)}
                        ORDER BY gfc.CATALOG_SCORE DESC, gfc.FILE_NAME
                    """), params).fetchall()

                df = pd.DataFrame(rows, columns=[
                    "file_path","file_name","extension","type_group",
                    "readiness","score",
                    "uwi","well_name","operator","well_field",
                    "state","county","latitude","longitude",
                    "total_depth","spud_date","rig_release",
                    "report_type","survey_type","contractor","confidence",
                ])
                st.session_state["wh2_df"] = df
            except Exception as e:
                st.error(f"Query failed: {e}")
                st.caption(
                    "Run Phase 2 extraction to populate FILE_WELL_HEADER.")

        df = st.session_state.get("wh2_df")
        if df is None:
            st.info("Click Query. Run Phase 2 extraction if empty.")
        else:
            m1,m2,m3 = st.columns(3)
            m1.metric("Files",       len(df))
            m2.metric("Has UWI",
                      int(df["uwi"].notna().sum()))
            m3.metric("Has Lat/Lon",
                      int((df["latitude"].notna() &
                           df["longitude"].notna()).sum()))
            st.dataframe(df, hide_index=True, use_container_width=True)
            st.download_button(
                "⬇ Export Well Header Flat File",
                data=df.to_csv(index=False),
                file_name="well_header_flat_file.csv",
                mime="text/csv", key="wh2_export",
            )

    # ── Seis Headers ──────────────────────────────────────────────────────────
    with sub[1]:
        st.markdown("**Seismic header flat file**")
        has_survey = st.checkbox("Has survey name only", key="sh2_has_survey")

        if st.button("🔍 Query", type="primary", key="sh2_query"):
            try:
                conds = ["1=1"]
                if has_survey:
                    conds.append(
                        "sh.SURVEY_NAME IS NOT NULL "
                        "AND sh.SURVEY_NAME!=''")
                with engine.connect() as con:
                    rows = con.execute(_t(f"""
                        SELECT
                            gfc.FILE_PATH, gfc.FILE_NAME, gfc.FILE_EXT,
                            gfc.FILE_TYPE_GROUP, gfc.CATALOG_READINESS,
                            sh.SURVEY_NAME, sh.LINE_NAME,
                            sh.SEIS_SET_TYPE, sh.SURVEY_DATE,
                            sh.CONTRACTOR,
                            sh.BBOX_MIN_LAT, sh.BBOX_MAX_LAT,
                            sh.BBOX_MIN_LON, sh.BBOX_MAX_LON,
                            sh.EPSG_CODE, sh.SAMPLE_INTERVAL,
                            sh.TRACE_COUNT, sh.SHOT_FIRST, sh.SHOT_LAST,
                            sh.SURVEY_OUTLINE,
                            sh.INVENTORY_ID
                        FROM file_catalog.FILE_SEIS_HEADER sh
                        JOIN file_catalog.GLOBAL_FILE_CATALOG gfc
                            ON gfc.INVENTORY_ID = sh.INVENTORY_ID
                        WHERE {" AND ".join(conds)}
                        ORDER BY gfc.FILE_NAME
                    """)).fetchall()

                df = pd.DataFrame(rows, columns=[
                    "file_path","file_name","extension","type_group",
                    "readiness","survey_name","line_name",
                    "seis_set_type","survey_date","contractor",
                    "bbox_min_lat","bbox_max_lat",
                    "bbox_min_lon","bbox_max_lon",
                    "epsg_code","sample_interval",
                    "trace_count","shot_first","shot_last",
                    "survey_outline",
                    "inventory_id",
                ])
                st.session_state["sh2_df"] = df
            except Exception as e:
                st.error(f"Query failed: {e}")
                st.caption(
                    "Run Phase 2 extraction to populate FILE_SEIS_HEADER.")

        df = st.session_state.get("sh2_df")
        if df is None:
            st.info("Click Query. Run Phase 2 extraction if empty.")
        else:
            m1,m2,m3 = st.columns(3)
            m1.metric("Seismic files", len(df))
            m2.metric("Has survey",
                      int(df["survey_name"].notna().sum()))
            m3.metric("Has bbox",
                      int(df["bbox_min_lat"].notna().sum()))

            st.dataframe(
                df.drop(columns=["inventory_id"]),
                hide_index=True, use_container_width=True,
            )

            c1, c2 = st.columns(2)
            c1.download_button(
                "⬇ Export Seis Header CSV",
                data=df.to_csv(index=False),
                file_name="seis_header_flat_file.csv",
                mime="text/csv", key="sh2_export",
            )
            if c2.button("🚀 Load to dv_seis_set",
                         key="sh2_load"):
                _load_seis(engine, dialect, df)


def _load_seis(engine, dialect, df):
    """Insert/update seis headers into dataview.dv_seis_set."""
    from sqlalchemy import text as _t

    to_load = df[
        df["survey_name"].notna() &
        (df["survey_name"].astype(str).str.strip() != "")
    ]
    if to_load.empty:
        st.warning("No rows have a survey name.")
        return

    loaded = errors = 0
    for _, row in to_load.iterrows():
        try:
            sid = uuid.uuid4().hex[:40].upper()
            with engine.begin() as con:
                ex = con.execute(_t("""
                    SELECT seis_set_id FROM dataview.dv_seis_set
                    WHERE seis_set_name=:n
                """), {"n": row["survey_name"]}).fetchone()

                if ex:
                    con.execute(_t("""
                        UPDATE dataview.dv_seis_set SET
                            file_path=:fp, catalog_id=:cid,
                            bbox_min_lat=:bmin_lat, bbox_max_lat=:bmax_lat,
                            bbox_min_lon=:bmin_lon, bbox_max_lon=:bmax_lon,
                            epsg_code=:epsg, remark=:remark,
                            geog = CASE WHEN :wkt IS NULL THEN geog ELSE (
                                CASE WHEN geography::STGeomFromText(:wkt,4326)
                                          .STArea()/1000000.0 > 255000000
                                     THEN geography::STGeomFromText(:wkt,4326)
                                          .ReorientObject()
                                     ELSE geography::STGeomFromText(:wkt,4326)
                                END) END,
                            row_changed_by='DataWrangler',
                            row_changed_date=GETUTCDATE()
                        WHERE seis_set_name=:n
                    """), {
                        "fp":       row["file_path"],
                        "cid":      row.get("inventory_id"),
                        "bmin_lat": _safe_num(row.get("bbox_min_lat")),
                        "bmax_lat": _safe_num(row.get("bbox_max_lat")),
                        "bmin_lon": _safe_num(row.get("bbox_min_lon")),
                        "bmax_lon": _safe_num(row.get("bbox_max_lon")),
                        "epsg":     _safe_int(row.get("epsg_code")),
                        "remark":   str(row.get("contractor",""))[:2000],
                        "wkt":      (row.get("survey_outline") or None),
                        "n":        row["survey_name"],
                    })
                else:
                    con.execute(_t("""
                        INSERT INTO dataview.dv_seis_set (
                            seis_set_id, seis_set_name, seis_set_type,
                            file_path, catalog_id,
                            bbox_min_lat, bbox_max_lat,
                            bbox_min_lon, bbox_max_lon,
                            epsg_code, remark, geog, active_ind, source,
                            row_created_by, row_created_date,
                            row_changed_by, row_changed_date
                        ) VALUES (
                            :sid,:sn,:stype,:fp,:cid,
                            :bmin_lat,:bmax_lat,:bmin_lon,:bmax_lon,
                            :epsg,:remark,
                            CASE WHEN :wkt IS NULL THEN NULL ELSE (
                                CASE WHEN geography::STGeomFromText(:wkt,4326)
                                          .STArea()/1000000.0 > 255000000
                                     THEN geography::STGeomFromText(:wkt,4326)
                                          .ReorientObject()
                                     ELSE geography::STGeomFromText(:wkt,4326)
                                END) END,
                            'Y','FILE_CATALOG',
                            'DataWrangler',GETUTCDATE(),
                            'DataWrangler',GETUTCDATE()
                        )
                    """), {
                        "sid":      sid,
                        "sn":       row["survey_name"],
                        "stype":    str(row.get("seis_set_type","2D"))[:40],
                        "fp":       row["file_path"],
                        "cid":      row.get("inventory_id"),
                        "bmin_lat": _safe_num(row.get("bbox_min_lat")),
                        "bmax_lat": _safe_num(row.get("bbox_max_lat")),
                        "bmin_lon": _safe_num(row.get("bbox_min_lon")),
                        "bmax_lon": _safe_num(row.get("bbox_max_lon")),
                        "epsg":     _safe_int(row.get("epsg_code")),
                        "remark":   str(row.get("contractor",""))[:2000],
                        "wkt":      (row.get("survey_outline") or None),
                    })
            loaded += 1
        except Exception:
            errors += 1

    if errors:
        st.error(f"Loaded {loaded}, {errors} errors.")
    else:
        st.success(f"✅ Loaded {loaded} rows to dv_seis_set.")


# =============================================================================
# Tab 5 -- Pipeline  (post-catalog operations: enrich, vault, collect,
#          key wells/surveys, deep catalog). Each button runs the same code
#          as its standalone CLI script; scripts live in ROOT next to this.
# =============================================================================

def _run_op(engine, fn, a, spinner_label):
    """Run a pipeline core function against the app DB, capturing its log.
    Passes a raw pyodbc connection (what enrich/vault/collect expect).
    Returns (lines, result_or_None, error_or_None)."""
    lines, res, err = [], None, None
    raw = engine.raw_connection()
    try:
        with st.spinner(spinner_label):
            res = fn(raw, a, log=lines.append)
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
    finally:
        try:
            raw.close()
        except Exception:
            pass
    return lines, res, err


def _run_op_engine(engine, fn, a, spinner_label):
    """Like _run_op, but passes the SQLAlchemy *engine* (not a raw connection)
    to the core function. Used by deep_catalog, whose cataloger modules need
    the engine to open their own pooled connections."""
    lines, res, err = [], None, None
    try:
        with st.spinner(spinner_label):
            res = fn(engine, a, log=lines.append)
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
    return lines, res, err


def _pipeline_formats():
    """Supported file-format reference, shown behind an expander on the page."""
    with st.expander("📋 Supported formats — what the scanner ingests",
                     expanded=False):
        rows = [
            ("PDF", ", ".join(sorted(PDF_EXTS)),
             "Scout tickets, directional surveys, mud logs, completion reports"),
            ("Well Log", ", ".join(sorted(LOG_EXTS)),
             "Curves / channels \u2192 dv_log_curve (LAS, DLIS, LIS)"),
            ("Seismic", ", ".join(sorted(SEGY_EXTS | P190_EXTS)),
             "Survey identity, bounding box, geometry \u2192 dv_seis_set"),
            ("Shapefile", ", ".join(sorted(SHP_EXTS)),
             "Well / lease / survey geometries"),
            ("Office", ", ".join(sorted(OFFICE_EXTS)),
             "Tabular and document data (Excel, Word, CSV / TSV)"),
            ("WITSML", ", ".join(sorted(WITSML_EXTS)) + "  (1.3.1 / 1.4.1)",
             "Wells, trajectories, log curves"),
            ("OSDU / JSON Well Log", ", ".join(sorted(JSON_LOG_EXTS)),
             "Well, WellLog, Trajectory, Markers, Pressure, SCAL, Field, "
             "Reservoir, Seismic"),
            ("Image", ", ".join(sorted(IMAGE_EXTS)),
             "Catalogued only \u2014 no content extractor"),
        ]
        df = pd.DataFrame(
            rows, columns=["Format group", "Extensions", "What we capture"])
        st.dataframe(df, hide_index=True, use_container_width=True)


def _tab_pipeline(engine, dialect):
    """File Catalog / Pipeline page — hero run, advanced stages, clear tools."""
    # Run Full Pipeline is the first thing on the page; explanatory prose stays
    # tucked behind expanders.
    _pipeline_run_hero(engine, dialect)

    with st.expander("ℹ️ About this pipeline", expanded=False):
        st.caption(
            "The system scans structured and unstructured documents on a shared "
            "drive, extracts the contents, associates the contents to the correct "
            "well and/or seismic survey, and loads to the database as a validated "
            "and trustworthy source of data. Optionally files the qualifying "
            "documents in a digital vault.")

    _pipeline_formats()

    _pipeline_report(engine)

    _seismic_coverage(engine)

    st.divider()
    _pipeline_clear(engine, dialect)


# Words that signal a CRS statement in a free-text seismic header. SEG-Y Rev 0/1
# have NO CRS field, so the projection — if recorded at all — is prose in the
# 3200-byte textual header, phrased however the processor felt like it.
_CRS_HINT_WORDS = (
    "UTM", "ZONE", "DATUM", "PROJECTION", "SPHEROID", "ELLIPSOID", "MERIDIAN",
    "EPSG", "WGS", "ED50", "ED 50", "NAD", "GDA", "AGD", "NZGD", "AMG", "MGA",
    "CLARKE", "BESSEL", "INTERNATIONAL", "GEODETIC", "GRID", "EASTING",
    "NORTHING", "LAMBERT", "MERCATOR", "STEREOGRAPHIC", "COORDINATE",
)


def _seis_text_header(path, max_lines=45):
    """The file's own header text, as lines — the primary place a CRS hides.

    SEG-Y: the 3200-byte textual header, 40 lines x 80 chars. It is EBCDIC by
    the standard but ASCII in plenty of real files, so both are decoded and the
    one yielding more printable characters wins.

    P1/90: the leading 'H' records, which unlike SEG-Y DO state the CRS
    formally — H1400/H1500 datum + spheroid, H1800 projection code, H1900 zone.
    """
    import os as _os
    ext = _os.path.splitext(path)[1].lower()
    if ext in (".p190", ".p90", ".p1"):
        out = []
        with open(path, "r", errors="replace") as fh:
            for line in fh:
                if line[:1].upper() != "H":
                    break
                out.append(line.rstrip("\r\n"))
                if len(out) >= max_lines * 2:
                    break
        return out
    with open(path, "rb") as fh:
        raw = fh.read(3200)
    best, best_score = "", -1
    for enc in ("cp037", "ascii", "latin-1"):      # cp037 = EBCDIC
        try:
            s = raw.decode(enc, errors="replace")
        except Exception:
            continue
        score = sum(1 for ch in s if ch.isprintable() or ch in " \n\t")
        if score > best_score:
            best, best_score = s, score
    lines = [best[i:i + 80].rstrip() for i in range(0, len(best), 80)]
    return [ln for ln in lines][:max_lines]


def _seismic_coverage(engine):
    """Georeferencing status for every SEG-Y / P190 in the catalog, plus the two
    controls needed to act on it.

    WHY THIS EXISTS: a seismic file only reaches the map if FILE_SEIS_HEADER
    holds geometry — a SURVEY_OUTLINE polygon or a complete BBOX_* set. When it
    doesn't, there are two very different causes that look identical in the UI
    but need opposite fixes, and the extractor leaves a fingerprint that tells
    them apart:

      EPSG_CODE NULL + no bbox  -> the trace headers DO carry coordinates, but
                                   they're projected and nothing declares the
                                   CRS. Supply one and re-extract.
      EPSG_CODE 4326 + no bbox  -> the headers carry NO coordinates at all (with
                                   no points sampled the geographic branch runs
                                   vacuously and stamps 4326). A CRS cannot help;
                                   this needs a companion P190 or shapefile.

    Grouped by FOLDER because that is the unit a CRS applies to — one UTM zone
    per directory is the normal shape of a seismic archive, and DV_SEGY_EPSG is
    global to a run.
    """
    import os as _os
    import pandas as pd
    from sqlalchemy import text as _t

    with st.expander("📡 Seismic coverage — georeferencing status", expanded=False):
        _SEIS_EXTS = ("'.segy'", "'.sgy'", "'.seg'", "'.p190'", "'.p90'", "'.p1'")
        try:
            with engine.connect() as _c:
                df = pd.read_sql(_t(f"""
                    SELECT
                      LEFT(g.FILE_PATH,
                           LEN(g.FILE_PATH) - CHARINDEX('\\', REVERSE(g.FILE_PATH))
                          ) AS folder,
                      COUNT(*) AS files,
                      SUM(CASE WHEN sh.SURVEY_OUTLINE IS NOT NULL
                                 OR TRY_CAST(sh.BBOX_MIN_LAT AS float) IS NOT NULL
                               THEN 1 ELSE 0 END) AS mapped,
                      SUM(CASE WHEN sh.INVENTORY_ID IS NOT NULL
                                AND sh.SURVEY_OUTLINE IS NULL
                                AND TRY_CAST(sh.BBOX_MIN_LAT AS float) IS NULL
                                AND sh.EPSG_CODE IS NULL
                               THEN 1 ELSE 0 END) AS needs_crs,
                      SUM(CASE WHEN sh.SURVEY_OUTLINE IS NULL
                                AND TRY_CAST(sh.BBOX_MIN_LAT AS float) IS NULL
                                AND sh.EPSG_CODE IS NOT NULL
                               THEN 1 ELSE 0 END) AS no_coords,
                      SUM(CASE WHEN sh.INVENTORY_ID IS NULL
                               THEN 1 ELSE 0 END) AS never_extracted,
                      MIN(sh.EPSG_CODE)            AS epsg_min,
                      COUNT(DISTINCT sh.EPSG_CODE) AS epsg_n
                    FROM file_catalog.GLOBAL_FILE_CATALOG g
                    LEFT JOIN file_catalog.FILE_SEIS_HEADER sh
                           ON sh.INVENTORY_ID = g.INVENTORY_ID
                    WHERE LOWER(g.FILE_EXT) IN ({",".join(_SEIS_EXTS)})
                    GROUP BY LEFT(g.FILE_PATH,
                             LEN(g.FILE_PATH) - CHARINDEX('\\', REVERSE(g.FILE_PATH)))
                """), _c)
        except Exception as _e:
            st.caption(f"(coverage unavailable: {str(_e)[:140]})")
            return

        if df.empty:
            st.caption("No seismic files catalogued yet.")
            return

        for _c2 in ("files", "mapped", "needs_crs", "no_coords", "never_extracted"):
            df[_c2] = pd.to_numeric(df[_c2], errors="coerce").fillna(0).astype(int)
        df = df.sort_values(["needs_crs", "files"], ascending=False).reset_index(drop=True)

        tot, mp = int(df["files"].sum()), int(df["mapped"].sum())
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Seismic files", f"{tot:,}")
        m2.metric("On the map", f"{mp:,}",
                  delta=f"{(mp / tot * 100):.0f}%" if tot else None,
                  delta_color="off")
        m3.metric("Need a CRS", f"{int(df['needs_crs'].sum()):,}")
        m4.metric("No coordinates", f"{int(df['no_coords'].sum()):,}")

        _show = df.copy()
        # The full path is often 150+ chars and swamps the grid, so lead with
        # the leaf folder name and keep the full path as the last column for
        # copy/paste.
        # RENAME FIRST: the SQL already aliases a column "folder", and
        # DataFrame.insert refuses a name that exists — that raises
        # "cannot insert folder, already exists".
        _show = _show.rename(columns={"folder": "path"})
        _show.insert(0, "folder",
                     [str(p).rstrip("\\/").split("\\")[-1].split("/")[-1] or str(p)
                      for p in _show["path"]])
        def _epsg_cell(v, n):
            """NaN-safe: pandas turns a NULL EPSG into float('nan'), which is
            TRUTHY — so a plain `if not v` prints the literal 'nan', and
            int(nan) raises. Both are guarded here."""
            try:
                blank = v is None or pd.isna(v) or not str(v).strip()
            except (TypeError, ValueError):
                blank = not str(v).strip()
            if blank:
                return "—"
            try:
                n = int(n)
            except (TypeError, ValueError):
                n = 1
            return f"{v}" if n <= 1 else f"{v} (+{n - 1} more)"

        _show["epsg"] = [_epsg_cell(r.epsg_min, r.epsg_n) for r in df.itertuples()]
        _show.insert(0, "", ["✅" if r.needs_crs == 0 and r.no_coords == 0
                             else "🧭" if r.needs_crs else "∅"
                             for r in df.itertuples()])
        _cols = ["", "folder", "files", "mapped", "needs_crs", "no_coords",
                 "never_extracted", "epsg", "path"]
        st.dataframe(_show[[c for c in _cols if c in _show.columns]],
                     hide_index=True, use_container_width=True)
        st.caption("✅ fully mapped · 🧭 has coordinates but no CRS — fixable below · "
                   "∅ no coordinates in the headers at all — needs a companion "
                   "P190 or shapefile, a CRS won't help. "
                   "**epsg** is what extraction actually resolved; blank means "
                   "nothing declared one.")

        st.divider()

        # ── control 0: find out WHAT the CRS is, by reading the file ─────────
        # The fallback field below is useless until you know which EPSG to type.
        # SEG-Y Rev 0/1 have no CRS field, so if it's recorded anywhere it is
        # prose in the 3200-byte textual header — which the _epsg_hint regex may
        # not match even when a human can read it plainly. P1/90 is better: its
        # H records state datum and projection formally.
        st.markdown("**① Find the CRS** — read the header of a file that needs one")
        try:
            with engine.connect() as _c4:
                _cand = [r[0] for r in _c4.execute(_t(f"""
                    SELECT TOP 200 g.FILE_PATH
                      FROM file_catalog.GLOBAL_FILE_CATALOG g
                      LEFT JOIN file_catalog.FILE_SEIS_HEADER sh
                             ON sh.INVENTORY_ID = g.INVENTORY_ID
                     WHERE LOWER(g.FILE_EXT) IN ({",".join(_SEIS_EXTS)})
                     ORDER BY CASE WHEN sh.EPSG_CODE IS NULL THEN 0 ELSE 1 END,
                              g.FILE_PATH
                """)).fetchall()]
        except Exception:
            _cand = []
        if _cand:
            _pickf = st.selectbox(
                "File to inspect (those missing a CRS are listed first)",
                _cand, key="wb_segy_peek_file",
                format_func=lambda p: str(p).split("\\")[-1].split("/")[-1])
            if st.button("🔍 Show text header", key="wb_segy_peek"):
                # ONLY the read is inside the try. Rendering used to be in
                # here too, so any exception raised while DRAWING was
                # reported as "Could not read header" — blaming the SEG-Y
                # file for a bug in this function. That is precisely what
                # happened: the nested st.expander("Full header") that used
                # to sit below raised StreamlitAPIException ("Expanders may
                # not be nested inside other expanders"), and all the user
                # was ever told was that their file could not be read.
                _lines, _pe = None, None
                try:
                    _lines = _seis_text_header(_pickf)
                except Exception as _e:
                    _pe = _e
                if _pe is not None:
                    st.error(f"Could not read header: {str(_pe)[:160]}")
                else:
                    _hits = [ln for ln in _lines
                             if any(w in ln.upper() for w in _CRS_HINT_WORDS)]
                    if _hits:
                        st.markdown("**Lines mentioning a coordinate system:**")
                        st.code("\n".join(_hits), language=None)
                    else:
                        st.warning("No coordinate-system wording found in this "
                                   "header — the CRS will have to come from the "
                                   "survey report, a companion P190/shapefile, "
                                   "or the data provider.")
                    # Inline, not an expander: this block runs inside the
                    # "📡 Seismic coverage" expander and Streamlit forbids
                    # nesting. A toggle cannot stand in for it either — the
                    # body renders only on the run where the button was
                    # pressed, and the rerun a toggle triggers would lose
                    # it. st.code scrolls on its own, so dropping the second
                    # disclosure hides nothing.
                    st.markdown("**Full header**")
                    st.code("\n".join(_lines), language=None)
                    st.caption(str(_pickf))

        st.divider()
        st.markdown("**② Arm the CRS** for files whose header declares none")

        # ── control 1: the fallback CRS, as a visible field ──────────────────
        # This replaces setting DV_SEGY_EPSG in the shell. An environment
        # variable can only be read by a process that already had it at launch,
        # so setting it after Streamlit started silently did nothing — a control
        # whose failure is invisible. Writing os.environ here means the value is
        # in place before the run spawns its pool, and children inherit it.
        c1, c2 = st.columns([2, 1])
        _cur = st.session_state.get("wb_segy_epsg", _os.environ.get("DV_SEGY_EPSG", ""))
        _val = c1.text_input(
            "Fallback CRS for SEG-Y with no CRS in the header (EPSG)",
            value=_cur, key="wb_segy_epsg",
            placeholder="e.g. 32754  (WGS84 / UTM zone 54S)",
            help="Applied ONLY to files whose text header declares no CRS — a "
                 "declared one always wins. This is normal for SEG-Y: Rev 0/1 "
                 "have no CRS field at all, so the projection usually comes "
                 "from the survey report, not the file. One CRS per run, so "
                 "re-extract one zone's folder at a time.")
        _clean = "".join(ch for ch in str(_val or "") if ch.isdigit())
        if _clean:
            _os.environ["DV_SEGY_EPSG"] = _clean
            c2.success(f"EPSG {_clean} armed")
        else:
            _os.environ.pop("DV_SEGY_EPSG", None)
            c2.caption("No fallback CRS set.")

        # ── control 2: force re-extract ──────────────────────────────────────
        # HEADER_EXTRACTED='Y' is CONTENT-based idempotency, so a parser change
        # or a new CRS never invalidates it — without this the next run silently
        # skips every file you just fixed.
        st.markdown("**③ Re-extract** so the change takes effect")
        _folders = ["(all seismic)"] + list(df["folder"])
        f1, f2 = st.columns([3, 1])
        _pick = f1.selectbox("Re-extract which folder?", _folders,
                             key="wb_segy_reextract_folder",
                             help="Clears HEADER_EXTRACTED so the next run "
                                  "re-parses these files. Needed after any "
                                  "extractor change or CRS change — the stamp "
                                  "keys on file content, not on your settings.")
        if f2.button("🔄 Arm re-extract", key="wb_segy_reextract"):
            try:
                _sql = (f"UPDATE file_catalog.GLOBAL_FILE_CATALOG "
                        f"SET HEADER_EXTRACTED='N', ROW_CHANGED_DATE=GETUTCDATE() "
                        f"WHERE LOWER(FILE_EXT) IN ({','.join(_SEIS_EXTS)})")
                _p = {}
                if _pick != "(all seismic)":
                    # ESCAPE '\' so a folder containing _ or % (common in survey
                    # names) is matched literally rather than as a wildcard.
                    _sql += " AND FILE_PATH LIKE :fp ESCAPE '\\'"
                    _p["fp"] = (_pick.replace("\\", "\\\\").replace("%", "\\%")
                                .replace("_", "\\_").replace("[", "\\[")) + "%"
                with engine.begin() as _c3:
                    _n = _c3.execute(_t(_sql), _p).rowcount
                st.success(f"{_n:,} file(s) armed — run the pipeline with "
                           f"Inventory + Capture off and Extract on to re-parse.")
            except Exception as _e2:
                st.error(f"Could not arm re-extract: {str(_e2)[:160]}")


def _pipeline_report(engine):
    """Per-file audit of what the pipeline did to each catalogued file —
    path · file_name · uwi · well/survey name · action taken. A testing aid:
    scope it to the scan path and see at a glance how each file landed."""
    import pandas as pd
    from sqlalchemy import text as _t
    with st.expander("📋 Stage scorecard — extract · capture · promote per file",
                     expanded=False):
        # Scope: default to THIS crawl (files scanned today), with an option to
        # widen to a scan root or the whole catalog.
        try:
            with engine.connect() as _c:
                _roots = [r[0] for r in _c.execute(_t(
                    "SELECT DISTINCT ROOT_PATH "
                    "FROM file_catalog.GLOBAL_FILE_CATALOG "
                    "WHERE NULLIF(LTRIM(RTRIM(ROOT_PATH)),'') IS NOT NULL "
                    "ORDER BY ROOT_PATH")).fetchall()]
        except Exception:
            _roots = []
        sc1, sc2 = st.columns([3, 1])
        _ALL = "(whole catalog)"
        scsel = sc1.selectbox(
            "Scan root", [_ALL] + _roots, index=0, key="score_root_sel",
            help="Limit to a scan root, or the whole catalog.")
        this_crawl = sc2.checkbox(
            "This crawl only", value=True, key="score_this_crawl",
            help="Limit to files scanned today (the current crawl).")
        if not st.button("Run scorecard", key="score_run",
                         use_container_width=True):
            return

        # Lineage comes from promotion_lineage, so this report, the aggregate
        # scorecard above and pipeline_run's run report all answer "did this
        # file's data land?" identically. It used to keep its own table list,
        # which drifted: this one counted dv_well and dv_well_petro_interp, the
        # aggregate counted dv_prod_entity and dv_well_dir_srvy_hdr, and neither
        # counted what the other did.
        #
        # It is also ONE query now. The previous version ran two COUNT(*) per
        # table per file — ~24 round-trips a file, ~38,000 on a 1,600-file
        # catalog — which is why it needed a button and a this-crawl-only
        # default. Each lineage table is now aggregated once and LEFT JOINed.
        from dataview.file_catalog import promotion_lineage as _lin
        try:
            df = _lin.file_detail(
                engine,
                root=(None if (not scsel or scsel == _ALL) else scsel),
                this_crawl=bool(this_crawl))
        except Exception as e:
            st.error(f"Scorecard failed: {type(e).__name__}: {e}")
            return

        if df.empty:
            st.info("No files in scope. Run a crawl, or widen the scope.")
            return

        n_ext = int((df["extract"] == "Y").sum())
        n_cap = int((df["capture"] == "Y").sum())
        n_prom = int((df["promote"] == "Y").sum())
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Files", len(df))
        m2.metric("Extracted", n_ext)
        m3.metric("Captured", n_cap)
        m4.metric("Promoted", n_prom)
        # Per-extension rollup first — it's the view that answers "did the LAS
        # files load?" without scrolling a thousand rows.
        try:
            _roll = df.groupby("ext").agg(
                files=("file", "size"),
                extracted=("extract", lambda c: int((c == "Y").sum())),
                captured=("capture", lambda c: int((c == "Y").sum())),
                promoted=("promote", lambda c: int((c == "Y").sum())),
            ).reset_index().sort_values("files", ascending=False)
            st.markdown("**By extension**")
            st.dataframe(_roll, use_container_width=True, hide_index=True)
        except Exception:
            pass
        st.markdown("**Per file**")
        # `path` is carried for the Excel hyperlink, not for reading on screen —
        # a full Windows path per row would push every useful column off the
        # right edge. It's in the download.
        st.dataframe(df.drop(columns=["path"], errors="ignore"),
                     use_container_width=True, hide_index=True)
        st.caption("captured=Y means rows reached cat_* or dv_*; promoted=Y means "
                   "they reached dv_*. Credit follows INVENTORY_ID lineage, so "
                   "LAS/DLIS/LIS show promoted from dv_well_log(_curve) and SEG-Y "
                   "from its survey in dv_seis_set, neither of which ever gets a "
                   "PROMOTED_AT stamp. A promoted file shows captured=Y even "
                   "though promote drains cat_* — that's expected.")

        import time as _tm
        _ts = _tm.strftime("%Y%m%d_%H%M%S")
        _rdir = st.session_state.get("fp_report", r"C:\Bulk\reports")
        try:
            os.makedirs(_rdir, exist_ok=True)
            _path = os.path.join(_rdir, f"stage_scorecard_{_ts}.csv")
            df.to_csv(_path, index=False)
            st.success(f"Scorecard written to `{_path}`")
        except Exception as e:
            st.warning(f"Shown above; CSV write failed: {type(e).__name__}: {e}")

        # ── Excel export, file name clickable ────────────────────────────────
        # Same convention as the Browse & View exports: the FILE NAME cell is
        # the link and the path column stays plain text, so the sheet reads
        # cleanly and the long path is still there to copy. Offered as a
        # download rather than only a disk write, because the reports folder is
        # on the server and whoever is reading the screen may not be.
        try:
            _xl = _scorecard_xlsx(df)
        except ImportError:
            _xl = None
            st.caption("Excel export needs openpyxl — `pip install openpyxl`. "
                       "The CSV above is unaffected.")
        except Exception as _xe:
            _xl = None
            st.caption(f"Excel export unavailable: {type(_xe).__name__}: {_xe}")
        if _xl:
            st.download_button(
                "⬇ Download Excel (file names clickable)", data=_xl,
                file_name=f"stage_scorecard_{_ts}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
                key="score_xlsx", use_container_width=True)


_PIPE_STAGES = ["scan", "extract", "enrich", "triage",
                "capture", "vault", "promote", "report"]


def _expected_stages(do_scan, do_capture, do_vault, do_promote):
    """The stages a run will actually execute, given the toggles — used as the
    denominator for the progress bar so it reaches 100% on a real finish."""
    exp = ["scan"] if do_scan else []
    exp += ["extract", "enrich", "triage"]
    if do_capture:
        exp.append("capture")
    if do_vault:
        exp.append("vault")
    if do_promote:
        exp.append("promote")
    exp.append("report")
    return exp


def _pipeline_progress(log_lines, expected):
    """Parse the streamed stage banners ('[stage] ▶ starting…' / '[stage] ✓ done')
    into (fraction, active_stage, per-stage checklist marks)."""
    started, done = set(), set()
    for ln in log_lines:
        for s in _PIPE_STAGES:
            if ln.startswith(f"[{s}]"):
                if "✓ done" in ln:
                    done.add(s)
                elif "▶" in ln or "starting" in ln:
                    started.add(s)
    current = None
    for s in expected:
        if s in started and s not in done:
            current = s
    if current is None:
        for s in expected:
            if s in done:
                current = s
    frac = (len(done & set(expected)) / len(expected)) if expected else 0.0
    marks = [f"{'✅' if s in done else ('⏳' if s == current else '▫')} {s}"
             for s in expected]
    return frac, current, marks


def _inventory_report_df(engine):
    """Full file-level inventory status straight from GLOBAL_FILE_CATALOG — every
    inventoried file, with a synthesized `action` (what actually happened to it),
    plus `readiness`, `duplicate`/`copied`/`promoted` flags and the content
    `file_hash` so duplicates can be traced to their canonical. `promoted`/
    `cataloged` count seismic once its survey reaches dv_seis_set. Written to the
    Report root after a run."""
    from sqlalchemy import text as _t
    import pandas as pd
    with engine.connect() as con:
        try:
            _seis_ok = con.execute(_t(
                "SELECT CASE WHEN OBJECT_ID('dataview.dv_seis_set') IS NOT NULL "
                "AND OBJECT_ID('file_catalog.FILE_SEIS_HEADER') IS NOT NULL "
                "THEN 1 ELSE 0 END")).scalar() == 1
        except Exception:
            _seis_ok = False
        if _seis_ok:
            _cte = ("WITH seis_done AS ("
                    "SELECT DISTINCT sh.INVENTORY_ID "
                    "FROM file_catalog.FILE_SEIS_HEADER sh "
                    "JOIN dataview.dv_seis_set ss "
                    "ON ss.seis_set_name = sh.SURVEY_NAME) ")
            _join = "LEFT JOIN seis_done sd ON sd.INVENTORY_ID = g.INVENTORY_ID "
            _seis = "OR sd.INVENTORY_ID IS NOT NULL "
        else:
            _cte = _join = _seis = ""
        sql = f"""
            {_cte}SELECT g.FILE_NAME AS file_name, g.FILE_EXT AS ext,
                   g.MATCHED_UWI AS uwi,
                   CASE
                     WHEN g.DUPLICATE_GROUP IS NOT NULL THEN 'duplicate — skipped'
                     WHEN g.HEADER_EXTRACTED='S' OR g.CATALOG_READINESS='SKIPPED'
                          THEN 'skipped'
                     WHEN (g.PROMOTED_AT IS NOT NULL {_seis}) AND g.VAULTED_AT IS NOT NULL
                          THEN 'vaulted + promoted'
                     WHEN (g.PROMOTED_AT IS NOT NULL {_seis})
                          THEN 'promoted'
                     WHEN g.VAULTED_AT IS NOT NULL THEN 'vaulted'
                     WHEN g.CATALOG_READINESS='CATALOGED' {_seis}THEN 'cataloged'
                     WHEN g.HEADER_EXTRACTED='Y' THEN 'extracted'
                     ELSE 'inventoried (pending)'
                   END AS action,
                   g.CATALOG_READINESS AS readiness,
                   g.HEADER_EXTRACTED AS extracted,
                   CASE WHEN g.DUPLICATE_GROUP IS NOT NULL THEN 'Y' ELSE '' END AS duplicate,
                   CASE WHEN g.VAULTED_AT  IS NOT NULL THEN 'Y' ELSE '' END AS copied,
                   CASE WHEN g.PROMOTED_AT IS NOT NULL {_seis}THEN 'Y' ELSE '' END AS promoted,
                   g.FILE_HASH AS file_hash,
                   g.VAULTED_AT AS vaulted_at, g.PROMOTED_AT AS promoted_at,
                   g.SCAN_DATE AS scanned, g.FILE_PATH AS path
            FROM file_catalog.GLOBAL_FILE_CATALOG g
            {_join}ORDER BY g.SCAN_DATE DESC, g.FILE_NAME
        """
        res = con.execute(_t(sql))
        return pd.DataFrame(res.fetchall(), columns=list(res.keys()))


def _pipeline_run_hero(engine, dialect):
    """The one-click full-pipeline run, shown as the hero at the top."""
    REF = _wb_ref()
    # ── Run Full Pipeline (always visible — not behind an expander) ──────────
    with st.container():
        # Demo control: pick which reference master enrich/triage resolve
        # against. Defaults to the production master; the mini is for testing.
        st.selectbox(
            "Reference master (enrich / triage lookup)",
            options=list(_WB_REF_OPTIONS),
            format_func=lambda x: f"{_WB_REF_OPTIONS.get(x, x)}  ·  {x}",
            key="wb_ref_choice")
        REF = _wb_ref()
        st.caption(f"Resolving against **{REF}**")
        with st.expander("❓ How to use — fields, the Apply toggle & the stages",
                         expanded=False):
            st.markdown(r"""
**Filling out the fields**

- **Scan root folder** — the top folder to crawl. Every supported file beneath
  it (LAS, DLIS, SEG-Y, PDF, shapefile, Excel/Word, WITSML, JSON…) is inventoried.
  Subfolders are included. *Required.*
- **Vault root** — where qualifying documents are filed when **Vault** + **Apply**
  are both on (default `C:\Bulk\Vault`). Files are copied, so the vault is a
  self-contained archive independent of the originals.
- **Report root** — where the run log, enrich report, and inventory report are
  written (default `C:\Bulk\reports`).
- **Formats to scan** — leave blank to scan every supported type, or list a few
  (e.g. `.las, .pdf`) to limit the whole run. The leading dot is optional.
- **Inventory / Capture / Vault / Promote** — which stages run (see below).
- **Parse workers / ⚡ Use all CPU cores** — parsing parallelism. Multi-core runs
  the parse in a detached process pool (fastest for SEG-Y/LAS/PDF); off keeps it
  in-app on one core.

**The Apply toggle — dry-run vs. execute**

This is the safety switch. The metadata stages (**scan, extract, enrich, triage,
capture**) always write — they build the inventory and the `cat_*` staging
mirrors regardless. The **destructive** stages behave differently:

- **Apply OFF (dry-run)** — Vault and Promote only *plan and count*: you see how
  many files would be filed and how many rows would move into `dv_*`, but nothing
  is copied and nothing is written to the golden tables. Safe to preview.
- **Apply ON (execute)** — Vault actually copies files into the vault, and Promote
  actually moves rows from `cat_*` up into the `dv_*` golden tables.

Run with **Apply off first** to review the counts, then turn it on to commit.

**The stages, in order**

1. **Scan** — crawl the folder, fingerprint each file, build/update the inventory
   (`GLOBAL_FILE_CATALOG`), and skip duplicates.
2. **Extract** — parse each file's header/metadata into the file-header tables
   (multi-core).
3. **Enrich** — resolve missing UWIs and fill blank attributes against the
   reference (`WELL_MASTER`); writes the enrich report.
4. **Triage** — normalize UWI/name, cross-fill from inventory and reference, then
   score and tier every file (HIGH / REVIEW / LOW).
5. **Capture** — for files with a resolved UWI, parse the content (PDF surveys &
   scout tickets, shapefiles, LAS curves) into the `cat_*` staging mirrors.
6. **Vault** — file qualifying documents into the curated vault *(plan only unless
   Apply)*.
7. **Promote** — move `cat_*` rows up into the `dv_*` golden tables, holding any
   row whose reference value (UOM, status, datum…) isn't seeded *(plan/count
   unless Apply)*.
8. **Report** — roll up the run, write a markdown summary, and record a
   `PIPELINE_RUN` row.
""")
        # Clear inventory is an ACTION button, which Streamlit forbids inside a
        # form — so it sits above the run-config form below.
        ci1, ci2 = st.columns([1, 4])
        _arm = ci1.checkbox(
            "Arm clear", value=False, key="fp_clearinv_arm",
            help="Safety latch — tick this, then press Clear inventory.")
        if ci2.button("🧹 Clear inventory", key="fp_clearinv", disabled=not _arm,
                      help="Empty the inventory list (GLOBAL_FILE_CATALOG). File IDs "
                           "are deterministic, so re-scanning rebuilds it and any "
                           "extracted headers re-link. dv_* / reference data left "
                           "untouched."):
            try:
                from sqlalchemy import text as _t
                with engine.begin() as _c:
                    _r = _c.execute(_t("DELETE FROM file_catalog.GLOBAL_FILE_CATALOG"))
                st.session_state.pop("fp_scorecard", None)   # force recompute
                st.success(f"Inventory cleared ({_r.rowcount:,} row(s)). Re-scan to "
                           "rebuild.")
            except Exception as e:
                st.error(f"Clear failed: {type(e).__name__}: {e}")

        running = st.session_state.get("fp_running", False)

        # Run config sits in a plain container — NOT a form. A form would make
        # Enter in the scan-root field submit and start the crawl; a container
        # keeps the grouped layout while leaving Run as the only trigger.
        with st.container():
            fp1, fp2, fp3 = st.columns(3)
            fp_root  = fp1.text_input("Scan root folder", value="", key="fp_root",
                                      placeholder=r"D:\data",
                                      help="Quotes and doubled separators are cleaned. A path pasted from JSON or a SQL result often has \\\\ in it, which Windows opens fine but which would catalog every file a second time.")
            # NORMALISE BEFORE IT REACHES ANYTHING. A doubled-separator root
            # scans correctly and silently duplicates the whole catalog,
            # because the id is a hash of the path string. Explorer's
            # "Copy as path" quotes are stripped for the same reason the
            # loader boxes strip them.
            fp_root = _canon_root(fp_root)
            fp_vault = fp2.text_input("Vault root", value=r"C:\Bulk\Vault",
                                      key="fp_vault")
            fp_report = fp3.text_input("Report root", value=r"C:\Bulk\reports",
                                       key="fp_report",
                                       help="Folder for the run log, enrich "
                                            "report, and inventory report.")
            fp_exts_raw = st.text_input(
                "Formats to scan", value="", key="fp_exts",
                placeholder="all supported types — or e.g.  .las, .dlis, .pdf",
                help="Blank = scan every supported type. To narrow the whole run "
                     "(scan · extract · capture), type a comma-separated list of "
                     "extensions, e.g.  .las, .dlis  — the leading dot is optional.")
            s1, s2, s4, s5 = st.columns(4)
            fp_inventory = s1.checkbox(
                "Inventory", value=True, key="fp_inventory",
                help="Scan the folder and build/update the catalog. On = (re)inventory "
                     "first; off = skip the scan and process the existing catalog.")
            fp_capture = s2.checkbox(
                "Capture", value=True, key="fp_capture",
                help="After inventory, parse documents (PDF surveys / scout tickets + "
                     "shapefiles) for every file with a resolved UWI. Off = stop after "
                     "inventory.")
            fp_vaulton = False   # Vault removed from the pipeline; run it from
            #                      the "📦 Vault" section of this page.
            fp_promote = s4.checkbox("Promote", value=True,  key="fp_promote")
            fp_apply   = s5.checkbox("Apply",   value=False, key="fp_apply",
                                     help="Move promote rows into dv_*. "
                                          "Off = plan/count only.")
            tcol, ccol = st.columns([1.5, 4])
            fp_workers = tcol.number_input(
                "Parse workers", min_value=1, max_value=64,
                value=int(st.session_state.get("fp_workers", 6)),
                step=1, key="fp_workers",
                help="Parallel parse workers for the extract stage.")
            # THE RECOGNISER STAGE HAS EXISTED SINCE JULY AND NOTHING TURNED
            # IT ON. pipeline_run._stage_recognise is complete and wired into
            # run_pipeline — but the parameter defaults to False and this page
            # never passed it, so every run used the classifier plus the
            # per-format extractors. That is why a scout ticket reads its
            # casing table 8/8 in the Document Assistant and reported "no
            # detail rows" here: two readers on the same document, and the
            # good one was never asked.
            fp_recognise = ccol.checkbox(
                "🔍 Use the recogniser for capture",
                value=bool(st.session_state.get("fp_recognise", True)),
                key="fp_recognise",
                help="Read tables with the document vocabulary instead of the "
                     "per-format extractors. Covers scout tickets, casing "
                     "records and end-of-well reports — the extractors only "
                     "handle the formats somebody wrote a handler for.")
            fp_force = ccol.checkbox(
                "♻ Force re-extract (ignore what is already catalogued)",
                value=bool(st.session_state.get("fp_force", False)),
                key="fp_force",
                help="Normally a file that is already CATALOGED with an "
                     "unchanged hash is passed over — right for a re-run over "
                     "a big tree, wrong the moment the CODE changes. 1,638 LAS "
                     "files sat skipped as 'already done' while no stage had "
                     "ever processed them, and the only way back in was a "
                     "hand-written DELETE. Files you explicitly SKIPPED stay "
                     "skipped: this ignores 'already processed', not 'leave "
                     "this alone'. Bounded by the scope below, like every "
                     "other stage — forcing decides WHETHER done files are "
                     "redone, not WHICH files are in scope.")
            fp_scope = ccol.radio(
                "Which files does this run process?",
                options=["path", "queue"],
                index=0 if st.session_state.get("fp_scope", "path") == "path"
                        else 1,
                key="fp_scope",
                horizontal=True,
                format_func=lambda v: ("Only under the scan root"
                                       if v == "path"
                                       else "The whole pending inventory"),
                help="Only the SCAN stage was ever scoped to the folder you "
                     "give it. Every stage after it claimed from the whole "
                     "catalog's pending queue, so a run pointed at one folder "
                     "still extracted and captured files from every other tree "
                     "ever scanned — including rows for files you have since "
                     "moved away. 'Only under the scan root' bounds the whole "
                     "run to that folder. 'The whole pending inventory' is the "
                     "old behaviour, kept for finishing work already scanned.")
            fp_multicore = ccol.checkbox(
                "⚡ Use all CPU cores (multi-core parse)",
                value=bool(st.session_state.get("fp_multicore", True)),  # default: multi-core ON
                key="fp_multicore",
                help="ON: run the pipeline in a detached process so the extract "
                     "stage uses a true process pool across every core — same speed "
                     "as the CLI's --parse-mode process. Best for the SEG-Y / LAS / "
                     "PDF parsers, which are GIL-bound otherwise. OFF: run in-app on "
                     "threads (safe default; parsing is capped to one core).")
            bcol1, bcol2 = st.columns([1.5, 4])
            fp_batch = bcol2.checkbox(
                "📦 Batch mode — inventory all, then process N at a time",
                value=bool(st.session_state.get("fp_batch", False)),
                key="fp_batch",
                help="ON: walk the whole tree once to inventory it, then process "
                     "the catalog in batches until the queue is clear. Best for a "
                     "large corpus — bounds memory, gives resume points, and the "
                     "filesystem is only walked once. OFF: a single straight-through "
                     "run.")
            fp_batch_size = bcol1.number_input(
                "Batch size", min_value=50, max_value=100000,
                value=int(st.session_state.get("fp_batch_size", 1000)),
                step=50, key="fp_batch_size",
                disabled=not fp_batch,
                help="Files processed per batch (extract→…→promote), looped until "
                     "the inventory is cleared.")
            run_clicked = st.button(
                "▶ Run pipeline", type="primary", key="fp_run",
                use_container_width=True,
                disabled=(running or not fp_root.strip()
                          or not (fp_inventory or fp_capture)))

        fp_exts_sel = {
            "." + t.strip().lower().lstrip(".")
            for t in fp_exts_raw.replace(";", ",").split(",")
            if t.strip()
        } or None
        if fp_exts_sel:
            # Tabular types are refused outright, not merely warned about. The
            # File Catalog has no CSV/Excel extractor, so honouring the request
            # would inventory rows that can never be extracted — they'd show as
            # "pending" on every future run and never clear. Drop them from the
            # set so the run proceeds with whatever else was asked for.
            _tab = sorted(e for e in fp_exts_sel if e in TABULAR_EXTS)
            if _tab:
                fp_exts_sel = {e for e in fp_exts_sel if e not in TABULAR_EXTS}
                st.error(
                    "Not scanned here: " + ", ".join(_tab) + ". Delimited and "
                    "spreadsheet tables load through the **Bulk Tabular "
                    "Loader**, which maps columns and resolves foreign keys. "
                    "Cataloguing them would only create inventory rows with no "
                    "extractor behind them.")
                if not fp_exts_sel:
                    fp_exts_sel = None
            _unknown = sorted(e for e in (fp_exts_sel or ()) if e not in KNOWN_EXTS)
            if _unknown:
                st.warning("Not a known type (will match nothing): "
                           + ", ".join(_unknown))
            if fp_exts_sel:
                st.caption("This run is limited to: "
                           + ", ".join(sorted(fp_exts_sel)))

        # Enrich always runs (core to the chain). The per-batch stall watchdog is
        # fixed at 180s; Max-files / Sample-N were retired from the UI (CLI only).
        _STALL_TIMEOUT = 180
        import threading, time as _t_mod
        # Surface the REAL write target so a stale sidebar selection (the engine
        # only rebuilds on Connect, not on selectbox change) can't silently send
        # a run to the wrong database.
        try:
            from dataview.import_data import pipeline_run as _plc
            _tgt = (st.session_state.get("dw_conn_spec")
                    or _plc._engine_spec(engine))
            _tdb, _tsrv = _tgt.get("database") or "?", _tgt.get("server") or "?"
            st.caption(f"Target: **{_tsrv} / {_tdb}** — reconnect from the sidebar "
                       f"to change. Promote writes into `dataview.dv_*` here.")
        except Exception:
            pass

        stop_clicked = st.button(
            "⏹ Stop", key="fp_stop", use_container_width=True,
            disabled=not running,
            help="Abort after the current stage finishes — partial results kept.")

        # Guard against re-entry: don't start a second run if one is already
        # going (e.g. a stray rerun while running).
        if run_clicked and not st.session_state.get("fp_running"):
            from dataview.import_data import pipeline_run as _pl
            _ev = threading.Event()
            _log_buf, _result = [], {}
            _stopfile = None

            if fp_multicore:
                # Multi-core: run the pipeline in a DETACHED process (spawn-safe)
                # so the extract stage's process pool can use every core. A
                # reader thread tails the child's stdout into the same log list
                # the poller renders, and loads the child's state file on exit —
                # so timing panel / scorecard keep working. Stop is cooperative
                # via a stop-file the child polls at each stage boundary.
                import tempfile, json as _json, subprocess, sys as _sys
                _runner = os.path.join(
                    os.path.dirname(os.path.abspath(_pl.__file__)),
                    "pipeline_proc_runner.py")
                _tmpd = tempfile.mkdtemp(prefix="dwpipe_")
                _cfgp = os.path.join(_tmpd, "cfg.json")
                _statep = os.path.join(_tmpd, "state.json")
                _stopfile = os.path.join(_tmpd, "stop.flag")
                _spec = (st.session_state.get("dw_conn_spec")
                         or _pl._engine_spec(engine))
                _cfg = {
                    "url":  engine.url.render_as_string(hide_password=False),
                    "server":   _spec.get("server"),
                    "database": _spec.get("database"),
                    "driver":   _spec.get("driver"),
                    "root": fp_root.strip(),
                    "exts": sorted(fp_exts_sel) if fp_exts_sel else None,
                    "workers": int(fp_workers),
                    "dialect": str(dialect),
                    "do_scan":  bool(fp_inventory),
                    "do_enrich": True,
                    "do_capture": bool(fp_capture),
                    "recognise": bool(fp_recognise),
                    "pack": "petroleum",
                    "force": bool(fp_force),
                    "scope": str(fp_scope or "path"),
                    "inventory_only": not bool(fp_capture),
                    "do_vault": bool(fp_vaulton),
                    "vault_root": fp_vault.strip(),
                    "report_root": fp_report.strip() or None,
                    "vault_apply": bool(fp_apply and fp_vaulton),
                    "vault_mode": "copy",
                    "do_promote": bool(fp_promote),
                    "promote_apply": bool(fp_apply and fp_promote),
                    "do_deep": False,
                    "parse_mode": "process" if fp_multicore else "thread",
                    "batch_size": int(fp_batch_size) if fp_batch else None,
                    "stall_timeout": _STALL_TIMEOUT,
                    "ref": REF,
                    "state_out": _statep,
                    "stop_file": _stopfile,
                }
                _logfile = os.path.join(_tmpd, "console.log")
                _cfg["console_log"] = _logfile
                with open(_cfgp, "w", encoding="utf-8") as _f:
                    _json.dump(_cfg, _f)

                # DETACHED launch (page_run.py model): child writes stdout to a LOG
                # FILE; no pipe, no daemon reader thread, no rerun-held handles. The
                # poll loop reads the file. This is what stops orphan/respawn trees.
                if not os.path.exists(_runner):
                    _result["ok"] = False
                    _result["err"] = ("pipeline_proc_runner.py not deployed "
                                      "next to pipeline_run.py")
                    _result["done"] = True
                    _th = None
                else:
                    _env = dict(os.environ, PYTHONIOENCODING="utf-8", PYTHONUTF8="1")
                    _CREATE_NO_WINDOW = 0x08000000
                    _fh = open(_logfile, "w", encoding="utf-8")
                    _repo_root = os.path.dirname(os.path.dirname(
                        os.path.dirname(os.path.abspath(__file__))))
                    _proc = subprocess.Popen(
                        [_sys.executable, "-u", "-m",
                         "dataview.import_data.pipeline_proc_runner", _cfgp],
                        stdout=_fh, stderr=subprocess.STDOUT,
                        cwd=_repo_root, env=_env,
                        creationflags=_CREATE_NO_WINDOW)
                    _result["proc"] = _proc
                    st.session_state["fp_logfile"] = _logfile
                    st.session_state["fp_proc"] = _proc
                    st.session_state["fp_statep"] = _statep
                    _th = None      # no reader thread — the poll loop reads the file
            else:
                # In-app thread path (safe default): parsing stays GIL-bound to
                # one core, but no subprocess. The worker touches no Streamlit
                # APIs; it only appends to the shared log list and watches _ev.
                def _worker(_engine=engine, _root=fp_root.strip(), _exts=fp_exts_sel,
                            _vault=fp_vault.strip(), _cap=fp_capture,
                            _von=fp_vaulton, _prom=fp_promote, _apply=fp_apply,
                            _inv=fp_inventory, _workers=int(fp_workers),
                            _report=fp_report.strip() or None,
                            _batch=(int(fp_batch_size) if fp_batch else None),
                            _dialect=dialect,
                            _rec=bool(fp_recognise),
                            _force=bool(fp_force),
                            _scope=str(fp_scope or "path")):
                    try:
                        _common = dict(
                            workers=_workers, do_enrich=True,
                            do_capture=_cap, dialect=_dialect, do_deep=False,
                            do_vault=_von, vault_root=_vault,
                            vault_apply=_apply and _von, vault_mode="copy",
                            do_promote=_prom, promote_apply=_apply and _prom,
                            per_type_cap=None, stall_timeout=_STALL_TIMEOUT,
                            should_abort=_ev.is_set, ref=REF,
                            parse_mode=("process" if fp_multicore else "thread"),
                            force=_force, scope=_scope,
                            report_root=_report, log=_log_buf.append)
                        if _batch:
                            _state = _pl.run_pipeline_batched(
                                _engine, _root, exts=_exts,
                                batch_size=_batch, scan_first=_inv, **_common)
                        else:
                            _state = _pl.run_pipeline(
                                _engine, _root, exts=_exts,
                                do_scan=_inv,
                                inventory_only=not _cap,  # Capture off ⇒ stop after inventory
                                recognise=_rec, pack="petroleum",
                                max_files=None, **_common)   # force is in _common
                        _result["ok"] = True
                        _result["state"] = _state or {}
                    except Exception as e:
                        _result["ok"] = False
                        _result["err"] = f"{type(e).__name__}: {e}"
                    _result["done"] = True

                _th = threading.Thread(target=_worker, daemon=True)

            st.session_state["fp_abort"]     = _ev
            st.session_state["fp_stopfile"]  = _stopfile
            st.session_state["fp_log"]       = _log_buf
            st.session_state["fp_result"]    = _result
            st.session_state["fp_thread"]    = _th
            st.session_state["fp_running"]   = True
            # per-run scorecard anchor: mark when THIS run began (UTC), so the
            # current-run scorecard can scope to files scanned since now.
            from datetime import datetime as _dtu, timezone as _tzu
            st.session_state["fp_run_started"] = _dtu.now(_tzu.utc).strftime("%Y-%m-%d %H:%M:%S")
            st.session_state["fp_apply_run"] = fp_apply
            st.session_state["fp_vault_run"] = fp_vault.strip()
            st.session_state["fp_report_run"] = fp_report.strip() or r"C:\Bulk\reports"
            if _th is not None:          # detached run has no thread to start
                _th.start()
            # NB: no st.rerun() here. Calling it in the same script run as the
            # form submission can re-deliver the submit, re-enter this handler,
            # and spawn a restart loop ("uncontrolled reruns"). The poll loop at
            # the end (gated on fp_running) drives the live refresh instead, so
            # the progress view still appears on this very run.

        if stop_clicked:
            _ev = st.session_state.get("fp_abort")
            if _ev:
                _ev.set()
            # Multi-core runs in a separate process that can't see the Event, so
            # signal it via the stop-file it polls at each stage boundary.
            _sf = st.session_state.get("fp_stopfile")
            if _sf:
                try:
                    open(_sf, "w").close()
                except Exception:
                    pass
            st.session_state.get("fp_log", []).append(
                "⏹ abort requested — stopping after the current stage…")

        # Live log + completion handling. While the worker runs, poll a few times
        # a second so the log streams and the Stop click can land.
        if st.session_state.get("fp_running"):
            # fp_logfile poll: detached multi-core run writes to a log file — read it
            # fresh each cycle so the log/scorecard update without a reader thread.
            _lf = st.session_state.get("fp_logfile")
            if _lf:
                try:
                    with open(_lf, "r", encoding="utf-8", errors="replace") as _lfh:
                        st.session_state["fp_log"] = _lfh.read().splitlines()
                except Exception:
                    pass
                _pr = st.session_state.get("fp_proc")
                if _pr is not None and _pr.poll() is not None:
                    # process exited — load state file, mark done
                    _res2 = st.session_state.get("fp_result", {})
                    import json as _json2
                    try:
                        with open(st.session_state.get("fp_statep",""), "r",
                                  encoding="utf-8") as _sf2:
                            _res2["state"] = _json2.load(_sf2)
                    except Exception:
                        pass
                    _res2["ok"] = (_pr.returncode == 0)
                    if _pr.returncode != 0 and "err" not in _res2:
                        _res2["err"] = f"runner exit {_pr.returncode}"
                    _res2["done"] = True
                    st.session_state["fp_result"] = _res2
            _log = st.session_state.get("fp_log", [])
            _exp = _expected_stages(fp_inventory, fp_capture, fp_vaulton,
                                    fp_promote)
            _frac, _cur, _marks = _pipeline_progress(_log, _exp)
            _res = st.session_state.get("fp_result", {})
            _th = st.session_state.get("fp_thread")
            _proc_live = st.session_state.get("fp_proc")
            _done = bool(_res.get("done")
                         or (_th is not None and not _th.is_alive())
                         or (_proc_live is not None and _proc_live.poll() is not None))
            st.progress(1.0 if _done else _frac,
                        text=(f"▶ {(_cur or 'starting').upper()} · "
                              f"{int((1.0 if _done else _frac) * 100)}%"))
            st.caption("   ".join(_marks))
            if _done:
                st.session_state["fp_running"] = False
                for _k in ("fp_logfile", "fp_proc", "fp_statep"):
                    st.session_state.pop(_k, None)
                st.session_state["fp_stage_times"] = \
                    (_res.get("state") or {}).get("stage_times") or {}
                try:
                    st.session_state["fp_scorecard"] = _inventory_scorecard(engine)
                except Exception:
                    pass
                # Tee the run log + inventory report to the Report root
                _rr = st.session_state.get("fp_report_run", r"C:\Bulk\reports")
                try:
                    import time as _tm
                    _rdir = _rr
                    os.makedirs(_rdir, exist_ok=True)
                    _ts = _tm.strftime("%Y%m%d_%H%M%S")
                    _logp = os.path.join(_rdir, f"pipeline_{_ts}.log")
                    with open(_logp, "w", encoding="utf-8") as _lf:
                        _lf.write("\n".join(st.session_state.get("fp_log", [])))
                    _csvp = os.path.join(_rdir, f"inventory_{_ts}.csv")
                    _inventory_report_df(engine).to_csv(_csvp, index=False)
                    st.session_state["fp_report_paths"] = (_logp, _csvp)
                except Exception as _e:
                    st.session_state["fp_report_paths"] = None
                    st.warning(f"Couldn't write reports to {_rr}: "
                               f"{type(_e).__name__}: {_e}")
                _aborted = bool(st.session_state.get("fp_abort")
                                and st.session_state["fp_abort"].is_set())
                if _res.get("ok"):
                    if _aborted:
                        st.warning("Pipeline aborted — stages completed before the "
                                   "stop are kept; the rest were skipped.")
                    else:
                        _ap = st.session_state.get("fp_apply_run")
                        st.success(
                            f"Pipeline finished ({'APPLY' if _ap else 'dry-run'}). "
                            f"All reports saved under {_rr}.")
                        # accurate 'what actually promoted' report (reads dv_* directly,
                        # unlike the per-file 'promoted' flag which only covers seismic)
                        if _ap:
                            try:
                                from modules.promote_report_ui import render as _render_promote_report
                                _render_promote_report(engine, st)
                            except Exception as _rx:
                                st.caption(f"(report unavailable: {str(_rx)[:100]})")
                    _paths = st.session_state.get("fp_report_paths")
                    if _paths:
                        st.caption(f"Reports in `{_rr}`  ·  "
                                   f"run report, enrich report, "
                                   f"{os.path.basename(_paths[0])} · "
                                   f"{os.path.basename(_paths[1])}")
                else:
                    st.error(f"Pipeline failed: {_res.get('err', '?')}")
            else:
                # live scorecard: refresh at most every ~5s so the funnel updates
                # without re-aggregating the catalog often enough to slow the run
                _now = _t_mod.monotonic()
                if _now - st.session_state.get("fp_scorecard_at", 0.0) >= 5.0:
                    try:
                        st.session_state["fp_scorecard"] = _inventory_scorecard(engine)
                        st.session_state["fp_scorecard_at"] = _now
                    except Exception:
                        pass
        elif st.session_state.get("fp_report_paths"):
            st.caption(
                f"Last run's reports are in "
                f"`{st.session_state.get('fp_report_run', r'C:\Bulk\reports')}`.")

        _render_stage_timing()
        _render_scorecard(engine)        # rendered every cycle, including mid-run
        # per-run scorecard: what THIS run just did (cumulative table is all crawls)
        try:
            from dataview.file_catalog.current_run_scorecard import render as _render_run_scorecard
            _render_run_scorecard(engine, st, since=st.session_state.get("fp_run_started"))
        except Exception as _rsx:
            st.caption(f"(per-run scorecard unavailable: {str(_rsx)[:100]})")

        # keep polling AFTER the scorecard renders, so each cycle shows fresh numbers
        #
        # POLL INTERVAL — this is what makes the page pulse. st.rerun() redraws
        # the WHOLE hero (controls, toggles, timing, both scorecards) and
        # Streamlit dims stale elements for the duration of every rerun, so a
        # 0.5s loop means the page dims twice a second for the length of the
        # run. Nothing here needs half-second granularity: a run is minutes
        # long and the scorecard behind it only re-aggregates every 5s anyway.
        #
        # Fast for the first few seconds so starting the run feels responsive,
        # then back off — 4x fewer redraws, and the numbers still move well
        # inside human reaction time.
        if st.session_state.get("fp_running"):
            _started = st.session_state.get("fp_poll_t0")
            if _started is None:
                _started = st.session_state["fp_poll_t0"] = _t_mod.monotonic()
            _t_mod.sleep(0.5 if (_t_mod.monotonic() - _started) < 5.0 else 2.0)
            st.rerun()
        else:
            st.session_state.pop("fp_poll_t0", None)   # reset for the next run



_CATALOG_COLS_ENSURED = False


def _inventory_scorecard(engine):
    """Per-file-type funnel: how many files were inventoried vs how far each got
    (extracted → cataloged → vaulted → promoted) and how many are still pending.
    Reads only GLOBAL_FILE_CATALOG with NOLOCK so it can poll live during a run
    without blocking — or being blocked by — the writing worker (dirty reads are
    fine for a progress view; the post-run refresh reads a settled catalog)."""
    global _CATALOG_COLS_ENSURED
    import pandas as pd
    from sqlalchemy import text as _t
    if not _CATALOG_COLS_ENSURED:                 # ensure stamp columns once/process
        with engine.begin() as con:
            con.execute(_t(
                "IF COL_LENGTH('file_catalog.GLOBAL_FILE_CATALOG','VAULTED_AT') IS NULL "
                "ALTER TABLE file_catalog.GLOBAL_FILE_CATALOG ADD VAULTED_AT DATETIME2 NULL;"))
            con.execute(_t(
                "IF COL_LENGTH('file_catalog.GLOBAL_FILE_CATALOG','PROMOTED_AT') IS NULL "
                "ALTER TABLE file_catalog.GLOBAL_FILE_CATALOG ADD PROMOTED_AT DATETIME2 NULL;"))
        try:
            # Covering index so the per-type aggregate is an index-only scan (no
            # base-table scan) — keeps the live read cheap and off the writer's back.
            with engine.begin() as con:
                con.execute(_t(
                    "IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='IX_GFC_scorecard' "
                    "AND object_id=OBJECT_ID('file_catalog.GLOBAL_FILE_CATALOG')) "
                    "CREATE NONCLUSTERED INDEX IX_GFC_scorecard "
                    "ON file_catalog.GLOBAL_FILE_CATALOG (FILE_EXT) "
                    "INCLUDE (HEADER_EXTRACTED, CATALOG_READINESS, VAULTED_AT, PROMOTED_AT);"))
        except Exception:
            pass                                  # perf aid only — never block on it
        _CATALOG_COLS_ENSURED = True
    # WHAT COUNTS AS CATALOGED / PROMOTED lives in promotion_lineage, not here.
    # Formats reach the database by different routes and the per-file stamps
    # only describe one of them: documents capture into cat_* and get
    # PROMOTED_AT stamped, but LAS/DLIS/LIS write dv_well_log(_curve) directly
    # and SEG-Y merges into dv_seis_set — neither is ever stamped. Reading the
    # stamp alone reports the deep-path formats as never promoted when their
    # data is in dv_* and queryable.
    #
    # The honest test is lineage: every dv_ detail table carries the
    # INVENTORY_ID of the file its rows came from. promotion_lineage builds the
    # CTE + LEFT JOIN that expresses it, and the per-file stage scorecard and
    # the run report use the SAME definition, so the three can no longer
    # disagree about the same file.
    from dataview.file_catalog import promotion_lineage as _lin
    with engine.connect() as con:
        _cte, _join, _dprom = _lin.promoted_sql(con, alias="g")
        _ccte, _cjoin, _dcap = _lin.captured_sql(con, alias="g")

        # Merge the two CTE sets into one WITH clause. Both builders emit a
        # complete "WITH ... " string, so strip the keyword off the second and
        # comma-join the bodies.
        _bodies = [c[len("WITH "):].rstrip() for c in (_cte, _ccte) if c]
        _cte = ("WITH " + ", ".join(_bodies) + " ") if _bodies else ""
        _join = _join + _cjoin

        sql = _t(f"""
            {_cte}SELECT
                ISNULL(NULLIF(g.FILE_EXT,''),'(none)')                           AS [type],
                COUNT(*)                                                          AS inventoried,
                SUM(CASE WHEN g.HEADER_EXTRACTED='Y' THEN 1 ELSE 0 END)           AS extracted,
                -- captured: the CAPTURED_HASH stamp (durable — survives the
                -- drain of cat_* on promote) OR rows present in cat_*/dv_* by
                -- lineage. The lineage half is what makes Office and LAS
                -- correct: neither gets CAPTURED_HASH, yet both put rows in
                -- dv_*. Without it .xlsx read 0 cataloged while simultaneously
                -- reporting 126 promoted, which cannot both be true.
                SUM(CASE WHEN g.CAPTURED_HASH IS NOT NULL
                         OR g.CATALOG_READINESS='CATALOGED'
                         OR g.PROMOTED_AT IS NOT NULL {_dcap}THEN 1 ELSE 0 END)   AS cataloged,
                SUM(CASE WHEN g.VAULTED_AT  IS NOT NULL THEN 1 ELSE 0 END)        AS vaulted,
                SUM(CASE WHEN g.PROMOTED_AT IS NOT NULL {_dprom}THEN 1 ELSE 0 END) AS promoted,
                -- pending: nothing happened to it — not extracted, not
                -- captured, no rows anywhere. A file whose data reached dv_*
                -- is done regardless of which route took it there.
                SUM(CASE WHEN (g.HEADER_EXTRACTED IS NULL OR g.HEADER_EXTRACTED IN ('N',''))
                          AND g.CAPTURED_HASH IS NULL
                          AND g.PROMOTED_AT IS NULL
                          AND NOT (1=0 {_dcap})
                         THEN 1 ELSE 0 END)                                       AS pending,
                SUM(CASE WHEN g.HEADER_EXTRACTED='S' THEN 1 ELSE 0 END)           AS skipped
            FROM file_catalog.GLOBAL_FILE_CATALOG g WITH (NOLOCK)
            {_join}GROUP BY ISNULL(NULLIF(g.FILE_EXT,''),'(none)')
            ORDER BY pending DESC, inventoried DESC
        """)
        rows = con.execute(sql).fetchall()
    return pd.DataFrame(rows, columns=["type", "inventoried", "extracted",
                                       "cataloged", "vaulted", "promoted",
                                       "pending", "skipped"])


def _render_stage_timing():
    """Show the last run's per-stage wall-clock, slowest first, so the bottleneck
    is obvious at a glance. Populated from run_pipeline's stage_times on finish."""
    _times = st.session_state.get("fp_stage_times") or {}
    if not _times:
        return
    import pandas as pd
    total = sum(_times.values()) or 1.0
    rows = [{"stage": k, "seconds": round(v, 1), "% of run": round(v / total * 100)}
            for k, v in sorted(_times.items(), key=lambda kv: -kv[1])]
    st.divider()
    st.markdown(f"**⏱ Last run — stage timing (slowest first) · {total:.1f}s total**")
    st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)


def _render_scorecard(engine):
    """Render the inventory-vs-processed scorecard from the cached snapshot in
    session_state. Refreshes happen elsewhere (live throttle while running,
    on-complete, manual button, first open) so this just displays."""
    st.divider()
    live = bool(st.session_state.get("fp_running"))
    sc1, sc2 = st.columns([3, 1])
    sc1.markdown("**📊 Inventory vs processed — by file type**"
                 + ("  🟢 live" if live else ""))
    if sc2.button("↻ Refresh", key="fp_scorecard_refresh",
                  use_container_width=True, disabled=live):
        try:
            st.session_state["fp_scorecard"] = _inventory_scorecard(engine)
        except Exception as e:
            st.caption(f"scorecard unavailable: {type(e).__name__}: {e}")
    if "fp_scorecard" not in st.session_state and not live:   # first open, idle
        try:
            st.session_state["fp_scorecard"] = _inventory_scorecard(engine)
        except Exception as e:
            st.caption(f"scorecard unavailable: {type(e).__name__}: {e}")
    _sc = st.session_state.get("fp_scorecard")
    if _sc is not None and not _sc.empty:
        _tot = _sc.sum(numeric_only=True)
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Inventoried", f"{int(_tot['inventoried']):,}")
        m2.metric("Cataloged",   f"{int(_tot['cataloged']):,}")
        m3.metric("Vaulted",     f"{int(_tot['vaulted']):,}")
        m4.metric("Pending",     f"{int(_tot['pending']):,}")
        st.dataframe(_sc, hide_index=True, use_container_width=True)
        st.caption("**Pending** = nothing reached the database — not extracted, "
                   "no rows in cat_* or dv_* · **Cataloged** = rows staged or "
                   "landed · **Vaulted** = file placed in the vault · "
                   "**Promoted** = the file\'s data is in dv_*, by whichever "
                   "route: documents via cat_*, LAS/DLIS/LIS straight into "
                   "dv_well_log(_curve), SEG-Y via its survey in dv_seis_set. "
                   "Credit is by INVENTORY_ID lineage, not the PROMOTED_AT "
                   "stamp, which the deep-path formats never receive.")
    elif _sc is not None:
        st.caption("Inventory is empty — run an inventory to populate this.")


def _pipeline_stages(engine, dialect):
    """The individual ①–⑧ stages, shown inside the Advanced expander."""
    import types as _types
    try:
        _db   = engine.url.database or "?"
        _host = engine.url.host or "?"
    except Exception:
        _db = _host = "?"
    REF = _wb_ref()
    # ── ① Enrich Headers ───────────────────────────────────────────────────
    with st.expander(
        "① Enrich Headers — curate UWI14 · resolve by name · fill blanks",
        expanded=True,
    ):
        st.caption("Adds/refreshes the canonical UWI14 on FILE_WELL_HEADER, "
                   "resolves blank UWIs by name against the reference, fills blank "
                   "attributes, and records document→reference contributions.")
        c1, c2 = st.columns([1, 3])
        en_dry = c1.checkbox("Dry run", value=True, key="pl_en_dry")
        en_rev = c1.checkbox(
            "Reverse-capture", value=False, key="pl_en_rev",
            help="Write document values back into the reference where it's "
                 "missing them. Slow (full WELL_MASTER scan) — off by default.")
        if c2.button("Run enrichment", type="primary", key="pl_en_run",
                     use_container_width=True):
            from dataview.file_catalog import enrich_file_headers as _en
            a = _types.SimpleNamespace(
                server=_host, database=_db, odbc_driver="",
                ref=REF, depth_tol=50.0,
                no_well=False, no_seis=False, no_reverse=not en_rev,
                dry_run=en_dry, report=None, reverse_report=None)

            _bar   = st.progress(0.0, text="Starting enrichment…")
            _logbox = st.empty()
            _lines = []

            def _log(line):
                _lines.append(str(line))
                # stream live; keep the tail so the box doesn't grow unbounded
                _logbox.code("\n".join(_lines[-400:]))

            def _prog(step, total, label):
                _bar.progress(min(1.0, step / max(total, 1)),
                              text=f"[{step}/{total}] {label}")

            err = None
            raw = engine.raw_connection()
            try:
                _en.enrich(raw, a, log=_log, progress=_prog)
            except Exception as e:
                err = f"{type(e).__name__}: {e}"
            finally:
                try:
                    raw.close()
                except Exception:
                    pass
            _bar.empty()
            if err:
                st.error(err)
            _logbox.code("\n".join(_lines) or "(no output)")

    # ── ② Vault Copy ───────────────────────────────────────────────────────
    with st.expander(
        "② Vault Copy — file catalogued wells & seismic into the vault tree"
    ):
        st.caption(r"Wells → <vault>\curated\wells\<STATE>\<COUNTY>\<UWI>__<NAME>\<class> · "
                   r"Seismic → <vault>\curated\seismic\<2D|3D>\<SURVEY>\<class> · "
                   "spatial / _unmatched as applicable. Files are copied, never moved. "
                   "Same routing as `python vault_organizer.py`.")
        v1, v2 = st.columns(2)
        v_vault = v1.text_input("Vault root", value=r"C:\Bulk\Vault", key="pl_v_vault")
        v_mode  = v2.selectbox("Mode", ["copy", "hardlink", "symlink"], index=0,
                               key="pl_v_mode",
                               help="copy = duplicate bytes (any volume). "
                                    "hardlink / symlink = no extra disk, same "
                                    "volume only.")
        v3, v4 = st.columns([1, 3])
        v_dry = v3.checkbox("Dry run", value=True, key="pl_v_dry")
        if v4.button("Run vault copy", type="primary", key="pl_v_run",
                     use_container_width=True):
            import os as _os
            from dataview.file_catalog import vault_organizer as _vo
            from collections import Counter as _Counter
            try:
                with engine.connect() as _con:
                    rows = _vo.fetch_rows(_con, "file_catalog", None)
                plan, _carried = _vo.build_plan(
                    rows, _os.path.join(v_vault, "curated"))
                if not plan:
                    st.info("Nothing to place — catalog is empty or no files have "
                            "resolved identities yet.")
                else:
                    buckets = _Counter(p[2].split("/")[0] for p in plan)
                    detail = " · ".join(f"{b} {n:,}"
                                        for b, n in sorted(buckets.items()))
                    if v_dry:
                        st.info(f"Planned {len(plan):,} placements (dry-run) — "
                                f"{detail}. Uncheck **Dry run** to write them.")
                    else:
                        placed = exists = errored = 0
                        bar = st.progress(0.0, text="Copying to vault…")
                        for i, (src, dst, _b) in enumerate(plan, 1):
                            try:
                                r = _vo.place(src, dst, v_mode)
                                placed += (r == "ok")
                                exists += (r == "exists")
                            except Exception:
                                errored += 1
                            if i % 50 == 0 or i == len(plan):
                                bar.progress(i / len(plan))
                        bar.empty()
                        st.success(f"Placed {placed:,} ({v_mode}) · {exists:,} "
                                   f"already present · {errored:,} errored — of "
                                   f"{len(plan):,} total. {detail}")
            except Exception as e:
                st.error(f"Vault step failed: {e}")

    # ── ③ Collect Final Documents ──────────────────────────────────────────
    with st.expander(
        "③ Collect Final Documents — gather files whose path contains a keyword"
    ):
        st.caption("Copies catalogued documents whose path contains the keyword "
                   "(as a whole word) into the vault, classified well/seismic/other.")
        d1, d2 = st.columns(2)
        d_vault = d1.text_input("Vault root", value=r"C:\Bulk\Vault", key="pl_d_vault")
        d_word  = d2.text_input("Keyword", value="final", key="pl_d_word")
        d3, d4 = st.columns([1, 3])
        d_dry = d3.checkbox("Dry run", value=True, key="pl_d_dry")
        if d4.button("Run collect", type="primary", key="pl_d_run",
                     use_container_width=True):
            from dataview.file_catalog import collect_final_documents as _cf
            a = _types.SimpleNamespace(
                server=_host, database=_db, odbc_driver="",
                vault=d_vault, dest=None, word=d_word, types=None,
                limit=0, all_ext=False, substring=False, dry_run=d_dry)
            lines, res, err = _run_op(engine, _cf.collect, a, "Collecting documents…")
            if err:
                st.error(err)
            st.code("\n".join(lines) or "(no output)")

    # ── ④ Key Wells & Surveys — review & assign from filename ───────────────
    with st.expander(
        "④ Key Wells & Surveys — review filename-derived UWIs / survey names and assign"
    ):
        st.caption("Files the extractor couldn't key internally, with a guess "
                   "parsed from the path. Edit the assign column and Save — wells "
                   "write UWI + UWI14 (so re-enrichment keeps them), seismic writes "
                   "SURVEY_NAME.")
        sub = st.radio("Review set", ["Wells (need UWI)", "Seismic (need survey)"],
                       horizontal=True, key="pl_rev_kind")
        if sub.startswith("Wells"):
            _well_key_grid(engine)
        else:
            _seis_survey_grid(engine)

    # ── ⑤ Deep Catalog ─────────────────────────────────────────────────────
    with st.expander(
        "⑤ Deep Catalog — parse LAS/DLIS/LIS/SEG-Y/P190 into las_catalog detail tables"
    ):
        st.caption(
            "Runs the per-format deep cataloger over files already in the "
            "catalog, populating las_catalog.* (LAS curves, DLIS frames/channels, "
            "LIS channels, seismic headers). Separate from Phase 2 extraction — "
            "it reads each file once for the detail tables and can run over files "
            "extracted earlier. Failures are tallied per file, never fatal."
        )
        import deep_catalog as _dc
        _groups = list(_dc.DEEP_GROUPS.keys())
        g_sel = st.multiselect("Formats", _groups, default=_groups,
                               key="pl_dc_groups")
        dc1, dc2, dc3 = st.columns([1, 1, 1])
        dc_limit = dc1.number_input("Limit (0 = all)", min_value=0, value=0,
                                    step=10, key="pl_dc_limit")
        dc_workers = dc2.number_input(
            "Workers", min_value=1, max_value=16, value=1, step=1,
            key="pl_dc_workers",
            help="1 = serial (safe). >1 only if the cataloger modules open "
                 "their own connection per call.")
        dc_dry = dc3.checkbox("Dry run", value=True, key="pl_dc_dry")
        dc_hdr = st.checkbox(
            "Headers only (DLIS/LIS skip the channel walk — much faster; "
            "LAS keeps its curve definitions)",
            value=True, key="pl_dc_hdr")
        if st.button("Run deep catalog", type="primary", key="pl_dc_run",
                     use_container_width=True):
            if not g_sel:
                st.warning("Select at least one format.")
            else:
                exts = sorted({e for g in g_sel for e in _dc.DEEP_GROUPS[g]})
                a = _types.SimpleNamespace(
                    server=_host, database=_db, odbc_driver="",
                    exts=",".join(exts), limit=int(dc_limit),
                    workers=int(dc_workers), dry_run=dc_dry,
                    header_only=dc_hdr, report=None)

                # Live progress: a bar + tally driven by deep_catalog's
                # per-file `progress` callback, so a long serial run shows
                # movement instead of an opaque spinner.
                _bar   = st.progress(0.0, text="Starting deep catalog…")
                _tally = st.empty()
                _lines = []

                def _prog(done, total, ok, failed, skipped, name):
                    _bar.progress(
                        min(1.0, done / max(total, 1)),
                        text=f"🧩 Deep catalog — {done:,}/{total:,} · {name}")
                    _tally.caption(
                        f"ok {ok:,} · failed {failed:,} · skipped {skipped:,}")

                res, err = None, None
                try:
                    res = _dc.deep_catalog(
                        engine, a, log=_lines.append, progress=_prog)
                except Exception as e:
                    err = f"{type(e).__name__}: {e}"

                _bar.empty()
                if err:
                    st.error(err)
                if isinstance(res, dict):
                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("Scanned", f"{res.get('scanned', 0):,}")
                    m2.metric("OK", f"{res.get('ok', 0):,}")
                    m3.metric("Failed", f"{res.get('failed', 0):,}")
                    m4.metric("Skipped", f"{res.get('skipped', 0):,}")
                    _errs = res.get("errors") or []
                    if _errs:
                        st.warning(f"⚠️ {len(_errs)} failure(s):")
                        st.code("\n".join(_errs))
                st.code("\n".join(_lines) or "(no output)")

    # ── ⑥ Load Scorecard ────────────────────────────────────────────────────
    with st.expander(
        "⑥ Load Scorecard — files & rows loaded, cataloged, vaulted, by data type"
    ):
        st.caption("Runs catalog_scorecard.py against this database: files and "
                   "rows captured into the cat_* mirrors (to date / this month), "
                   "files cataloged and vaulted, and a per-mirror data-type "
                   "breakdown. Same report as the command-line script.")
        sc1, sc2 = st.columns([1, 3])
        sc_month = sc1.text_input("Month (YYYY-MM · blank = current)", value="",
                                  key="pl_sc_month", placeholder="2026-06")
        if sc2.button("Generate scorecard", type="primary", key="pl_sc_run",
                      use_container_width=True):
            from dataview.file_catalog import catalog_scorecard as _sc
            raw = engine.raw_connection()
            data, err = None, None
            try:
                with st.spinner("Building scorecard…"):
                    data = _sc.gather(raw.cursor(), sc_month.strip() or None)
            except Exception as e:
                err = f"{type(e).__name__}: {e}"
            finally:
                try:
                    raw.close()
                except Exception:
                    pass
            if err:
                st.session_state["pl_sc_err"] = err
                st.session_state.pop("pl_sc_html", None)
            else:
                st.session_state["pl_sc_html"] = _sc.render_html(data, _host, _db)
                st.session_state["pl_sc_text"] = _sc.render_report(data, _host, _db)
                st.session_state["pl_sc_label"] = data["month_label"]
                st.session_state.pop("pl_sc_err", None)

        if st.session_state.get("pl_sc_err"):
            st.error(st.session_state["pl_sc_err"])
        if st.session_state.get("pl_sc_html"):
            import streamlit.components.v1 as _components
            _components.html(st.session_state["pl_sc_html"], height=1500,
                             scrolling=True)
            st.download_button(
                "⬇ Download scorecard HTML",
                data=st.session_state["pl_sc_html"].encode("utf-8"),
                file_name=f"scorecard_{st.session_state.get('pl_sc_label','')}.html",
                mime="text/html", key="pl_sc_dl")
            # A toggle, not an expander: this runs inside the ⑥ expander and
            # Streamlit forbids nesting. A toggle works here because the text
            # lives in session_state, so the rerun it triggers redraws it —
            # the same pattern as ⑦ Triage & Review below.
            if st.toggle("Show plain-text version", key="pl_sc_txt_open"):
                st.code(st.session_state["pl_sc_text"])

    # ── ⑦ Triage & Review ──────────────────────────────────────────────────
    # Not an expander: page_triage.render() has its own expanders, and Streamlit
    # forbids nesting. A toggle reveals it inline instead.
    st.markdown("---")
    st.markdown("**⑦ Triage & Review** — tier the catalog · work the "
                "REVIEW / LOW queues")
    st.caption("The full triage panel, on demand: run a triage pass, then "
               "resolve the REVIEW and LOW worklists. It loads only when you "
               "open it, so it never scans the reference on its own.")
    if st.toggle("Open Triage & Review", key="pl_triage_open"):
        from dataview.file_catalog import page_triage as _ptri
        _ptri.render(engine, dialect)

    # ── ⑧ Promote to Database ───────────────────────────────────────────────
    with st.expander(
        "⑧ Promote to Database — lift cat_* mirror rows up into the dv_* tables"
    ):
        st.caption("Moves captured rows from the file_catalog.cat_* mirrors into "
                   "dataview.dv_*, then clears the promoted rows from the mirrors. "
                   "Same code as `python promote_catalog.py`. Dry run reports "
                   "eligible counts without moving anything.")
        p1, p2 = st.columns(2)
        p_uwi = p1.text_input("Only this UWI (blank = all)", value="",
                              key="pl_pr_uwi", placeholder="42…")
        p_dry = p2.checkbox("Dry run", value=True, key="pl_pr_dry")
        if st.button("Run promote", type="primary", key="pl_pr_run",
                     use_container_width=True):
            from dataview.file_catalog import promote_catalog as _pc
            apply = not p_dry
            raw = engine.raw_connection()
            lines, err = [], None
            try:
                cur = raw.cursor()
                with st.spinner("Promoting…" if apply else "Counting eligible…"):
                    _pc.run_promote(cur, (p_uwi.strip() or None), apply,
                                    log=lines.append)
                if apply:
                    raw.commit()
                else:
                    raw.rollback()      # nothing written; release read locks
            except Exception as e:
                try:
                    raw.rollback()
                except Exception:
                    pass
                err = f"{type(e).__name__}: {e}"
            finally:
                try:
                    raw.close()
                except Exception:
                    pass
            if err:
                st.error(err)
            st.code("\n".join(lines) or "(no output)")
            if not err and apply:
                st.success("Promote complete — rows moved into dv_* and cleared "
                           "from the cat_* mirrors.")

        # ── reference FK review — resolve rows promote HELD on unseeded codes ──
        # Rows whose source/uom/etc. code isn't in its dv_r_* reference are parked
        # (not crashed). This grid surfaces those held codes so you can Add them as
        # new vocabulary or Map them to an existing code, then re-run promote.
        st.divider()
        try:
            from dataview.file_catalog.promote_fk_review import render as _render_promote_fk
            _render_promote_fk(engine, st)
        except Exception as _fkx:
            st.caption(f"(FK review unavailable: {str(_fkx)[:100]})")

    # ── ⑨ Clear / Reset ─────────────────────────────────────────────────────
    with st.expander(
        "⑨ Clear / Reset — wipe file_catalog, las_catalog & dv_* catalog rows "
        "(+ optional vault)"
    ):
        st.caption("Deletes what the document pipeline produced: the "
                   "file_catalog + las_catalog tables and the catalog-derived "
                   "dv_* tables. Reference / spatial tables are left intact. "
                   "How much of that actually goes is the Scope below — the "
                   "default spares every LAS/CSV row. Same code as "
                   "`python clear_catalog.py`. Destructive — type CLEAR to "
                   "enable.")
        # SCOPE IS ASKED, NOT ASSUMED. gather() defaults to 'documents', under
        # which GLOBAL_FILE_CATALOG — which carries INVENTORY_ID — is delete-
        # scoped to pdf/docx/html rows only. Every LAS, CSV, XLSX, DLIS and
        # SEG-Y entry survives. Neither page passed `scope` at all, so the
        # caption above promised a wipe, the clear delivered a document-only
        # delete, and a full wipe was reachable from the CLI and nowhere else.
        from dataview.file_catalog import clear_catalog as _cc_scopes
        x_scope = st.radio(
            "Scope",
            options=_cc_scopes.SCOPES,
            format_func=_cc_scopes.scope_label,
            key="pl_x_scope", horizontal=True)
        if x_scope == "documents":
            st.caption("LAS, CSV, XLSX, DLIS and SEG-Y entries stay in "
                       "GLOBAL_FILE_CATALOG, along with the dv_* rows they "
                       "produced — provenance is kept with the rows it explains.")
        elif x_scope == "documents+las":
            st.caption("Adds the log family (.las / .lis / .dlis): las_catalog, "
                       "their GLOBAL_FILE_CATALOG entries and the dv_* rows they "
                       "produced (dv_well_log, dv_well_log_curve, …) clear "
                       "together, so no log row is left citing a source that is "
                       "gone. CSV, XLSX and SEG-Y entries still stay.")
        else:
            # The same warning main() prints for --scope all, and for the same
            # reason: gather()'s dv_* block stays document-scoped regardless of
            # `scope`, so a wholesale catalog wipe strands the rest.
            st.warning(
                "Wipes the catalog wholesale while the dv_* deletes stay "
                "document-scoped, so every non-document dv_* row (CSV/LAS-"
                "derived) will be left citing a source that no longer exists. "
                "That is orphaned provenance, and selftest's invariants tier "
                "will report it.")
        x1, x2 = st.columns(2)
        x_dv    = x1.checkbox("Include dv_* catalog tables", value=True,
                              key="pl_x_dv")
        x_vault = x2.checkbox(r"Also delete vault\curated", value=False,
                              key="pl_x_vault")
        x_vroot = st.text_input("Vault root (for the vault delete)",
                                value=r"C:\Bulk\Vault", key="pl_x_vroot")
        x_conf = st.text_input(
            "⚠ Type CLEAR in this box to enable the delete button",
            value="", key="pl_x_conf", placeholder="CLEAR")
        bcol1, bcol2 = st.columns(2)
        run_clear = bcol2.button(
            "🗑 Clear now", type="primary", key="pl_x_apply",
            use_container_width=True,
            disabled=(x_conf.strip().upper() != "CLEAR"))

        if bcol1.button("Show counts (dry run)", key="pl_x_dry",
                        use_container_width=True):
            from dataview.file_catalog import clear_catalog as _cc
            raw = engine.raw_connection()
            out = []
            try:
                cur = raw.cursor()
                cur.execute(_cc._SET_OPTS)
                # CAPTURE FIRST, EVERY TIME. _DOC_IDS is a module global, and
                # Streamlit keeps the module loaded across reruns — so gather()'s
                # `_DOC_IDS or capture_doc_ids(...)` fallback silently reuses the
                # ids read by an EARLIER clear in this session, missing every
                # document scanned since. The CLI never hit this because each run
                # is a fresh process that calls capture_doc_ids explicitly.
                #
                # SAME SCOPE TO BOTH CALLS. capture decides WHICH ids are in
                # scope (documents+las adds the log family); gather decides
                # which tables are id-scoped. Pass it to one only and the run
                # reports a scope it isn't performing — gather() raises on the
                # mismatch rather than deleting the narrower set.
                _ids = _cc.capture_doc_ids(cur, out.append, scope=x_scope)
                tbls = _cc.gather(cur, do_dv=x_dv, keep=[],
                                  doc_ids=_ids, scope=x_scope)
                total = sum(n for _, _, n, sc in tbls
                            if n > 0 and sc not in ("skip", "protected"))
                out.append(f"{'table':46} {'rows':>10}  scope")
                out.append("-" * 70)
                for sch, t, n, sc in tbls:
                    rows = ("   (skip)" if sc == "skip"
                            else "  (keep)" if sc == "protected" else f"{n:>10,}")
                    # KEYED OFF gather()'s ACTUAL scope names. This map still said
                    # "inventory" after that scope was renamed "document" and
                    # "protected" was added, so the first table carrying an
                    # INVENTORY_ID raised KeyError — swallowed by the except below,
                    # which rendered a half-drawn table and a bare "ERROR: 'document'"
                    # in place of the counts.
                    tag = {"all": "all rows",
                           "document": _cc.row_label(x_scope),
                           "protected": "PROTECTED — learned state, never cleared",
                           "skip": "no INVENTORY_ID — left intact"}[sc]
                    out.append(f"{sch + '.' + t:46} {rows}  {tag}")
                out.append("-" * 70)
                out.append(f"{'TOTAL rows to delete':46} {total:>10,}")
                raw.rollback()
            except Exception as e:
                out.append(f"ERROR: {type(e).__name__}: {e}")
            finally:
                try:
                    raw.close()
                except Exception:
                    pass
            st.code("\n".join(out) or "(nothing)")

        if run_clear:
            from dataview.file_catalog import clear_catalog as _cc
            raw = engine.raw_connection()
            out, err = [], None
            try:
                cur = raw.cursor()
                cur.execute(_cc._SET_OPTS)
                # Capture before anything deletes: the dv_* scope is derived from
                # the catalog, and this clears the catalog in the same transaction.
                _ids = _cc.capture_doc_ids(cur, out.append, scope=x_scope)
                tbls = _cc.gather(cur, do_dv=x_dv, keep=[],
                                  doc_ids=_ids, scope=x_scope)
                _cc.clear(cur, tbls, out.append, doc_ids=_ids, scope=x_scope)
                raw.commit()
            except Exception as e:
                try:
                    raw.rollback()
                except Exception:
                    pass
                err = f"{type(e).__name__}: {e}"
            finally:
                try:
                    raw.close()
                except Exception:
                    pass
            if not err and x_vault:
                try:
                    _cc.clear_vault(x_vroot.strip(), True, out.append)
                except Exception as e:
                    out.append(f"vault delete failed: {e}")
            if err:
                st.error(err)
            else:
                st.success("Cleared.")
            st.code("\n".join(out) or "(no output)")


_REVIEW_PAGE = 200   # cap rows rendered at once (text_inputs are cheap but not free)



def _pipeline_clear(engine, dialect):
    """Destructive maintenance — empty the file catalog and remove only the
    catalog-promoted rows from the dv_* tables (INVENTORY_ID-scoped). Bulk- or
    hand-loaded rows and every reference / spatial table are left intact. Files
    on disk are never touched."""
    st.markdown("##### \U0001f9f9 Clear catalog & data rows")
    st.caption(
        "Clears the file catalog (GLOBAL_FILE_CATALOG, headers, cat_*) and the "
        "deep las_catalog tables, and removes only the catalog-promoted rows from "
        "the dv_* tables. Bulk-loaded data and all reference / spatial tables are "
        "left intact. Files on disk are never touched. The Scope below decides "
        "how far it reaches — the default leaves every LAS / CSV row standing."
    )
    from dataview.file_catalog import clear_catalog as _cc

    # See the note in the ⑨ Clear / Reset expander: 'documents' scoping keeps
    # every LAS / CSV / XLSX / SEG-Y row in GLOBAL_FILE_CATALOG, which is why
    # "Clear now" left the inventory populated. The choice is surfaced rather
    # than hardcoded, and it defaults to the safe side.
    _scope = st.radio(
        "Scope",
        options=_cc.SCOPES,
        format_func=_cc.scope_label,
        key="clr_scope", horizontal=True)
    if _scope == "documents+las":
        st.caption("Adds the log family (.las / .lis / .dlis): las_catalog, "
                   "their catalog entries and the dv_* rows they produced clear "
                   "together. CSV / XLSX / SEG-Y entries are untouched.")
    if _scope == "all":
        st.warning(
            "Wipes the catalog wholesale while the dv_* deletes stay "
            "document-scoped, so every non-document dv_* row will be left "
            "citing a source that no longer exists (orphaned provenance).")

    if st.button("Preview what would be cleared", key="clr_preview"):
        try:
            raw = engine.raw_connection(); cur = raw.cursor()
            _ids = _cc.capture_doc_ids(cur, lambda *_a: None, scope=_scope)
            rows = _cc.gather(cur, do_dv=True, keep=set(),
                              doc_ids=_ids, scope=_scope)
            raw.close()
            import pandas as _pd
            df = _pd.DataFrame(
                [(f"{s}.{t}", "" if n < 0 else f"{n:,}", scope)
                 for s, t, n, scope in rows],
                columns=["table", "rows", "scope"])
            st.dataframe(df, hide_index=True, use_container_width=True)
        except Exception as e:
            st.error(f"{type(e).__name__}: {e}")

    cc1, cc2 = st.columns([3, 1])
    confirm = cc1.checkbox(
        "Yes, clear the catalog and catalog-derived dv_* rows", key="clr_ok")
    if cc2.button("Clear now", type="primary", key="clr_run",
                  use_container_width=True, disabled=not confirm):
        _box, _lines = st.empty(), []
        def _log(m):
            _lines.append(str(m)); _box.code("\n".join(_lines[-400:]))
        try:
            raw = engine.raw_connection(); cur = raw.cursor()
            cur.execute(_cc._SET_OPTS)          # required for dv_well spatial DML
            # Capture before the deletes, and pass the ids explicitly — the
            # module-level _DOC_IDS fallback goes stale across Streamlit reruns.
            _ids = _cc.capture_doc_ids(cur, _log, scope=_scope)
            rows = _cc.gather(cur, do_dv=True, keep=set(),
                              doc_ids=_ids, scope=_scope)
            _cc.clear(cur, rows, _log, doc_ids=_ids, scope=_scope)
            # reset capture stamps on clear: clearing cat_*/dv_* leaves CAPTURED_HASH
            # stamped, which makes the next crawl's capture skip these files as
            # "already captured" so they never re-capture. Null the stamps so a
            # cleared catalog truly re-processes every file.
            try:
                _rc = cur.execute(
                    "UPDATE file_catalog.GLOBAL_FILE_CATALOG "
                    "SET CAPTURED_HASH = NULL, VAULTED_AT = NULL, PROMOTED_AT = NULL "
                    "WHERE CAPTURED_HASH IS NOT NULL OR VAULTED_AT IS NOT NULL "
                    "OR PROMOTED_AT IS NOT NULL").rowcount
                _log(f"reset capture/vault/promote stamps on {_rc or 0} catalog row(s)")
            except Exception as _se:
                _log(f"(stamp reset skipped: {str(_se)[:80]})")
            raw.commit(); raw.close()
            st.success("Catalog and catalog-derived dv_* rows cleared "
                       "(capture stamps reset — files will re-capture on next run).")
        except Exception as e:
            st.error(f"Clear failed: {type(e).__name__}: {e}")


def _well_key_grid(engine):
    from sqlalchemy import text as _t
    from dataview.core import path_identity as _pi

    with engine.connect() as con:
        has = con.execute(_t(
            "SELECT 1 FROM sys.columns WHERE name='UWI14' "
            "AND object_id=OBJECT_ID('file_catalog.FILE_WELL_HEADER')")).fetchone()
        if not has:
            st.info("No UWI14 column yet — run ① Enrich Headers first.")
            return
        rows = con.execute(_t("""
            SELECT h.WELL_HEADER_ID AS id, h.INVENTORY_ID AS inv,
                   g.FILE_NAME AS fname, g.FILE_PATH AS path,
                   h.WELL_NAME AS well_name, h.UWI AS internal_uwi,
                   h.OPERATOR AS operator, h.TOTAL_DEPTH AS td, h.SPUD_DATE AS spud
            FROM file_catalog.FILE_WELL_HEADER h
            JOIN file_catalog.GLOBAL_FILE_CATALOG g
                   ON g.INVENTORY_ID = h.INVENTORY_ID
            WHERE h.UWI14 IS NULL OR h.UWI14 = '00000000000000'
            ORDER BY g.FILE_NAME""")).fetchall()

    if not rows:
        st.success("Every well header already has a valid UWI14 — nothing to key.")
        return

    total = len(rows)
    rows = rows[:_REVIEW_PAGE]
    if total > _REVIEW_PAGE:
        st.caption(f"{total} files need a UWI — showing the first {_REVIEW_PAGE}. "
                   "Save these, then the next batch appears.")
    else:
        st.caption(f"{total} file(s) need a UWI. Edit the value, then Save.")

    # ---- match well names against the reference, corroborated by total
    #      depth / spud date / operator, so a shared name alone never
    #      auto-fills the wrong well ---------------------------------------- #
    from sqlalchemy import bindparam
    REF = _wb_ref()

    def _norm_op(s):
        return "".join(ch for ch in (s or "").upper() if ch.isalnum())

    def _td_close(a, b, tol=50.0):
        try:
            a, b = float(a), float(b)
        except (TypeError, ValueError):
            return False
        if a <= 0 or b <= 0:
            return False
        return abs(a - b) <= max(tol, 0.01 * max(a, b))

    def _same_spud(a, b):
        if not a or not b:
            return False
        try:
            da = a.date() if hasattr(a, "date") else a
            db = b.date() if hasattr(b, "date") else b
            return da == db
        except Exception:
            return False

    st.caption("Both **Well name** and **UWI** are editable guesses. Fix them, "
               "then **Match against reference** to look each row up — by UWI if "
               "you gave one, otherwise by name — and see what the reference "
               "returns. **Save** writes the UWI to the file's header.")

    # A confirmed UWI from a prior Match is promoted into the UWI field here,
    # before the widget is created (Streamlit forbids setting a widget's state
    # after it's been instantiated in the same run).
    _pending = st.session_state.pop("_pl_well_fill", {})

    h1, h2, h3, h4 = st.columns([3, 2, 2, 3])
    h1.markdown("**File**")
    h2.markdown("**Well name (guess)**")
    h3.markdown("**UWI (guess)**")
    h4.markdown("**Match result**")

    inputs = []   # (id, name_key, uwi_key, result_key, row)
    for r in rows:
        nkey = f"pl_wn_{r.id}"
        ukey = f"pl_wu_{r.id}"
        rkey = f"pl_wr_{r.id}"
        if nkey not in st.session_state:
            st.session_state[nkey] = (r.well_name or "").strip()
        if ukey not in st.session_state:
            st.session_state[ukey] = (_pi.norm_uwi14(r.internal_uwi or "")
                                      or (_pi.uwi14_from_path(r.path or "")[0] or ""))
        if r.id in _pending:
            st.session_state[ukey] = _pending[r.id]   # promote confirmed UWI
        c1, c2, c3, c4 = st.columns([3, 2, 2, 3])
        c1.write(r.fname or _pi._basename(r.path or "") or f"(inventory {r.inv})")
        c2.text_input("name", key=nkey, label_visibility="collapsed",
                      placeholder="well name")
        c3.text_input("uwi", key=ukey, label_visibility="collapsed",
                      placeholder="UWI / API")
        c4.write(st.session_state.get(rkey, "—"))
        inputs.append((r.id, nkey, ukey, rkey, r))

    b1, b2 = st.columns(2)
    do_match = b1.button("🔍 Match against reference", key="pl_well_match",
                         use_container_width=True)
    do_save = b2.button("💾 Save UWIs", type="primary", key="pl_well_save",
                        use_container_width=True)

    if do_match:
        cur = [(rid, nkey, ukey, rkey, r,
                st.session_state.get(nkey, "").strip(),
                _pi.norm_uwi14(st.session_state.get(ukey, "")))
               for rid, nkey, ukey, rkey, r in inputs]
        name_set = sorted({nm for *_x, nm, _u in cur if nm})
        uwi_set = sorted({u for *_x, _nm, u in cur if u})

        cols = _wb_ref_select(engine, REF)
        name_map, uwi_map, ok = {}, {}, True
        try:
            with engine.connect() as con:
                if name_set:
                    qn = _t(f"SELECT {cols} FROM {REF} WHERE WELL_NAME IN :v"
                            ).bindparams(bindparam("v", expanding=True))
                    for m in con.execute(qn, {"v": name_set}).fetchall():
                        name_map.setdefault((m.WELL_NAME or "").strip().upper(),
                                            []).append(m)
                if uwi_set:
                    qu = _t(f"SELECT {cols} FROM {REF} WHERE UWI14 IN :v"
                            ).bindparams(bindparam("v", expanding=True))
                    for m in con.execute(qu, {"v": uwi_set}).fetchall():
                        uwi_map[(m.UWI14 or "").strip()] = m
        except Exception as e:
            st.error(f"Reference lookup failed: {type(e).__name__}: {e}")
            ok = False

        if ok:
            fills = {}
            for rid, nkey, ukey, rkey, r, nm, uw in cur:
                txt, fill = "enter a name or UWI", ""
                if uw and uw in uwi_map:
                    m = uwi_map[uw]
                    loc = " · ".join(x for x in (m.COUNTY, m.PROVINCE_STATE) if x)
                    txt = (f"✓ {uw} · in reference · {m.WELL_NAME or ''}"
                           + (f" ({loc})" if loc else ""))
                    fill = uw
                elif uw:
                    # Same rule as the other grid: the file's own UWI is an
                    # answer, so the result IS the UWI and it fills. Both
                    # copies of this block changed together -- they are the
                    # same screen twice and drifting them is how one gets
                    # fixed and the other keeps the old behaviour.
                    txt = f"● {uw} · not in reference — using the file's own"
                    fill = uw
                elif nm:
                    cands = name_map.get(nm.upper(), [])
                    scored = []
                    for m in cands:
                        sigs = []
                        if _td_close(r.td, m.TOTAL_DEPTH):
                            sigs.append("TD")
                        if _same_spud(r.spud, m.SPUD_DATE):
                            sigs.append("spud")
                        if r.operator and _norm_op(r.operator) == _norm_op(m.OPERATOR_NAME):
                            sigs.append("oper")
                        scored.append((len(sigs), sigs, m))
                    scored.sort(key=lambda x: x[0], reverse=True)
                    if not scored:
                        txt = "✗ name not found in reference"
                    else:
                        top_n, top_sigs, top_m = scored[0]
                        tie = len(scored) > 1 and scored[1][0] == top_n
                        u14 = top_m.UWI14 or top_m.UWI or top_m.API_NUM
                        loc = " · ".join(x for x in (top_m.COUNTY, top_m.PROVINCE_STATE) if x)
                        if top_n >= 1 and not tie:
                            txt = (f"✓ {u14} · name+{'+'.join(top_sigs)}"
                                   + (f" ({loc})" if loc else ""))
                            fill = _pi.norm_uwi14(u14 or "")
                        elif top_n >= 1 and tie:
                            txt = f"⚠ {len(scored)} wells match name+attrs — verify"
                        elif len(cands) == 1:
                            txt = (f"? {u14} · name only — verify"
                                   + (f" ({loc})" if loc else ""))
                        else:
                            txt = f"⚠ {len(cands)} share this name — verify"
                st.session_state[rkey] = txt
                if fill:
                    fills[rid] = fill
            st.session_state["_pl_well_fill"] = fills
            st.rerun()

    if do_save:
        ups = []
        for rid, nkey, ukey, rkey, r in inputs:
            u = _pi.norm_uwi14(st.session_state.get(ukey, ""))
            if u:
                ups.append({"id": rid, "u": u})
        if not ups:
            st.warning("No valid 10–14 digit UWIs to write.")
        else:
            with engine.begin() as con:
                for up in ups:
                    con.execute(_t("UPDATE file_catalog.FILE_WELL_HEADER "
                                   "SET UWI=:u, UWI14=:u WHERE WELL_HEADER_ID=:id"), up)
            st.success(f"Wrote {len(ups)} UWI(s). They'll flow into vault/enrich next run.")
            st.rerun()


def _seis_survey_grid(engine):
    """Delegates to the ONE implementation — see seis_survey_assign.

    This page and its twin each carried a private copy and they had already
    DRIFTED: a data_editor version querying different columns with a LEFT JOIN
    here, a paged text_input grid there. Two spellings of one feature is the
    shape that let the escapechar bug return through a fourth writer.

    The shared module also adds GROUP ASSIGN, which is the common case: 2D
    lines arrive as a set and one survey spans all of them.
    """
    from dataview.file_catalog.seis_survey_assign import seis_survey_grid
    return seis_survey_grid(engine)
