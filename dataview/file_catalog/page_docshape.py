"""
dataview/file_catalog/page_docshape.py
=====================================
Read a document, see what the recogniser made of it, see what would be STORED,
correct it, then check the correction across a batch of similar files.

THREE QUESTIONS THIS PAGE ANSWERS, IN ORDER
-------------------------------------------
    what did it match?      shape, score, destination
    what would be stored?   the actual rows, in destination columns, coerced
    is it right generally?  the same rules across a folder of similar files

The second matters more than it looks. A mapping can be correct at the column
level and still produce rubbish — a depth that maps cleanly but arrives as
"1,240 ft" rather than 1240.0, or an identity that resolves for one table in a
document and not the rest. Only the stored rows show that.

The third is what makes correcting worth doing. Twenty files from one vendor
share one unrecognised table; defining a shape once should fix all twenty, and
the batch view is where you see whether it did.

CORRECTIONS GO TO AN OVERLAY, never to the pack. The pack is hand-written and
version-controlled; the overlay is JSON beside it, learned per deployment and
promotable after review. A correction records that a header WORDING means a
field — an alias — so it applies to every future document, not just this one.
"""
from __future__ import annotations

import os
import sys

import streamlit as st

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)


# --------------------------------------------------------------------------- #
# What would actually be stored
# --------------------------------------------------------------------------- #
def _stored_rows(pack, shape, res, identity, limit=12):
    """Rows as the store would write them: destination columns, coerced.

    Mirrors docshape.store — same coercion, same transform, same identity rule
    — so the page shows what the loader will do rather than a prettier
    approximation of it.
    """
    from docshape.engine.recognise import to_number, INTERNAL_KEYS

    rows = res["rows"]
    tf = (getattr(pack, "transforms", {}) or {}).get(shape)
    if tf:
        rows = tf(rows)

    colmap = (getattr(pack, "columns", {}) or {}).get(shape, {})
    ident_field = getattr(pack, "identity_field", None) or "identity"
    out, unmapped = [], set()

    for r in rows[:limit]:
        rec_out = {}
        for k, v in r.items():
            if k in INTERNAL_KEYS:
                continue
            val = to_number(v) if k in pack.numeric else (
                None if v is None or str(v).strip() == "" else str(v).strip())
            cands = colmap.get(k)
            if cands:
                rec_out[cands[0]] = val
            elif k == ident_field:
                rec_out[ident_field] = val
            else:
                unmapped.add(k)
                rec_out["(" + k + ")"] = val
        iv = rec_out.get(ident_field) or identity
        if iv:
            rec_out[ident_field] = pack.normalise_identity(iv) or iv
        out.append(rec_out)
    return out, sorted(unmapped), bool(tf), len(rows)


def _identity_for(pack, path, results):
    """The store's three-step resolution: header table, document text, filename."""
    ident_field = getattr(pack, "identity_field", None) or "identity"
    for res in results:
        if res["shape"] == "UNKNOWN":
            continue
        for r in res["rows"]:
            v = r.get(ident_field)
            if v and str(v).strip():
                got = pack.normalise_identity(v)
                if got:
                    return got, "a header table in this document"
    if hasattr(pack, "identity_from_text"):
        from docshape.readers import read_text
        got = pack.identity_from_text(read_text(path))
        if got:
            return got, "the document text"
    got = pack.identity_from_name(path)
    if got:
        return got, "the file name"
    return None, None


def _scan(rec, path):
    """[(table_name, rows, result)] for one document."""
    from docshape.readers import read_tables
    out = []
    for name, rows in (read_tables(path) or {}).items():
        if not rows:
            continue
        header = list(rows[0].keys())
        res = rec.read_table(header, [[r.get(k) for k in header] for r in rows])
        res["table"] = name
        out.append((name, rows, res))
    return out



# --------------------------------------------------------------------------- #
# Export — what a capture run would produce, as a workbook
# --------------------------------------------------------------------------- #
def _collect_stored(pack, rec, path, limit=None):
    """(index_entries, {destination: [rows]}) for one document.

    The same code path the store uses, so the workbook is a DRY RUN of a
    capture rather than a separate rendering of it. Nothing is written to any
    database.
    """
    base = os.path.basename(path)
    found = _scan(rec, path)
    identity, ident_src = _identity_for(pack, path,
                                        [r for _a, _b, r in found])
    index, by_table = [], {}
    for tname, rows, res in found:
        dest = res["target"] or (res["shape"] if res["shape"] != "UNKNOWN"
                                 else "(unrecognised)")
        entry = {"document": base, "table": tname, "shape": res["shape"],
                 "score": res["score"], "destination": res["target"] or "",
                 "rows_in": len(rows), "rows_out": 0,
                 "identity": identity or "", "identity_from": ident_src or ""}
        if res["shape"] == "UNKNOWN":
            entry["headers"] = " | ".join(list(rows[0].keys())[:10])
            index.append(entry)
            continue
        stored, unmapped, transformed, n_out = _stored_rows(
            pack, res["shape"], res, identity, limit=limit or 100000)
        entry["rows_out"] = n_out
        entry["unmapped"] = ", ".join(unmapped)
        index.append(entry)
        for r in stored:
            by_table.setdefault(dest, []).append(
                {"_document": base, "_table": tname, **r})
    return index, by_table


def _write_workbook(index_rows, by_table, out_path):
    """INDEX sheet plus one sheet per destination, pooled across documents."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    def sheet(wb, title, rows, used):
        bad = set('[]:*?/\\')
        name = "".join(c if c not in bad else "_" for c in str(title))[:31] or "s"
        n, i = name, 2
        while n.lower() in used:
            sfx = "~" + str(i)
            n = name[:31 - len(sfx)] + sfx
            i += 1
        used.add(n.lower())
        ws = wb.create_sheet(n)
        if not rows:
            ws["A1"] = "(no rows)"
            return ws
        keys = []
        for r in rows:
            for k in r:
                if k not in keys:
                    keys.append(k)
        for j, k in enumerate(keys, 1):
            ws.cell(row=1, column=j, value=k).font = Font(bold=True)
        for i2, r in enumerate(rows, 2):
            for j, k in enumerate(keys, 1):
                v = r.get(k)
                ws.cell(row=i2, column=j,
                        value=v if isinstance(v, (int, float))
                        else ("" if v is None else str(v)))
        ws.freeze_panes = "A2"
        for j, k in enumerate(keys, 1):
            w = max([len(str(k))] + [len(str(r.get(k, ""))) for r in rows[:300]])
            ws.column_dimensions[get_column_letter(j)].width = min(44, max(9, w + 2))
        return ws

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    sheet(wb, "INDEX", index_rows, used)
    for dest in sorted(by_table, key=lambda d: -len(by_table[d])):
        sheet(wb, dest, by_table[dest], used)
    wb.save(out_path)
    return out_path


def _export_controls(index_rows, by_table, default_name, key):
    """Write the workbook, offer download, and offer to open it locally."""
    import tempfile
    total = sum(len(v) for v in by_table.values())
    if not total and not index_rows:
        st.caption("Nothing to export.")
        return
    out_dir = st.session_state.get("ds_outdir")
    if not out_dir:
        out_dir = tempfile.mkdtemp()
        st.session_state["ds_outdir"] = out_dir

    c1, c2, c3 = st.columns([1.2, 1.2, 3])
    if c1.button("⬇ Build spreadsheet", key="ds_xl_" + key,
                 type="primary"):
        path = os.path.join(out_dir, default_name)
        try:
            _write_workbook(index_rows, by_table, path)
            st.session_state["ds_xlsx_" + key] = path
        except Exception as e:
            st.error("Could not write: " + type(e).__name__ + ": " + str(e))

    made = st.session_state.get("ds_xlsx_" + key)
    if made and os.path.exists(made):
        with open(made, "rb") as f:
            c2.download_button(
                "💾 Download", f.read(), file_name=os.path.basename(made),
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet", key="ds_dl_" + key)
        # Streamlit runs on the same machine here, so opening it directly is
        # the shortest path to actually looking at the data.
        if c3.button("📂 Open in Excel", key="ds_open_" + key):
            try:
                if hasattr(os, "startfile"):
                    os.startfile(made)          # Windows
                else:
                    import subprocess, sys as _s
                    subprocess.Popen(
                        ["open" if _s.platform == "darwin" else "xdg-open", made])
                st.caption("Opened " + os.path.basename(made))
            except Exception as e:
                st.warning("Could not open it here: " + str(e)
                           + " — use Download instead.")
        st.caption(str(total) + " row(s) across " + str(len(by_table))
                   + " destination(s)  ·  " + made)


# --------------------------------------------------------------------------- #
def render(engine=None):
    from docshape import Recogniser
    from docshape.packs import available
    from docshape.packs.overlay import (empty, load_layered, promote_sandbox,
                                        save_overlay)

    st.markdown("## 🔍 Document recogniser — bench")
    st.caption("Reads documents and reports. Writes no data anywhere — the "
               "only thing it changes is the VOCABULARY.")

    packs = available()
    c1, c2, c3 = st.columns([3.2, 1.2, 1.1])
    mode = c1.radio("Mode", ["Single document", "Batch"], horizontal=True,
                    key="ds_mode", label_visibility="collapsed")
    pack_name = c2.selectbox("Vocabulary", packs,
                             index=packs.index("petroleum")
                             if "petroleum" in packs else 0, key="ds_pack")
    use_sandbox = c3.toggle("🧪 Sandbox", key="ds_sandbox")

    pack, overlay, overlay_path, sandbox, sb_path = load_layered(
        pack_name, use_sandbox=use_sandbox)
    if overlay is None:
        overlay = empty(pack_name)
    if use_sandbox and sandbox is None:
        sandbox = empty(pack_name)
    # Corrections land wherever we are working. The sandbox sits ON TOP of the
    # overlay, so an experiment is judged against the established vocabulary
    # rather than against the bare pack.
    target_overlay = sandbox if use_sandbox else overlay
    target_path = sb_path if use_sandbox else overlay_path

    rec = Recogniser(pack)
    bits = [str(len(pack.fields)) + " fields", str(len(pack.shapes)) + " shapes",
            "overlay: " + str(len(overlay.get("log", []))) + " correction(s)"]
    if use_sandbox:
        bits.append("sandbox: " + str(len(sandbox.get("log", []))) + " pending")
    st.caption(" · ".join(bits))

    if use_sandbox:
        n_pending = (len(sandbox.get("fields", {}))
                     + len(sandbox.get("shapes", {})))
        b1, b2, b3 = st.columns([1.3, 1, 3])
        if b1.button("⬆ Promote to vocabulary", type="primary",
                     disabled=not n_pending, key="ds_promote"):
            overlay, n = promote_sandbox(sandbox, overlay)
            save_overlay(overlay, overlay_path)
            if os.path.exists(sb_path):
                os.remove(sb_path)
            st.success("Promoted " + str(n) + " entr(y/ies) into the "
                       "vocabulary. Sandbox cleared.")
            st.rerun()
        if b2.button("🗑 Discard sandbox", disabled=not n_pending,
                     key="ds_drop_sb"):
            if os.path.exists(sb_path):
                os.remove(sb_path)
            st.rerun()
        if n_pending:
            b3.info("Sandbox is active — corrections are speculative and do "
                    "NOT affect capture runs until promoted.")
        else:
            b3.caption("Sandbox is empty. Corrections made now go here, not to "
                       "the live vocabulary.")

    if mode == "Batch":
        _batch(rec, pack)
    else:
        _single(rec, pack, target_overlay, target_path, use_sandbox)


# --------------------------------------------------------------------------- #
def _single(rec, pack, overlay, overlay_path, use_sandbox=False):
    import pandas as pd
    from docshape.packs.overlay import add_alias, add_shape, save_overlay, promote_summary
    from docshape.readers import SUPPORTED, NATIVE_EXTS

    c1, c2 = st.columns([3, 1])
    path = c1.text_input("Document path", key="ds_path",
                         placeholder=r"C:\docs\Survey_PIONEER_2H.pdf")
    up = c2.file_uploader("or upload", type=[e.lstrip(".") for e in SUPPORTED],
                          label_visibility="collapsed", key="ds_upload")
    if up is not None:
        import tempfile
        if not st.session_state.get("ds_tmpdir"):
            st.session_state["ds_tmpdir"] = tempfile.mkdtemp()
        path = os.path.join(st.session_state["ds_tmpdir"], up.name)
        with open(path, "wb") as f:
            f.write(up.getbuffer())

    if not path:
        st.info("Give a document path, or upload one.")
        return
    if not os.path.exists(path):
        st.error("Not found: " + path)
        return
    if os.path.splitext(path)[1].lower() in NATIVE_EXTS:
        st.info("LAS and SEG-Y are parsed natively — no tables to correct.")
        return

    with st.expander("📄 Document", expanded=False):
        try:
            from dataview.file_catalog.file_viewer import view as _view
            _view(path, os.path.splitext(path)[1].lower())
        except ImportError:
            st.caption("File viewer not available.")
        except Exception as e:
            st.warning("Viewer: " + type(e).__name__ + ": " + str(e))

    try:
        found = _scan(rec, path)
    except Exception as e:
        st.error("Could not read: " + type(e).__name__ + ": " + str(e))
        return
    if not found:
        st.warning("No tables found. If the document plainly has them it may "
                   "have no text layer — printing to PDF flattens text to "
                   "outlines and leaves nothing extractable.")
        return

    results = [r for _n, _rw, r in found]
    identity, ident_src = _identity_for(pack, path, results)
    ident_field = getattr(pack, "identity_field", None) or "identity"
    if identity:
        st.success("**" + ident_field + " = " + str(identity) + "** — from " + ident_src)
    else:
        st.warning("No " + ident_field + " found. Rows would be stored without "
                   "one and cannot be joined to a subject until migration "
                   "resolves it.")

    nonce = st.session_state.get("ds_nonce", 0)
    all_fields = ["— none —"] + sorted(pack.fields)
    pending = st.session_state.setdefault("ds_pending", {})

    # A correction you have not saved yet must still show its effect, or the
    # preview is answering a question you have already moved on from. Pending
    # changes become a throwaway overlay layered on top, so "what gets stored"
    # reflects the mapping ON SCREEN rather than the mapping on disk.
    live_pack, live_rec = pack, rec
    if pending:
        from docshape.packs.overlay import empty as _empty, add_alias as _aa, \
            apply_overlay as _ap
        _tmp = _empty("preview")
        for (_t, _cell), _field in pending.items():
            _aa(_tmp, _field, _cell)
        live_pack = _ap(pack, _tmp)
        from docshape import Recogniser as _R
        live_rec = _R(live_pack)
        found = _scan(live_rec, path)
        st.info("Preview reflects " + str(len(pending)) + " unsaved "
                "correction(s). Save or discard below to keep them.")

    for tname, rows, res in found:
        matched = res["shape"] != "UNKNOWN"
        icon = "✅" if matched and res["target"] else ("🟡" if matched else "❓")
        title = (icon + "  " + tname + " — " + res["shape"]
                 + (" (" + format(res["score"], ".2f") + ")" if matched else "")
                 + "  ·  " + str(len(rows)) + " row(s)")
        with st.expander(title, expanded=not matched):
            t_map, t_store, t_raw = st.tabs(["Columns", "What gets stored", "As read"])
            header = list(rows[0].keys())

            with t_map:
                by_index = {header.index(c): f for f, c in res["columns"].items()}
                for i, cell in enumerate(header):
                    cur = by_index.get(i)
                    samples = [str(r.get(cell, ""))[:20] for r in rows[:3]
                               if str(r.get(cell, "")).strip()]
                    a, b = st.columns([1, 1])
                    a.markdown("`" + str(cell) + "`")
                    a.caption(", ".join(samples) or "—")
                    choice = b.selectbox(
                        "f", all_fields,
                        index=all_fields.index(cur) if cur in all_fields else 0,
                        key="ds_" + str(nonce) + "_" + tname + "_" + str(i),
                        label_visibility="collapsed")
                    if choice != "— none —" and choice != cur:
                        pending[(tname, cell)] = choice
                    elif choice == cur:
                        pending.pop((tname, cell), None)

            with t_store:
                if not matched:
                    st.caption("Nothing matched, so nothing would be stored.")
                elif not res["target"]:
                    st.info("**" + res["shape"] + "** is recognised but has no "
                            "destination table. Rows accumulate in the capture "
                            "store and go no further until a target is set.")
                else:
                    st.markdown("→ `" + res["target"] + "`")
                    stored, unmapped, transformed, n_out = _stored_rows(
                        live_pack, res["shape"], res, identity)
                    if transformed:
                        st.caption("A transform applies: " + str(len(res["rows"]))
                                   + " document row(s) become " + str(n_out)
                                   + " database row(s).")
                    if stored:
                        st.dataframe(pd.DataFrame(stored), hide_index=True,
                                     use_container_width=True)
                    if unmapped:
                        st.caption("in parentheses: recognised, but the pack "
                                   "gives no destination column — "
                                   + ", ".join(unmapped))
                    st.caption("Plus provenance (file, path, content hash, "
                               "shape, score, captured) and review (status NEW, "
                               "confidence, extra_json).")

            with t_raw:
                st.dataframe(pd.DataFrame(rows[:15]), hide_index=True,
                             use_container_width=True)
                st.caption("What the parser returned before any mapping — so a "
                           "column the extractor missed can be told from one "
                           "that was never in the document.")

            if not matched:
                st.divider()
                known = [f for _i, _c, f in rec.header_fields(header) if f]
                other_req = set()
                for spec in pack.shapes.values():
                    other_req.update(spec.get("required", ()))
                s1, s2 = st.columns(2)
                sname = s1.text_input("Shape name", key="ds_sn_" + str(nonce) + tname)
                target = s2.text_input("Destination table (optional)",
                                       key="ds_tg_" + str(nonce) + tname)
                req = st.multiselect(
                    "Required — pick fields no other shape requires, or a "
                    "general shape keeps winning", sorted(pack.fields),
                    default=[f for f in known if f not in other_req][:2],
                    key="ds_rq_" + str(nonce) + tname)
                opt = st.multiselect(
                    "Optional", sorted(pack.fields),
                    default=[f for f in known if f not in req],
                    key="ds_op_" + str(nonce) + tname)
                if req:
                    clash = [n for n, sp in pack.shapes.items()
                             if set(sp.get("required", ())) <= set(req)]
                    if clash:
                        st.warning("Those required fields are a superset of "
                                   + ", ".join(clash) + " — it may still win.")
                if st.button("💾 Save shape", key="ds_ss_" + str(nonce) + tname,
                             disabled=not (sname and req)):
                    add_shape(overlay, sname, req, opt, target=target or None,
                              by="ui")
                    save_overlay(overlay, overlay_path)
                    st.session_state["ds_nonce"] = nonce + 1
                    st.rerun()

    if pending:
        st.divider()
        st.markdown("### " + str(len(pending)) + " pending correction(s)")
        for (t, cell), field in pending.items():
            st.markdown("- `" + str(cell) + "` → **" + field + "**  _(from " + t + ")_")
        where = ("the SANDBOX — speculative, no effect on capture until "
                 "promoted" if use_sandbox else "the live vocabulary")
        st.caption("Saved as an alias — this header wording means this field — "
                   "so every future document with that wording is right without "
                   "being touched. Going to " + where + ".")
        b1, b2, _ = st.columns([1, 1, 3])
        if b1.button("💾 Save corrections", type="primary", key="ds_save"):
            for (_t, cell), field in pending.items():
                add_alias(overlay, field, cell, by="ui",
                          note=os.path.basename(path))
            save_overlay(overlay, overlay_path)
            st.session_state["ds_pending"] = {}
            st.session_state["ds_nonce"] = nonce + 1
            st.rerun()
        if b2.button("✕ Discard", key="ds_discard"):
            st.session_state["ds_pending"] = {}
            st.rerun()

    # ── export what this document would give ──────────────────────────────
    st.divider()
    st.markdown("### Export")
    st.caption("A dry run: the rows a capture would produce from this "
               "document under the vocabulary as it stands, including any "
               "unsaved corrections above. Nothing is written to a database.")
    try:
        idx, by_table = _collect_stored(live_pack, live_rec, path)
        _export_controls(
            idx, by_table,
            os.path.splitext(os.path.basename(path))[0] + "_extract.xlsx",
            "single")
    except Exception as e:
        st.error("Export failed: " + type(e).__name__ + ": " + str(e))

    # ── create an attribute the pack does not have ────────────────────────
    with st.expander("＋ New attribute"):
        st.caption("Use this when a column names something the vocabulary has "
                   "no concept of — not when it is an existing attribute worded "
                   "differently. For that, just re-point the column above.")
        n1, n2 = st.columns([1, 2])
        new_field = n1.text_input("Field name", key="ds_nf_" + str(nonce),
                                  placeholder="cement_top")
        new_aliases = n2.text_input(
            "Wordings, comma separated", key="ds_na_" + str(nonce),
            placeholder="cement top, top of cement, toc")
        is_num = st.checkbox("Numeric — values are measurements",
                             key="ds_nn_" + str(nonce))
        if new_field:
            clash = [f for f in pack.fields if f == new_field.strip()]
            if clash:
                st.warning("That field already exists — re-point the column "
                           "above instead, which adds the new wording to it.")
            for a in [x.strip() for x in (new_aliases or "").split(",") if x.strip()]:
                hit = rec.field_for(a)
                if hit:
                    st.warning("`" + a + "` already resolves to **" + hit
                               + "**. Adding it here would create a second "
                                 "claim on the same wording; the longer alias "
                                 "wins, which may not be the one you want.")
        if st.button("＋ Add attribute", key="ds_nadd_" + str(nonce),
                     disabled=not (new_field and new_aliases)):
            from docshape.packs.overlay import set_numeric
            for a in [x.strip() for x in new_aliases.split(",") if x.strip()]:
                add_alias(overlay, new_field.strip(), a, by="ui")
            if is_num:
                set_numeric(overlay, [new_field.strip()])
            save_overlay(overlay, overlay_path)
            st.session_state["ds_nonce"] = nonce + 1
            st.rerun()

    with st.expander("📄 Overlay contents"):
        lines = []
        promote_summary(overlay, log=lines.append)
        st.code("\n".join(lines) or "(empty)", language="text")


# --------------------------------------------------------------------------- #
def _batch(rec, pack):
    """Run the same rules across a folder — where a correction proves itself."""
    import pandas as pd
    from docshape.readers import collect, TABLE_EXTS

    c1, c2, c3 = st.columns([3, 0.9, 0.9])
    folder = c1.text_input("Folder", key="ds_dir", placeholder=r"C:\docs\surveys")
    limit = c2.number_input("Max files", 1, 2000, 50, key="ds_limit")
    go = c3.button("▶ Run batch", type="primary", use_container_width=True)

    # EVERY Streamlit button reruns the script. Without holding the results,
    # clicking Build spreadsheet re-entered here with go=False and returned
    # early — the whole results block disappeared and the page looked as if it
    # had jumped back to the top. Results live in session state and are
    # re-rendered until a new run replaces them.
    cached = st.session_state.get("ds_batch_results")

    if not go and cached:
        if st.button("✕ Clear results", key="ds_clear_batch"):
            st.session_state.pop("ds_batch_results", None)
            st.rerun()
        _batch_results(pack, *cached)
        return

    if not go:
        st.info("Point at a folder of similar documents. After correcting one, "
                "re-run to see whether the correction generalises.")
        return
    if not folder or not os.path.exists(folder):
        st.error("Folder not found.")
        return

    paths = [p for p in collect(folder)
             if os.path.splitext(p)[1].lower() in TABLE_EXTS][:int(limit)]
    if not paths:
        st.warning("No table-bearing documents there (.pdf, .docx, .xlsx).")
        return

    rows, shape_counts, unknown_sigs = [], {}, {}
    all_index, all_tables = [], {}
    bar = st.progress(0.0, text="0 / " + str(len(paths)))
    for n, p in enumerate(paths, 1):
        base = os.path.basename(p)
        try:
            found = _scan(rec, p)
            # Collect the ROWS as well as the counts — a dry run should show
            # the data, not a description of it.
            idx, by_table = _collect_stored(pack, rec, p)
            all_index.extend(idx)
            for dest, rws in by_table.items():
                all_tables.setdefault(dest, []).extend(rws)
        except Exception as e:
            rows.append({"file": base, "tables": 0, "matched": 0, "unknown": 0,
                         "rows": 0, "identity": "—",
                         "note": type(e).__name__ + ": " + str(e)[:60]})
            bar.progress(n / len(paths), text=str(n) + " / " + str(len(paths)))
            continue
        ident, _src = _identity_for(pack, p, [r for _a, _b, r in found])
        matched = unknown = nrows = 0
        for tname, trows, res in found:
            if res["shape"] == "UNKNOWN":
                unknown += 1
                sig = " | ".join(list(trows[0].keys())[:8])
                e = unknown_sigs.setdefault(sig, {"files": set(), "rows": 0})
                e["files"].add(base)
                e["rows"] += len(trows)
            else:
                matched += 1
                nrows += len(trows)
                shape_counts[res["shape"]] = shape_counts.get(res["shape"], 0) + len(trows)
        rows.append({"file": base, "tables": len(found), "matched": matched,
                     "unknown": unknown, "rows": nrows,
                     "identity": ident or "—", "note": ""})
        bar.progress(n / len(paths), text=str(n) + " / " + str(len(paths)))
    bar.empty()
    st.session_state["ds_batch_results"] = (
        rows, shape_counts,
        {k: {"files": sorted(v["files"]), "rows": v["rows"]}
         for k, v in unknown_sigs.items()},
        all_index, all_tables,
        os.path.basename(os.path.normpath(folder)) or "batch")
    _batch_results(pack, *st.session_state["ds_batch_results"])


def _batch_results(pack, rows, shape_counts, unknown_sigs, all_index,
                   all_tables, folder_name):
    """Render a completed batch. Separate from the run so it can be shown
    again on any rerun without re-reading every document."""
    import pandas as pd

    tot_rows = sum(r["rows"] for r in rows)
    tot_unknown = sum(r["unknown"] for r in rows)
    no_ident = sum(1 for r in rows if r.get("identity") == "—")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Documents", len(rows))
    m2.metric("Rows recognised", format(tot_rows, ","))
    m3.metric("Unrecognised tables", tot_unknown)
    m4.metric("No identity", no_ident)

    if shape_counts:
        st.markdown("**Rows by shape**")
        st.dataframe(pd.DataFrame(
            [{"shape": k, "rows": v,
              "target": pack.shapes.get(k, {}).get("target") or "(none)"}
             for k, v in sorted(shape_counts.items(), key=lambda kv: -kv[1])]),
            hide_index=True, use_container_width=True)

    if unknown_sigs:
        st.markdown("**Unrecognised tables, grouped by header**")
        st.caption("One row here is one shape to define — not one per file. "
                   "Start at the top.")
        st.dataframe(pd.DataFrame(
            [{"header": sig, "documents": len(v["files"]), "rows": v["rows"],
              "example": v["files"][0] if v["files"] else ""}
             for sig, v in sorted(unknown_sigs.items(),
                                  key=lambda kv: -len(kv[1]["files"]))]),
            hide_index=True, use_container_width=True)
        st.caption("Copy an example filename into Single document mode to "
                   "define its shape, then re-run this batch.")

    st.markdown("**Per document**")
    st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

    # ── what would actually be extracted, pooled ──────────────────────────
    if all_tables:
        st.divider()
        st.markdown("### What would be extracted")
        st.caption("Rows pooled across every document, by destination. This is "
                   "a dry run — nothing is written to a database.")
        dest = st.selectbox("Destination", sorted(
            all_tables, key=lambda d: -len(all_tables[d])), key="ds_bdest")
        st.dataframe(pd.DataFrame(all_tables[dest][:300]), hide_index=True,
                     use_container_width=True)
        if len(all_tables[dest]) > 300:
            st.caption("showing the first 300 of "
                       + format(len(all_tables[dest]), ",")
                       + " — the spreadsheet has all of them")

        st.markdown("### Export")
        _export_controls(all_index, all_tables,
                         folder_name + "_extract.xlsx", "batch")
