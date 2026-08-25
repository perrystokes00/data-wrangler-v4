"""Did the load actually land, and if not, what is fixable.

ONE DEFINITION, TWO SURFACES. tools/held_rows.py and the Data Assistant's
"After the load" panel both call this. Answering the same question two ways is
how MIRROR_TABLES and LINEAGE drifted, how demo_reset and clear_catalog
drifted, and how two loaders came to mint provenance differently -- three times
in one week. So the queries live here once.

Everything is keyed on what CANNOT be reset:

  * HELD is arithmetic. stg keeps its rows after promote (unlike file_catalog's
    cat_*, which is a drain), so a staged row whose key is absent from the
    target did not land -- whatever any status column says. A resettable flag
    put 1,701 loaded files back in the queue this week; a flag is not evidence.

  * PROVENANCE is a join. A row's INVENTORY_ID either resolves to a catalogued
    file or it does not.

The staging table carries SOURCE headers -- API_NUMBER, FORMATION -- and
promote renames them, so the key columns are resolved through dv_column_map,
which is where those decisions were persisted. And the key EXPRESSIONS are
imported from the promote path rather than reimplemented: the UWI-14 pad
already has three sites that must agree and it pads with ZEROS, not spaces.
"""
import re

from sqlalchemy import text

GFC = "file_catalog.GLOBAL_FILE_CATALOG"
_HASH_SUFFIX = re.compile(r"_[0-9a-f]{8}$")


def _key(alias, col, target_col=None):
    """Promote's own expression for one key column. Imported, never copied."""
    from dataview.import_data.bulk_dir_loader import _val_expr, _uwi14_sql, _IDENT
    tgt = (target_col or col).lower()
    expr = _val_expr(alias, col, tgt in _IDENT or col.lower() in _IDENT)
    return _uwi14_sql(expr) if tgt == "uwi" else expr


def _staged_tables(cx):
    """{staging table: row count} for everything with rows in stg."""
    out = {}
    for r in cx.execute(text(
        "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
        "WHERE TABLE_SCHEMA='stg' AND TABLE_TYPE='BASE TABLE' "
        "ORDER BY TABLE_NAME")).fetchall():
        n = cx.execute(text("SELECT COUNT(*) FROM stg.[%s]" % r[0])).scalar()
        if n:
            out[r[0]] = n
    return out


def _stg_columns(cx):
    out = {}
    for r in cx.execute(text(
        "SELECT TABLE_NAME, COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_SCHEMA='stg'")).fetchall():
        out.setdefault(r[0], set()).add(r[1].lower())
    return out


def _pk(cx, table):
    return [r[0] for r in cx.execute(text(
        "SELECT kc.COLUMN_NAME FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc "
        "JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE kc "
        "  ON kc.CONSTRAINT_NAME = tc.CONSTRAINT_NAME "
        "WHERE tc.TABLE_SCHEMA='dataview' AND tc.TABLE_NAME = :t "
        "AND tc.CONSTRAINT_TYPE='PRIMARY KEY' "
        "ORDER BY kc.ORDINAL_POSITION"), {"t": table}).fetchall()]


def _pk_feeders(cx, target, pk, have):
    """{pk column: the staging column that fills it}, via dv_column_map.

    Without this every table reports "no comparable key", because the staging
    column is called API_NUMBER and the target column is called uwi.
    """
    out = {}
    for p in pk:
        if p.lower() in have:
            out[p] = p
            continue
        for r in cx.execute(text(
            "SELECT DISTINCT source_column FROM dataview.dv_column_map "
            "WHERE UPPER(target_table) = :t AND LOWER(target_column) = :c "
            "AND ISNULL(active_ind,'Y') = 'Y'"),
            {"t": target.upper(), "c": p.lower()}).fetchall():
            if str(r[0]).lower() in have:
                out[p] = str(r[0])
                break
    return out


def held_report(engine):
    """[{table, target, staged, landed, held, partial, note}], one per staged table."""
    rows = []
    with engine.connect() as cx:
        stg = _staged_tables(cx)
        cols = _stg_columns(cx)
        for s in sorted(stg):
            tgt = _HASH_SUFFIX.sub("", s)
            have = cols.get(s, set())
            pk = _pk(cx, tgt)
            feeders = _pk_feeders(cx, tgt, pk, have) if pk else {}
            if not feeders:
                rows.append({"table": s, "target": tgt, "staged": stg[s],
                             "landed": None, "held": None, "partial": False,
                             "note": "no column feeds %s"
                                     % ("+".join(pk) if pk else "any key")})
                continue
            on = " AND ".join("%s = %s" % (_key("s", feeders[p], p), _key("d", p, p))
                              for p in pk if p in feeders)
            held = cx.execute(text(
                "SELECT COUNT(*) FROM stg.[%s] s WHERE NOT EXISTS "
                "(SELECT 1 FROM dataview.[%s] d WHERE %s)" % (s, tgt, on))).scalar()
            partial = len(feeders) < len(pk)
            rows.append({
                "table": s, "target": tgt, "staged": stg[s],
                "landed": stg[s] - held, "held": held, "partial": partial,
                # A PARTIAL KEY OVER-MATCHES, so the honest note says which way
                # the error runs rather than presenting a smaller number as fact.
                "note": ("keyed on %s only — under-counts"
                         % "+".join(p for p in pk if p in feeders)) if partial else "",
            })
    return rows


def held_causes(engine, stg_table):
    """[{child_col, parent, unmatched, examples}] — which missing parent explains it.

    An empty list means every FK resolves, so the rows were refused for another
    reason. Duplicate keys are far and away the usual one: promote is
    insert-only, so a repeated key is skipped without a word.
    """
    out = []
    tgt = _HASH_SUFFIX.sub("", stg_table)
    with engine.connect() as cx:
        have = _stg_columns(cx).get(stg_table, set())
        fks = cx.execute(text(
            "SELECT cc.name, OBJECT_NAME(fk.referenced_object_id), pc.name "
            "FROM sys.foreign_keys fk "
            "JOIN sys.foreign_key_columns f ON f.constraint_object_id = fk.object_id "
            "JOIN sys.columns cc ON cc.object_id = fk.parent_object_id "
            "  AND cc.column_id = f.parent_column_id "
            "JOIN sys.columns pc ON pc.object_id = fk.referenced_object_id "
            "  AND pc.column_id = f.referenced_column_id "
            "WHERE fk.parent_object_id = OBJECT_ID('dataview.' + :t)"),
            {"t": tgt}).fetchall()
        for child_col, parent, parent_col in fks:
            if child_col.lower() not in have:
                continue
            where = ("s.[%s] IS NOT NULL AND NOT EXISTS (SELECT 1 FROM dataview.[%s] p "
                     "WHERE %s = %s)"
                     % (child_col, parent,
                        _key("p", parent_col, parent_col),
                        _key("s", child_col, parent_col)))
            n = cx.execute(text(
                "SELECT COUNT(*) FROM stg.[%s] s WHERE %s" % (stg_table, where))).scalar()
            if not n:
                continue
            vals = cx.execute(text(
                "SELECT DISTINCT TOP 4 %s FROM stg.[%s] s WHERE %s"
                % (_key("s", child_col, parent_col), stg_table, where))).fetchall()
            out.append({"child_col": child_col, "parent": parent, "unmatched": n,
                        "examples": [str(v[0]) for v in vals]})
    return out


def orphan_provenance(engine):
    """[{table, inventory_id, rows, guess}] — rows citing a file nothing can resolve.

    `guess` is the workbook a sidecar id WOULD have come from, when the id
    matches one. That is what makes this repairable rather than merely
    reportable: the sidecar is deleted, but its path is recomputable.
    """
    out = []
    with engine.connect() as cx:
        tables = [r[0] for r in cx.execute(text(
            "SELECT DISTINCT OBJECT_NAME(c.object_id) FROM sys.columns c "
            "WHERE c.name='INVENTORY_ID' "
            "AND OBJECT_SCHEMA_NAME(c.object_id)='dataview' "
            "AND OBJECTPROPERTY(c.object_id,'IsUserTable')=1 ORDER BY 1")).fetchall()]
        for t in tables:
            for r in cx.execute(text(
                "SELECT x.INVENTORY_ID, COUNT(*) n FROM dataview.[%s] x "
                "WHERE x.INVENTORY_ID IS NOT NULL AND NOT EXISTS "
                "(SELECT 1 FROM %s g WHERE g.INVENTORY_ID = x.INVENTORY_ID) "
                "GROUP BY x.INVENTORY_ID ORDER BY COUNT(*) DESC" % (t, GFC))).fetchall():
                out.append({"table": t, "inventory_id": r[0], "rows": r[1]})
    return out


def sidecar_guess(inventory_id, search_dirs):
    """The workbook whose sidecar hashes to `inventory_id`, or None.

    inventory_id is SHA1(UPPER(path)), so the id of a DELETED sidecar is still
    computable from the path it would have had -- <dir>/_xl_sheets/<book>__<sheet>.csv.
    """
    import glob
    import hashlib
    import os

    def _iid(p):
        return hashlib.sha1(str(p).upper().strip().encode("utf-16-le")).hexdigest().upper()

    want = str(inventory_id or "").upper()
    for d in search_dirs or []:
        if not d or not os.path.isdir(d):
            continue
        for book in glob.glob(os.path.join(d, "*.xls*")):
            if os.path.basename(book).startswith("~$"):
                continue
            stem = os.path.splitext(os.path.basename(book))[0]
            try:
                from openpyxl import load_workbook
                sheets = load_workbook(book, read_only=True).sheetnames
            except Exception:
                sheets = ["Sheet1"]
            for sh in sheets:
                side = os.path.join(d, "_xl_sheets", "%s__%s.csv" % (stem, sh))
                if _iid(os.path.abspath(side)) == want:
                    return book
    return None
