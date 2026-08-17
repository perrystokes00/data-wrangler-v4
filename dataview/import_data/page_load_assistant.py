"""
dataview/import_data/page_load_assistant.py
═══════════════════════════════════════════
Load Assistant — the AI-FIRST entry to loading (Perry's picture, July 29:
"The AI is the first step. I ask it: please load this to the well header
table. Then I drop a sample of the file in.")

The conversation leads; the machinery follows. One instruction + one file →
a reviewable LOAD PLAN:

    1. recognize   what the file is (headers + sample rows, delimiter-sniffed)
    2. ground      against the LIVE schema (columns, NOT-NULLs, FK graph)
    3. plan        table · column map · required-gap rules · parent
                   prerequisites · row-shape transforms (e.g. the tops pivot)
    4. verify      the hard checks deterministically (parent-well match uses
                   the same UWI-14 pad promote uses — never the AI's word)
    5. teach       Adopt writes dv_column_map, so the Bulk Tabular Loader's
                   fingerprint recall auto-assigns the table and auto-maps
                   the columns on the real run — hands-free from then on

Design law (written in blood this week): AI PROPOSES, the deterministic core
VERIFIES and EXECUTES, the human CONFIRMS, the stores REMEMBER. This page
executes nothing destructive: it teaches the stores and writes derived CSVs;
staging and promote stay in the Bulk Tabular Loader's engine room.

Wire-in (nav router):  from dataview.import_data.page_load_assistant import run
                       run()
"""
from __future__ import annotations

import os
import re

import streamlit as st

# The engine room — everything heavy is REUSED, never re-implemented.
try:
    from dataview.import_data import synonym_store as _store
except Exception:                                    # not deployed yet
    _store = None

from dataview.import_data.bulk_dir_loader import (
    get_engine, ai_suggest_table_map, seed_vendor_synonyms, _uwi14,
    _live_catalog_parsed,
)
from dataview.import_data import page_dir_loader as pdl


# ─────────────────────────── sample reading ────────────────────────────────
def _read_sample(path, n_rows=8):
    """(columns, sample_rows, n_total_rows_or_None). Delimiter-sniffed —
    RMOTC 'CSVs' are frequently tab/;/| separated (the tops file was), and a
    comma-only read collapses the header into one mangled column."""
    import pandas as pd
    if str(path).lower().endswith((".xlsx", ".xls", ".xlsm")):
        df = pd.read_excel(path, dtype=str).fillna("")
    else:
        df = pd.read_csv(path, sep=None, engine="python", dtype=str,
                         keep_default_na=False)
    cols = [str(c) for c in df.columns]
    rows = [tuple(r) for r in df.head(n_rows).itertuples(index=False)]
    return cols, rows, len(df)


# ───────────────────── the transform shelf (for the plan) ───────────────────
# What the AI may SELECT and PARAMETERIZE — it never invents transforms.
# Param values naming columns are validated against the file's real columns;
# invalid ones are dropped so a hallucinated column can only cost a blank
# picker, never a wrong run.
_TRANSFORM_CATALOG = {
    "pivot_picks_to_intervals": {
        "when": "one pick per row (Top/Base types) but the table wants "
                "top_depth+base_depth interval rows",
        "params": {"uwi_col": "column holding API/UWI",
                   "unit_col": "stratigraphic unit NAME column",
                   "type_col": "pick type column (Top/Base/Lower)",
                   "depth_col": "pick depth column",
                   "code_col": "unit CODE column (optional)"}},
    "derive_prod_entities": {
        "when": "by-well production rows need their missing DV_PROD_ENTITY "
                "parent derived (one entity per distinct well)",
        "params": {"uwi_col": "column holding API/UWI",
                   "date_col": "production period/date column (optional)",
                   "name_col": "well name column (optional)"}},
    "unpivot_prod_volumes": {
        "when": "wide well-month rows with Oil/Water/Gas as COLUMNS but the "
                "table wants one row per well-month-fluid",
        "params": {"uwi_col": "column holding API/UWI",
                   "date_col": "period/date column",
                   "oil_col": "oil volume column (optional)",
                   "water_col": "water volume column (optional)",
                   "gas_col": "gas volume column (optional)",
                   "days_col": "days-produced column (optional)"}},
    "derive_dir_surveys": {
        "when": "station-per-row survey file needs a derived "
                "DV_WELL_DIR_SRVY_HDR parent + numbered stations",
        "params": {"uwi_col": "column holding API/UWI",
                   "md_col": "measured depth column",
                   "incl_col": "inclination column (optional)",
                   "azim_col": "azimuth column (optional)",
                   "tvd_col": "TVD column (optional)"}},
}


def _validate_transform(prop, cols):
    """AI transform proposal -> (name, {param: value}) with column-params
    validated against the file's REAL columns, or (None, {})."""
    if not isinstance(prop, dict):
        return None, {}
    name = str(prop.get("name") or "")
    spec = _TRANSFORM_CATALOG.get(name)
    if not spec:
        return None, {}
    ok = {}
    colset = set(cols)
    for p, v in (prop.get("params") or {}).items():
        if p not in spec["params"] or v in (None, ""):
            continue
        if str(p).endswith("_col"):
            if str(v) in colset:
                ok[p] = str(v)
        else:
            ok[p] = str(v)
    return name, ok


def _catalog_cached(ss, engine, schema):
    """(FKC, COLS, KIND) once per session.

    This was called on EVERY rerun of the shape grid — so every checkbox
    click re-read the whole FK catalog, and the screen sat dimmed with no
    explanation (Perry, July 31). Cached here; a fresh scan drops it.
    """
    key = "la_cat::" + str(schema)
    if key not in ss:
        ss[key] = _live_catalog_parsed(engine, schema)
    return ss[key]


# ─────────────── store-first mapping + fit pre-flight (wiring) ─────────────
def apply_store(engine, schema, table, cols, ai_cmap, rows=None,
                bypass=False):
    """THE WIRING (July 31). The column-level store maps what it knows;
    the AI only fills what is left over.

    Order matters and is deliberate: a synonym is a fact somebody already
    confirmed, an AI mapping is a proposal. So the store goes first, the AI
    fills the remainder, and nothing may claim a target column twice.

    Returns (cmap, provenance, fit_issues, unmatched) where provenance is
    {source: 'store'|'ai'} so the plan can show where each row came from.
    """
    if _store is None or engine is None or bypass:
        # bypass = the AI's answer, unaided. Kept as a deliberate experiment
        # so the store's value can be MEASURED rather than argued about
        # (Perry, July 31: "let's try it, I am curious").
        cmap0 = {s: str(t).lower() for s, t in (ai_cmap or {}).items() if t}
        left0 = [c for c in cols if c not in cmap0]
        issues0 = []
        if rows and _store is not None and engine is not None:
            try:
                issues0 = _store.check_fit(engine, schema, table, cmap0, rows)
            except Exception:
                issues0 = []
        return cmap0, {s: "ai" for s in cmap0}, issues0, left0
    try:
        store_map, unmatched, _notes = _store.suggest_map(
            engine, schema, table, cols)
    except Exception:
        return dict(ai_cmap or {}), {s: "ai" for s in (ai_cmap or {})}, [], []

    cmap = dict(store_map)
    prov = {s: "store" for s in store_map}
    taken = {t for t in cmap.values()}
    for s, t in (ai_cmap or {}).items():
        if s in cmap or not t:
            continue                       # the store already spoke
        tl = str(t).lower()
        if tl in taken or _store.is_system_column(tl):
            continue                       # no double claims, no audit cols
        cmap[s] = tl
        prov[s] = "ai"
        taken.add(tl)
    issues = []
    if rows:
        try:
            issues = _store.check_fit(engine, schema, table, cmap, rows)
        except Exception:
            issues = []
    left = [c for c in cols if c not in cmap]
    return cmap, prov, issues, left


def render_fit(issues, where="plan"):
    """Show the pre-flight. Errors are things that WILL fail at promote —
    the period_date nvarchar(7) class of failure, caught before staging."""
    errs = [i for i in issues if i[0] == "error"]
    warns = [i for i in issues if i[0] != "error"]
    for _s, col, msg in errs:
        st.error(f"✗ **{col}** — {msg}")
    for _s, col, msg in warns:
        st.warning(f"⚠ **{col}** — {msg}")
    return len(errs)


def sample_rows(path, n=300):
    """Rows as dicts for the fit check, without reading a whole workbook."""
    try:
        df = _read_any(path)
        return df.head(n).astype(str).to_dict("records")
    except Exception:
        return []


# ───────────────────── foreign keys: check, then resolve ───────────────────
def fk_scan(engine, schema, table, cmap, path, max_vals=300):
    """Which mapped values have no parent row yet.

    Returns [{child_col, source_col, parent, parent_col, kind, missing:
    [(value, rows)], options: [parent values]}] — one entry per FK whose
    child column the map actually fills. Checked BEFORE staging, so the
    operator resolves violations while the file is still a file.
    """
    from sqlalchemy import text
    from dataview.import_data.bulk_dir_loader import (
        _live_catalog_parsed, _table_pk_live)
    import pandas as pd

    FKC, COLS, _KIND = _live_catalog_parsed(engine, schema)
    tu = table.upper()
    inv = {str(t).lower(): s for s, t in (cmap or {}).items()}
    out = []
    df = None
    for fk in FKC.get(tu, []):
        ccols = [c.lower() for c in fk.get("child_cols", [])]
        if len(ccols) != 1:
            continue
        ccol = ccols[0]
        src_col = inv.get(ccol)
        if not src_col:
            continue                      # not filled by this file -> no risk
        parent = str(fk.get("parent_table", "")).upper()
        if not parent or parent == tu:
            continue
        info = pdl._fk_of(tu, ccol, FKC)
        kind = info[1] if info else "parent"
        ppk = _table_pk_live(engine, parent, schema) or []
        pcol = ppk[0] if len(ppk) == 1 else (
            ccol if ccol in {c.lower() for c in COLS.get(parent, set())}
            else None)
        if not pcol:
            continue
        if df is None:
            df = _read_any(path)
        if src_col not in df.columns:
            continue
        vals = (df[src_col].astype(str).str.strip()
                .replace({"nan": "", "None": ""}))
        counts = vals[vals != ""].value_counts()
        if counts.empty:
            continue
        with engine.connect() as cx:
            have = {str(r[0]).strip() for r in cx.execute(text(
                f"SELECT DISTINCT [{pcol}] FROM {schema}.{parent.lower()}"
            )).fetchall() if r[0] is not None}
            options = sorted(have)[:max_vals]
        missing = [(v, int(n)) for v, n in counts.items()
                   if v not in have][:max_vals]
        if missing:
            out.append({"child_col": ccol, "source_col": src_col,
                        "parent": parent, "parent_col": pcol, "kind": kind,
                        "missing": missing, "options": options,
                        "n_have": len(have)})
    return out


def resolve_from_grid(edited):
    """Perry's two-checkbox state machine (July 31):
        add ticked            -> ADD the value to the parent
        replace ticked        -> REPLACE it with the chosen standard value
        neither               -> NULL the value out
        both                  -> a contradiction; reported, never guessed
    Returns (resolutions, errors) where each resolution is
    (value, 'add'|'remap'|'null', target_or_None).
    """
    res, errs = [], []
    for r in edited.to_dict("records"):
        v = str(r.get("value", ""))
        add = bool(r.get("add"))
        rep = bool(r.get("replace with →"))
        std = str(r.get("standard value") or "").strip()
        if add and rep:
            errs.append(f"'{v}': both boxes ticked — add it, or replace it, "
                        f"not both.")
            continue
        if rep and not std:
            errs.append(f"'{v}': replace is ticked but no standard value is "
                        f"chosen.")
            continue
        if add:
            res.append((v, "add", None))
        elif rep:
            res.append((v, "remap", std))
        else:
            res.append((v, "null", None))
    return res, errs


def apply_fk_fixes(engine, schema, stg, fixes, say=None):
    """Apply the operator's decisions to STAGING (set-based, one statement
    per action) and seed parents for 'add'. Runs after staging, before
    promote — so promote sees data that already satisfies its own FKs."""
    from sqlalchemy import text
    n_add = n_remap = n_null = 0
    with engine.begin() as cx:
        for f in fixes:
            src_col, parent, pcol = f["source_col"], f["parent"], f["parent_col"]
            adds = [v for v, a, _t in f["resolutions"] if a == "add"]
            remaps = [(v, t) for v, a, t in f["resolutions"] if a == "remap"]
            nulls = [v for v, a, _t in f["resolutions"] if a == "null"]
            if adds:
                cols_p = {c.lower() for c in cx.execute(text(
                    "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA=:s AND TABLE_NAME=:t"),
                    {"s": schema, "t": parent.lower()}).scalars().all()}
                extra_cols, extra_vals = [], []
                if "active_ind" in cols_p:
                    extra_cols.append("[active_ind]"); extra_vals.append("'Y'")
                if "row_created_by" in cols_p:
                    extra_cols.append("[row_created_by]")
                    extra_vals.append("'LOAD_ASSISTANT'")
                if "row_created_date" in cols_p:
                    extra_cols.append("[row_created_date]")
                    extra_vals.append("SYSUTCDATETIME()")
                for v in adds:
                    cx.execute(text(
                        f"INSERT INTO {schema}.{parent.lower()} "
                        f"([{pcol}]{''.join(',' + c for c in extra_cols)}) "
                        f"SELECT :v{''.join(',' + e for e in extra_vals)} "
                        f"WHERE NOT EXISTS (SELECT 1 FROM "
                        f"{schema}.{parent.lower()} WHERE [{pcol}] = :v)"),
                        {"v": v})
                    n_add += 1
            for v, t in remaps:
                r = cx.execute(text(
                    f"UPDATE {stg} SET [{src_col}] = :t "
                    f"WHERE LTRIM(RTRIM([{src_col}])) = :v"),
                    {"t": t, "v": v})
                n_remap += (r.rowcount or 0)
            if nulls:
                marks = ", ".join(f":n{i}" for i in range(len(nulls)))
                p = {f"n{i}": v for i, v in enumerate(nulls)}
                r = cx.execute(text(
                    f"UPDATE {stg} SET [{src_col}] = NULL "
                    f"WHERE LTRIM(RTRIM([{src_col}])) IN ({marks})"), p)
                n_null += (r.rowcount or 0)
    if say:
        say(f"FK fixes: {n_add} parent row(s) added, {n_remap} value(s) "
            f"remapped, {n_null} value(s) nulled")
    return {"added": n_add, "remapped": n_remap, "nulled": n_null}


def _advance_queue(ss, loaded_path):
    """Mark the loaded file done and tee up the next one.

    Auto-advance is the point of a queue — finishing a file should not
    require a human to go find the next one (Perry, July 31). The next
    Analyze fires on the following rerun via la_autorun.
    """
    q = ss.get("la_queue") or []
    if not q:
        return
    for i, r in enumerate(q):
        if r["path"] == loaded_path or _os_basename(r["path"]) == \
                _os_basename(loaded_path):
            r["status"] = "loaded"
            if i == int(ss.get("la_q_idx", 0)):
                # DO NOT auto-fire the next Analyze. Advancing immediately
                # wiped the load result off the screen before it could be
                # read — "a green message popped up and disappeared and I was
                # kicked back to the beginning" (Perry, July 31). Mark the
                # queue, park the result, and wait for a click.
                ss["la_pending_next"] = i + 1
            break


def _os_basename(p):
    """Basename that works whichever separator the path uses — os.path is
    platform-bound and this compares Windows paths from anywhere."""
    return str(p).replace("\\", "/").rstrip("/").rsplit("/", 1)[-1]


# ────────────────────────── whole-file reader ──────────────────────────────
def _read_any(path):
    """Full-file read, Excel-aware. Old-format .xls starts with byte 0xd0 —
    reading it as text throws "utf-8 codec can't decode byte 0xd0" (the
    Verify button did exactly that on the Teapot tops .xls, July 29). Every
    full-file read goes through here; only _read_prod_workbook differs (it
    additionally merges the NPR-3 two-row multi-sheet header)."""
    import pandas as pd
    if str(path).lower().endswith((".xlsx", ".xls", ".xlsm")):
        return pd.read_excel(path, dtype=str).fillna("")
    return pd.read_csv(path, sep=None, engine="python", dtype=str,
                       keep_default_na=False)


_HDR_CACHE = {}


def _read_header(path):
    """Column names ONLY, as cheaply as the format allows.

    A shape scan needs headers, not data — but pd.read_excel parses the
    whole workbook to hand back column names, which on a 210k-row NPR-3
    file is seconds per file (Perry, July 31: "scanning shapes should be
    nearly instantaneous"). csv/txt: one line. xlsx/xlsm: openpyxl in
    read_only mode, first row. xls: pandas with nrows=1.
    """
    import csv as _csv
    p = str(path).lower()
    try:
        if p.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb[wb.sheetnames[0]]
                for row in ws.iter_rows(min_row=1, max_row=1,
                                        values_only=True):
                    return [("" if v is None else str(v)).strip()
                            for v in row]
                return []
            finally:
                wb.close()
        if p.endswith(".xls"):
            import pandas as pd
            return [str(c) for c in
                    pd.read_excel(path, nrows=1, dtype=str).columns]
        with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
            first = fh.readline()
        if not first:
            return []
        try:
            sep = _csv.Sniffer().sniff(first, delimiters=",\t;|").delimiter
        except Exception:
            sep = "\t" if "\t" in first else ","
        return [c.strip().strip('"') for c in first.rstrip("\r\n").split(sep)]
    except Exception:
        cols, _rows, _n = _read_sample(path, n_rows=0)     # last resort
        return cols


# ───────────────────────── directory shape scan ────────────────────────────
def _scan_shapes(engine, directory, schema="dataview"):
    """Group a directory's CSVs by COLUMN-SHAPE fingerprint and check each
    shape against dv_column_map: a shape confirmed before is KNOWN (table
    recalled) and never queues. One plan per UNKNOWN shape — one decision
    covers every file of that shape."""
    import glob
    from sqlalchemy import text
    # A tabular file is rarely NAMED .csv in vendor drops — Teapot's are
    # .txt/.dat/Excel. _read_sample sniffs the delimiter (and reads Excel),
    # so every tabular extension is fair game here.
    _EXTS = ("*.csv", "*.txt", "*.dat", "*.tsv", "*.prn",
             "*.xlsx", "*.xls", "*.xlsm")
    files = sorted({p for pat in _EXTS
                    for p in glob.glob(os.path.join(directory, pat))
                    if not p.lower().endswith("__tabfix.csv")
                    and not p.lower().endswith("__intervals.csv")
                    and not p.lower().endswith("__prod_entity.csv")
                    and not p.lower().endswith("__volumes.csv")
                    and not p.lower().endswith("__srvy_hdr.csv")
                    and not p.lower().endswith("__srvy_sta.csv")})
    groups = {}
    for p in files:
        try:
            mt = os.path.getmtime(p)
        except Exception:
            mt = 0
        cached = _HDR_CACHE.get(p)
        if cached and cached[0] == mt:
            cols = cached[1]
        else:
            cols = _read_header(p)
            if not cols:
                continue
            _HDR_CACHE[p] = (mt, cols)
        fp = pdl.fingerprint_cols(sorted(cols))
        g = groups.setdefault(fp, {"fp": fp, "cols": cols, "files": [],
                                   "known": None})
        g["files"].append(p)
    if groups:
        # ONE query for every fingerprint, not one per shape
        try:
            fps = list(groups.keys())
            marks = ", ".join(f":f{i}" for i in range(len(fps)))
            params = {f"f{i}": v for i, v in enumerate(fps)}
            seen = {}
            with engine.connect() as cx:
                for fp, tt in cx.execute(text(
                        f"SELECT source_file_pattern, target_table "
                        f"FROM dataview.dv_column_map "
                        f"WHERE source_file_pattern IN ({marks}) "
                        f"AND confirmed_ind = 'Y' AND active_ind = 'Y' "
                        f"GROUP BY source_file_pattern, target_table"),
                        params).fetchall():
                    seen.setdefault(str(fp), set()).add(str(tt).upper())
            for fp, g in groups.items():
                tt = seen.get(fp) or set()
                if len(tt) == 1:                 # ambiguity = no recall
                    g["known"] = next(iter(tt))
        except Exception:
            pass
    return list(groups.values())


# ───────────────────── batch table proposal (one AI call) ───────────────────
def ai_propose_tables(engine, schema, shapes):
    """ONE call proposing a target table per unknown shape: {shape_idx: TABLE}.
    Headers + a sample row per shape in; validated against the live catalog
    out. Proposals FILL the grid — Adopt remains the explicit human act."""
    import json as _json
    import re as _re2
    import anthropic
    from dataview.import_data.bulk_dir_loader import _ai_api_key
    key = _ai_api_key()
    if not key or not shapes:
        return {}
    FKC, COLS, _ = _live_catalog_parsed(engine, schema)

    def _toks(s):
        return {t for t in _re2.split(r"[^a-z0-9]+", str(s).lower()) if len(t) > 2}
    cand_union = set()
    for sh in shapes:
        want = set()
        for c in sh["cols"]:
            want |= _toks(c)
        ranked = sorted(((len(want & (_toks(t) | {w for c in cols for w in _toks(c)})), t)
                         for t, cols in COLS.items()), reverse=True)
        cand_union |= {t for _, t in ranked[:6]}
    catalog = {t: sorted(c.lower() for c in COLS.get(t, set()))
               for t in sorted(cand_union)}
    payload = [{"shape": sh["idx"], "columns": sh["cols"],
                "sample_row": [str(v)[:24] for v in (sh.get("row") or [])]}
               for sh in shapes]
    prompt = ("You classify source data tables against a SQL Server schema.\n"
              "For EACH shape below, pick the best target table from the "
              "candidates, or null when none fits.\nShapes:\n"
              + _json.dumps(payload, indent=1)
              + "\nCandidate tables and their columns:\n"
              + _json.dumps(catalog, indent=1)
              + '\nRespond with ONLY JSON: {"assignments": {"<shape>": '
                '"<TABLE or null>"}}')
    import os as _os
    model = _os.environ.get("DATAVIEW_AI_MODEL", "claude-sonnet-5")
    client = anthropic.Anthropic(api_key=key)
    msg = client.messages.create(model=model, max_tokens=2000,
                                 messages=[{"role": "user", "content": prompt}])
    txt = "".join(b.text for b in msg.content
                  if getattr(b, "type", "") == "text").strip()
    if txt.startswith("```"):
        txt = txt.split("```", 2)[1]
        txt = txt[4:] if txt.lower().startswith("json") else txt
    out = {}
    try:
        data, _ = _json.JSONDecoder().raw_decode(txt.strip())
        for k, v in (data.get("assignments") or {}).items():
            t = str(v or "").upper()
            if t in COLS:
                out[int(k)] = t
    except Exception:
        pass
    return out


def _parents_of(t, FKC):
    """Full-graph DATA parents of a table (dv_r_* excluded)."""
    out = set()
    for fk in FKC.get(t, []):
        p = str(fk.get("parent_table", "")).upper()
        if p and p != t and not p.startswith("DV_R_"):
            out.add(p)
    return out


def _load_order(tables, FKC):
    """{TABLE: ordinal} — dv_r_* first, then data tables by FULL-graph FK
    depth (parents before children even when the middle of the chain is not
    in this load), alphabetical within a depth. The in-set-only version put
    DV_PROD_VOLUME at load #1 because its parents weren't among the files
    (July 29): with no visible dependencies it tied with DV_WELL for rank 1
    and won alphabetically. Depth over the WHOLE graph puts it after the
    wells where it belongs."""
    tset = {t.upper() for t in tables if t}
    refs = sorted(t for t in tset if t.startswith("DV_R_"))
    data = sorted(tset - set(refs))
    memo = {}

    def depth(t, seen=()):
        if t in memo:
            return memo[t]
        if t in seen:                       # cycle guard
            return 0
        ps = _parents_of(t, FKC)
        d = 0 if not ps else 1 + max(depth(p, seen + (t,)) for p in ps)
        memo[t] = d
        return d

    order = refs + sorted(data, key=lambda t: (depth(t), t))
    return {t: i + 1 for i, t in enumerate(order)}


def _absent_parents(tables, FKC):
    """{TABLE: [full-graph data parents not in the load set]} — the honest
    footnote to a load order: these children will HOLD at promote until the
    named parents are populated by some other load."""
    tset = {t.upper() for t in tables if t}
    out = {}
    for t in sorted(tset):
        miss = sorted(p for p in _parents_of(t, FKC) if p not in tset)
        if miss:
            out[t] = miss
    return out


# ──────────────── single-file load: plan → stage → promote ─────────────────
def load_single_file(engine, schema, path, table, cmap, progress=None,
                     fk_fixes=None):
    """Perry's spec, verbatim (July 30): "Load a file. Analyze and develop a
    load plan. Load the file." This is step 3 — ONE file, THE plan, no grid,
    no skip sets, no cascade, no form. Stages the file into its own stg
    table, runs the standard promote (same builder the bulk path uses: UWI
    pad, entity ids, HOLD filters), and returns the honest numbers.

    Staging: BULK INSERT first (fast, needs the SQL service to read the
    path); pandas+executemany fallback (always works, a bit slower)."""
    import os as _os
    import re as _re
    import pandas as pd
    from sqlalchemy import text
    from dataview.import_data.bulk_dir_loader import build_promote_sql

    def _say(m):
        if progress:
            progress(m)

    import time as _time
    _t = {}
    _m = _time.perf_counter()
    hdr = _read_any(path)
    cols = [str(c) for c in hdr.columns]
    _t["read file"] = _time.perf_counter() - _m
    stem = _re.sub(r"[^A-Za-z0-9]+", "_",
                   _os.path.splitext(_os.path.basename(path))[0]).strip("_")[:80]
    stg = f"stg.one_{stem}".lower()
    _say(f"Creating staging table {stg}…")
    # RIGHT-SIZE THE STAGING COLUMNS. Every column used to be NVARCHAR(4000),
    # and pyodbc's fast_executemany pre-allocates buffers at the DECLARED
    # width: 8 columns x 4000 chars x a 20k-row batch is >1 GB of churn to
    # move 7,200 short values. That, not the driver, was the slowness
    # (Perry, July 31). Width comes from the data, floor 32, ceiling 4000.
    _m = _time.perf_counter()
    widths = {}
    for c in cols:
        try:
            longest = int(hdr[c].astype(str).str.len().max() or 0)
        except Exception:
            longest = 255
        widths[c] = max(32, min(4000, longest + 16))
    with engine.begin() as cx:
        cx.execute(text(f"IF OBJECT_ID('{stg}') IS NOT NULL DROP TABLE {stg}"))
        cx.execute(text(f"CREATE TABLE {stg} ("
                        + ", ".join(f"[{c}] NVARCHAR({widths[c]}) NULL"
                                    for c in cols)
                        + ")"))
    _t["create staging"] = _time.perf_counter() - _m
    n_staged = 0
    _m = _time.perf_counter()
    # Excel has no delimited form for BULK INSERT to read, but the dataframe
    # is already in hand from _read_any — stage it directly. Before this, a
    # plain .xlsx fell through a gap: too simple to need a transform, wrong
    # format for the loader, so it had no Load button at all (Perry, July 31).
    _is_excel = str(path).lower().endswith((".xlsx", ".xls", ".xlsm"))
    try:
        if _is_excel:
            raise RuntimeError("Excel source — staging from the dataframe")
        _say("Staging via BULK INSERT…")
        with engine.begin() as cx:
            cx.execute(text(
                f"BULK INSERT {stg} FROM '{path}' WITH (FORMAT='CSV', "
                f"FIRSTROW=2, CODEPAGE='65001', TABLOCK)"))
            n_staged = cx.execute(text(f"SELECT COUNT(*) FROM {stg}")).scalar()
        _t["stage (BULK INSERT)"] = _time.perf_counter() - _m
    except Exception as _bi:
        _t["BULK INSERT attempt (failed)"] = _time.perf_counter() - _m
        _say("Staging the workbook rows directly…" if _is_excel else
             f"BULK INSERT unavailable ({str(_bi)[:60]}…) — batched inserts "
             f"instead…")
        _m = _time.perf_counter()
        raw = engine.raw_connection()
        try:
            cur = raw.cursor()
            fast = False
            try:
                cur.fast_executemany = True
                fast = True
            except Exception:
                pass
            ins = (f"INSERT INTO {stg} ("
                   + ", ".join(f"[{c}]" for c in cols) + ") VALUES ("
                   + ", ".join("?" for _ in cols) + ")")
            vals = hdr.astype(str).where(hdr.notna(), None).values.tolist()
            # smaller batches: buffer size is batch x columns x width, so a
            # modest batch keeps the driver in cache instead of in swap
            B = 1000
            for i in range(0, len(vals), B):
                cur.executemany(ins, vals[i:i + B])
                if (i // B) % 5 == 0 or i + B >= len(vals):
                    _say(f"Staged {min(i + B, len(vals)):,} / "
                         f"{len(vals):,} rows…")
            raw.commit()
            n_staged = len(vals)
        finally:
            raw.close()
        _t[f"stage (executemany, fast={fast})"] = _time.perf_counter() - _m
    if fk_fixes:
        _say("Applying your FK resolutions to staging…")
        _m = _time.perf_counter()
        apply_fk_fixes(engine, schema, stg, fk_fixes, say=_say)
        _t["apply FK resolutions"] = _time.perf_counter() - _m
    _say("Building promote SQL (pad, ids, hold filters)…")
    _m = _time.perf_counter()
    holds = []
    # parsed catalog passed in from the session cache when available: this
    # step measured 10.95s on Perry's box, per load, for information the
    # session already had (July 31)
    _parsed = None
    try:
        import streamlit as _stx
        _parsed = _stx.session_state.get("la_cat::" + str(schema))
    except Exception:
        _parsed = None
    # PROVENANCE. The canonical DataView id — SHA1(UPPER(path), UTF-16-LE),
    # the same one file_gate mints for every extracted file, and the same
    # value HASHBYTES computes in T-SQL. Minted here so it goes INTO the
    # promoted rows; the two catalog writes happen after the counts are
    # known, because an unverified load has no business claiming anything.
    _inv = None
    try:
        from dataview.import_data.file_gate import inventory_id as _iid
        _inv = _iid(_os.path.abspath(path))
    except Exception:
        pass
    if _inv:
        # THE COMPUTED ID WINS over a mapped source column, for this one
        # target only. Normally a mapped column beats a stamp — the file is
        # the authority on its own data — but inventory_id is not data
        # about the well, it is a claim about WHICH FILE the row came from,
        # and only the loader reading the file can know that. Copying the
        # column through means trusting whatever the producer wrote there:
        # these synthetic exports carry "INVENTORY_ID-641", which resolves
        # to nothing and makes a provenance query look answerable when it
        # is not. Dropping the pair here also keeps build_promote_sql's
        # "already claimed" rule intact rather than special-casing it.
        _drop = [s for s, t in cmap.items()
                 if str(t).lower() == "inventory_id"]
        for _s in _drop:
            cmap = {k: v for k, v in cmap.items() if k != _s}
    sql, ins_cols, _pk = build_promote_sql(engine, table.upper(), cmap, [],
                                           schema, stg=stg, holds_out=holds,
                                           parsed=_parsed, inventory_id=_inv)
    _t["build promote SQL (schema introspection)"] = _time.perf_counter() - _m
    from sqlalchemy import text as _t2
    with engine.begin() as cx:
        before = cx.execute(_t2(
            f"SELECT COUNT(*) FROM {schema}.{table.lower()}")).scalar()
        _say(f"Promoting into {table}…")
        _m = _time.perf_counter()
        cx.execute(_t2(sql))
        _t["promote"] = _time.perf_counter() - _m
        after = cx.execute(_t2(
            f"SELECT COUNT(*) FROM {schema}.{table.lower()}")).scalar()
        held = []
        for parent, pcol, ccol, f in holds:
            try:
                n = cx.execute(_t2(f"SELECT COUNT(*) FROM {stg} s "
                                   f"WHERE NOT {f}")).scalar() or 0
            except Exception:
                n = 0
            if n:
                held.append((parent, int(n)))
    _say("— timings —")
    for k, v in _t.items():
        _say(f"    {k:44s} {v:7.2f}s")
    # The ledger is written HERE — after the counts are known — because a
    # load that failed half way is not a fact about the database. Best
    # effort: record_load swallows its own errors so bookkeeping can never
    # fail a load that already succeeded.
    try:
        from dataview.import_data.load_ledger import record_load
        # Registers the file in file_catalog.GLOBAL_FILE_CATALOG (identity,
        # via file_gate — the same path a LAS takes) AND records the result
        # in dataview.dv_global_file_catalog (staged/promoted/held), both
        # keyed by the id already stamped on the rows.
        record_load(engine, path, target=table.upper(),
                    staged=int(n_staged or 0), promoted=int(after - before),
                    held=sum(int(n) for _p, n in held) if held else 0,
                    fingerprint=None, register=True, log=_say)
    except Exception:
        pass
    return {"staged": int(n_staged or 0), "inserted": int(after - before),
            "present": int(after), "held": held, "stg": stg,
            "inventory_id": _inv, "timings": _t}


# ─────────────────────── deterministic verification ────────────────────────
def _verify_parent_wells(engine, path, uwi_src_col, schema="dataview"):
    """(distinct_wells, matched, missing_sample) — the FILE's wells vs
    dv_well, compared 14-vs-14 exactly as promote will. This is the check the
    AI is never trusted with: 'DV_WELL must be populated first' is its
    judgment; THIS is the count."""
    from sqlalchemy import text
    df = _read_any(path)
    uwis = sorted({_uwi14(v) for v in df[uwi_src_col] if _uwi14(v)})
    matched = set()
    with engine.connect() as cx:
        for i in range(0, len(uwis), 500):
            chunk = uwis[i:i + 500]
            marks = ",".join(f":u{j}" for j in range(len(chunk)))
            rows = cx.execute(
                text(f"SELECT uwi FROM {schema}.dv_well WHERE uwi IN ({marks})"),
                {f"u{j}": u for j, u in enumerate(chunk)}).fetchall()
            matched |= {str(r[0]).strip() for r in rows}
    missing = [u for u in uwis if u not in matched]
    return len(uwis), len(matched), missing[:10]


# ───────────────────────── transform: tops pivot ───────────────────────────
def pivot_picks_to_intervals(df, uwi_col, unit_col, type_col, depth_col,
                             code_col=None, interp_id="RMOTC_TOPS",
                             depth_ouom="ft"):
    """Pick-per-row → interval rows, with STRATIGRAPHIC STACKING (Perry's
    feedback, July 29: where top and bottom sit together in the sequence,
    translate them too).

    Per well, picks sort by depth and three rules fill base_depth:
      1. NAME PAIR — a Base pick whose unit name minus a trailing
         'base'/'bs' equals a Top's name closes that top ("Sussex Base" →
         "Sussex"), wherever it sits. Strongest evidence; wins outright.
      2. STACK CLOSE — an unpaired Base pick closes the nearest SHALLOWER
         still-open top ("Shannon Base" at 1070.76 closes Shannon Lower at
         1005.26). The two sit together in the column; that adjacency is the
         evidence.
      3. IMPLIED BASE — a top with no Base pick runs to the NEXT DEEPER
         pick's depth (Shannon Upper's base = Shannon Lower's top). The
         deepest pick stays open-bottomed.
    Base picks are consumed by 1/2 — they translate into intervals, never
    emit rows of their own. Output headers are dv-named (uwi,
    formation_name, top_depth, base_depth, pick_type) so the loader
    exact-matches with no synonyms.

    Returns (out_df, stats): name_pairs, stack_closed, implied, open_bottom,
    bad_depth."""
    import pandas as pd

    def _norm_name(s):
        return re.sub(r"\s+", " ", str(s or "").strip())

    def _is_base(name, ptype):
        t = str(ptype or "").strip().lower()
        n = _norm_name(name).lower()
        return t.startswith("base") or bool(re.search(r"\bbase\b\s*$|bs\s*$", n))

    def _base_stem(name):
        return re.sub(r"[\s_]*(base|bs)[\s_]*$", "", _norm_name(name),
                      flags=re.I).strip()

    def _f(v):
        try:
            return float(str(v).replace(",", ""))
        except ValueError:
            return None

    stats = {"name_pairs": 0, "stack_closed": 0, "implied": 0,
             "open_bottom": 0, "bad_depth": 0}
    by_well = {}
    for r in df.itertuples(index=False):
        d = dict(zip(df.columns, r))
        u = _uwi14(d.get(uwi_col, ""))
        dep = _f(d.get(depth_col, ""))
        if dep is None:
            stats["bad_depth"] += 1
            continue
        by_well.setdefault(u, []).append({
            "unit": _norm_name(d.get(unit_col, "")),
            "ptype": str(d.get(type_col, "") or "").strip(),
            "code": str(d.get(code_col, "") or "").strip() if code_col else "",
            "depth": dep,
            "raw": str(d.get(depth_col, "") or "").strip(),
        })

    out = []
    for uwi, picks in by_well.items():
        picks.sort(key=lambda p: p["depth"])
        tops = [p for p in picks if not _is_base(p["unit"], p["ptype"])]
        bases = [p for p in picks if _is_base(p["unit"], p["ptype"])]

        # rule 1 — name pair
        by_name = {}
        for t in tops:
            by_name.setdefault(t["unit"].lower(), t)
        for b in list(bases):
            t = by_name.get(_base_stem(b["unit"]).lower())
            if t is not None and "base" not in t:
                t["base"] = b["raw"]
                bases.remove(b)
                stats["name_pairs"] += 1

        # rule 2 — stack close: nearest shallower still-open top
        for b in bases:
            cand = [t for t in tops if t["depth"] < b["depth"] and "base" not in t]
            if cand:
                cand[-1]["base"] = b["raw"]        # deepest of the shallower
                stats["stack_closed"] += 1
            # a base with nothing above it has nothing to close — dropped;
            # counted implicitly as neither pair nor close

        # rule 3 — implied base from the next deeper pick
        all_depths = sorted(p["depth"] for p in picks)
        for t in tops:
            if "base" in t:
                continue
            deeper = [d for d in all_depths if d > t["depth"]]
            if deeper:
                t["base"] = f"{deeper[0]:g}"
                stats["implied"] += 1
            else:
                t["base"] = ""
                stats["open_bottom"] += 1

        for t in tops:
            # REAL dv_well_formation_top columns (schema from Perry, July 29):
            # strat_unit_id NOT NULL <- the unit CODE the file itself carries
            # (SHNNu, SSXS...) with the unit name as fallback; interp_id NOT
            # NULL <- constant; strat_unit_name; depth_ouom. Pick type goes
            # to remark as provenance — strat_unit_type in the schema means
            # the UNIT's type, not the pick's.
            out.append({"uwi": uwi,
                        "strat_unit_id": t["code"] or t["unit"].upper().replace(" ", "_"),
                        "strat_unit_name": t["unit"],
                        "top_depth": t["raw"],
                        "base_depth": t.get("base", ""),
                        "depth_ouom": depth_ouom,
                        "interp_id": interp_id,
                        "remark": f"pick_type={t['ptype']}" if t["ptype"] else ""})
    return pd.DataFrame(out, columns=["uwi", "strat_unit_id",
                                      "strat_unit_name", "top_depth",
                                      "base_depth", "depth_ouom", "interp_id",
                                      "remark"]), stats


# ─────────────────── transform: derive production entities ──────────────────
def derive_prod_entities(df, uwi_col, date_col=None, name_col=None):
    """One DV_PROD_ENTITY row per distinct well in a by-well production file.

    THE ID RULE (the whole point): prod_entity_id = the DE-SEPARATED RAW UWI,
    exactly what promote's identifier transform will mint from the volumes
    file's UWI column — so entity IDs and volumes' prod_entity_id agree BY
    CONSTRUCTION, no lookups. The entity's own `uwi` column carries the same
    value; promote pads THAT to UWI-14 (only columns named `uwi` pad), so it
    joins dv_well. Deterministic + idempotent: re-deriving yields identical
    rows.

    Optional: date_col → first/last_prod_date per well (min/max of parseable
    dates); name_col → prod_entity_name (else the raw UWI).
    Output headers are dv-named. Returns (out_df, stats)."""
    import pandas as pd

    def _desep(v):
        return "".join(ch for ch in str(v or "") if ch not in "-. ").strip()

    ent = {}
    bad = 0
    for r in df.itertuples(index=False):
        d = dict(zip(df.columns, r))
        rid = _desep(d.get(uwi_col, ""))
        if not rid:
            bad += 1
            continue
        e = ent.setdefault(rid, {"prod_entity_name": "", "dates": []})
        if name_col and not e["prod_entity_name"]:
            e["prod_entity_name"] = str(d.get(name_col, "") or "").strip()
        if date_col:
            e["dates"].append(str(d.get(date_col, "") or "").strip())
    out = []
    for rid, e in sorted(ent.items()):
        first = last = ""
        if e["dates"]:
            # SAME parser as the volumes unpivot (_prod_period) — a generic
            # parse reads "Dec-22" as 2022 and put first_prod after last_prod
            # for a 1922 well (caught in test, July 29). One date rule, two
            # transforms, no drift.
            ds = sorted(d for d in (_prod_period(v) for v in e["dates"]) if d)
            if ds:
                first, last = ds[0], ds[-1]
        out.append({"prod_entity_id": rid, "uwi": rid,
                    "prod_entity_type": "WELL",
                    "prod_entity_name": e["prod_entity_name"] or rid,
                    "first_prod_date": first, "last_prod_date": last})
    stats = {"entities": len(out), "rows_read": len(df), "bad_uwi_rows": bad}
    return pd.DataFrame(out, columns=["prod_entity_id", "uwi",
                                      "prod_entity_type", "prod_entity_name",
                                      "first_prod_date", "last_prod_date"]), stats


# ─────────────── production workbook reader + volumes unpivot ───────────────
def _read_prod_workbook(path):
    """Read a production workbook the way NPR-3 is actually shaped: MULTIPLE
    sheets (wells split alphabetically across tabs) and a TWO-ROW merged
    header (Oil/Water/Gas sit under a spanned "Monthly Production" cell).
    Header merge rule per column: row-2 name if present, else row-1. Sheets
    whose merged header yields <3 names are skipped (title/notes tabs).
    Non-Excel paths fall back to the ordinary sniffed read."""
    import pandas as pd
    if not str(path).lower().endswith((".xlsx", ".xls", ".xlsm")):
        return pd.read_csv(path, sep=None, engine="python", dtype=str,
                           keep_default_na=False)
    sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=str)
    frames = []
    for name, raw in sheets.items():
        if raw.shape[0] < 3:
            continue
        r0 = ["" if pd.isna(v) else str(v).strip() for v in raw.iloc[0]]
        r1 = ["" if pd.isna(v) else str(v).strip() for v in raw.iloc[1]]
        names, seen = [], {}
        for a, b in zip(r0, r1):
            n = (b or a).strip()
            if n in seen:                       # e.g. Section on both rows
                seen[n] += 1
                n = f"{n}_{seen[n]}"
            elif n:
                seen[n] = 0
            names.append(n)
        if sum(1 for n in names if n) < 3:
            continue
        body = raw.iloc[2:].copy()
        body.columns = [n or f"col{j}" for j, n in enumerate(names)]
        body = body.dropna(how="all")
        frames.append(body.fillna(""))
    if not frames:
        raise RuntimeError("no data sheets found (two-row header expected)")
    return pd.concat(frames, ignore_index=True).astype(str)


def _prod_period(v, century_pivot=2005):
    """'Dec-22' → 1922-12-31 (month END). Teapot produced from the 1920s, so
    2-digit years pivot at `century_pivot`: <= its last two digits → 2000s,
    else 1900s — a generic date parse would put Dec-22 in 2022. Full dates
    fall through to pandas."""
    import re as _r
    import pandas as pd
    s = str(v or "").strip()
    m = _r.match(r"^([A-Za-z]{3,9})[- ]?(\d{2})$", s)
    if m:
        yy = int(m.group(2))
        year = 2000 + yy if yy <= century_pivot % 100 else 1900 + yy
        try:
            p = pd.Period(f"{m.group(1)} {year}", freq="M")
            return p.end_time.strftime("%Y-%m-%d")
        except Exception:
            return None
    try:
        d = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(d) else d.strftime("%Y-%m-%d")
    except Exception:
        return None


def unpivot_prod_volumes(df, uwi_col, date_col, fluid_cols, days_col=None,
                         drop_zero=False, century_pivot=2005):
    """WIDE by-well-month production → TALL dv_prod_volume rows: one row per
    (well, month, fluid). fluid_cols = {"OIL": col, "WATER": col, "GAS": col}
    (missing fluids just omitted). prod_entity_id = de-separated raw UWI —
    the SAME rule derive_prod_entities uses, so the FK matches by
    construction. volume_ouom: bbl for OIL/WATER, mcf for GAS. Output headers
    are dv-named. Returns (out_df, stats)."""
    import pandas as pd

    def _desep(v):
        return "".join(ch for ch in str(v or "") if ch not in "-. ").strip()

    def _num(v):
        try:
            return float(str(v).replace(",", ""))
        except ValueError:
            return None

    out = []
    stats = {"rows_in": len(df), "rows_out": 0, "bad_date": 0, "bad_uwi": 0,
             "zero_dropped": 0, "non_numeric": 0}
    for r in df.itertuples(index=False):
        d = dict(zip(df.columns, r))
        rid = _desep(d.get(uwi_col, ""))
        if not rid:
            stats["bad_uwi"] += 1
            continue
        period = _prod_period(d.get(date_col, ""), century_pivot)
        if not period:
            stats["bad_date"] += 1
            continue
        # dv_prod_volume.period_date is nvarchar(7) = YYYY-MM (Perry's live
        # schema, July 30) — a MONTH period, which is the truer shape for
        # monthly production anyway. Full dates stay in the ENTITY file,
        # whose first/last_prod_date are datetime2.
        period = period[:7]
        days = str(d.get(days_col, "") or "").strip() if days_col else ""
        for fluid, col in fluid_cols.items():
            if not col:
                continue
            v = _num(d.get(col, ""))
            if v is None:
                stats["non_numeric"] += 1
                continue
            if drop_zero and v == 0:
                stats["zero_dropped"] += 1
                continue
            out.append({"prod_entity_id": rid, "period_date": period,
                        "fluid_type": fluid, "volume": f"{v:g}",
                        "volume_ouom": "mcf" if fluid == "GAS" else "bbl",
                        "days_on_prod": days})
    stats["rows_out"] = len(out)
    return pd.DataFrame(out, columns=["prod_entity_id", "period_date",
                                      "fluid_type", "volume", "volume_ouom",
                                      "days_on_prod"]), stats


# ──────────── transform: directional surveys → header + stations ────────────
def derive_dir_surveys(df, uwi_col, md_col, incl_col=None, azim_col=None,
                       tvd_col=None, ns_col=None, ew_col=None,
                       survey_id="DS1", depth_ouom="ft"):
    """Station-per-row survey file → TWO dv-named files: a derived
    DV_WELL_DIR_SRVY_HDR row per well and DV_WELL_DIR_SRVY_STA rows.

    ID RULES (match by construction, the prod-entity doctrine): survey_id =
    the SAME constant on both files — the hdr/sta FK cannot miss; station_id
    = 1..N per well ordered by MD (NOT NULL in the schema, so generated HERE,
    explicitly, not left to promote). uwi = de-separated raw; promote pads
    both files identically. Header carries survey_top/base_depth = min/max
    MD, survey_type='DIRECTIONAL'. Absent wells are handled downstream by
    promote's hold filter — their hdr AND sta rows park together.
    Returns (hdr_df, sta_df, stats)."""
    import pandas as pd

    def _desep(v):
        return "".join(ch for ch in str(v or "") if ch not in "-. ").strip()

    def _num(v):
        try:
            return float(str(v).replace(",", ""))
        except ValueError:
            return None

    by_well, bad = {}, 0
    for r in df.itertuples(index=False):
        d = dict(zip(df.columns, r))
        rid = _desep(d.get(uwi_col, ""))
        md = _num(d.get(md_col, ""))
        if not rid or md is None:
            bad += 1
            continue
        by_well.setdefault(rid, []).append((md, d))
    hdr, sta = [], []
    for rid, rows in sorted(by_well.items()):
        rows.sort(key=lambda x: x[0])
        hdr.append({"uwi": rid, "survey_id": survey_id,
                    "survey_type": "DIRECTIONAL",
                    "survey_top_depth": f"{rows[0][0]:g}",
                    "survey_base_depth": f"{rows[-1][0]:g}",
                    "depth_ouom": depth_ouom})
        for i, (md, d) in enumerate(rows, start=1):
            def g(col):
                return str(d.get(col, "") or "").strip() if col else ""
            sta.append({"uwi": rid, "survey_id": survey_id,
                        "station_id": str(i), "md": f"{md:g}",
                        "incl": g(incl_col), "azim": g(azim_col),
                        "tvd": g(tvd_col), "ns_offset": g(ns_col),
                        "ew_offset": g(ew_col), "depth_ouom": depth_ouom})
    stats = {"wells": len(hdr), "stations": len(sta), "bad_rows": bad}
    return (pd.DataFrame(hdr, columns=["uwi", "survey_id", "survey_type",
                                       "survey_top_depth", "survey_base_depth",
                                       "depth_ouom"]),
            pd.DataFrame(sta, columns=["uwi", "survey_id", "station_id", "md",
                                       "incl", "azim", "tvd", "ns_offset",
                                       "ew_offset", "depth_ouom"]), stats)


# ─────────────────────────────── the page ──────────────────────────────────
def run():
    """Standalone entry (run_load_assistant.py) — draws its own connection
    inputs, then the panel. The Bulk Tabular Loader embeds render() directly
    as Phase 0 instead."""
    ss = st.session_state
    st.header("🧭 Load Assistant")
    c1, c2 = st.columns(2)
    server = c1.text_input("Server", value=ss.get("bdl_server", r"localhost\SQLEXPRESS"),
                           key="la_server")
    database = c2.text_input("Database", value=ss.get("bdl_db", "DataView_Demo"),
                             key="la_db")
    render(ss, server, database, ss.get("bdl_schema", "dataview"),
           directory=ss.get("bdl_dir"))


def render(ss, server, database, schema="dataview", directory=None):
    """`directory` is passed in by the host page. It used to be read out of
    session state, which broke the moment the host returned early before
    writing it — three rounds of "Directory not found" (July 31). An
    argument cannot go stale, and if it is empty the box below is always
    there as a fallback."""
    """The assistant panel — embeddable, no header or connection inputs of
    its own. Hosted by the Bulk Tabular Loader as Phase 0 (the AI is the
    first step) and by run() standalone."""
    import pandas as pd
    st.caption("Tell it what you want loaded, point it at the file. It reads the "
               "file, grounds itself in the live schema, and lays out the load "
               "plan — table, column map, required-gap rules, parent "
               "prerequisites, and any row-shape transform — for your review. "
               "**Adopt** teaches the loader, whose fingerprint recall then "
               "runs the real load hands-free. **↺ Refine** talks back.")

    # NO INTAKE MODE. A folder and a file are both just a path, and the
    # difference is something os.path can answer — asking the operator to
    # declare it first was a mode where none was needed (Perry, July 31:
    # "why is there a separate single file option… allow a directory or a
    # single file"). One box, both meanings, and skipping handles anything
    # in a folder you don't want.
    def _clean0(p):
        return str(p or "").strip().strip('"').strip("'")

    _inherited = _clean0(directory) or _clean0(ss.get("bdl_dir"))
    _target = _clean0(st.text_input(
        "Folder or file", value=_clean0(ss.get("la_target")) or _inherited,
        key="la_target",
        help="A folder plans every file in it (skip what you don't want); "
             "a single file plans just that one. Quotes from Explorer's "
             "'Copy as path' are stripped for you."))
    _is_dir = bool(_target) and os.path.isdir(_target)
    _is_file = bool(_target) and os.path.isfile(_target)
    if _target and not (_is_dir or _is_file):
        st.error(f"Not found: `{_target}`")
    elif _is_file:
        st.caption(f"📄 single file: `{os.path.basename(_target)}`")
    elif _is_dir:
        st.caption(f"📁 folder: `{_target}`")

    # ── 📄 DOCUMENTS: NOT THIS PAGE'S JOB ────────────────────────────────
    # The Document Assistant is its own nav page now. This panel keeps only
    # a SIGNPOST: pdf/docx land here often enough (they look like "files in
    # a folder") that silently ignoring them reads as a bug. Detect, name
    # the right door, don't embed — one tool, one page.
    _DOC_EXTS = {".pdf", ".docx"}

    if _is_file and os.path.splitext(_target)[1].lower() in _DOC_EXTS:
        st.info("📄 That's a document, not a tabular file — open the "
                "**Document Assistant** in the left nav and point it here. "
                "This page loads CSV/Excel.")
        return                      # nothing below applies to a document

    if _is_dir:
        try:
            import glob as _glob
            # Flat, like the tabular scan below it — one folder, one rule.
            _docs = sorted(
                p for _pat in ("*.pdf", "*.docx")
                for p in _glob.glob(os.path.join(_target, _pat))
                if not os.path.basename(p).startswith(("~$", "._")))
        except Exception:
            _docs = []
        if _docs:
            st.caption(f"📄 {len(_docs)} document(s) here (pdf/docx) — those "
                       f"belong to the **Document Assistant**; the scan "
                       f"below reads only tabular files.")

    mode = "Directory" if _is_dir else "Single file"
    rep_path, rep_note = "", ""
    if _is_dir:
        d = _target
        if st.button("🔍 Scan shapes", key="la_scan_shapes"):
            if not d:
                st.error("Paste a folder path in the box above.")
            elif not os.path.isdir(d):
                st.error(f"Not a folder: `{d}` — check the path. Explorer's "
                         f"'Copy as path' adds quotes; those are stripped "
                         f"automatically, but a file path instead of a "
                         f"folder path will land here.")
            else:
                _eng_s = get_engine(server, database)
                with st.spinner("🔍 Reading headers and checking the "
                                "mapping store…"):
                    gs = _scan_shapes(_eng_s, d, schema)
                # ONE batch AI call proposes tables for the unknown shapes —
                # proposals only; Adopt stays the explicit act.
                # The AI call is NOT part of scanning. A scan reads headers
                # and the store — both local and fast; a network round trip
                # does not belong on that path (Perry, July 31). Proposals
                # moved to their own button below.
                for i, g in enumerate(gs):
                    g["proposed"] = None
                    g["assigned"] = g["known"] or ""
                    g["flag"] = False
                _nf = sum(len(g["files"]) for g in gs)
                if gs:
                    st.caption(f"{_nf} tabular file(s) found → {len(gs)} "
                               f"shape(s). Every csv/txt/dat/tsv/prn/Excel in "
                               f"the folder is listed — including earlier "
                               f"conversions and copies. Tick skip on the "
                               f"ones you aren't loading.")
                ss.pop("la_cat::" + str(schema), None)   # re-read once
                ss["la_groups"] = gs
                ss["la_grid_ver"] = int(ss.get("la_grid_ver", 0)) + 1
                ss.pop("la_plan", None)
                if not gs:
                    st.warning("No readable tabular files found. Looked for "
                               "csv / txt / dat / tsv / prn / xlsx / xls / "
                               "xlsm directly in that folder (no "
                               "subdirectories). Files that failed to parse "
                               "are skipped silently — tell me an example "
                               "filename if some should have appeared.")
        groups = ss.get("la_groups") or []
        if groups:
            import pandas as _pd
            # live table list for the dropdown, cached per session
            if "la_tables" not in ss:
                try:
                    _, _COLS, _ = _catalog_cached(
                        ss, get_engine(server, database), schema)
                    ss["la_tables"] = sorted(_COLS.keys())
                except Exception:
                    ss["la_tables"] = []
            # ── Perry's grid: every file, its table, its plan status ─────────
            # One row per FILE for visibility; the working unit stays the
            # SHAPE — assigning a table to any file assigns its whole shape,
            # and one AI plan covers every file of that shape.
            try:
                _FKC, _, _ = _catalog_cached(ss, get_engine(server, database),
                                             schema)
            except Exception:
                _FKC = {}
            _ords = _load_order([g.get("assigned") or "" for g in groups], _FKC)
            _abs = _absent_parents([g.get("assigned") or "" for g in groups],
                                   _FKC)
            for _t, _ps in _abs.items():
                st.caption(f"⛓ {_t}: parent {', '.join(_ps)} is not among "
                           f"these files — its rows will hold at promote "
                           f"until that parent is populated.")
            _rows_g = []
            for gi, g in enumerate(groups):
                t_cur = g.get("assigned") or ""
                # "remembered" ≠ "tested": the store recalls where this shape
                # goes (from a previous Save/Adopt of the OPERATOR'S), but only
                # a successful load + verify proves the mapping. A green
                # "known" overstated that (Perry, July 29).
                stat = ("📋 planned" if g.get("_adopted")
                        else "📇 remembered" if g["known"]
                        else "🤖 proposed" if g.get("proposed")
                        else "⚠ unassigned")
                _skips = g.setdefault("skips", set())
                _ord_eff = (g.get("load_override")
                            or _ords.get(t_cur.upper(), 999))
                for p in g["files"]:
                    _rows_g.append({"skip": p in _skips,
                                    "needs plan": bool(g.get("flag")),
                                    "load #": _ord_eff,
                                    "file": os.path.basename(p),
                                    "shape": gi + 1,
                                    "→ table": t_cur,
                                    "status": ("⏭ skipped" if p in _skips
                                               else stat)})
            _rows_g.sort(key=lambda r: (r["load #"], r["shape"], r["file"]))
            _only_un = st.checkbox(
                "⚠ Show unassigned only", key="la_grid_unonly",
                help="Filter to files whose shape has no table yet (skipped "
                     "files hidden too). Edits made in the filtered view "
                     "apply on 🔃 Re-sort / 💾 Adopt like any other; hidden "
                     "rows keep their current settings.")
            if _only_un:
                _rows_g = [r for r in _rows_g
                           if not r["→ table"] and r["status"] != "⏭ skipped"]
                if not _rows_g:
                    st.success("Nothing unassigned — every shape has a table.")
            # IN A FORM: a data_editor outside one reruns the whole page on
            # every cell change. Inside, Streamlit holds the edits until the
            # submit button — which is also when "fully edited" actually
            # means something (Perry, July 31).
            _ver = int(ss.get("la_grid_ver", 0))
            _form_g = st.form(key=f"la_gridform_v{_ver}", border=False)
            edited_g = _form_g.data_editor(
                _pd.DataFrame(_rows_g), hide_index=True,
                use_container_width=True,
                key=f"la_dirgrid_v{_ver}_{int(_only_un)}",
                column_config={
                    "skip": st.column_config.CheckboxColumn(
                        width="small",
                        help="Leave this FILE out of the assistant — planning, "
                             "counts and the representative sample ignore it. "
                             "Phase-1 staging loads by TABLE, so park a file "
                             "elsewhere if it must never load at all."),
                    "needs plan": st.column_config.CheckboxColumn(
                        width="small",
                        help="Flag this SHAPE for a full AI plan below (map, "
                             "gaps, parents, transforms — e.g. the tops "
                             "pivot). One plan covers all its files."),
                    "load #": st.column_config.NumberColumn(
                        width="small", min_value=1, step=1,
                        help="Computed from the FK graph (dv_r_* first, "
                             "parents before children; 999 = unassigned). "
                             "Type a number to OVERRIDE — one number per "
                             "shape, applied on 🔃 Re-sort; typing the "
                             "computed value returns the shape to "
                             "automatic. Advisory: it orders this grid and "
                             "your work sequence; the FK gate still holds "
                             "children whose parents are missing."),
                    "file": st.column_config.TextColumn(disabled=True),
                    "shape": st.column_config.NumberColumn(
                        disabled=True, width="small",
                        help="Files sharing a shape share one decision"),
                    "→ table": st.column_config.SelectboxColumn(
                        options=[""] + ss["la_tables"], required=False,
                        help="✅ recalled · 🤖 proposed by the AI · or set it "
                             "yourself. Override freely, then 🔃 Re-sort."),
                    "status": st.column_config.TextColumn(
                        disabled=True, width="small",
                        help="📇 remembered = the store recalls this shape's "
                             "table from a previous Save/Adopt of yours — "
                             "NOT yet proven; the load + verify query is the "
                             "test. 🤖 proposed = AI suggestion, adopt to "
                             "remember. 📋 planned = full AI plan adopted "
                             "this session.")})
            def _harvest(ed):
                """Pull ALL pending edits (tables, plan flags, per-file skips)
                out of the editor into `groups`. EVERY button that bumps the
                grid version must call this FIRST — Adopt used to rebuild
                without harvesting and silently discarded pending skip ticks
                (July 29). Returns conflict shape list."""
                _conf = []
                _tbl, _flg, _skp = {}, {}, {}
                _n2p = {os.path.basename(p): p
                        for g in groups for p in g["files"]}
                for _rr in ed.to_dict("records"):
                    t = str(_rr["→ table"] or "").strip().upper()
                    si = _rr["shape"]
                    if t:
                        prev = _tbl.setdefault(si, t)
                        if prev != t:
                            _conf.append(si)
                    _flg[si] = _flg.get(si, False) or bool(_rr["needs plan"])
                    if bool(_rr["skip"]):
                        _skp.setdefault(si, set()).add(
                            _n2p.get(_rr["file"], _rr["file"]))
                _lod = {}
                for _rr in ed.to_dict("records"):
                    try:
                        _lod.setdefault(_rr["shape"], set()).add(
                            int(_rr["load #"]))
                    except (TypeError, ValueError):
                        pass
                msgs = []
                if _conf:
                    msgs.append("One table per shape: shape(s) "
                                + ", ".join(str(s) for s in sorted(set(_conf)))
                                + " have two different tables set.")
                _lconf = sorted(si for si, v in _lod.items() if len(v) > 1)
                if _lconf:
                    msgs.append("One load # per shape: shape(s) "
                                + ", ".join(str(s) for s in _lconf)
                                + " have two different numbers.")
                if msgs:
                    return msgs
                _seen = {r["shape"] for r in ed.to_dict("records")}
                for si, g in enumerate(groups, start=1):
                    if si not in _seen:
                        continue        # hidden by the filter — keep as-is
                    g["assigned"] = _tbl.get(si, g.get("assigned") or "")
                    g["flag"] = _flg.get(si, False)
                    g["skips"] = _skp.get(si, set())
                _ords2 = _load_order([g.get("assigned") or "" for g in groups],
                                     _FKC)
                for si, g in enumerate(groups, start=1):
                    vals = _lod.get(si)
                    if not vals:
                        continue
                    v = vals.pop()
                    comp = _ords2.get((g.get("assigned") or "").upper(), 999)
                    g["load_override"] = None if v == comp else v
                return []

            c_r, c_a, c_b = _form_g.columns([1, 1, 2])
            if c_r.form_submit_button("🔃 Re-sort by load order",
                          help="Applies your table overrides, plan flags and "
                               "skips, then re-orders the grid parents-first."):
                _conf = _harvest(edited_g)
                if _conf:
                    st.error(" ".join(_conf))
                else:
                    ss["la_grid_ver"] = _ver + 1
                    st.rerun()
            if c_a.form_submit_button("💾 Adopt table assignments",
                          help="Remembers shape → table for every row where "
                               "you set one (all files of the shape). Columns "
                               "map in Phase 2 via synonyms/exact matches — "
                               "or plan the shape with the AI for a full map."):
                _eng_g = get_engine(server, database)
                _conflict = _harvest(edited_g)      # skips/flags survive Adopt
                if _conflict:
                    st.error(" ".join(_conflict))
                else:
                    _n_new = 0
                    for _si, g in enumerate(groups, start=1):
                        t = g.get("assigned") or ""
                        if not t or g["known"] == t:
                            continue
                        # a MARKER row, not fake column confirmations: it
                        # gives fingerprint→table recall without pretending
                        # any column mapping was human-reviewed. Phase 2
                        # still proposes columns via synonyms/exact.
                        pdl._remember_mapping(_eng_g, t, g["fp"],
                                              {"__SHAPE__": "__table__"})
                        g["known"] = t
                        _n_new += 1
                    ss["la_grid_ver"] = _ver + 1
                    if _n_new:
                        st.success(f"{_n_new} shape assignment(s) remembered.")
                    st.rerun()
            def _live_files(g):
                return [p for p in g["files"] if p not in g.get("skips", set())]
            n_known = sum(1 for g in groups if g["known"])
            unknown = [(i, g) for i, g in enumerate(groups)
                       if not g["known"] and _live_files(g)]
            _un_now = [(i, g) for i, g in enumerate(groups)
                       if not (g.get("assigned") or "")]
            if _un_now and st.button(
                    f"🤖 Propose tables for {len(_un_now)} unassigned "
                    f"shape(s)", key="la_propose",
                    help="One batch call. Scanning stays local and instant; "
                         "asking the AI is a separate, deliberate step."):
                _eng_p = get_engine(server, database)
                _unk = []
                for i, g in _un_now:
                    try:
                        _, rows1, _ = _read_sample(g["files"][0], n_rows=1)
                    except Exception:
                        rows1 = []
                    _unk.append({"idx": i, "cols": g["cols"],
                                 "row": rows1[0] if rows1 else []})
                try:
                    with st.spinner("🤖 One batch call for every unknown "
                                    "shape…"):
                        props = ai_propose_tables(_eng_p, schema, _unk)
                    for i, g in enumerate(groups):
                        if i in props:
                            g["proposed"] = props[i]
                            g["assigned"] = props[i]
                    ss["la_grid_ver"] = int(ss.get("la_grid_ver", 0)) + 1
                    st.rerun()
                except Exception as _pe:
                    st.error(f"Proposal failed: {str(_pe)[:160]}")

            if not unknown:
                st.success(f"All {len(groups)} shape(s) remembered — scan the "
                           f"directory in Phase 1 below and fingerprint recall "
                           f"auto-assigns every file. Remembered ≠ tested: "
                           f"the promote run and its verify counts are the "
                           f"test of each mapping.")
            else:
                c_b.caption(f"{n_known} shape(s) assigned · {len(unknown)} "
                            f"open — ✅/🤖/⚠ in status; override tables, 🔃 "
                            f"re-sort, 💾 adopt, and flag shapes for a plan.")
                # A "needs plan" tick selects ANY file's shape — including
                # one already known/proposed (re-planning is legitimate:
                # a shape can be assigned right and still need its transform
                # or gap analysis). Unflagged pool = the unassigned ones.
                _flagged = [(i, g) for i, g in enumerate(groups)
                            if g.get("flag") and _live_files(g)]
                _pool = _flagged or unknown
                labels = {f"shape {i + 1}"
                          + (" 🤖" if g.get("flag") else "")
                          + f" — {len(_live_files(g))} file(s), "
                          f"e.g. {os.path.basename(_live_files(g)[0])}": g
                          for i, g in _pool}
                pick = st.selectbox("Plan a shape with the AI"
                                    + (" (flagged first)" if _flagged else ""),
                                    list(labels), key="la_shape_pick")
                rep_path = _live_files(labels[pick])[0]
                rep_note = (f"representative file: `{os.path.basename(rep_path)}` "
                            f"— adopting covers all "
                            f"{len(_live_files(labels[pick]))} unskipped "
                            f"file(s) of this shape")

    # ── last load result: stays on screen until you move on ────────────────
    _res = ss.get("la_last_result")
    if _res:
        st.success(f"✅ **{_res['file']}** → {_res['table']}: "
                   f"{_res['staged']:,} staged, {_res['inserted']:,} inserted "
                   f"(table now holds {_res['present']:,}).")
        for _p, _n in _res.get("held", []):
            st.warning(f"⏸ {_n:,} row(s) held — no match in {_p}.")
        if _res.get("learned"):
            st.caption("📇 learned: " + ", ".join(f"{s}→{t}"
                                                  for s, t in _res["learned"]))
        _nx = ss.get("la_pending_next")
        _qq = ss.get("la_queue") or []
        _cols_r = st.columns([1, 1, 3])
        if _nx is not None and _nx < len(_qq):
            if _cols_r[0].button(
                    f"▶ Next: {os.path.basename(_qq[_nx]['path'])}",
                    type="primary", key="la_next_file"):
                ss["la_q_idx"] = _nx
                ss.pop("la_pending_next", None)
                ss.pop("la_last_result", None)
                ss.pop("la_plan", None)
                ss.pop("la_fk", None)
                ss["la_autorun"] = True
                st.rerun()
        elif _nx is not None:
            _cols_r[0].success("Queue complete.")
        if _cols_r[1].button("✔ Dismiss", key="la_dismiss_result"):
            ss.pop("la_last_result", None)
            st.rerun()
        st.divider()

    # ── the work queue ──────────────────────────────────────────────────────
    _q = ss.get("la_queue") or []
    if (_is_dir or _q) and (ss.get("la_groups") or _q):
        st.markdown("**Queue**")
        if not _q and ss.get("la_groups"):
            if st.button("▶ Build the queue from this scan",
                         key="la_qbuild", type="primary",
                         help="Every unskipped file, in load order, parents "
                              "first. The assistant walks them itself."):
                # FK ORDER, not scan order. The first cut read only the
                # manual override, so every row tied at 999 and sorted by
                # scan sequence — which put stations first and wells last,
                # exactly backwards (Perry, July 31).
                try:
                    _FKC_q, _, _ = _catalog_cached(
                        ss, get_engine(server, database), schema)
                except Exception:
                    _FKC_q = {}
                _ord_q = _load_order([(_g.get("assigned") or "")
                                      for _g in ss["la_groups"]], _FKC_q)
                _rows = []
                for _gi, _g in enumerate(ss["la_groups"]):
                    _skips = _g.get("skips", set())
                    _tbl = (_g.get("assigned") or "").upper()
                    _ordinal = (_g.get("load_override")
                                or _ord_q.get(_tbl, 999))
                    for _p in _g["files"]:
                        if _p in _skips:
                            continue
                        _rows.append({
                            "path": _p, "shape": _gi + 1, "table": _tbl,
                            "order": _ordinal, "status": "pending"})
                _rows.sort(key=lambda r: (r["order"], r["shape"], r["path"]))
                ss["la_queue"] = _rows
                ss["la_q_idx"] = 0
                st.rerun()
        if _q:
            _idx = int(ss.get("la_q_idx", 0))
            _done = sum(1 for r in _q if r["status"] in ("loaded", "skipped"))
            st.progress(min(1.0, _done / max(1, len(_q))),
                        text=f"{_done} of {len(_q)} file(s) done")
            st.dataframe(pd.DataFrame([{
                "": "▶" if i == _idx else "",
                "load #": r.get("order", 999),
                "file": os.path.basename(r["path"]),
                "→ table": r["table"] or "(plan it)",
                "status": {"pending": "· pending", "loaded": "✅ loaded",
                           "skipped": "⏭ skipped",
                           "planned": "📋 planned"}.get(r["status"], "")}
                for i, r in enumerate(_q)]),
                hide_index=True, use_container_width=True, height=180)

            # move the arrow anywhere — the queue suggests an order, it does
            # not impose one (Perry: "how do you move the arrow")
            _labels = [f"{i + 1}. {os.path.basename(r['path'])}"
                       f"{'  ✅' if r['status'] == 'loaded' else ''}"
                       f"{'  ⏭' if r['status'] == 'skipped' else ''}"
                       for i, r in enumerate(_q)]
            _pick = st.selectbox(
                "▶ Work on", _labels,
                index=min(_idx, len(_labels) - 1), key=f"la_qpick_{len(_q)}",
                help="Jump the arrow to any file. Loading still follows FK "
                     "sense: a child loaded before its parent just holds.")
            _pi = _labels.index(_pick)
            if _pi != _idx:
                ss["la_q_idx"] = _pi
                ss.pop("la_plan", None)
                ss.pop("la_fk", None)
                st.rerun()

            _a, _b, _c = st.columns(3)
            if _idx < len(_q):
                if _a.button("▶ Analyze this one", key="la_qgo"):
                    ss["la_autorun"] = True
                    ss.pop("la_plan", None)
                    ss.pop("la_fk", None)
                    st.rerun()
                if _b.button("⏭ Skip it", key="la_qskip"):
                    _q[_idx]["status"] = "skipped"
                    ss["la_q_idx"] = _idx + 1
                    ss.pop("la_plan", None)
                    ss.pop("la_fk", None)
                    ss["la_autorun"] = _idx + 1 < len(_q)
                    st.rerun()
            if _c.button("✖ Clear queue", key="la_qclear"):
                ss.pop("la_queue", None)
                ss.pop("la_q_idx", None)
                st.rerun()
            if _idx >= len(_q):
                st.success("Queue finished — every file is loaded or "
                           "skipped.")

    instruction = st.text_input(
        "What should be loaded, and where?",
        placeholder="e.g.  please load this to the well header table   ·   "
                    "formation tops picks, one pick per row",
        key="la_instruction")
    _qcur = None
    _q2 = ss.get("la_queue") or []
    _qi = int(ss.get("la_q_idx", 0))
    if _q2 and _qi < len(_q2):
        _qcur = _q2[_qi]
    if _qcur:
        fpath = _qcur["path"]
        st.caption(f"▶ queue file {_qi + 1} of {len(_q2)}: "
                   f"`{os.path.basename(fpath)}`"
                   + (f" → {_qcur['table']}" if _qcur["table"] else "")
                   + " — nothing to type; the queue hands it over.")
    elif rep_path:
        fpath = rep_path
        st.caption(rep_note)
    elif _is_dir:
        # a folder run never types a file path: the shape grid or the queue
        # hands one over
        fpath = ""
        st.caption("📄 file comes from the grid or the queue above — pick a "
                   "shape to plan, or build the queue.")
    else:
        fpath = _target

    # The queue fires this too: "files should present themselves to the AI
    # without having to hand enter" (Perry, July 31). One flag, same body —
    # no duplicated analyze path to drift out of sync.
    _bypass = st.checkbox(
        "🔬 Bypass the synonym store (AI only) — for comparison",
        key="la_bypass",
        help="Analyze without the store, to see what the AI maps unaided. "
             "The comparison is printed under the grid. Turn it off for "
             "normal work — the store is faster, free, and deterministic.")
    _auto = bool(ss.pop("la_autorun", False))
    if st.button("🤖 Analyze and plan", type="primary", key="la_go") or _auto:
        if not fpath or not os.path.exists(fpath):
            st.error("File not found — paste the full path.")
            st.stop()
        try:
            cols, sample, n_total = _read_sample(fpath)
        except Exception as e:
            st.error(f"Could not read the file: {str(e)[:200]}")
            st.stop()
        eng = get_engine(server, database)
        if not ss.get("la_seeded"):
            seed_vendor_synonyms(eng, schema)   # vocabulary before anything
            ss["la_seeded"] = True
        try:
            with st.spinner("🤖 Reading the file, consulting the live "
                            "schema, drafting the plan…"):
                table, cmap, notes, extra = ai_suggest_table_map(
                    eng, schema, instruction, cols, sample, None,
                    transforms_catalog=_TRANSFORM_CATALOG)
        except Exception as e:
            st.error(f"AI analysis failed: {str(e)[:250]}")
            st.stop()
        _tn, _tp = _validate_transform((extra or {}).get("transform"), cols)
        _sample_dicts = [dict(zip(cols, r)) for r in (sample or [])]
        _cmap2, _prov, _fit, _left = apply_store(eng, schema, table, cols,
                                                 cmap, _sample_dicts,
                                                 bypass=_bypass)
        ss["la_plan"] = {"file": fpath, "n_rows": n_total, "cols": cols,
                         "instruction": instruction, "table": table,
                         "cmap": _cmap2, "ai_cmap": cmap, "map_src": _prov,
                         "fit": _fit, "unmapped": _left,
                         "sample": _sample_dicts,
                         "notes": notes, "extra": extra,
                         "transform": _tn, "t_params": _tp}
        ss["la_history"] = [("you", instruction or "(no instruction)"),
                            ("assistant", f"→ {table} · {len(cmap)} column(s) "
                                          f"mapped · {notes}")]
        ss["la_plan_ver"] = int(ss.get("la_plan_ver", 0)) + 1

    plan = ss.get("la_plan")
    if not plan:
        return

    # ── the conversation so far ─────────────────────────────────────────────
    hist = ss.get("la_history") or []
    if hist:
        with st.expander(f"💬 Conversation ({len(hist)} turn(s))",
                         expanded=False):
            for role, txt in hist:
                st.markdown(("**You:** " if role == "you" else "**Assistant:** ")
                            + str(txt))
    eng = get_engine(server, database)
    extra = plan.get("extra") or {}

    st.subheader(f"Plan — `{os.path.basename(plan['file'])}` "
                 f"({plan['n_rows']:,} rows) → **{plan['table']}**")
    if plan.get("notes"):
        st.caption(plan["notes"])

    # Every transform tool below pre-fills its pickers from the plan's
    # validated params; a Refine bumps la_plan_ver, which re-keys the pickers
    # so revised defaults actually TAKE (fixed keys keep stale choices —
    # "It didn't pick up my revision", July 30).
    _prop_p = plan.get("t_params") or {}
    _pver = int(ss.get("la_plan_ver", 0))
    def _pp(param, cols_list, fallback_idx, optional=False):
        v = _prop_p.get(param)
        if v and v in cols_list:
            return cols_list.index(v) + (1 if optional else 0)
        return fallback_idx

    # ── ① the mapping grid: one row per SOURCE column, editable ─────────────
    # Read-only lists made overriding impossible (the ' base' column that
    # holds unit codes had no way to be repointed). One grid, every source
    # column, a dropdown of real target columns, and where each match came
    # from. Perry, July 31: "a simple grid of source column and matched
    # columns would be better."
    # ── ⚡ FAST PATH ─────────────────────────────────────────────────────────
    # A file whose columns all came from the store, whose values fit the
    # target types and whose required columns are covered has nothing left
    # for a human to decide. Making the operator walk ①→⑤ to confirm five
    # green screens is ceremony, not care. The steps stay below, unchanged,
    # for anything that isn't in this state — and for anyone who wants to
    # look before loading.
    #
    # FKs are the one thing this cannot know without asking the database,
    # so the button ASKS: it runs the FK scan first, loads only if it comes
    # back clean, and drops into ④b if it doesn't. That keeps the promise
    # ("everything was already known") honest rather than optimistic.
    _ff = plan.get("fit") or []
    _fit_errs = sum(1 for i in _ff if i[0] == "error")
    _gaps0 = [g for g in (extra.get("required_gaps") or [])
              if not g.get("can_generate")]
    _prov0 = dict(plan.get("map_src") or {})
    _from_store = sum(1 for v in _prov0.values() if str(v).lower() == "store")
    _mapped0 = dict(plan.get("cmap") or {})
    _fk_clean = (ss.get("la_fk") is not None and not ss.get("la_fk"))
    _ready = (bool(_mapped0) and not _fit_errs and not _gaps0
              and not plan.get("derived")
              and _from_store == len(_prov0) and _from_store > 0
              and str(plan["file"]).lower().endswith(
                  (".csv", ".txt", ".tsv", ".dat", ".prn",
                   ".xlsx", ".xls", ".xlsm")))
    if _ready:
        st.success(f"⚡ **Ready to load** — all {_from_store} column(s) came "
                   f"from the store, values fit the target types, and no "
                   f"required column is missing."
                   + ("  Foreign keys already checked and clean."
                      if _fk_clean else ""))
        if st.button(f"🚀 Load {os.path.basename(plan['file'])} into "
                     f"{plan['table']} now",
                     key=f"la_fast_{_pver}", type="primary",
                     use_container_width=True,
                     help="Checks foreign keys, then loads. Nothing below "
                         "needs answering — the steps are there if you want "
                         "to look first."):
            _eng_f = get_engine(server, database)
            _stat_f = st.status("Checking foreign keys…", expanded=True)
            try:
                _fkres = fk_scan(_eng_f, schema, plan["table"], plan["cmap"],
                                 plan["file"])
            except Exception as _fe:
                _fkres = None
                _stat_f.write(f"FK check failed: {str(_fe)[:160]}")
            if _fkres:
                ss["la_fk"] = _fkres
                ss["la_fk_ver"] = int(ss.get("la_fk_ver", 0)) + 1
                _stat_f.update(label="Foreign keys need a decision",
                               state="error")
                st.warning(f"{sum(len(f['missing']) for f in _fkres)} value(s) "
                           f"across {len(_fkres)} parent(s) have no parent row "
                           f"— resolve them in ④b below, then load.")
            elif _fkres is None:
                _stat_f.update(label="Could not check foreign keys",
                               state="error")
            else:
                ss["la_fk"] = []
                _stat_f.write("Foreign keys clean — loading.")
                try:
                    _r = load_single_file(_eng_f, schema, plan["file"],
                                          plan["table"], plan["cmap"],
                                          progress=_stat_f.write)
                    _stat_f.update(label="Loaded", state="complete")
                    ss["la_result"] = {"table": plan["table"],
                                       "file": plan["file"], **_r}
                    st.rerun()
                except Exception as _le:
                    _stat_f.update(label="Load failed", state="error")
                    st.error(str(_le)[:400])
        st.caption("The steps below are unchanged — open them if you want to "
                   "see the mapping before it loads.")

    st.markdown("**① Column map**")
    _prov = dict(plan.get("map_src") or {})
    _cmap = dict(plan.get("cmap") or {})
    _eng_m = get_engine(server, database)

    if "la_tgt_cols::" + plan["table"] not in ss:
        _opts = []
        if _store is not None:
            try:
                _opts = sorted(c for c in _store.attributes(
                    _eng_m, schema, plan["table"]).keys()
                    if not _store.is_system_column(c))
            except Exception:
                _opts = []
        if not _opts:
            _opts = sorted({v for v in _cmap.values() if v})
        ss["la_tgt_cols::" + plan["table"]] = _opts
    _targets = ss["la_tgt_cols::" + plan["table"]]

    _VIA = {"store": "📇 store", "ai": "🤖 AI", "you": "✋ you"}
    _mver = int(ss.get("la_map_ver", 0))
    _form_m = st.form(key=f"la_mapform_{_pver}_{_mver}", border=False)
    _grid = pd.DataFrame([{
        "source": c,
        "→ column": _cmap.get(c, ""),
        "via": _VIA.get(_prov.get(c), "" if c not in _cmap else "📇 store"),
    } for c in plan["cols"]])
    _edited = _form_m.data_editor(
        _grid, hide_index=True, use_container_width=True,
        key=f"la_mapgrid_{_pver}_{_mver}",
        column_config={
            "source": st.column_config.TextColumn(disabled=True),
            "→ column": st.column_config.SelectboxColumn(
                options=[""] + _targets, required=False,
                help="Blank = not loaded. Change any row to override — your "
                     "choice always wins over the store and the AI."),
            "via": st.column_config.TextColumn(
                disabled=True, width="small",
                help="📇 store = a synonym confirmed before · 🤖 AI = this "
                     "run's proposal · ✋ you = your override"),
        })

    _n_store = sum(1 for v in _prov.values() if v == "store")
    _n_ai = sum(1 for v in _prov.values() if v == "ai")
    _n_you = sum(1 for v in _prov.values() if v == "you")
    st.caption(f"{_n_store} from the store · {_n_ai} from the AI · "
               f"{_n_you} yours · {len(plan['cols']) - len(_cmap)} unmapped. "
               f"Store beats AI; you beat both.")

    _c1, _c2 = _form_m.columns([1, 2])
    _teach = _c2.checkbox("Also teach my changes to the synonym store",
                          key=f"la_teach_{_pver}_{_mver}",
                          help="Writes each changed row as an operator-grade "
                               "synonym for this table, so the next file "
                               "with that header maps itself. Leave off for "
                               "a one-file quirk.")
    if _c1.form_submit_button("✔ Apply mapping",
                              help="Applies your edits and re-runs the fit "
                                   "pre-flight."):
        _new, _dups = {}, []
        _claimed = {}
        for _r in _edited.to_dict("records"):
            _s, _t = _r["source"], str(_r["→ column"] or "").strip().lower()
            if not _t:
                continue
            if _t in _claimed:
                _dups.append((_t, _claimed[_t], _s))
                continue
            _claimed[_t] = _s
            _new[_s] = _t
        if _dups:
            for _t, _a, _b in _dups:
                st.error(f"Two sources both map to **{_t}**: '{_a}' and "
                         f"'{_b}'. One column can only be filled once.")
        else:
            _changed = [(s, t) for s, t in _new.items() if _cmap.get(s) != t]
            for s, t in _changed:
                _prov[s] = "you"
            for s in list(_prov):
                if s not in _new:
                    _prov.pop(s, None)
            plan["cmap"], plan["map_src"] = _new, _prov
            plan["unmapped"] = [c for c in plan["cols"] if c not in _new]
            if _store is not None and plan.get("sample"):
                try:
                    plan["fit"] = _store.check_fit(_eng_m, schema,
                                                   plan["table"], _new,
                                                   plan["sample"])
                except Exception:
                    plan["fit"] = []
            if _teach and _changed and _store is not None:
                for s, t in _changed:
                    try:
                        _store.set_synonym(_eng_m, schema, plan["table"], s, t)
                    except Exception:
                        pass
                st.success(f"{len(_changed)} override(s) taught to the store.")
            ss["la_map_ver"] = _mver + 1
            ss["la_plan"] = plan
            st.rerun()

    # ── store vs AI, measured ───────────────────────────────────────────────
    if _store is not None and plan.get("ai_cmap"):
        with st.expander("🔬 Store vs AI — what each mapped", expanded=False):
            try:
                _sm, _su, _ = _store.suggest_map(_eng_m, schema,
                                                 plan["table"], plan["cols"])
            except Exception:
                _sm, _su = {}, []
            _ai = {s: str(t).lower()
                   for s, t in (plan.get("ai_cmap") or {}).items() if t}
            _rows_c = []
            for c in plan["cols"]:
                a, s_ = _ai.get(c, ""), _sm.get(c, "")
                if not a and not s_:
                    continue
                _rows_c.append({"source": c, "store said": s_ or "—",
                                "AI said": a or "—",
                                "": ("=" if a == s_ else
                                     "store only" if s_ and not a else
                                     "AI only" if a and not s_ else "DIFFER")})
            if _rows_c:
                st.dataframe(pd.DataFrame(_rows_c), hide_index=True,
                             use_container_width=True)
                _agree = sum(1 for r in _rows_c if r[""] == "=")
                _diff = [r for r in _rows_c if r[""] == "DIFFER"]
                _so = sum(1 for r in _rows_c if r[""] == "store only")
                _ao = sum(1 for r in _rows_c if r[""] == "AI only")
                st.caption(f"agree on {_agree} · store only {_so} · AI only "
                           f"{_ao} · disagree {len(_diff)}. Agreement means "
                           f"the store bought speed and determinism; "
                           f"disagreement is where it bought correctness — "
                           f"check those rows.")
                for r in _diff:
                    st.warning(f"**{r['source']}** — store says "
                               f"`{r['store said']}`, AI says `{r['AI said']}`")

    _fit = plan.get("fit") or []
    if _fit:
        st.markdown("**①b Fit pre-flight** — values checked against the live "
                    "column types before anything is staged")
        render_fit(_fit)

    # step 2 — required gaps → derived rules
    gaps = extra.get("required_gaps") or []
    if gaps:
        st.markdown("**② Required columns with no source**")
        st.dataframe(pd.DataFrame([{
            "column": g.get("column", ""),
            "generatable": "✅" if g.get("can_generate") else "✗",
            "how": g.get("how", "")} for g in gaps]),
            hide_index=True, use_container_width=True)

    # step 3 — parent prerequisites, VERIFIED not just asserted
    parents = extra.get("parents") or []
    if parents:
        st.markdown("**③ Parent prerequisites**")
        for p in parents:
            st.info(f"⛓ **{p.get('table')}** — {p.get('why', '')}")
        uwi_src = next((s for s, t in plan["cmap"].items() if t == "uwi"), None)
        if uwi_src and any(str(p.get("table", "")).upper() == "DV_WELL" for p in parents):
            if st.button("🔎 Verify against dv_well now (14-vs-14, exactly as "
                         "promote compares)", key="la_verify"):
                try:
                    n, m, miss = _verify_parent_wells(eng, plan["file"], uwi_src, schema)
                except Exception as e:
                    st.error(f"verify failed: {str(e)[:200]}")
                else:
                    if n == m:
                        st.success(f"✅ all {n} distinct wells in the file exist "
                                   f"in dv_well — nothing will be held.")
                    else:
                        st.warning(f"{n - m} of {n} wells are NOT in dv_well — "
                                   f"their rows will be held at promote. First "
                                   f"missing: {', '.join(miss)}")

    # step 4 — shape transform: now a PLAN STEP, not prose
    shape = (extra.get("shape_note") or "").strip()
    _tn = plan.get("transform")
    if _tn or shape:
        st.markdown("**④ Row-shape transform**")
    if _tn:
        _tp = plan.get("t_params") or {}
        st.info(f"🤖 Proposed: **{_tn}** — "
                + (", ".join(f"{k}=`{v}`" for k, v in _tp.items()) or "(no params)")
                + ". The matching tool below opens pre-filled; check the "
                  "pickers and press its Run/Derive button — running is "
                  "always your act.")
    if shape:
        st.warning("📐 " + shape)
    _mapped = set(plan["cmap"].values())
    _looks_tops = plan["table"].upper() == "DV_WELL_FORMATION_TOP"
    if _looks_tops:
        with st.expander("🔧 Pivot picks → intervals (writes a derived CSV "
                         "beside the file; the original is untouched)",
                         expanded=bool(shape)
                         or plan.get("transform") == "pivot_picks_to_intervals"):
            st.caption("Three rules, strongest first: ① name pair — 'Sussex "
                       "Base' closes 'Sussex'; ② stack close — an unpaired "
                       "Base pick closes the nearest shallower open top; "
                       "③ implied base — a top with no Base runs to the next "
                       "deeper pick's depth; the deepest stays open-bottomed. "
                       "Base picks translate into the intervals — they never "
                       "emit rows of their own. Output headers are dv-named, "
                       "so the loader exact-matches them.")
            cols = plan["cols"]
            def _guess(pred, default=0):
                for i, c in enumerate(cols):
                    if pred(c.lower().strip()):
                        return i
                return default
            g1, g2 = st.columns(2)
            uwi_c = g1.selectbox("UWI column", cols,
                                 index=_pp("uwi_col", cols,
                                           _guess(lambda c: "uwi" in c or "api" in c)),
                                 key=f"la_p_uwi_{_pver}")
            unit_c = g2.selectbox("Unit-name column", cols,
                                  index=_pp("unit_col", cols,
                                            _guess(lambda c: "strat" in c or "form" in c, 1)),
                                  key=f"la_p_unit_{_pver}")
            g3, g4 = st.columns(2)
            type_c = g3.selectbox("Pick-type column (Top/Base/…)", cols,
                                  index=_guess(lambda c: "type" in c, 2),
                                  key=f"la_p_type_{_pver}")
            depth_c = g4.selectbox("Depth column", cols,
                                   index=_guess(lambda c: "depth" in c or "md" in c, 3),
                                   key=f"la_p_depth_{_pver}")
            g5, g6, g7 = st.columns(3)
            code_c = g5.selectbox("Unit-code column (→ strat_unit_id)",
                                  ["—"] + cols,
                                  index=(_guess(lambda c: c in ("base", "code")
                                                or "code" in c, -1) + 1),
                                  key=f"la_p_code_{_pver}",
                                  help="The tops file carries unit codes "
                                       "(SHNNu, SSXS) — they become the NOT-"
                                       "NULL strat_unit_id; blank codes fall "
                                       "back to the condensed unit name.")
            interp_v = g6.text_input("interp_id (NOT NULL)",
                                     value="RMOTC_TOPS", key=f"la_p_interp_{_pver}")
            ouom_v = g7.text_input("depth_ouom", value="ft", key=f"la_p_ouom_{_pver}")
            if st.button("Run pivot", key="la_pivot"):
                with st.spinner("📖 Reading the file and stacking picks "
                                "into intervals…"):
                    df = _read_any(plan["file"])
                    out, stats = pivot_picks_to_intervals(
                        df, uwi_c, unit_c, type_c, depth_c,
                        None if code_c == "—" else code_c,
                        interp_v.strip() or "RMOTC_TOPS", ouom_v.strip() or "ft")
                dest = os.path.splitext(plan["file"])[0] + "__intervals.csv"
                out.to_csv(dest, index=False)
                _dv = plan.setdefault("derived", [])
                for _dp, _dt in [(dest, plan["table"])]:
                    _dv[:] = [e for e in _dv if e[0] != _dp]
                    _dv.append((_dp, _dt))
                st.success(f"{len(out):,} interval row(s) → `{dest}` · "
                           f"{stats['name_pairs']} name pair(s), "
                           f"{stats['stack_closed']} stack-closed, "
                           f"{stats['implied']} implied base(s), "
                           f"{stats['open_bottom']} open-bottomed, "
                           f"{stats['bad_depth']} bad depth(s) dropped")
                st.dataframe(out.head(12), hide_index=True,
                             use_container_width=True)
                st.caption("Point the Bulk Tabular Loader at the folder — the "
                           "derived file's headers exact-match dv columns.")

    # refine — the two-way part: object, and the plan revises in place
    st.markdown("**↺ Refine** — tell it what's wrong, or ask differently")
    fb = st.text_input("Feedback / new angle",
                       placeholder="e.g.  the base column is a code, not a depth  ·  "
                                   "this belongs in the picks table, not tops",
                       key="la_feedback")
    if st.button("↺ Revise the plan", key="la_refine") and fb.strip():
        eng2 = get_engine(server, database)
        try:
            cols0, sample0, _ = _read_sample(plan["file"])
            _spin_rev = st.spinner("🤖 Revising the plan with your feedback…")
            _spin_rev.__enter__()
            table, cmap, notes, extra = ai_suggest_table_map(
                eng2, schema, plan.get("instruction", ""), cols0, sample0, None,
                transforms_catalog=_TRANSFORM_CATALOG,
                prior={"plan": {"table": plan["table"], "colmap": plan["cmap"],
                                "required_gaps": (plan.get("extra") or {}).get("required_gaps"),
                                "parents": (plan.get("extra") or {}).get("parents"),
                                "shape_note": (plan.get("extra") or {}).get("shape_note")},
                       "feedback": fb.strip()})
        except Exception as e:
            st.error(f"revision failed: {str(e)[:250]}")
        else:
            _spin_rev.__exit__(None, None, None)
            _tn2, _tp2 = _validate_transform((extra or {}).get("transform"),
                                             plan["cols"])
            _sd2 = [dict(zip(cols0, r)) for r in (sample0 or [])]
            cmap, _prov2, _fit2, _left2 = apply_store(
                eng2, schema, table, plan["cols"], cmap, _sd2)
            plan["map_src"], plan["fit"], plan["unmapped"] = (
                _prov2, _fit2, _left2)
            plan["sample"] = _sd2
            plan.update({"table": table, "cmap": cmap, "notes": notes,
                         "extra": extra, "transform": _tn2, "t_params": _tp2})
            ss["la_plan"] = plan
            hist = ss.get("la_history") or []
            hist.append(("you", fb.strip()))
            hist.append(("assistant", f"→ {table} · {len(cmap)} column(s) "
                                      f"mapped · {notes}"))
            ss["la_history"] = hist
            ss["la_plan_ver"] = int(ss.get("la_plan_ver", 0)) + 1
            st.rerun()

    _is_prod = plan["table"].upper() in ("DV_PROD_VOLUME", "DV_PROD_ENTITY")
    if _is_prod:
        with st.expander("🏭 Derive production entities (writes a "
                         "__prod_entity.csv beside the file)",
                         expanded=plan.get("transform") == "derive_prod_entities"):
            st.caption("One DV_PROD_ENTITY row per distinct well — the absent "
                       "parent the ⛓ warning names. prod_entity_id = the "
                       "de-separated raw UWI, exactly what the volumes rows "
                       "will carry, so the FK matches by construction; the "
                       "entity's uwi joins dv_well via the UWI-14 pad. Load "
                       "the derived file AFTER wells, BEFORE volumes.")
            colsx = plan["cols"]
            def _gx(pred, default=0):
                for i, c in enumerate(colsx):
                    if pred(c.lower().strip()):
                        return i
                return default
            p1, p2, p3 = st.columns(3)
            uwi_cx = p1.selectbox("UWI column", colsx,
                                  index=_pp("uwi_col", colsx,
                                            _gx(lambda c: "uwi" in c or "api" in c)),
                                  key=f"la_pe_uwi_{_pver}")
            date_cx = p2.selectbox("Date column (optional)", ["—"] + colsx,
                                   index=_pp("date_col", colsx,
                                             _gx(lambda c: "date" in c or "month" in c,
                                                 -1) + 1, optional=True),
                                   key=f"la_pe_date_{_pver}")
            name_cx = p3.selectbox("Name column (optional)", ["—"] + colsx,
                                   index=(_gx(lambda c: "name" in c or "well" in c,
                                              -1) + 1),
                                   key=f"la_pe_name_{_pver}")
            if st.button("Derive entities", key=f"la_pe_go_{_pver}"):
                with st.spinner("📖 Reading the workbook (all sheets) and "
                                "deriving one entity per well…"):
                    df = _read_prod_workbook(plan["file"])
                    out, stats = derive_prod_entities(
                        df, uwi_cx,
                        None if date_cx == "—" else date_cx,
                        None if name_cx == "—" else name_cx)
                dest = os.path.splitext(plan["file"])[0] + "__prod_entity.csv"
                out.to_csv(dest, index=False)
                st.success(f"{stats['entities']:,} entit(ies) from "
                           f"{stats['rows_read']:,} row(s) "
                           f"({stats['bad_uwi_rows']} bad UWI row(s) skipped) "
                           f"→ `{dest}`")
                _dv = plan.setdefault("derived", [])
                for _dp, _dt in [(dest, "DV_PROD_ENTITY")]:
                    _dv[:] = [e for e in _dv if e[0] != _dp]
                    _dv.append((_dp, _dt))
                st.dataframe(out.head(10), hide_index=True,
                             use_container_width=True)
        with st.expander("🛢 Unpivot volumes — wide Oil/Water/Gas columns → "
                         "tall dv_prod_volume rows (writes __volumes.csv)",
                         expanded=plan.get("transform") == "unpivot_prod_volumes"):
            st.caption("The NPR-3 workbook is one row per well-month with "
                       "fluids as COLUMNS; dv_prod_volume wants one row per "
                       "well-month-FLUID. Handles the multi-sheet workbook, "
                       "the two-row merged header, and the 2-digit years "
                       "(Dec-22 = 1922 — this field predates your grandad's "
                       "truck). prod_entity_id uses the same rule as 🏭, so "
                       "the FK matches by construction.")
            _wbk = "la_wbcols::" + plan["file"]
            try:
                if _wbk not in ss:
                    # cached: this used to re-read the WHOLE workbook on
                    # every rerun just to fill dropdowns — the silent
                    # "is it frozen?" freeze (Perry, July 30)
                    with st.spinner("📖 First read of the workbook…"):
                        ss[_wbk] = list(_read_prod_workbook(plan["file"]).columns)
                _wb_cols = list(ss[_wbk])
            except Exception as _we:
                _wb_cols = plan["cols"]
                st.caption(f"(workbook read fell back to scanned columns — "
                           f"{str(_we)[:80]})")
            def _gv(pred, default=0):
                for i, c in enumerate(_wb_cols):
                    if pred(c.lower().strip()):
                        return i
                return default
            v1, v2, v3 = st.columns(3)
            uwi_cv = v1.selectbox("UWI column", _wb_cols,
                                  index=_pp("uwi_col", _wb_cols,
                                            _gv(lambda c: "api" in c or "uwi" in c)),
                                  key=f"la_pv_uwi_{_pver}")
            date_cv = v2.selectbox("Date column", _wb_cols,
                                   index=_pp("date_col", _wb_cols,
                                             _gv(lambda c: "date" in c, 2)),
                                   key=f"la_pv_date_{_pver}")
            days_cv = v3.selectbox("Days-produced column (optional)",
                                   ["—"] + _wb_cols,
                                   index=(_gv(lambda c: "day" in c, -1) + 1),
                                   key=f"la_pv_days_{_pver}")
            v4, v5, v6 = st.columns(3)
            oil_cv = v4.selectbox("Oil column", ["—"] + _wb_cols,
                                  index=_pp("oil_col", _wb_cols,
                                            _gv(lambda c: c == "oil", -1) + 1,
                                            optional=True),
                                  key=f"la_pv_oil_{_pver}")
            wat_cv = v5.selectbox("Water column", ["—"] + _wb_cols,
                                  index=_pp("water_col", _wb_cols,
                                            _gv(lambda c: "water" in c, -1) + 1,
                                            optional=True),
                                  key=f"la_pv_wat_{_pver}")
            gas_cv = v6.selectbox("Gas column", ["—"] + _wb_cols,
                                  index=_pp("gas_col", _wb_cols,
                                            _gv(lambda c: c == "gas", -1) + 1,
                                            optional=True),
                                  key=f"la_pv_gas_{_pver}")
            zdrop = st.checkbox("Drop zero volumes", value=False,
                                key=f"la_pv_zdrop_{_pver}",
                                help="Unticked keeps 0s (a well producing 0 "
                                     "gas is data). Ticking roughly halves "
                                     "the row count for this field.")
            if st.button("Unpivot volumes", key=f"la_pv_go_{_pver}"):
                fl = {"OIL": None if oil_cv == "—" else oil_cv,
                      "WATER": None if wat_cv == "—" else wat_cv,
                      "GAS": None if gas_cv == "—" else gas_cv}
                if not any(fl.values()):
                    st.error("Pick at least one fluid column. (If Oil/Water/"
                             "Gas aren't in the dropdowns, you're likely on "
                             "an exploded per-sheet CSV whose two-row header "
                             "wasn't merged — point the plan at the ORIGINAL "
                             ".xls workbook instead; the reader merges all "
                             "sheets and their headers there.)")
                    st.stop()
                with st.spinner("📖 Reading all sheets and unpivoting "
                                "fluids into rows…"):
                    df = _read_prod_workbook(plan["file"])
                    out, stats = unpivot_prod_volumes(
                        df, uwi_cv, date_cv, fl,
                        None if days_cv == "—" else days_cv, zdrop)
                dest = os.path.splitext(plan["file"])[0] + "__volumes.csv"
                out.to_csv(dest, index=False)
                _msg = (f"{stats['rows_in']:,} well-month row(s) → "
                        f"{stats['rows_out']:,} volume row(s) → `{dest}` · "
                        f"{stats['bad_date']} bad date(s), "
                        f"{stats['bad_uwi']} bad UWI(s), "
                        f"{stats['non_numeric']} non-numeric value(s) skipped, "
                        f"{stats['zero_dropped']} zero(s) dropped")
                if stats["rows_out"] == 0:
                    st.error("0 rows out — wrong fluid/date columns for this "
                             "file. " + _msg)
                else:
                    st.success(_msg)
                    _dv = plan.setdefault("derived", [])
                    _dv[:] = [e for e in _dv if e[0] != dest]
                    _dv.append((dest, "DV_PROD_VOLUME"))
                st.dataframe(out.head(9), hide_index=True,
                             use_container_width=True)
                st.caption("Load order: wells → __prod_entity.csv → this. "
                           "Then verify: row count vs the stats line, and "
                           "MIN(period_date) should read 19xx-xx (YYYY-MM), "
                           "not 20xx.")

    _is_srvy = plan["table"].upper() in ("DV_WELL_DIR_SRVY_HDR",
                                         "DV_WELL_DIR_SRVY_STA")
    if _is_srvy:
        with st.expander("🌀 Derive survey header + stations (writes "
                         "__srvy_hdr.csv and __srvy_sta.csv beside the file)",
                         expanded=plan.get("transform") == "derive_dir_surveys"):
            st.caption("Station-per-row file → one derived header row per "
                       "well + station rows with station_id = 1..N ordered "
                       "by MD. survey_id is the same constant on both files, "
                       "so the sta→hdr FK matches by construction; wells "
                       "absent from dv_well are ⏸ held by promote, hdr and "
                       "sta together. Load order: wells → __srvy_hdr → "
                       "__srvy_sta.")
            _svk = "la_wbcols::" + plan["file"]
            try:
                if _svk not in ss:
                    with st.spinner("📖 First read of the workbook…"):
                        ss[_svk] = list(_read_prod_workbook(plan["file"]).columns)
                _sv_cols = list(ss[_svk])
            except Exception:
                _sv_cols = plan["cols"]
            def _gs(pred, default=-1):
                for i, c in enumerate(_sv_cols):
                    if pred(c.lower().strip()):
                        return i
                return default
            s1, s2, s3 = st.columns(3)
            uwi_cs = s1.selectbox("UWI column", _sv_cols,
                                  index=_pp("uwi_col", _sv_cols,
                                            max(_gs(lambda c: "api" in c or "uwi" in c), 0)),
                                  key=f"la_sv_uwi_{_pver}")
            md_cs = s2.selectbox("MD column", _sv_cols,
                                 index=_pp("md_col", _sv_cols,
                                           max(_gs(lambda c: c.startswith("md")
                                                   or c == "depth"
                                                   or "measured" in c), 0)),
                                 key=f"la_sv_md_{_pver}")
            incl_cs = s3.selectbox("Inclination (optional)", ["—"] + _sv_cols,
                                   index=_gs(lambda c: "incl" in c or "drift" in c) + 1,
                                   key=f"la_sv_incl_{_pver}")
            s4, s5, s6 = st.columns(3)
            azim_cs = s4.selectbox("Azimuth (optional)", ["—"] + _sv_cols,
                                   index=_gs(lambda c: "azim" in c or "azi" in c) + 1,
                                   key=f"la_sv_azim_{_pver}")
            tvd_cs = s5.selectbox("TVD (optional)", ["—"] + _sv_cols,
                                  index=_gs(lambda c: "tvd" in c or "true vertical" in c) + 1,
                                  key=f"la_sv_tvd_{_pver}")
            sid_v = s6.text_input("survey_id constant", value="DS1",
                                  key=f"la_sv_sid_{_pver}")
            s7, s8, _ = st.columns(3)
            ns_cs = s7.selectbox("N/S offset (optional)", ["—"] + _sv_cols,
                                 index=_gs(lambda c: c.startswith("ns") or "north" in c) + 1,
                                 key=f"la_sv_ns_{_pver}")
            ew_cs = s8.selectbox("E/W offset (optional)", ["—"] + _sv_cols,
                                 index=_gs(lambda c: c.startswith("ew") or "east" in c) + 1,
                                 key=f"la_sv_ew_{_pver}")
            if st.button("Derive header + stations", key=f"la_sv_go_{_pver}"):
                with st.spinner("📖 Reading all sheets, numbering stations "
                                "by MD…"):
                    df = _read_prod_workbook(plan["file"])
                _opt = lambda v: None if v == "—" else v
                hdr, sta, stats = derive_dir_surveys(
                    df, uwi_cs, md_cs, _opt(incl_cs), _opt(azim_cs),
                    _opt(tvd_cs), _opt(ns_cs), _opt(ew_cs),
                    sid_v.strip() or "DS1")
                base = os.path.splitext(plan["file"])[0]
                hdr.to_csv(base + "__srvy_hdr.csv", index=False)
                sta.to_csv(base + "__srvy_sta.csv", index=False)
                st.success(f"{stats['wells']:,} survey header(s), "
                           f"{stats['stations']:,} station(s) "
                           f"({stats['bad_rows']} bad row(s) skipped) → "
                           f"`{base}__srvy_hdr.csv` + `__srvy_sta.csv`")
                _dv = plan.setdefault("derived", [])
                for _dp, _dt in [(base + "__srvy_hdr.csv", "DV_WELL_DIR_SRVY_HDR"), (base + "__srvy_sta.csv", "DV_WELL_DIR_SRVY_STA")]:
                    _dv[:] = [e for e in _dv if e[0] != _dp]
                    _dv.append((_dp, _dt))
                st.dataframe(sta.head(8), hide_index=True,
                             use_container_width=True)

    # ── ④b foreign keys — Perry's four-column grid (July 31) ────────────────
    # add ✔ = seed the parent · replace ✔ + a standard value = remap ·
    # neither = null it out. Checked BEFORE staging so violations are
    # resolved while the file is still a file.
    st.markdown("**④b Foreign key check**")
    _fkver = int(ss.get("la_fk_ver", 0))
    if st.button("🔗 Check foreign keys", key=f"la_fkscan_{_pver}_{_fkver}",
                 help="Reads the mapped values and finds any with no parent "
                      "row yet."):
        with st.spinner("Comparing mapped values against their parents…"):
            try:
                ss["la_fk"] = fk_scan(get_engine(server, database), schema,
                                      plan["table"], plan["cmap"],
                                      plan["file"])
                ss["la_fk_ver"] = _fkver + 1
            except Exception as _fe:
                st.error(f"FK check failed: {str(_fe)[:200]}")
        st.rerun()

    _fk = ss.get("la_fk")
    if _fk is not None and not _fk:
        st.success("Every mapped value has a parent row — nothing to resolve.")
    for _i, _f in enumerate(_fk or []):
        st.markdown(f"**{_f['parent']}** ← `{_f['source_col']}` → "
                    f"`{_f['child_col']}` · {len(_f['missing'])} value(s) "
                    f"with no parent row ({_f['n_have']:,} exist)")
        # CHECK ALL. Fifty unmatched status codes is fifty clicks, and the
        # answer is nearly always the same for every row in one parent —
        # seed them all, or none. Deliberately OUTSIDE the form: a form
        # only reports on submit, and this has to change the grid's
        # defaults before the operator submits anything. The editor key
        # carries its state so the grid actually re-defaults when it is
        # toggled (a fixed-key data_editor keeps its old values forever).
        if _f["kind"] != "reference":
            # Said BEFORE the check-all, not after the grid: "add all" on a
            # data parent creates a stub row per value with only the key
            # filled, and a warning underneath the thing it warns about is
            # read too late.
            st.warning(f"⚠ {_f['parent']} is a DATA table — adding creates a "
                       f"stub row with only the key filled. Usually the right "
                       f"move is to load {_f['parent']} first, or leave these "
                       f"unticked so the rows hold and promote later.")
        _allkey = f"la_fkall_{_i}_{_pver}_{_fkver}"
        _all = st.checkbox(
            f"☑ Add all {len(_f['missing'])} value(s) to {_f['parent']}",
            key=_allkey,
            help="Ticks every row's add box. Untick to clear them all; "
                 "individual rows can still be changed afterwards.")
        _fg = pd.DataFrame([{"add": bool(_all),
                             "value": v,
                             "rows": n,
                             "standard value": "",
                             "replace with →": False}
                            for v, n in _f["missing"]])
        _form_f = st.form(key=f"la_fkform_{_i}_{_pver}_{_fkver}",
                          border=False)
        _ed = _form_f.data_editor(
            _fg, hide_index=True, use_container_width=True,
            key=f"la_fkgrid_{_i}_{_pver}_{_fkver}_{int(bool(_all))}",
            column_config={
                "add": st.column_config.CheckboxColumn(
                    width="small",
                    help=f"Add this value to {_f['parent']} as a new row"),
                "value": st.column_config.TextColumn(disabled=True),
                "rows": st.column_config.NumberColumn(disabled=True,
                                                      width="small"),
                "standard value": st.column_config.SelectboxColumn(
                    options=[""] + _f["options"],
                    help="The existing parent value to use instead"),
                "replace with →": st.column_config.CheckboxColumn(
                    width="small",
                    help="Replace this value with the standard value chosen "
                         "on this row"),
            })
        _form_f.form_submit_button("✔ Apply these decisions")
        _f["_edited"] = _ed
    if _fk:
        st.caption("Ticked **add** = create it in the parent · ticked "
                   "**replace with →** = use the standard value on that row · "
                   "**neither** = null the value out. Both ticked is a "
                   "contradiction and will be reported.")

    # ── 🚀 the three-line path: this file, this plan, loaded now ─────────────
    st.markdown("**⑤ Load**")
    _derived = list(plan.get("derived") or [])
    if _derived:
        st.caption("Derived file(s) ready — load in this order (parents "
                   "first). Mapping is exact by construction: the derived "
                   "headers ARE the table's columns.")
        for _dpath, _dtable in _derived:
            if st.button(f"🚀 Load {os.path.basename(_dpath)} into "
                         f"{_dtable} now",
                         key=f"la_go_dv_{_dtable}_{_pver}", type="primary"):
                _stat = st.status("Loading…", expanded=True)
                try:
                    _eng_l = get_engine(server, database)
                    _hcols, _, _ = _read_sample(_dpath, n_rows=0)
                    _idmap = {c: c.lower() for c in _hcols}
                    _stat.write("Pre-flight: checking values against the "
                                "live column types…")
                    _iss = []
                    if _store is not None:
                        try:
                            _iss = _store.check_fit(_eng_l, schema, _dtable,
                                                    _idmap,
                                                    sample_rows(_dpath))
                        except Exception:
                            _iss = []
                    _errs = [i for i in _iss if i[0] == "error"]
                    if _errs:
                        _stat.update(label="Blocked by pre-flight",
                                     state="error")
                        render_fit(_iss)
                        st.error("Not staged — fix the derive (or the target "
                                 "column) and re-run. This is the check that "
                                 "would otherwise fail at promote.")
                        st.stop()
                    if _iss:
                        render_fit(_iss)
                    res = load_single_file(_eng_l,
                                           schema, _dpath, _dtable, _idmap,
                                           progress=lambda m: _stat.write(m))
                    _stat.update(label="Load complete", state="complete")
                    st.success(f"✅ {res['staged']:,} staged → "
                               f"{res['inserted']:,} inserted into {_dtable} "
                               f"(now {res['present']:,}).")
                    for parent, n in res["held"]:
                        st.warning(f"⏸ {n:,} held — no match in {parent}; "
                                   f"they load on a re-run once {parent} "
                                   f"has those rows.")
                    if res["inserted"] == 0 and not res["held"]:
                        st.info("0 inserted, nothing held = every row "
                                "already existed (clean re-run).")
                    ss["la_last_result"] = {
                        "file": os.path.basename(_dpath), "table": _dtable,
                        "staged": res["staged"], "inserted": res["inserted"],
                        "present": res["present"], "held": res["held"],
                        "learned": []}
                    _advance_queue(ss, _dpath)
                    # ── the store learns, but only from a load that WORKED ──
                    if _store is not None:
                        try:
                            _lr = _store.learn_from_load(_eng_l, schema,
                                                         _dtable, _idmap)
                            if ss.get("la_last_result"):
                                ss["la_last_result"]["learned"] = \
                                    _lr["learned"][:8]
                            if _lr["learned"]:
                                st.caption("📇 learned: " + ", ".join(
                                    f"{s}→{t}" for s, t in _lr["learned"][:8]))
                            for s, was, now in _lr["conflicts"]:
                                st.warning(f"'{s}' already means {was} in "
                                           f"{_dtable}; not overwritten with "
                                           f"{now} — your call.")
                        except Exception:
                            pass
                except Exception as _le:
                    _stat.update(label="Load failed", state="error")
                    st.error(f"Load failed: {str(_le)[:300]}")
    _can_load = (bool(plan.get("cmap"))
                 and not _derived            # derived buttons supersede —
                                             # two buttons for one table
                                             # confused Perry (July 30)
                 and str(plan["file"]).lower().endswith(
                     (".csv", ".txt", ".tsv", ".dat", ".prn",
                      ".xlsx", ".xls", ".xlsm")))
    if not _can_load and not _derived:
        st.caption("Nothing to load directly here — the mapping is empty, or "
                   "this file needs a transform first (run its derive tool "
                   "above; each derived CSV gets its own Load button).")
    elif _can_load and str(plan["file"]).lower().endswith(
            (".xlsx", ".xls", ".xlsm")):
        st.caption("Excel source: the first sheet is staged directly, with "
                   "no CSV conversion.")
    if _can_load and st.button(f"🚀 Load this file into {plan['table']} now",
                 key=f"la_go_{_pver}", type="primary",
                 help="Stages THIS file, promotes with THIS plan's mapping "
                      "(same engine as the bulk path: UWI pad, id rules, "
                      "hold filters), reports the counts. No grid, no "
                      "skips."):
        _stat = st.status("Loading…", expanded=True)
        try:
            _eng_l = get_engine(server, database)
            _errs0 = [i for i in (plan.get("fit") or []) if i[0] == "error"]
            if _errs0:
                _stat.update(label="Blocked by pre-flight", state="error")
                render_fit(plan.get("fit") or [])
                st.error("Not staged — the values do not fit the target "
                         "columns. Fix the mapping or the source, then "
                         "Analyze again.")
                st.stop()
            _fixes, _fkerrs = [], []
            for _f in (ss.get("la_fk") or []):
                if _f.get("_edited") is None:
                    continue
                _r, _e = resolve_from_grid(_f["_edited"])
                _fkerrs += _e
                if _r:
                    _fixes.append({"source_col": _f["source_col"],
                                   "parent": _f["parent"],
                                   "parent_col": _f["parent_col"],
                                   "resolutions": _r})
            if _fkerrs:
                _stat.update(label="Fix the FK grid first", state="error")
                for _e in _fkerrs:
                    st.error(_e)
                st.stop()
            res = load_single_file(_eng_l, schema,
                                   plan["file"], plan["table"], plan["cmap"],
                                   progress=lambda m: _stat.write(m),
                                   fk_fixes=_fixes)
            _held_n = sum(n for _, n in res["held"])
            _stat.update(label="Load complete", state="complete")
            st.success(f"✅ {res['staged']:,} row(s) staged → "
                       f"{res['inserted']:,} inserted into {plan['table']} "
                       f"(table now holds {res['present']:,}).")
            for parent, n in res["held"]:
                st.warning(f"⏸ {n:,} row(s) held — no match in {parent}; "
                           f"they load automatically on a re-run once "
                           f"{parent} has those rows.")
            if res["inserted"] == 0 and not res["held"]:
                st.info("0 inserted with nothing held = every row already "
                        "existed (a clean re-run).")
            ss["la_last_result"] = {
                "file": os.path.basename(plan["file"]),
                "table": plan["table"], "staged": res["staged"],
                "inserted": res["inserted"], "present": res["present"],
                "held": res["held"], "learned": []}
            _advance_queue(ss, plan["file"])
            if _store is not None:
                try:
                    _lr = _store.learn_from_load(_eng_l, schema,
                                                 plan["table"], plan["cmap"])
                    if ss.get("la_last_result"):
                        ss["la_last_result"]["learned"] = _lr["learned"][:8]
                    if _lr["learned"]:
                        st.caption("📇 learned: " + ", ".join(
                            f"{s}→{t}" for s, t in _lr["learned"][:8]))
                    for s, was, now in _lr["conflicts"]:
                        st.warning(f"'{s}' already means {was} in "
                                   f"{plan['table']}; not overwritten with "
                                   f"{now} — your call.")
                except Exception:
                    pass
        except Exception as _le:
            _stat.update(label="Load failed", state="error")
            st.error(f"Load failed: {str(_le)[:300]}")

    # step 5 — adopt: teach the stores
    st.markdown("**⑥ Adopt** — remember this so the real run is hands-free")
    st.caption("Writes the column map to dv_column_map keyed by this file's "
               "column fingerprint (your click IS the confirmation). The Bulk "
               "Tabular Loader's fingerprint recall then assigns the table at "
               "100% and pre-maps every column on scan.")
    if st.button("✅ Adopt plan (teach the loader)", key="la_adopt"):
        try:
            fp = pdl.fingerprint_cols(sorted(plan["cols"]))
            pdl._remember_mapping(eng, plan["table"], fp, plan["cmap"])
            st.success(f"Remembered: shape {fp} → {plan['table']} + "
                       f"{len(plan['cmap'])} column mapping(s). Scan the folder "
                       f"in the Bulk Tabular Loader; this file auto-assigns.")
            for g in (ss.get("la_groups") or []):
                if g["fp"] == fp:
                    g["known"] = plan["table"]      # retire from the queue
                    g["_adopted"] = True            # grid shows 📋 planned
                    ss["la_grid_ver"] = int(ss.get("la_grid_ver", 0)) + 1
                    ss.pop("la_plan", None)
                    st.rerun()
        except Exception as e:
            st.error(f"adopt failed: {str(e)[:200]}")
