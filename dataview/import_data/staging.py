"""
staging.py  —  PPDM Loader · Source File Staging
=================================================
Ingests CSV/TSV source files with minimal Python — reads only the header
row to build the CREATE TABLE statement, saves the file to disk, then
lets SQL Server BULK INSERT do all the heavy lifting.

No pandas. No row iteration. Fast even for 500k+ row files.

Excel files are the one exception — openpyxl converts them to CSV first,
then the same BULK INSERT path is used.
"""

from __future__ import annotations

import csv
import io
import os
import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


# ═══════════════════════════════════════════════════════════════════════
# RESULT TYPES
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class IngestResult:
    ok:             bool
    message:        str
    filename:       str       = ""
    staging_name:   str       = ""
    row_count:      int       = 0     # -1 until after BULK INSERT
    col_count:      int       = 0
    columns:        list[str] = field(default_factory=list)
    warnings:       list[str] = field(default_factory=list)
    csv_path:       str       = ""    # path to saved file for BULK INSERT
    delimiter:      str       = ","   # output delimiter (always comma after normalise)
    encoding:       str       = "utf-8"
    bad_lines_path: str       = ""    # path to bad lines CSV report if any
    # Small preview df — populated only if preview requested (optional)
    preview_df:     object    = field(default=None, repr=False)


@dataclass
class StagingResult:
    ok:          bool
    message:     str
    rows_loaded: int = 0
    table_name:  str = ""


# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

_SAFE_RE = re.compile(r"[^a-zA-Z0-9_]")

# The staging directory is config.scratch_dir("bulk") now, resolved per call so
# DW_SCRATCH can redirect it. It was r"C:\Bulk" -- which is also the VAULT root,
# so throwaway staging CSVs shared a directory with curated documents, and a
# distribution had six separate literals to find before it could move any of
# them. Still falls back to system temp if the probe below fails.


def _get_bulk_path(staging_name: str) -> str:
    """
    Return a writable path for the BULK INSERT staging CSV.

    Priority:
      1. config.scratch_dir("bulk") — the one configurable root. The SQL
         Server service account must be able to READ it, because BULK INSERT
         opens the file server-side, not us. Checked 4 Sep rather than
         assumed: NT Service\\MSSQL$SQLEXPRESS reads the LOCALAPPDATA default
         here. DW_SCRATCH is the override where it cannot -- a hardened
         service account, or a server on another host.
      2. System temp dir — always writable by US, and not necessarily
         readable by the SERVER, which is why it is the fallback and not the
         default.
    """
    filename = f"stage_{staging_name}.csv"
    # Try the configured scratch root first
    try:
        from dataview.core.config import scratch_dir
        candidate = os.path.join(scratch_dir("bulk"), filename)
        # Quick write-access probe
        with open(candidate, "w") as _probe:
            pass
        os.unlink(candidate)
        return candidate
    except OSError:
        # Fall back to system temp — note this may not be accessible by the
        # SQL Server service account; user may need to grant access or move
        # the file to a shared location.
        tmp_dir = tempfile.gettempdir()
        return os.path.join(tmp_dir, filename)


def _staging_name(filename: str) -> str:
    stem = Path(filename).stem
    safe = _SAFE_RE.sub("_", stem).lower().strip("_")
    return f"stg_{safe[:50]}"


def _sanitize_col(name: str) -> str:
    return _SAFE_RE.sub("_", str(name).strip()).strip("_") or "col"


def _dedupe_cols(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            result.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 1
            result.append(c)
    return result


def _detect_delimiter(sample: str) -> str:
    # Count candidates per first line — most consistent delimiter wins
    first_line = sample.split('\n')[0] if '\n' in sample else sample[:500]
    counts = {d: first_line.count(d) for d in ('|', '\t', ',', ';')}
    # Prefer comma and tab (most common CSV formats) over pipe/semicolon
    for d in (',', '\t', '|', ';'):
        if counts[d] > 0:
            return d
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
        return dialect.delimiter
    except Exception:
        return ","


def _detect_encoding(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            data.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"


# ═══════════════════════════════════════════════════════════════════════
# INGEST  —  header only, save to disk
# ═══════════════════════════════════════════════════════════════════════

def ingest_file(
    data:       bytes,
    filename:   str,
    delimiter:  str = "",    # "" = auto-detect
    quotechar:  str = '"',
    encoding:   str = "",    # "" = auto-detect
    sheet_name: str | None = None,  # Excel: sheet name (None = active)
    header:     int = 0,            # Excel: 0-based header row index
    skiprows:   int = 0,            # Excel: rows to skip after header
) -> IngestResult:
    """
    Read only what's needed from Python's side:
      - Detect encoding and delimiter
      - Read the header row
      - Sanitize column names
      - Save the full file to a temp CSV for BULK INSERT

    For Excel: convert to CSV via openpyxl (header + all rows), then same path.
    Row count is -1 until after load_to_staging() completes.
    """
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return _ingest_excel(data, filename,
                             sheet_name=sheet_name,
                             header_row=header,
                             skip_rows=skiprows)
    return _ingest_csv(data, filename, delimiter=delimiter,
                       quotechar=quotechar, encoding=encoding)


def _ingest_csv(
    data:      bytes,
    filename:  str,
    delimiter: str = "",
    quotechar: str = '"',
    encoding:  str = "",
) -> IngestResult:
    warnings: list[str] = []

    # ── Encoding ─────────────────────────────────────────────────────
    used_enc = encoding if encoding else _detect_encoding(data)
    try:
        text = data.decode(used_enc)
    except (UnicodeDecodeError, LookupError):
        text = data.decode("latin-1")
        used_enc = "latin-1"
        warnings.append("Encoding fallback to Latin-1.")

    # Strip BOM
    text = text.lstrip("\ufeff")

    if used_enc not in ("utf-8", "utf-8-sig"):
        warnings.append(f"File encoding: {used_enc} — will be converted to UTF-8.")

    # ── Delimiter ────────────────────────────────────────────────────
    detected = delimiter if delimiter else _detect_delimiter(text[:4096])

    # ── Header only ──────────────────────────────────────────────────
    reader = csv.reader(io.StringIO(text), delimiter=detected, quotechar=quotechar)
    try:
        raw_headers = next(reader)
    except StopIteration:
        return IngestResult(ok=False, message="File appears to be empty.",
                            filename=filename)

    raw_headers = [h.strip() for h in raw_headers]
    columns     = _dedupe_cols([_sanitize_col(h) for h in raw_headers])

    # Remove empty/unnamed trailing columns from trailing delimiters
    while columns and columns[-1] in ('', 'col'):
        columns.pop()

    if any(c != _sanitize_col(h) for c, h in zip(columns, raw_headers)):
        warnings.append("Some column names were sanitized.")

    # ── Save to temp CSV (pipe-delimited) ───────────────────────────
    tmp_path = _get_bulk_path(_staging_name(filename))
    _bad_rows = []

    with open(tmp_path, mode="w", encoding="utf-8", newline="") as tmp:
        tmp.write("|".join(columns) + "\r\n")

        reader2 = csv.reader(io.StringIO(text), delimiter=detected, quotechar=quotechar)
        next(reader2)  # skip header
        for _lineno, row in enumerate(reader2, 2):
            _raw_len = len(row)
            if _raw_len != len(columns) and _raw_len > 0:
                _bad_rows.append({
                    "line": _lineno,
                    "reason": f"Expected {len(columns)} fields, got {_raw_len}",
                    "data": detected.join(str(v) for v in row),
                })
            row = row[:len(columns)]
            row += [""] * (len(columns) - len(row))
            row = [str(v).replace('|', ' ').replace('\r', '').replace('\n', ' ').replace('"', '').replace('\\', '')
                   for v in row]
            tmp.write("|".join(row) + "\r\n")

    # Write bad rows report if any
    _bad_path = None
    if _bad_rows:
        import csv as _csv2
        _bad_path = tmp_path.replace(".csv", "_bad_lines.csv")
        with open(_bad_path, mode="w", encoding="utf-8", newline="") as _bf:
            _bw = _csv2.writer(_bf, quoting=_csv2.QUOTE_ALL)
            _bw.writerow(["LINE_NUMBER", "REASON", "DATA"])
            for _b in _bad_rows:
                _bw.writerow([_b["line"], _b["reason"], _b["data"]])
        warnings.append(f"{len(_bad_rows)} row(s) had wrong field count — see {_bad_path}")

    return IngestResult(
        ok=True,
        message=f"{len(columns)} columns detected in {filename}. "
                f"File saved — row count available after load.",
        filename=filename,
        staging_name=_staging_name(filename),
        row_count=-1,   # filled in by load_to_staging
        col_count=len(columns),
        columns=columns,
        warnings=warnings,
        csv_path=tmp_path,
        bad_lines_path=_bad_path or "",
        delimiter=detected,
        encoding="utf-8",
    )


def _ingest_excel(data: bytes, filename: str,
                   sheet_name: str | None = None,
                   header_row: int = 0,
                   skip_rows:  int = 0) -> IngestResult:
    try:
        import openpyxl
    except ImportError:
        return IngestResult(
            ok=False,
            message="openpyxl is required for Excel files: pip install openpyxl",
            filename=filename,
        )

    warnings = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            warnings.append(f"Sheet: '{sheet_name}'")
        else:
            ws = wb.active
            if sheet_name:
                warnings.append(
                    f"Sheet '{sheet_name}' not found — using active sheet '{ws.title}'.")
            else:
                warnings.append(f"Sheet: '{ws.title}'")
        rows = ws.iter_rows(values_only=True)
    except Exception as exc:
        return IngestResult(ok=False, message=f"Could not read Excel: {exc}",
                            filename=filename)

    # Skip to header row (0-based index)
    for _ in range(header_row):
        try:
            next(rows)
        except StopIteration:
            return IngestResult(ok=False,
                                message=f"Header row {header_row+1} exceeds sheet length.",
                                filename=filename)

    try:
        raw_headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return IngestResult(ok=False, message="Excel sheet appears empty.",
                            filename=filename)

    # Skip extra rows after header
    for _ in range(skip_rows):
        try:
            next(rows)
        except StopIteration:
            break

    if header_row > 0 or skip_rows > 0:
        warnings.append(
            f"Header row: {header_row+1}, skipped {skip_rows} row(s) after header.")

    columns = _dedupe_cols([_sanitize_col(h) for h in raw_headers])
    tmp_path = _get_bulk_path(_staging_name(filename))

    with open(tmp_path, mode="w", encoding="utf-8", newline="") as tmp:
        writer = csv.writer(tmp, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(columns)
        for row in rows:
            row = [str(v) if v is not None else "" for v in row]
            row = row[:len(columns)]
            row += [""] * (len(columns) - len(row))
            writer.writerow(row)
    wb.close()

    # Detect delimiter from the written CSV so parse options can be pre-populated
    try:
        with open(tmp_path, encoding="utf-8") as _tf:
            _sample = _tf.read(4096)
        _detected_delim = _detect_delimiter(_sample)
    except Exception:
        _detected_delim = ","

    return IngestResult(
        ok=True,
        message=f"{len(columns)} columns detected in {filename}. "
                f"File converted to CSV — row count available after load.",
        filename=filename,
        staging_name=_staging_name(filename),
        row_count=-1,
        col_count=len(columns),
        columns=columns,
        warnings=warnings,
        csv_path=tmp_path,
        delimiter=_detected_delim,
        encoding="utf-8",
    )


# ═══════════════════════════════════════════════════════════════════════
# PREVIEW  —  read N rows from saved CSV for display only
# ═══════════════════════════════════════════════════════════════════════

def preview_csv(ingest: IngestResult, n: int = 20) -> Optional[object]:
    """
    Read the first n data rows from the saved CSV for display.
    Returns a list of dicts (no pandas needed, but caller may wrap in pd.DataFrame).
    """
    if not ingest.csv_path or not os.path.exists(ingest.csv_path):
        return None
    rows = []
    try:
        with open(ingest.csv_path, encoding="utf-8", newline="") as f:
            delim = ingest.delimiter if ingest.delimiter else ','
            reader = csv.DictReader(f, delimiter=delim)
            for i, row in enumerate(reader):
                if i >= n:
                    break
                # Strip None/empty keys from trailing empty columns
                rows.append({k: v for k, v in row.items()
                             if k is not None and str(k).strip() != ''})
    except Exception:
        return None
    return rows


# ═══════════════════════════════════════════════════════════════════════
# LOAD TO STAGING  —  server-side BULK INSERT (SQL Server) or executemany (Oracle)
# ═══════════════════════════════════════════════════════════════════════

def load_to_staging(
    engine,
    ingest: IngestResult,
    schema: str = "stg",
    table_name: str = "",
) -> StagingResult:
    """
    Load source file into a staging table.

    SQL Server: uses BULK INSERT for maximum speed.
    Oracle:     reads CSV with pandas, inserts via executemany.

    Steps:
      1. Ensure schema exists (SQL Server only — Oracle uses connected user schema)
      2. DROP + CREATE staging table (all VARCHAR)
      3. Load data
      4. Read back row count
      5. Clean up temp file on success
    """
    from sqlalchemy import text
    from dataview.core.db import _detect_dialect

    table   = table_name if table_name else (ingest.staging_name or "raw_data")
    columns = ingest.columns
    dialect = _detect_dialect(engine)

    if dialect == "oracle":
        return _load_to_staging_oracle(engine, ingest, schema, table, columns)
    elif dialect == "snowflake":
        return _load_to_staging_snowflake(engine, ingest, schema, table, columns)
    else:
        return _load_to_staging_sqlserver(engine, ingest, schema, table, columns)


def _find_bcp() -> str:
    """Locate the newest bcp.exe on the system."""
    import glob
    candidates = glob.glob(
        r"C:\Program Files\Microsoft SQL Server\Client SDK\ODBC\*\Tools\Binn\bcp.exe"
    )
    if candidates:
        return sorted(candidates)[-1]  # newest version
    # Fallback: try PATH
    import shutil
    found = shutil.which("bcp")
    if found:
        return found
    raise FileNotFoundError(
        "bcp.exe not found. Install SQL Server command-line tools or ensure "
        "C:\\Program Files\\Microsoft SQL Server\\Client SDK\\ODBC\\...\\Tools\\Binn is present."
    )


def _load_to_staging_sqlserver(engine, ingest, schema, table, columns) -> StagingResult:
    """SQL Server staging via BCP (bulk copy program)."""
    import subprocess
    from sqlalchemy import text

    csv_path = ingest.csv_path
    col_defs = ",\n    ".join(f"[{c}] NVARCHAR(4000) NULL" for c in columns)

    sql_schema  = f"IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name = '{schema}') EXEC('CREATE SCHEMA [{schema}]')"
    sql_drop    = f"IF OBJECT_ID('[{schema}].[{table}]', 'U') IS NOT NULL DROP TABLE [{schema}].[{table}]"
    sql_create  = f"CREATE TABLE [{schema}].[{table}] (\n    {col_defs}\n)"
    sql_add_col = f"ALTER TABLE [{schema}].[{table}] ADD [_batch_loaded_at] DATETIME2 NULL"
    sql_stamp   = f"UPDATE [{schema}].[{table}] SET [_batch_loaded_at] = GETUTCDATE()"

    # Pre-flight: verify CSV has data rows
    try:
        with open(csv_path, encoding="utf-8") as _pf:
            _line_count = sum(1 for _ in _pf) - 1
        if _line_count <= 0:
            return StagingResult(ok=False,
                                 message="Staging CSV is empty — no data rows found.",
                                 table_name=f"{schema}.{table}")
    except Exception:
        pass

    # Find BCP
    try:
        bcp_exe = _find_bcp()
    except FileNotFoundError as e:
        return StagingResult(ok=False, message=str(e), table_name=f"{schema}.{table}")

    # Get server and database from the live connection
    try:
        with engine.connect() as con:
            row = con.execute(text("SELECT @@SERVERNAME, DB_NAME()")).fetchone()
            server, database = str(row[0]), str(row[1])
    except Exception as exc:
        return StagingResult(ok=False,
                             message=f"Could not determine server/database: {exc}",
                             table_name=f"{schema}.{table}")

    # Create schema + table
    try:
        with engine.begin() as con:
            con.execute(text(sql_schema))
            con.execute(text(sql_drop))
            con.execute(text(sql_create))
    except Exception as exc:
        return StagingResult(ok=False,
                             message=f"Table creation failed: {exc}",
                             table_name=f"{schema}.{table}")

    # BCP load
    # Temp file is ALWAYS pipe-delimited (see _ingest_csv)
    _delim = "|"
    _err_path = csv_path.replace(".csv", "_errors.txt")
    bcp_cmd = [
        bcp_exe,
        f"{database}.{schema}.{table}",
        "in", csv_path,
        "-S", server,
        "-T",                     # trusted connection
        "-c",                     # character mode
        "-C", "65001",            # UTF-8
        "-t", _delim,             # field terminator
        "-F", "2",                # skip header row
        "-e", _err_path,          # error file
    ]

    try:
        result = subprocess.run(
            bcp_cmd, capture_output=True, text=True, timeout=600
        )
        if result.returncode != 0:
            detail = (result.stdout + "\n" + result.stderr).strip()
            return StagingResult(
                ok=False,
                message=f"BCP failed (exit {result.returncode}): {detail}",
                table_name=f"{schema}.{table}",
            )
    except subprocess.TimeoutExpired:
        return StagingResult(ok=False, message="BCP timed out after 600s.",
                             table_name=f"{schema}.{table}")
    except Exception as exc:
        return StagingResult(ok=False, message=f"BCP error: {exc}",
                             table_name=f"{schema}.{table}")

    # Add timestamp column
    try:
        with engine.begin() as con:
            con.execute(text(sql_add_col))
            con.execute(text(sql_stamp))
    except Exception:
        pass  # non-critical

    # Count rows
    try:
        with engine.connect() as con:
            n = con.execute(text(
                "SELECT SUM(p.rows) FROM sys.partitions p "
                "JOIN sys.tables t ON t.object_id = p.object_id "
                "JOIN sys.schemas s ON s.schema_id = t.schema_id "
                f"WHERE p.index_id IN (0,1) AND t.name = '{table}' AND s.name = '{schema}'"
            )).scalar() or 0
    except Exception:
        n = 0

    ingest.row_count = n

    # Log BCP output for diagnostics
    if n == 0:
        _bcp_out = (result.stdout + "\n" + result.stderr).strip()
        return StagingResult(
            ok=False,
            message=f"BCP completed but 0 rows loaded. BCP output:\n{_bcp_out}",
            table_name=f"{schema}.{table}",
        )

    return StagingResult(
        ok=True,
        message=f"Loaded {n:,} rows × {len(columns)} cols into [{schema}].[{table}] via BCP",
        rows_loaded=n,
        table_name=f"{schema}.{table}",
    )



def _load_to_staging_snowflake(engine, ingest, schema, table, columns) -> StagingResult:
    """
    Snowflake staging via executemany.
    Uses the connected database/schema — no separate stg schema needed.
    """
    from sqlalchemy import text
    import pandas as _pd, io as _io

    # Use the connected schema (DEMO) — Snowflake doesn't need a separate stg schema
    try:
        with engine.connect() as _sc:
            sf_schema = _sc.execute(text("SELECT CURRENT_SCHEMA()")).scalar() or "DEMO"
    except Exception:
        sf_schema = "DEMO"  # fallback — never use stg schema for Snowflake

    col_defs   = ",\n    ".join(f'"{c.upper()}" VARCHAR(4000)' for c in columns)
    tbl_full   = f'"{sf_schema}"."{table.upper()}"'

    sql_drop   = f"DROP TABLE IF EXISTS {tbl_full}"
    sql_create = f"CREATE TABLE {tbl_full} (\n    {col_defs}\n)"
    sql_add_col= f"ALTER TABLE {tbl_full} ADD \"_BATCH_LOADED_AT\" TIMESTAMP_NTZ"
    sql_stamp  = f"UPDATE {tbl_full} SET \"_BATCH_LOADED_AT\" = CURRENT_TIMESTAMP()::TIMESTAMP_NTZ"

    try:
        with engine.begin() as con:
            con.execute(text(sql_drop))
            con.execute(text(sql_create))

        # Load data via executemany (Snowflake has no server-side BULK INSERT from local file)
        # Find the pipe-delimited temp CSV written by _ingest_csv
        import tempfile as _tf, glob as _glob
        _csv_to_read = ingest.csv_path
        # If path points to C:\Bulk which failed, check tempdir
        if not _csv_to_read or not os.path.exists(_csv_to_read):
            _stem = f"stage_{ingest.staging_name}.csv"
            _tmp_candidate = os.path.join(_tf.gettempdir(), _stem)
            if os.path.exists(_tmp_candidate):
                _csv_to_read = _tmp_candidate
            else:
                return StagingResult(ok=False,
                    message=f"Snowflake staging: temp CSV not found at {ingest.csv_path}",
                    table_name=f"DEMO.{table.upper()}")
        df = _pd.read_csv(
            _csv_to_read,
            dtype=str,
            keep_default_na=False,
            encoding="utf-8",
            sep="|",
        )
        df.columns = [c.upper() for c in df.columns]

        # Use write_pandas for fast bulk load via Snowflake internal stage
        try:
            from snowflake.connector.pandas_tools import write_pandas as _wp
            with engine.begin() as con:
                raw_con = con.connection.dbapi_connection
                success, nchunks, nrows, _ = _wp(
                    raw_con, df,
                    table_name=table.upper(),
                    schema=sf_schema,
                    database=raw_con.database,
                    quote_identifiers=True,
                    auto_create_table=False,
                    overwrite=False,
                )
            if not success:
                raise RuntimeError(f"write_pandas reported failure after {nrows} rows")
        except ImportError:
            # Fallback to executemany if write_pandas not available
            col_names  = ", ".join(f'"{c}"' for c in df.columns)
            holders    = ", ".join(["%s"] * len(df.columns))
            sql_insert = f"INSERT INTO {tbl_full} ({col_names}) VALUES ({holders})"
            rows_list = [tuple(None if (_pd.isna(v) or v == "") else v
                               for v in row)
                         for row in df.itertuples(index=False, name=None)]
            BATCH = 5000
            with engine.begin() as con:
                raw_con = con.connection.dbapi_connection
                cur = raw_con.cursor()
                for i in range(0, len(rows_list), BATCH):
                    cur.executemany(sql_insert, rows_list[i:i+BATCH])
                cur.close()

        with engine.begin() as con:
            con.execute(text(sql_add_col))
            con.execute(text(sql_stamp))

        n = len(df)
        ingest.row_count = n

        try:
            if os.path.exists(ingest.csv_path):
                os.unlink(ingest.csv_path)
        except Exception:
            pass

        return StagingResult(
            ok=True,
            message=f"Loaded {n:,} rows × {len(columns)} cols into {tbl_full}",
            rows_loaded=n,
            table_name=f"{sf_schema}.{table.upper()}",
        )
    except Exception as exc:
        return StagingResult(
            ok=False,
            message=f"Snowflake staging failed: {exc}",
            table_name=f"{sf_schema}.{table.upper()}",
        )


def _load_to_staging_oracle(engine, ingest, schema, table, columns) -> StagingResult:
    """
    Oracle staging via pandas executemany.
    Oracle schemas = users, so no schema creation needed.
    The connected user owns all objects created in this session.
    """
    from sqlalchemy import text
    import pandas as _pd, io as _io

    # Oracle: use connected user as schema (ignore the 'stg' schema arg)
    try:
        with engine.connect() as _sc:
            ora_schema = _sc.execute(text(
                "SELECT SYS_CONTEXT('USERENV','CURRENT_SCHEMA') FROM dual"
            )).scalar() or schema.upper()
    except Exception:
        ora_schema = schema.upper()

    # Quoted identifiers for Oracle
    q_schema = f'"{ora_schema}"'
    q_table  = f'"{table.upper()}"'
    full     = f"{q_schema}.{q_table}"

    # VARCHAR2(4000) for all staging columns + audit timestamp
    # All columns VARCHAR2 including audit stamp — avoids ORA-01843 date parsing
    col_defs = ",\n    ".join(f'"{c.upper()}" VARCHAR2(4000)' for c in columns)

    sql_drop = (
        f"BEGIN\n"
        f"  EXECUTE IMMEDIATE 'DROP TABLE {full}'\n;"
        f"EXCEPTION WHEN OTHERS THEN\n"
        f"  IF SQLCODE != -942 THEN RAISE; END IF;\n"
        f"END;"
    )
    sql_create = (
        f"CREATE TABLE {full} (\n"
        f"    {col_defs},\n"
        f"    \"_batch_loaded_at\" VARCHAR2(30)\n"
        f")"
    )

    try:
        with engine.begin() as con:
            con.execute(text(sql_drop))
            con.execute(text(sql_create))

        # Read CSV with pandas and insert via executemany
        csv_path = ingest.csv_path
        _df = _pd.read_csv(
            csv_path, dtype=str, sep="|",
            encoding="utf-8-sig", keep_default_na=False,
        )
        _df.columns = [c.upper() for c in _df.columns]
        # Add audit column
        import datetime as _dt
        _df["_batch_loaded_at"] = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

        # Build INSERT with Oracle bind variables (:1, :2, ...)
        all_cols   = list(_df.columns)
        placeholders = ", ".join(f":{i+1}" for i in range(len(all_cols)))
        col_list     = ", ".join(f'"{c}"' for c in all_cols)
        sql_insert   = f"INSERT INTO {full} ({col_list}) VALUES ({placeholders})"

        # Force every value to str — prevents Oracle from inferring date/numeric
        # types from values that look like dates (ORA-01843)
        rows = [
            tuple(str(v) if v is not None and str(v) != "" else None
                  for v in r)
            for r in _df.itertuples(index=False, name=None)
        ]

        BATCH = 1000
        with engine.begin() as con:
            raw = con.connection
            cur = raw.cursor()
            import oracledb as _odb
            cur.setinputsizes(*[_odb.DB_TYPE_VARCHAR] * len(all_cols))
            cur.executemany(sql_insert, rows, batcherrors=True)
            _errs = cur.getbatcherrors()
            if _errs:
                raise Exception(
                    f"{len(_errs)} batch error(s); first: {_errs[0].message}"
                )
            cur.close()

        n = len(_df)
        ingest.row_count = n
        try:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
        except Exception:
            pass

        return StagingResult(
            ok=True,
            message=f"Loaded {n:,} rows × {len(columns)} cols into {full}",
            rows_loaded=n,
            table_name=f"{ora_schema}.{table.upper()}",
        )
    except Exception as exc:
        return StagingResult(
            ok=False,
            message=f"Oracle staging failed: {exc}",
            table_name=f"{ora_schema}.{table.upper()}",
        )


def load_to_staging_demo(ingest: IngestResult, schema: str = "stg") -> StagingResult:
    return StagingResult(
        ok=True,
        message=f"[Demo] Would BULK INSERT into [{schema}].[raw_data]",
        rows_loaded=0,
        table_name=f"{schema}.raw_data",
    )


# ═══════════════════════════════════════════════════════════════════════
# DIRECT PATH INGEST  —  zero Python I/O for large files
# ═══════════════════════════════════════════════════════════════════════

def ingest_from_path(
    file_path:  str,
    delimiter:  str = "",   # "" = auto-detect from first line
    encoding:   str = "65001",  # SQL Server codepage — 65001=UTF-8, 1252=Windows
    quotechar:  str = '"',
    skip_cols:  list[str] | None = None,   # column names to drop before staging
) -> IngestResult:
    """
    Ingest a large file directly from disk — reads ONLY the header row in
    Python, then hands the full file path to BULK INSERT unchanged.

    No file copy. No row iteration. Works for files of any size.

    The file must be on a path readable by the SQL Server service account.
    C:\\Bulk\\ is the recommended location.

    Args:
        file_path:  Absolute path to the source file on disk
        delimiter:  Field separator — auto-detected from first line if ""
        encoding:   SQL Server CODEPAGE value (65001=UTF-8, 1252=Windows-1252)
        quotechar:  Quote character (default ")
        skip_cols:  Column names to exclude (e.g. trailing garbage columns)
    """
    p = Path(file_path)
    if not p.exists():
        return IngestResult(
            ok=False,
            message=f"File not found: {file_path}",
            filename=p.name,
        )

    warnings: list[str] = []

    # ── Read header only ─────────────────────────────────────────────
    try:
        with open(p, encoding="utf-8-sig", errors="replace", newline="") as f:
            first_line = f.readline()
    except Exception as exc:
        return IngestResult(ok=False, message=f"Could not read header: {exc}",
                            filename=p.name)

    # Auto-detect delimiter from header line
    if not delimiter:
        delimiter = _detect_delimiter(first_line)

    raw_headers = [h.strip() for h in first_line.split(delimiter)]
    columns     = _dedupe_cols([_sanitize_col(h) for h in raw_headers])

    # Remove empty/unnamed trailing columns
    while columns and columns[-1] in ('', 'col'):
        columns.pop()
        raw_headers.pop()

    # Drop explicitly skipped columns
    if skip_cols:
        skip_upper = {s.upper() for s in skip_cols}
        keep_idx   = [i for i, c in enumerate(columns)
                      if c.upper() not in skip_upper]
        columns    = [columns[i] for i in keep_idx]
        if len(keep_idx) < len(raw_headers):
            dropped = len(raw_headers) - len(keep_idx)
            warnings.append(f"Dropped {dropped} excluded column(s).")

    # ── Count rows (fast line count — no pandas) ─────────────────────
    try:
        with open(p, encoding="utf-8-sig", errors="replace") as f:
            row_count = sum(1 for _ in f) - 1   # subtract header
        warnings.append(f"Row count: {row_count:,}")
    except Exception:
        row_count = -1

    return IngestResult(
        ok=True,
        message=f"{len(columns)} columns detected. "
                f"Ready for direct BULK INSERT from {file_path}",
        filename=p.name,
        staging_name=_staging_name(p.name),
        row_count=row_count,
        col_count=len(columns),
        columns=columns,
        warnings=warnings,
        csv_path=str(p),        # points at original file — no copy needed
        delimiter=delimiter,
        encoding="utf-8",
    )


def load_from_path(
    engine,
    file_path:  str,
    schema:     str  = "stg",
    table_name: str  = "",
    delimiter:  str  = "",
    codepage:   str  = "65001",
    skip_cols:  list[str] | None = None,
) -> tuple[IngestResult, StagingResult]:
    """
    One-call convenience wrapper:
      1. ingest_from_path()  — header only, zero file copy
      2. load_to_staging()   — BULK INSERT direct from source path

    Returns (IngestResult, StagingResult).

    Example — load 400K KGS wells:
        ir, sr = load_from_path(
            engine,
            r"C:\\Bulk\\kgs_wells_400k.csv",
            schema="stg",
            delimiter="|",
            skip_cols=["STAUS_TYPE"],   # trailing garbage column
        )
        print(sr.message)  # "Loaded 400,000 rows × 40 cols into [stg].[stg_kgs_wells_400k]"
    """
    ir = ingest_from_path(
        file_path=file_path,
        delimiter=delimiter,
        encoding=codepage,
        skip_cols=skip_cols,
    )
    if not ir.ok:
        return ir, StagingResult(ok=False, message=ir.message)

    table = table_name if table_name else ir.staging_name
    sr    = load_to_staging(engine, ir, schema=schema, table_name=table)
    return ir, sr



def preview_staging_table(
    engine,
    table:  str = "raw_data",
    schema: str = "stg",
    n:      int = 20,
) -> tuple[bool, str, Optional[object]]:
    """
    Read the first n rows from the staging table for display.
    Returns (ok, message, list_of_dicts).
    Dialect-aware: supports SQL Server and Oracle.
    """
    from sqlalchemy import text
    from dataview.core.db import _detect_dialect

    dialect = _detect_dialect(engine)

    # Oracle: resolve actual schema from connected user
    if dialect == "oracle":
        try:
            with engine.connect() as _sc:
                schema = _sc.execute(text(
                    "SELECT SYS_CONTEXT('USERENV','CURRENT_SCHEMA') FROM dual"
                )).scalar() or schema.upper()
        except Exception:
            schema = schema.upper()
        table_upper = table.upper()
        q = lambda name: f'"{name}"'
        exists_sql  = ("SELECT 1 FROM all_tables "
                       "WHERE owner = :sch AND table_name = :tbl")
        exists_params = {"sch": schema, "tbl": table_upper}
        rows_sql    = f"SELECT * FROM {q(schema)}.{q(table_upper)} FETCH FIRST {n} ROWS ONLY"
        count_sql   = f"SELECT COUNT(*) FROM {q(schema)}.{q(table_upper)}"
        not_exists_msg = f'"{schema}"."{table_upper}" does not exist yet.'
        label       = f'"{schema}"."{table_upper}"'
    elif dialect == "snowflake":
        try:
            with engine.connect() as _sc:
                schema = _sc.execute(text("SELECT CURRENT_SCHEMA()")).scalar() or schema
        except Exception:
            schema = schema.upper()
        table_upper = table.upper()
        exists_sql  = ("SELECT 1 FROM INFORMATION_SCHEMA.TABLES "
                       "WHERE TABLE_SCHEMA = :sch AND TABLE_NAME = :tbl")
        exists_params = {"sch": schema.upper(), "tbl": table_upper}
        rows_sql    = f'SELECT * FROM "{schema}"."{table_upper}" LIMIT {n}'
        count_sql   = f'SELECT COUNT(*) FROM "{schema}"."{table_upper}"'
        not_exists_msg = f'"{schema}"."{table_upper}" does not exist yet.'
        label       = f'"{schema}"."{table_upper}"'
    else:
        table_upper = table
        exists_sql  = ("SELECT 1 FROM INFORMATION_SCHEMA.TABLES "
                       "WHERE TABLE_SCHEMA = :sch AND TABLE_NAME = :tbl")
        exists_params = {"sch": schema, "tbl": table}
        rows_sql    = f"SELECT TOP {n} * FROM [{schema}].[{table}]"
        count_sql   = f"SELECT COUNT(*) FROM [{schema}].[{table}]"
        not_exists_msg = f"[{schema}].[{table}] does not exist yet."
        label       = f"[{schema}].[{table}]"

    try:
        with engine.connect() as con:
            exists = con.execute(text(exists_sql), exists_params).fetchone()
            if not exists:
                return False, not_exists_msg, None
            rows  = con.execute(text(rows_sql)).mappings().all()
            total = con.execute(text(count_sql)).scalar()
        return True, f"{total:,} rows in {label}", [dict(r) for r in rows]

    except Exception as exc:
        return False, f"Could not read {label}: {exc}", None


# ═══════════════════════════════════════════════════════════════════════
# SELF-TEST
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("staging.py  --  self-test")
    print("=" * 60)

    # [1] CSV — basic
    csv_data = b"Well Name,API Number,Status,Depth\nWell A,4200100001,ACTIVE,5000\nWell B,4200100002,INACTIVE,7500\n"
    r = ingest_file(csv_data, "well_header.csv")
    assert r.ok, r.message
    assert r.col_count == 4
    assert r.columns == ["Well_Name", "API_Number", "Status", "Depth"]
    assert r.staging_name == "stg_well_header"
    assert os.path.exists(r.csv_path)
    print(f"  [1] CSV basic: {r.col_count} cols, staging={r.staging_name}")

    # [2] Preview
    rows = preview_csv(r, n=2)
    assert rows is not None and len(rows) == 2
    assert rows[0]["Well_Name"] == "Well A"
    print(f"  [2] Preview: {len(rows)} rows, first={rows[0]['Well_Name']}")
    os.unlink(r.csv_path)

    # [3] TSV auto-detect
    tsv_data = "UWI\tWell Class\tFormation\nABC123\tWILDCAT\tNIOBRARA\n".encode()
    r2 = ingest_file(tsv_data, "wells.tsv")
    assert r2.ok and r2.columns == ["UWI", "Well_Class", "Formation"]
    os.unlink(r2.csv_path)
    print(f"  [3] TSV auto-detect: delimiter detected, cols={r2.columns}")

    # [4] BOM handling
    bom_data = b"\xef\xbb\xbfCol1,Col2\nA,B\n"
    r3 = ingest_file(bom_data, "bom_test.csv")
    assert r3.ok and r3.columns == ["Col1", "Col2"], f"Got {r3.columns}"
    os.unlink(r3.csv_path)
    print(f"  [4] BOM stripped: cols={r3.columns}")

    # [5] Empty file
    r4 = ingest_file(b"", "empty.csv")
    assert not r4.ok
    print(f"  [5] Empty file: {r4.message}")

    # [6] Demo mode
    r5 = ingest_file(csv_data, "well_header.csv")
    sr = load_to_staging_demo(r5)
    assert sr.ok
    os.unlink(r5.csv_path)
    print(f"  [6] Demo: {sr.message}")

    # [7] Staging name
    assert _staging_name("My Well Header (2024).csv") == "stg_my_well_header__2024"
    assert _staging_name("wells.tsv") == "stg_wells"
    print(f"  [7] Staging names: OK")

    print("\nAll tests passed")
