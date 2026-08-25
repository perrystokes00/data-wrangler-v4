r"""Build the WHOLE Teapot Field model into one self-contained folder tree.

WHY A SEPARATE TREE. The synthetic data was scattered through the REAL NPR-3
dataset -- SEG-Y under Teapot_Dome\DataSets\Seismic, documents under
Teapot_Dome\DataSets\Synthetic_Field -- so anyone pointing a loader at the
Teapot folder got both, and the only thing keeping them apart was a UWI block
and a date range. Real production runs 1922-2005 on 49025063xxx; the model runs
2014-2024 on 49025900xxx. That is a distinction you have to KNOW to see, which
makes it the wrong kind of separation.

One root, nothing of the original in it, and a README that says so.

    python tools/make_teapot_model.py                 # plan only
    python tools/make_teapot_model.py --apply
    python tools/make_teapot_model.py --apply --skip-seismic   # 324 MB skipped

The database side (horizons, geography) is loaded by its own tools, but this
also EXPORTS both to CSV so the folder is a complete dataset rather than half
of one -- a tree that cannot rebuild what it describes is a backup with a hole
in it.
"""
import argparse
import os
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except Exception:
        pass

# WHICH EXTENSIONS THE FILE CATALOG WILL NOT CRAWL. Imported rather than
# retyped: if the app ever changes what it treats as tabular, this folder
# split has to change with it or the corpus quietly misfiles again.
try:
    from dataview.file_catalog.promotion_lineage import TABULAR_EXTS
except Exception:
    TABULAR_EXTS = {".csv", ".tsv", ".xlsx", ".xls", ".xlsm", ".xlsb"}

DEFAULT_ROOT = (r"C:\Users\perry\OneDrive\Documents\PPDM\claude_use_ai"
                r"\data_wrangler\training\Teapot_Field_Model")

README = """Teapot Field Model - synthetic dataset
=====================================

EVERYTHING HERE IS GENERATED. It is not NPR-3 field data and must not be
mixed with it. The real Teapot Dome data lives under training\\Teapot_Dome
and is untouched by the tools that build this tree.

Telling them apart:

    this model      UWI 49025900xxxxxx    production 2014-2024
    real NPR-3      UWI 490250xxxxxxxx    production 1922-2005

What is here
------------

  wells/                  {n_docs} documents describing {n_wells} wells
    sample_pdfs/          scout tickets, end-of-well, casing and cement,
                          well tests, petrophysics, directional surveys,
                          monthly production reports
    sample_office/        completion reports (.docx)
    las_files/            wireline logs in LAS 1.2, 2.0 and 3.0
    teapot_field_wells.csv    the well list the documents describe
    MANIFEST.csv          ground truth: which UWI each document belongs to,
                          including the deliberately unreadable cases

  seismic/
    2d_segy/              {n_segy} 2D lines, SEG-Y rev 1, IBM floats
    horizons/             four interpreted horizons, as grids and contours

  tabular/                production and core analysis workbooks (.xlsx).
                          THESE DO NOT GO THROUGH THE FILE CATALOG -- .xlsx
                          is in TABULAR_EXTS and a scan skips it silently.
                          Point the Data Assistant here instead.

  geography/              field outline, leases, reserve boundary, pipelines

How it holds together
---------------------

One structural model underlies all of it. The dome that generated the
seismic also generated the horizons, the formation tops in every document,
and the production profile of every well - so a horizon pick lands on its
reflector, a top in a scout ticket matches the top in that well's LAS, and
the wells nearest the crest are the ones that produce. Nothing here was
drawn independently and made to agree afterwards.

Loading it
----------

  1. File Catalog: scan  wells/  then run the pipeline and Promote.
     That folder holds only what the catalog can read - PDF, LAS, DOCX.
  1b. Data Assistant: point it at  tabular/  for the workbooks. Kept in
     its own folder because the File Catalog skips .xlsx WITHOUT SAYING
     SO, and the Data Assistant cannot use the .docx beside them.
  2. File Catalog: scan  seismic/2d_segy/  for the seismic lines.
  3. Horizons and geography are loaded straight into dataview.* by
     tools/make_teapot_horizons.py and tools/make_teapot_geography.py.
     The CSVs here are an export of the same content.

Rebuilding
----------

  python tools/make_teapot_model.py --apply

Deterministic: the same seeds produce the same files, byte for byte.
"""


def _export_horizons(root, log):
    """Horizon grids and contours to CSV, so the tree can rebuild them."""
    import csv
    from dataview.migration.synth_horizons import teapot_horizons
    d = os.path.join(root, "seismic", "horizons")
    os.makedirs(d, exist_ok=True)
    hz = teapot_horizons()
    meta_p = os.path.join(d, "horizons.csv")
    with open(meta_p, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["horizon_id", "horizon_name", "horizon_type",
                    "strat_unit_name", "seq_no", "pick_domain", "pick_uom",
                    "min_value", "max_value", "display_colour"])
        for meta, _g, _s in hz:
            w.writerow([meta[k] for k in
                        ("horizon_id", "horizon_name", "horizon_type",
                         "strat_unit_name", "seq_no", "pick_domain",
                         "pick_uom", "min_value", "max_value",
                         "display_colour")])
    n_nodes = n_seg = 0
    for meta, (lats, lons, vals), segs in hz:
        gp = os.path.join(d, f"{meta['horizon_id']}_grid.csv")
        with open(gp, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["horizon_id", "row_no", "col_no", "latitude",
                        "longitude", "twt_ms"])
            for i in range(vals.shape[0]):
                for j in range(vals.shape[1]):
                    w.writerow([meta["horizon_id"], i, j,
                                f"{lats[i]:.7f}", f"{lons[j]:.7f}",
                                f"{vals[i, j]:.4f}"])
                    n_nodes += 1
        cp = os.path.join(d, f"{meta['horizon_id']}_contours.csv")
        with open(cp, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["horizon_id", "contour_id", "contour_value", "wkt"])
            for k, (value, pts) in enumerate(segs):
                wkt = "LINESTRING(" + ", ".join(
                    f"{lo:.7f} {la:.7f}" for la, lo in pts) + ")"
                w.writerow([meta["horizon_id"],
                            f"{meta['horizon_id']}_C{k:04d}",
                            f"{value:.4f}", wkt])
                n_seg += 1
    log(f"   horizons  {len(hz)} horizon(s), {n_nodes:,} grid node(s), "
        f"{n_seg} contour(s)")
    return len(hz)


def _export_geography(root, log):
    """Field, leases, boundary and pipelines to CSV with WKT geometry."""
    import csv
    from dataview.migration.synth_field import Surfaces, plan_field
    from dataview.migration.synth_geography import (
        field_outline, gathering_system, lease_parcels, reserve_boundary,
        wkt_line, wkt_polygon)
    d = os.path.join(root, "geography")
    os.makedirs(d, exist_ok=True)
    S = Surfaces()
    wells = plan_field(surfaces=S)
    outline, level = field_outline(S)
    bnd = reserve_boundary()
    leases = lease_parcels(bnd)
    pipes = gathering_system(wells)

    with open(os.path.join(d, "field.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["field_id", "field_name", "field_type", "county",
                    "province_state", "basin_name", "remark", "wkt"])
        w.writerow(["TEAPOT_DOME", "Teapot Dome (NPR-3)", "OIL", "NATRONA",
                    "WY", "Powder River Basin",
                    f"{level:,.0f} ms closing contour of the Tensleep horizon",
                    wkt_polygon(outline)])
    with open(os.path.join(d, "boundary.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["boundary_id", "boundary_name", "boundary_type",
                    "province_state", "wkt"])
        w.writerow(["NPR3_RESERVE", "Naval Petroleum Reserve No. 3",
                    "FEDERAL RESERVE", "WY", wkt_polygon(bnd)])
    with open(os.path.join(d, "leases.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["land_tract_id", "tract_name", "lease_number",
                    "operator_name", "legal_description", "wkt"])
        for i, (nm, legal, ring, owner, _col) in enumerate(leases, start=1):
            w.writerow([f"NPR3_LSE_{i:03d}", nm, f"WYW-{160000 + i}",
                        owner, legal, wkt_polygon(ring)])
    with open(os.path.join(d, "pipelines.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["pipeline_id", "pipeline_name", "commodity", "wkt"])
        for i, (nm, pts) in enumerate(pipes, start=1):
            w.writerow([f"NPR3_PL_{i:03d}", nm,
                        "OIL" if nm.startswith("Flowline") else "CRUDE",
                        wkt_line(pts)])
    log(f"   geography 1 field, {len(leases)} lease(s), 1 boundary, "
        f"{len(pipes)} pipeline(s)")
    return len(leases) + len(pipes) + 2


# section -> (target table, [(display column, database column)], constants)
#
# COLUMN NAMES ARE THE LOADER'S INTERFACE. The first cut used the display
# headers the scout ticket carries -- "Top MD", "Rec %", "SPF" -- because the
# rows are shared with the PDF. The reader wants those; the Data Assistant maps
# by column NAME, so it saw nothing it recognised and guessed: CHECKSHOTS,
# CORE_RUNS, COMPLETION_SUMMARY and PRODUCTION_SUMMARY all landed on
# DV_WELL_STIMULATION, and stg does DROP+CREATE, so four of the five would have
# been silently overwritten by the fifth.
#
# The rows stay shared. Only the headers differ, because the two consumers want
# different things from the same data -- and the file is NAMED for its target
# so the mapping is obvious before anyone opens it.
TABULAR_MAP = {
    "Stratigraphy": ("dv_well_formation_top", [
        ("Formation", "strat_unit_name"), ("Top MD (ft)", "top_depth"),
        ("Base MD (ft)", "base_depth"), ("Gross (ft)", "gross_thickness"),
        ("Lithology", "lithology")], {"depth_ouom": "ft"}),
    "Directional Survey": ("dv_well_dir_srvy_sta", [
        ("MD (ft)", "md"), ("Inc", "incl"), ("Azi", "azim"),
        ("TVD (ft)", "tvd"), ("N/S (ft)", "ns_offset"),
        ("E/W (ft)", "ew_offset"), ("DLS", "dls")], {"depth_ouom": "ft"}),
    "DST": ("dv_well_dst", [
        ("Test Date", "test_date"), ("Type", "test_type"),
        ("Top MD", "top_depth"), ("Base MD", "base_depth"),
        ("Result", "test_result"), ("Max Oil", "max_oil_rate"),
        ("Max Gas", "max_gas_rate"), ("API Grav", "api_gravity")],
        {"depth_ouom": "ft", "rate_ouom": "BBL/D"}),
    "Core Runs": ("dv_well_core", [
        ("#", "core_num"), ("Type", "core_type"), ("Show", "core_show"),
        ("Formation", "strat_unit_name"), ("Top MD", "top_depth"),
        ("Base MD", "base_depth"), ("Length", "core_length"),
        ("Rec %", "recovery_pct"), ("Date", "core_date"),
        ("Photos", "photo_count")], {"depth_ouom": "ft"}),
    "Core Sample": ("dv_well_core_sample", [
        ("Sample", "sample_id"), ("Type", "sample_type"),
        ("Depth", "sample_depth"), ("Lithology", "lithology"),
        ("Show", "hydrocarbon_show"), ("Por %", "porosity_frac"),
        ("Perm", "permeability_air_md"), ("Bulk Den", "bulk_density_g_cc"),
        ("Sw", "water_saturation_frac"), ("So", "oil_saturation_frac")],
        {"depth_ouom": "ft"}),
    "Completion Summary": ("dv_well_completion", [
        ("Completion Date", "completion_date"), ("Type", "completion_type"),
        ("Orientation", "well_orientation"), ("Formation", "strat_unit_name"),
        ("Lateral (ft)", "lateral_length_ft"), ("Stages", "stage_count"),
        ("Fluid (bbl)", "total_fluid_bbl"),
        ("Proppant (lbs)", "total_proppant_lbs"),
        ("Prop Intensity", "proppant_intensity_lbs_ft"),
        ("Fluid System", "frac_fluid_system")], {"depth_ouom": "ft"}),
    "Frac Stages": ("dv_well_stimulation", [
        ("Stage", "stage_num"), ("Top MD", "stage_top_depth"),
        ("Base MD", "stage_base_depth"), ("Clusters", "num_clusters"),
        ("Cluster Sp", "cluster_spacing_ft"),
        ("Fluid (bbl)", "fluid_volume_bbl"),
        ("Proppant (lbs)", "proppant_mass_lbs"), ("ISIP", "isip_psi"),
        ("Avg Treat", "avg_treating_pressure_psi"),
        ("Max Rate", "max_rate_bpm")], {"stim_type": "HYDRAULIC FRACTURE"}),
    "Checkshots": ("dv_well_checkshot", [
        ("Station", "station_id"), ("MD (ft)", "md"), ("TVD (ft)", "tvd"),
        ("TWT (ms)", "twt_ms"), ("OWT (ms)", "owt_ms"),
        ("Avg Vel", "avg_velocity"), ("Int Vel", "interval_velocity")],
        {"depth_ouom": "ft", "time_ouom": "ms", "velocity_ouom": "ft/s",
         "depth_datum": "KB"}),
    "Perforations": ("dv_well_perforation", [
        ("Perf Date", "perf_date"), ("Top MD", "top_depth"),
        ("Base MD", "base_depth"), ("Shots", "shot_count"),
        ("SPF", "shot_density"), ("Gun", "gun_type"),
        ("Phasing", "phasing_deg"), ("Formation", "strat_unit_name"),
        ("Status", "perf_status")],
        {"depth_ouom": "ft", "shot_density_ouom": "SPF",
         "perf_diameter_in": "0.42"}),
}

# A PARENT KEY THE LOADER CAN RESOLVE. dv_well_dir_srvy_sta, _core_sample,
# _stimulation and _perforation all hang off a parent row, and a workbook that
# omits the key leaves the operator to invent one in the mapping UI. Derived
# from the UWI so the same well gets the same id in every workbook AND the same
# id promote would mint from a document.
_PARENT_KEY = {
    "dv_well_dir_srvy_hdr": ("survey_id", "SRVY_"),
    "dv_well_dir_srvy_sta": ("survey_id", "SRVY_"),
    "dv_well_core":         ("core_id", "CORE_"),
    "dv_well_core_sample":  ("core_id", "CORE_"),
    "dv_well_completion":   ("completion_id", "COMP_"),
    "dv_well_stimulation":  ("completion_id", "COMP_"),
    "dv_well_perforation":  ("completion_id", "COMP_"),
    "dv_well_checkshot":    ("checkshot_id", "CKSH_"),
    "dv_well_dst":          ("dst_id", "DST_"),
}
# Row-level keys, unique within the parent.
_ROW_KEY = {"dv_well_dir_srvy_sta": "station_id",
            "dv_well_core_sample": "sample_id",
            "dv_well_stimulation": "stim_id",
            "dv_well_perforation": "perf_id",
            "dv_well_checkshot": "station_id",
            "dv_well_core": "core_id",
            "dv_well_dst": "dst_id"}

# Sections that must sample the SAME wells as their parent. The share still
# applies -- it is applied once per family instead of once per table, so a
# well whose cores arrive as a workbook has its core SAMPLES there too.
_DRAW_GROUP = {"Core Sample":  "Core Runs",
               "Frac Stages":  "Completion Summary",
               "Perforations": "Completion Summary"}

# Provenance on every row. `source` is NOT NULL on four of these targets and
# nullable on the rest, so omitting it made the loader offer a `constant` rule
# with a BLANK argument -- which stamps '' and fails the dv_r_source FK. SYNTH
# is already registered and is what the rest of this model carries.
SOURCE_TAG = "SYNTH"


def _denum(v):
    """A DISPLAY string like '1,224' back to a real number.

    well_tables formats its rows for the PDF -- the scout ticket and the
    workbook share them -- so depths, pressures and volumes carry thousands
    separators. Written into a workbook they reach a numeric column as text,
    and the loader's TRY_CONVERT turns '1,224' into NULL rather than raising:
    ~15,000 values across eight mirrors would have loaded BLANK and reported
    success. It surfaced only because dv_well_perforation.base_depth is NOT
    NULL, which turned one silent wrong answer into a loud one.

    A value with no comma is returned untouched, so a 14-digit uwi is never
    turned into a float -- an identifier read as a number stops being an
    identifier, and that is the older bug this must not trade itself for.
    """
    if not isinstance(v, str) or "," not in v:
        return v
    try:
        f = float(v.replace(",", "").strip())
    except ValueError:
        return v                      # "Oil, gas to surface" is not a number
    return int(f) if f.is_integer() else f


def _month_start(period):
    """'2015-01' -> '2015-01-01'.

    dv_prod_volume.period_date is nvarchar and keeps the YYYY-MM label, but
    dv_prod_entity.first/last_prod_date are datetime2, where TRY_CONVERT
    returns NULL for a month with no day -- the same silent blank the
    thousands separator produced. An entity whose first and last production
    dates are both NULL cannot be reasoned about at all.
    """
    s = str(period or "").strip()
    if len(s) == 7 and s[4] == "-":
        return s + "-01"
    return s or None


def _unit_id(name):
    """A stratigraphic unit CODE from its display name.

    dv_well_formation_top requires strat_unit_id, and the workbook carried only
    the name. The convention already in the table is the name upper-cased with
    runs of punctuation collapsed to underscores -- "Council Grove" is
    COUNCIL_GROVE -- so the workbook path and the document path agree on the id
    instead of minting two codes for one unit.
    """
    s = _re_nonword.sub("_", str(name or "").upper()).strip("_")
    return s or None


_re_nonword = __import__("re").compile(r"[^A-Z0-9]+")
_DERIVED = {"dv_well_formation_top": {
    "strat_unit_id": lambda row: _unit_id(row.get("strat_unit_name"))}}


def _num(row, idx, key):
    """A float from a display cell, or None. Handles the thousands separator."""
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    try:
        return float(str(row[i]).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _write_tabular(wells, tab_dir, share, seed):
    """One workbook per TARGET TABLE, each holding `share` of the wells.

    Named for the table and keyed by its column names, so the Data Assistant's
    auto-mapping has nothing to guess. Five of these previously collapsed onto
    DV_WELL_STIMULATION because they were named and headed for a human.
    """
    import random
    import zlib
    from openpyxl import Workbook
    from dataview.migration.synth_tables import well_header_row, well_tables

    def _take(name):
        # crc32, not hash(): Python salts string hashing per process, so the
        # 75% draw would pick different wells on every run.
        r = random.Random(zlib.crc32(f"{seed}:{name}".encode("utf-8")))
        return {w["uwi"] for w in wells if r.random() < share}

    def _pid(uwi, prefix):
        return f"{prefix}{uwi}"

    total = 0

    # ---- the well header, its own workbook ----------------------------
    # EVERY WELL, NOT `share` OF THEM. The header workbook is the only thing
    # in tabular/ that CREATES a well; every other workbook hangs off one. Drawn
    # at 75% it named 93 wells while its own detail files referenced 120, so the
    # 27 that arrived only as a LAS header left 1,744 detail rows with no parent
    # -- held on promote, and surfaced as 27 FK "decisions" that were really one
    # load-order fact. The split still applies to the detail mirrors below; it is
    # the PARENT that cannot be partial, because a row cannot be filed under a
    # well that does not exist yet.
    rows = [well_header_row(w) for w in wells]
    if rows:
        wb = Workbook(); ws = wb.active; ws.title = "dv_well"
        ws.append(["uwi", "well_name", "operator_name", "field_name", "county",
                   "province_state", "country", "well_type", "well_status",
                   "well_profile_type", "spud_date", "completion_date",
                   "final_td", "kb_elevation", "ground_elevation",
                   "surface_latitude", "surface_longitude",
                   "bottom_hole_latitude", "bottom_hole_longitude",
                   "source"])
        for r in rows:
            ws.append([_denum(_x) for _x in [r["UWI"], r["WELL_NAME"], r["OPERATOR"], r["FIELD"],
                       r["COUNTY"], r["STATE"], r["COUNTRY"], r["WELL_TYPE"],
                       r["WELL_STATUS"], r["WELL_PROFILE"], r["SPUD_DATE"],
                       r["COMPLETION_DATE"], r["TOTAL_DEPTH"],
                       r["KB_ELEVATION"], r["GROUND_ELEVATION"],
                       r["SURFACE_LATITUDE"], r["SURFACE_LONGITUDE"],
                       r["BOTTOM_HOLE_LATITUDE"], r["BOTTOM_HOLE_LONGITUDE"],
                       SOURCE_TAG]])
        wb.save(os.path.join(tab_dir, "DV_WELL.xlsx"))
        total += len(rows)

    # ---- everything else, one workbook per target table ----------------
    by_section = {}
    for w in wells:
        for name, (cols, rws, _wd) in well_tables(w):
            by_section.setdefault(name, (cols, []))[1].append((w["uwi"], rws))

    # A CHILD MIRROR CANNOT BE DRAWN INDEPENDENTLY OF ITS PARENT. Each section
    # took its own 75% sample, so a well could land in Core Sample without
    # landing in Core Runs -- 57 parent keys across three pairs referenced rows
    # that were never written, and the composite FKs that catch it
    # (fk_dv_well_core_sample_uwi_core_id) are exactly the ones Phase 3 skips,
    # so it passed analysis and failed on promote. Families share one draw.
    core_ids = {}
    if "Core Runs" in by_section:
        _ccols, _cper = by_section["Core Runs"]
        _ci = {c: i for i, c in enumerate(_ccols)}
        for _u, _rws in _cper:
            core_ids[_u] = [(f"CORE_{_u}_{k:02d}",
                             _num(r, _ci, "Top MD"), _num(r, _ci, "Base MD"))
                            for k, r in enumerate(_rws, start=1)]

    def _parent_id(table, uwi, k, r, idx):
        """The FK parent key for ONE row.

        dv_well_core is the exception: core_id is BOTH the FK parent key and
        half its own PK (uwi, core_id), so one id per WELL makes every core
        after the first a duplicate -- 187 rows collapsing to 92 keys, 95 of
        them rejected. Cores get a sequence, and a sample is tied to the core
        whose interval CONTAINS its depth, because that is the only real link
        the data has; picking one arbitrarily would file samples under a core
        that never cut them.
        """
        lst = core_ids.get(uwi) or []
        # A table whose ROW key IS its parent key needs a per-row sequence.
        # dv_well_core and dv_well_dst both key on the id they also hang
        # off, so one id per WELL made every run after the first a duplicate
        # PK -- and an insert-only promote drops those without a word: 94 of
        # 190 DSTs never arrived and the load reported success.
        if table == "dv_well_dst":
            return f"DST_{uwi}_{k:02d}"
        if table == "dv_well_core":
            return lst[k - 1][0] if k - 1 < len(lst) else f"CORE_{uwi}_{k:02d}"
        if table == "dv_well_core_sample":
            if not lst:
                return f"CORE_{uwi}_01"
            d0 = _num(r, idx, "Depth")
            if d0 is not None:
                for cid, top, base in lst:
                    if top is not None and base is not None and top <= d0 <= base:
                        return cid
            return lst[0][0]
        return _pid(uwi, _PARENT_KEY[table][1])

    prod = []
    for name, (cols, per_well) in by_section.items():
        if name == "Production Summary":
            prod = per_well
            continue
        spec = TABULAR_MAP.get(name)
        if not spec:
            continue
        table, pairs, consts = spec
        keep = _take(_DRAW_GROUP.get(name, name))
        idx = {c: i for i, c in enumerate(cols)}
        parent = _PARENT_KEY.get(table)
        rowkey = _ROW_KEY.get(table)
        _mapped = {db for _disp, db in pairs} | set(consts)
        _need_rowkey = bool(rowkey) and rowkey not in _mapped and (
            not parent or rowkey != parent[0])
        hdr = (["uwi"]
               + ([parent[0]] if parent else [])
               + ([rowkey] if _need_rowkey else [])
               + [db for _disp, db in pairs] + list(consts)
               + list(_DERIVED.get(table, {})) + ["source"])
        out = []
        for uwi, rws in per_well:
            if uwi not in keep:
                continue
            for k, r in enumerate(rws, start=1):
                vals = [uwi]
                if parent:
                    vals.append(_parent_id(table, uwi, k, r, idx))
                if _need_rowkey:
                    vals.append(f"{rowkey.upper()[:4]}{k:04d}")
                _row = {}
                for disp, _db in pairs:
                    i = idx.get(disp)
                    _v = _denum(r[i] if i is not None and i < len(r) else None)
                    _row[_db] = _v
                    vals.append(_v)
                _row.update(consts)
                vals += list(consts.values())
                vals += [fn(_row) for fn in _DERIVED.get(table, {}).values()]
                vals.append(SOURCE_TAG)
                out.append(vals)
        if not out:
            continue
        wb = Workbook(); ws = wb.active; ws.title = table[:31]
        ws.append(hdr)
        for v in out:
            ws.append(v)
        wb.save(os.path.join(tab_dir, table.upper() + ".xlsx"))
        total += len(out)

    # ---- the survey header the stations hang off ----------------------
    # dv_well_dir_srvy_sta CARRIES A FOREIGN KEY to dv_well_dir_srvy_hdr
    # (fk_srvy_sta_hdr -- the same constraint that dictated the delete order
    # earlier). A workbook of stations with no header workbook loads into
    # nothing: every station is an orphan the loader cannot place. Derived
    # from the stations themselves so the depths agree.
    sta = by_section.get("Directional Survey")
    if sta:
        cols, per_well = sta
        keep = _take("Directional Survey")
        idx = {c: i for i, c in enumerate(cols)}
        hdr_rows = []
        for uwi, rws in per_well:
            if uwi not in keep or not rws:
                continue
            def _f(r, name):
                i = idx.get(name)
                try:
                    return float(str(r[i]).replace(",", "")) if i is not None else None
                except (ValueError, TypeError):
                    return None
            mds = [x for x in (_f(r, "MD (ft)") for r in rws) if x is not None]
            hdr_rows.append([uwi, _pid(uwi, "SRVY_"), "MWD",
                             min(mds) if mds else None,
                             max(mds) if mds else None, "ft", "KB"])
        if hdr_rows:
            wb = Workbook(); ws = wb.active; ws.title = "dv_well_dir_srvy_hdr"
            ws.append(["uwi", "survey_id", "survey_type", "survey_top_depth",
                       "survey_base_depth", "depth_ouom", "depth_datum",
                       "source"])
            for r in hdr_rows:
                ws.append([_denum(x) for x in r] + [SOURCE_TAG])
            wb.save(os.path.join(tab_dir, "DV_WELL_DIR_SRVY_HDR.xlsx"))
            total += len(hdr_rows)

    # ---- production: entity + volume, and UNPIVOTED --------------------
    # dv_prod_volume is TALL -- one row per fluid per period -- while the
    # report is wide. Loading the wide shape would put gas into a column that
    # means oil, so it is unpivoted here rather than left for a mapping UI
    # that has no way to express it.
    if prod:
        keep = _take("Production Summary")
        ent, vol = [], []
        for uwi, rws in prod:
            if uwi not in keep:
                continue
            per = [r for r in rws]
            if not per:
                continue
            eid = _pid(uwi, "PENT_")
            ent.append([eid, uwi, "WELL",
                        _month_start(per[0][0]), _month_start(per[-1][0]),
                        "OIL"])
            for r in per:
                period = str(r[0])
                for col, fluid, uom in ((1, "OIL", "bbl"), (2, "GAS", "mcf"),
                                        (3, "WATER", "bbl")):
                    raw = str(r[col]).replace(",", "") if col < len(r) else ""
                    try:
                        v = float(raw)
                    except ValueError:
                        continue
                    vol.append([eid, period, fluid, v, uom])
        if ent:
            wb = Workbook(); ws = wb.active; ws.title = "dv_prod_entity"
            ws.append(["prod_entity_id", "uwi", "prod_entity_type",
                       "first_prod_date", "last_prod_date", "primary_fluid",
                       "source"])
            for r in ent:
                ws.append([_denum(x) for x in r] + [SOURCE_TAG])
            wb.save(os.path.join(tab_dir, "DV_PROD_ENTITY.xlsx"))
            total += len(ent)
        if vol:
            wb = Workbook(); ws = wb.active; ws.title = "dv_prod_volume"
            ws.append(["prod_entity_id", "period_date", "fluid_type",
                       "volume", "volume_ouom", "source"])
            for r in vol:
                ws.append([_denum(x) for x in r] + [SOURCE_TAG])
            wb.save(os.path.join(tab_dir, "DV_PROD_VOLUME.xlsx"))
            total += len(vol)
    return total


def _tree_size(p):
    n = b = 0
    for root, _d, fs in os.walk(p):
        for f in fs:
            try:
                b += os.path.getsize(os.path.join(root, f))
                n += 1
            except OSError:
                pass
    return n, b


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Build the whole Teapot Field model into one folder tree.")
    ap.add_argument("--root", default=DEFAULT_ROOT)
    ap.add_argument("--wells", type=int, default=120)
    ap.add_argument("--per-well", type=int, default=2)
    ap.add_argument("--lines", type=int, default=250)
    ap.add_argument("--variants", type=int, default=4)
    ap.add_argument("--seed", type=int, default=90210,
                    help="seed for the field and the tabular draw")
    ap.add_argument("--tabular-share", type=float, default=0.75,
                    help="share of each mirror that arrives as a "
                         "workbook rather than a document")
    ap.add_argument("--skip-seismic", action="store_true",
                    help="skip the 1,000 SEG-Y (324 MB) -- everything else")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args(argv)

    wells_dir = os.path.join(a.root, "wells")
    segy_dir = os.path.join(a.root, "seismic", "2d_segy")

    print("Teapot Field Model")
    print(f"   root      {a.root}")
    print(f"   wells     {a.wells} well(s), {a.per_well} document(s) each")
    print(f"   seismic   {'SKIPPED' if a.skip_seismic else str(a.lines * a.variants) + ' SEG-Y line file(s)'}")
    print(f"   plus      horizons and geography, exported to CSV")
    if not a.apply:
        print("\nPLAN ONLY -- nothing written. Re-run with --apply.")
        return 0

    t0 = time.perf_counter()
    os.makedirs(wells_dir, exist_ok=True)

    print("\n-- wells ------------------------------------------------")
    from dataview.migration import synth_docs
    from dataview.migration.synth_field import (
        Surfaces, plan_field, write_csv)
    S = Surfaces()
    n_expl = max(1, round(a.wells * 0.025))
    n_delin = max(1, round(a.wells * 0.067))
    wells = plan_field(n_expl=n_expl, n_delin=n_delin,
                       n_dev=max(1, a.wells - n_expl - n_delin), surfaces=S)
    write_csv(wells, os.path.join(wells_dir, "teapot_field_wells.csv"))
    n_docs = synth_docs.generate(wells, wells_dir, a.per_well,
                                 log=lambda m: None)
    print(f"   {n_docs:,} document(s) for {len(wells)} well(s)")

    # TWO LOADERS, TWO FOLDERS. .xlsx is in TABULAR_EXTS, so the File
    # Catalog does not crawl it -- spreadsheets belong to the Bulk
    # Tabular Loader. Left beside the .docx, the workbooks are SILENTLY
    # skipped by a scan of wells/ and the Data Assistant is shown .docx
    # it cannot use. Neither tool complains; the rows just never arrive.
    # A folder each makes the boundary visible instead of implicit.
    tab_dir = os.path.join(a.root, "tabular")
    os.makedirs(tab_dir, exist_ok=True)
    moved = 0
    off = os.path.join(wells_dir, "sample_office")
    if os.path.isdir(off):
        import shutil
        for f in sorted(os.listdir(off)):
            if os.path.splitext(f)[1].lower() in TABULAR_EXTS:
                shutil.move(os.path.join(off, f),
                            os.path.join(tab_dir, f))
                moved += 1
    print(f"   {moved:,} workbook(s) -> tabular/ (Data Assistant, not "
          f"the File Catalog)")

    # -- the tabular share -------------------------------------------
    # SPREADSHEET FIRST, DOCUMENT SECOND. Perry's split: the well header and
    # every other mirror arrive 75% from a workbook and 25% from a log or a
    # report. That is how a real data-management job looks -- the same entity
    # turning up from more than one place -- and it is the only thing that
    # exercises BOTH loaders, since .xlsx is in TABULAR_EXTS and the File
    # Catalog will not crawl it.
    #
    # The mechanism is promote's own rule: it is insert-only with NOT EXISTS,
    # so THE FIRST SOURCE IN WINS. Load the workbooks first and 75% of each
    # mirror is theirs; the documents then fill the remaining 25% and add
    # nothing to the rows already there.
    #
    # The 75% is drawn INDEPENDENTLY PER MIRROR, so a well can have its header
    # off a spreadsheet and its tops off a scout ticket. Splitting whole wells
    # instead would produce two clean populations and never test a merge.
    _scratch = os.path.join(tab_dir, "_xl_sheets")
    if os.path.isdir(_scratch):
        import shutil
        shutil.rmtree(_scratch, ignore_errors=True)
        print("   cleared _xl_sheets/ -- the Data Assistant's scratch "
              "from a previous, differently-shaped run")
    _stale = [f for f in os.listdir(tab_dir)
              if f.lower().endswith(".xlsx")
              and f.upper().startswith(("PRODUCTION_", "CORE_ANALYSIS_"))
              and f.upper() not in ("PRODUCTION_SUMMARY.XLSX",)]
    for f in _stale:
        os.remove(os.path.join(tab_dir, f))
    if _stale:
        print(f"   {len(_stale):,} per-well workbook(s) removed -- their "
              f"content is in the consolidated mirror workbooks")
    n_tab = _write_tabular(wells, tab_dir, a.tabular_share, a.seed)
    print(f"   {n_tab:,} workbook row(s) -> tabular/ "
          f"({a.tabular_share:.0%} of each mirror)")


    if not a.skip_seismic:
        print("\n-- seismic ----------------------------------------------")
        os.makedirs(segy_dir, exist_ok=True)
        from tools.make_teapot_2d import main as segy_main
        segy_main(["--apply", "--dir", segy_dir,
                   "--count", str(a.lines * a.variants),
                   "--variants", str(a.variants)])

    print("\n-- horizons and geography -------------------------------")
    _export_horizons(a.root, print)
    _export_geography(a.root, print)

    with open(os.path.join(a.root, "README.md"), "w", encoding="utf-8") as fh:
        fh.write(README.format(
            n_docs=n_docs, n_wells=len(wells),
            n_segy=("0 (skipped)" if a.skip_seismic
                    else f"{a.lines * a.variants:,}")))

    n, b = _tree_size(a.root)
    print(f"\n{n:,} file(s), {b / 1048576:,.0f} MB in {time.perf_counter() - t0:,.0f}s")
    print(f"   {a.root}")
    print("\nNothing of the original NPR-3 data is in this tree. See README.md.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
