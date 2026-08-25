r"""Repoint rows stamped with a sidecar CSV's inventory_id at their WORKBOOK.

An .xlsx is exploded into <dir>/_xl_sheets/<workbook>__<sheet>.csv and loaded
from there, and the promote stamped the SIDECAR's id onto every row. The
sidecar is scratch -- never catalogued, rewritten on every scan -- so those
rows cite a source nothing can resolve, while the workbook sits in
GLOBAL_FILE_CATALOG under a different id. selftest's "no dv_well row cites a
missing catalog entry" is what notices.

The loader no longer does this (page_dir_loader.workbook_for_sheet_csv); this
repairs rows loaded before that fix. inventory_id is SHA1(UPPER(path)), so the
old id is recomputable from the sidecar path the workbook WOULD have produced
-- the file itself is long gone.

    python tools/repair_sidecar_provenance.py --dir <tabular folder>
    python tools/repair_sidecar_provenance.py --dir <tabular folder> --apply
"""
import argparse
import glob
import hashlib
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except Exception:
        pass

SHEET_DIR = "_xl_sheets"


def iid(path):
    """SHA1(UPPER(path), UTF-16-LE) -- file_gate.inventory_id, byte for byte."""
    return hashlib.sha1(str(path).upper().strip()
                        .encode("utf-16-le")).hexdigest().upper()


def sidecar_map(directory):
    """{sidecar_id: (workbook_id, workbook_name)} for every sheet of every
    workbook in `directory`."""
    from openpyxl import load_workbook
    out = {}
    for book in sorted(glob.glob(os.path.join(directory, "*.xlsx"))):
        if os.path.basename(book).startswith("~$"):
            continue
        stem = os.path.splitext(os.path.basename(book))[0]
        try:
            wb = load_workbook(book, read_only=True)
        except Exception as e:
            print(f"   ! {os.path.basename(book)} unreadable: {e}")
            continue
        for sheet in wb.sheetnames:
            side = os.path.join(directory, SHEET_DIR, f"{stem}__{sheet}.csv")
            out[iid(os.path.abspath(side))] = (iid(os.path.abspath(book)),
                                               os.path.basename(book))
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Repoint sidecar-stamped provenance at the workbook. "
                    "Counts only unless --apply.")
    ap.add_argument("--dir", required=True, help="the folder holding the workbooks")
    ap.add_argument("--database", default="DataView_Demo")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args(argv)

    import sqlalchemy as sa
    from dataview.core.dw_utils import make_engine

    m = sidecar_map(a.dir)
    if not m:
        print(f"No workbooks in {a.dir}")
        return 2
    print(f"{len(m)} sheet(s) across "
          f"{len({v[1] for v in m.values()})} workbook(s)\n")

    engine = make_engine(a.database)
    with engine.connect() as cx:
        tables = [r[0] for r in cx.execute(sa.text(
            "SELECT DISTINCT OBJECT_NAME(c.object_id) FROM sys.columns c "
            "WHERE c.name='INVENTORY_ID' "
            "AND OBJECT_SCHEMA_NAME(c.object_id)='dataview' "
            "AND OBJECTPROPERTY(c.object_id,'IsUserTable')=1"
        )).fetchall()]
        known = {r[0] for r in cx.execute(sa.text(
            "SELECT INVENTORY_ID FROM file_catalog.GLOBAL_FILE_CATALOG"
        )).fetchall()}

        plan, total = [], 0
        for t in sorted(tables):
            for old, (new, name) in m.items():
                n = cx.execute(sa.text(
                    f"SELECT COUNT(*) FROM dataview.[{t}] WHERE INVENTORY_ID=:o"),
                    {"o": old}).scalar()
                if n:
                    plan.append((t, old, new, name, n))
                    total += n
        if not plan:
            print("Nothing stamped with a sidecar id. Nothing to repair.")
            return 0
        # REFUSE TO POINT AT SOMETHING ALSO MISSING. Rewriting one dangling id
        # to another dangling id looks like a repair and fixes nothing.
        unreg = {new for _t, _o, new, _n, _c in plan if new not in known}
        for t, old, new, name, n in plan:
            mark = "  <-- TARGET NOT IN CATALOG" if new in unreg else ""
            print(f"   {t:26s} {n:>6,} row(s)  {old[:10]} -> {new[:10]}  {name}{mark}")
        print(f"\n   {total:,} row(s) across {len({p[0] for p in plan})} table(s)")
        if unreg:
            print(f"\nREFUSED: {len(unreg)} workbook(s) are not registered in "
                  f"GLOBAL_FILE_CATALOG, so the repair would swap one unresolvable "
                  f"id for another. Load them first, or register them.")
            return 2

    if not a.apply:
        print("\nCOUNTS ONLY -- nothing written. Re-run with --apply.")
        return 0

    done = 0
    with engine.begin() as cx:
        for t, old, new, _name, _n in plan:
            done += cx.execute(sa.text(
                f"UPDATE dataview.[{t}] SET INVENTORY_ID=:n WHERE INVENTORY_ID=:o"),
                {"n": new, "o": old}).rowcount
    print(f"\n   {done:,} row(s) repointed.")

    with engine.connect() as cx:
        left = cx.execute(sa.text("""
            SELECT COUNT(*) FROM dataview.dv_well w
             WHERE w.inventory_id IS NOT NULL
               AND NOT EXISTS (SELECT 1 FROM file_catalog.GLOBAL_FILE_CATALOG g
                               WHERE g.INVENTORY_ID = w.inventory_id)""")).scalar()
    print(f"   dv_well rows still citing a missing catalog entry: {left}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
